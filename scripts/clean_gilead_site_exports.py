#!/usr/bin/env python3
"""Prepare Gilead site export files for zero-review Migration Cloud import.

Reads the legacy export folder, writes a ``cleaned/`` sibling directory with:

* One canonical file per entity (drops CSV/XLSX duplicates).
* No ``school_stats*.pdf`` (PDF lines mis-classify as custom_fields and flood review).
* ``school_stats*.xlsx`` kept as reference-only (reports domain — skipped at apply).
* Student workbook: ``nan``/``None`` sentinels blanked; ``Religion`` renamed to avoid
  bad embedding-recall mapping to ``admission_number``.
* Teachers export: ``GIVEN PASSWORD`` column removed (never import plaintext passwords).
* Specialties: empty CODE/DEPARTMENT rows filled from NAME.

Usage::

    python scripts/clean_gilead_site_exports.py \\
        --source "C:/path/to/FILES FROM SITE" \\
        [--in-place]

Without ``--in-place``, writes to ``<source>/cleaned/`` and leaves originals untouched.
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook

_NULLISH = frozenset({"", "none", "nan", "n/a", "na", "null", "-"})


def _blank_null(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value != value:
        return ""
    text = str(value).strip()
    if text.lower() in _NULLISH:
        return ""
    return text


def _clean_specialties_rows(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return rows
    header = rows[0]
    out = [header]
    name_i = header.index("NAME") if "NAME" in header else 0
    code_i = header.index("CODE") if "CODE" in header else 1
    dept_i = header.index("DEPARTMENT") if "DEPARTMENT" in header else 2
    for row in rows[1:]:
        cells = list(row) + [""] * (len(header) - len(row))
        name = _blank_null(cells[name_i])
        if not name:
            continue
        if not _blank_null(cells[code_i]):
            base = re.sub(r"[^A-Z0-9]", "", name.upper())[:8] or "SPEC"
            cells[code_i] = base
        if not _blank_null(cells[dept_i]):
            cells[dept_i] = name
        out.append(cells[: len(header)])
    return out


def _clean_student_rows(header: list, rows: list[list]) -> tuple[list, list[list]]:
    header = list(header)
    if "Religion" in header:
        idx = header.index("Religion")
        header[idx] = "Student Religion"
    out_rows: list[list] = []
    for row in rows:
        cells = [_blank_null(v) for v in row]
        if len(cells) < len(header):
            cells.extend([""] * (len(header) - len(cells)))
        out_rows.append(cells[: len(header)])
    return header, out_rows


def _read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    text = path.read_text(encoding="utf-8-sig")
    reader = csv.reader(text.splitlines())
    rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _write_csv(path: Path, header: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(header)
        writer.writerows(rows)


def _write_xlsx(path: Path, header: list, rows: list[list]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _copy_or_clean_teachers(src: Path, dst: Path) -> None:
    header, rows = _read_csv(src)
    drop = {"GIVEN PASSWORD", "given password"}
    keep_idx = [i for i, h in enumerate(header) if h.strip().upper() not in drop]
    new_header = [header[i] for i in keep_idx]
    new_rows = [[row[i] if i < len(row) else "" for i in keep_idx] for row in rows]
    _write_csv(dst, new_header, new_rows)


def _process(source: Path, dest: Path) -> list[str]:
    notes: list[str] = []
    dest.mkdir(parents=True, exist_ok=True)

    # Canonical picks — prefer xlsx where both exist (no UTF-8 BOM issues).
    picks = {
        "subjects": source / "subjects_20260118_224355.xlsx",
        "specialties": source / "specialties_20260118_224259.xlsx",
        "students": source / "student_20260118_224125.xlsx",
        "teachers": source / "teachers_2026-01-18 22_42_32.778115.csv",
        "school_stats": source / "school_stats_2026-01-18 22_47_32.219490.xlsx",
    }
    skipped_dupes = [
        "subjects_2026-01-18 22_43_59.879388.csv",
        "specialties_20260118_224252.csv",
    ]
    skipped_pdf = "school_stats_2026-01-18 22_47_25.679938.pdf"

    for name in skipped_dupes + [skipped_pdf]:
        notes.append(f"skipped (duplicate or reference-only): {name}")

    # Subjects — copy xlsx as canonical name
    sub_src = picks["subjects"]
    sub_dst = dest / "subjects.xlsx"
    shutil.copy2(sub_src, sub_dst)
    notes.append(f"subjects -> {sub_dst.name}")

    # Specialties — clean and write xlsx
    spec_src = picks["specialties"]
    wb = load_workbook(spec_src, read_only=True, data_only=True)
    ws = wb.active
    spec_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if spec_rows:
        cleaned = _clean_specialties_rows(spec_rows)
        _write_xlsx(dest / "specialties.xlsx", cleaned[0], cleaned[1:])
    notes.append("specialties -> specialties.xlsx (empty dept/code rows filled)")

    # Students — clean xlsx
    stu_src = picks["students"]
    wb = load_workbook(stu_src, read_only=True, data_only=True)
    ws = wb.active
    stu_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if stu_rows:
        header, body = _clean_student_rows(stu_rows[0], stu_rows[1:])
        _write_xlsx(dest / "students.xlsx", header, body)
    notes.append("students -> students.xlsx (nan blanked; Religion -> Student Religion)")

    # Teachers — strip password column
    tea_src = picks["teachers"]
    _copy_or_clean_teachers(tea_src, dest / "teachers.csv")
    notes.append("teachers -> teachers.csv (password column removed)")

    # School stats — retain for operators but DO NOT upload (mis-routes as custom_fields
    # when mapping is empty; the PDF is worse). Reference copy lives under reference/.
    stats_src = picks["school_stats"]
    ref_dir = dest / "reference"
    ref_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(stats_src, ref_dir / "school_stats.xlsx")
    notes.append(
        "school_stats.xlsx moved to reference/ (do not upload — aggregate report only)"
    )

    readme = dest / "README.txt"
    readme.write_text(
        "\n".join(
            [
                "Gilead Tech High — cleaned Migration Cloud import bundle",
                "",
                "Upload ONLY the 4 files in this folder (not reference/ or parent duplicates).",
                "",
                "Changes applied:",
                *[f"  - {n}" for n in notes],
                "",
                "After upload, open held review once — autopilot dismisses any residual PDF-noise.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    return notes


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--source",
        type=Path,
        required=True,
        help="Path to FILES FROM SITE folder",
    )
    ap.add_argument(
        "--in-place",
        action="store_true",
        help="Replace source folder contents with cleaned files (backs up to _original/)",
    )
    args = ap.parse_args()
    source = args.source.resolve()
    if not source.is_dir():
        raise SystemExit(f"Not a directory: {source}")

    if args.in_place:
        backup = source / "_original"
        backup.mkdir(exist_ok=True)
        for p in source.iterdir():
            if p.name in {"cleaned", "_original", "README.txt"}:
                continue
            if p.is_file():
                shutil.copy2(p, backup / p.name)
        dest = source
        # Clear old files except backup/readme
        for p in list(dest.iterdir()):
            if p.is_file() and p.name not in {"README.txt"}:
                p.unlink()
    else:
        dest = source / "cleaned"

    notes = _process(source, dest)
    print(f"Wrote cleaned import set to: {dest}")
    for line in notes:
        print(f"  • {line}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
