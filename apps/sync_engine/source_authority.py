"""When two nodes disagree, an external roster is the only thing that can settle it.

WHY THIS EXISTS. Every tie-break the rail owns is a guess about WHICH NODE to believe.
Timestamps cannot do it -- ``updated_at`` is ``auto_now``, so a box's copy of a row is
newer than the cloud's by construction and last-write-wins always elects the box.
Field-level merge helps only while the two sides changed DIFFERENT columns. When the same
column holds two values, no property of the two rows distinguishes them: the engine has
two claims and no evidence.

A roster is evidence. It is the artefact the school actually keeps -- the register the
admission numbers were typed into -- and it predates both nodes, so it is not a third
opinion, it is the thing both opinions were supposed to be copies of.

WHAT IT CANNOT DO, WHICH MATTERS MORE. A roster settles a field only when it CARRIES that
field. A student code taken from an "Admission Number" column is settled. But a roster's
``Name`` is one string -- "ADA DECLAN LOVELACE" -- and ``first_name``/``last_name`` are
DERIVED from it by a splitter. Two nodes that ran different splitter variants hold
different decompositions of a string they agree about, and the roster cannot say which
token is the surname because it never said. Reading a decision out of a file that does not
contain one is exactly the guess this module exists to avoid, so a field the source does
not carry is left PENDING and reported as such.

THE MATCHING PROBLEM IS THE INTERESTING ONE. To ask the roster about a row you must first
find that row's line in it -- and the natural keys are the disputed values themselves. You
cannot key on a student code when the student code is what is in dispute, and you cannot
key on ``last_name`` when the two nodes put different tokens there. ``name_tokens`` exists
for that: the multiset of a name's tokens is INVARIANT under re-splitting, so it matches a
row whose split is the very thing being adjudicated. That is why matchers are a named,
extensible vocabulary rather than string equality -- the caller has to be able to say
which invariant survives the disagreement.

DETERMINISM IS THE POINT. Both nodes run this against the same file and reach the same
values, so they converge without exchanging anything. Nothing here may depend on row
order, dict order, wall clock, or randomness.
"""

from __future__ import annotations

import csv
import hashlib
import io
import pathlib
import re
from typing import Any, Callable, Iterable, Sequence

#: Values a source cell can hold that mean "this cell is empty". Spreadsheets exported
#: through pandas write the STRING "nan" for a missing cell, and a roster typed by hand
#: uses a dash. Treating either as a value would let this module overwrite a real name
#: with the word "nan" -- so they are absences, and an absence never settles anything.
#: A lone "." belongs here for the same reason, and it is not hypothetical: a real
#: 431-line register used it as the mark for "no admission number yet" on two students.
#: Read as a value it is worse than a blank, because two cells holding it collide on a
#: unique column and it looks like data in the audit trail.
_DEFAULT_BLANK_TOKENS = ("", ".", "nan", "none", "null", "n/a", "na", "-", "--", "?")

_WS = re.compile(r"\s+")
_NON_WORD = re.compile(r"[^0-9a-z]+")


def blank_tokens() -> frozenset[str]:
    """The casefolded strings that count as an empty cell.

    Deployments inherit a sensible default and override per-tenant, because which
    sentinel a school's export writes is a property of the school's old vendor.
    """
    from django.conf import settings

    raw = getattr(settings, "RMC_SOURCE_BLANK_TOKENS", None) or _DEFAULT_BLANK_TOKENS
    return frozenset(str(t).strip().casefold() for t in raw)


def is_blank(value: Any) -> bool:
    return str("" if value is None else value).strip().casefold() in blank_tokens()


# -- matchers ------------------------------------------------------------------
#
# A matcher turns one raw value -- from a spreadsheet cell or a model attribute -- into a
# comparable key, or None when the value cannot serve as a key at all. None is NOT a key:
# two rows that both fail to produce one have not been shown to be the same row, and
# indexing them together would silently merge two students.


def _m_exact(value: Any) -> str | None:
    if is_blank(value):
        return None
    return _WS.sub(" ", str(value).strip()).casefold()


def _m_digits(value: Any) -> str | None:
    if is_blank(value):
        return None
    digits = "".join(ch for ch in str(value) if ch.isdigit())
    return digits or None


def _m_alnum(value: Any) -> str | None:
    """Case- and punctuation-insensitive, so ``24GIL0202`` == ``24-gil-0202``."""
    if is_blank(value):
        return None
    key = _NON_WORD.sub("", str(value).casefold())
    return key or None


def _m_name_tokens(value: Any) -> tuple[str, ...] | None:
    """The token multiset of a name -- the one key a split disagreement cannot move.

    ``first_last`` and ``last_first`` disagree about which token is the surname, but they
    partition the SAME tokens; sorting them yields a key both nodes compute identically.
    Sorted rather than a set: a name with a repeated token must not collide with one that
    carries it once.
    """
    if is_blank(value):
        return None
    tokens = [t for t in (_NON_WORD.sub("", p) for p in _WS.split(str(value).casefold())) if t]
    return tuple(sorted(tokens)) or None


def _m_date(value: Any) -> str | None:
    """ISO date. A roster writes ``2012-11-16 00:00:00``; a model holds a ``date``."""
    if is_blank(value):
        return None
    import datetime as _dt

    if isinstance(value, _dt.datetime):
        return value.date().isoformat()
    if isinstance(value, _dt.date):
        return value.isoformat()
    text = str(value).strip()
    from django.utils.dateparse import parse_date, parse_datetime

    parsed = parse_datetime(text)
    if parsed is not None:
        return parsed.date().isoformat()
    parsed_date = parse_date(text[:10])
    return parsed_date.isoformat() if parsed_date is not None else None


MATCHERS: dict[str, Callable[[Any], Any]] = {
    "exact": _m_exact,
    "alnum": _m_alnum,
    "digits": _m_digits,
    "name_tokens": _m_name_tokens,
    "date": _m_date,
}


def matcher(name: str) -> Callable[[Any], Any]:
    try:
        return MATCHERS[str(name).strip().lower()]
    except KeyError:
        raise ValueError(
            "unknown matcher %r; known: %s" % (name, ", ".join(sorted(MATCHERS)))
        ) from None


# -- reading a roster ----------------------------------------------------------


def normalise_header(name: Any) -> str:
    """Headers are compared loosely because humans retype them.

    ``Admission Number``, ``admission number`` and ``ADMISSION_NUMBER`` are one column.
    """
    return _NON_WORD.sub("_", str("" if name is None else name).strip().casefold()).strip("_")


def load_source_rows(path: str | pathlib.Path, sheet: str | int | None = None) -> list[dict]:
    """Rows of a .xlsx or .csv as ``{normalised_header: text}``, in file order.

    Order is preserved but never depended on: it is kept so an operator reading the report
    can find the line, not so that anything here can decide by it.
    """
    p = pathlib.Path(path)
    if not p.is_file():
        raise ValueError("source file not found: %s" % p)
    suffix = p.suffix.lower()
    if suffix in (".csv", ".txt", ".tsv"):
        return _rows_from_csv(p)
    if suffix in (".xlsx", ".xlsm"):
        return _rows_from_xlsx(p, sheet)
    raise ValueError("unsupported source format %r (want .xlsx or .csv)" % suffix)


def _rows_from_csv(p: pathlib.Path) -> list[dict]:
    # utf-8-sig: every export sampled from this pipeline carries a BOM, and a BOM welded
    # to the first header turns "NAME" into a column no mapping can name.
    text = p.read_bytes().decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header = [normalise_header(c) for c in next(reader)]
    except StopIteration:
        return []
    out = []
    for raw in reader:
        if not any(str(c).strip() for c in raw):
            continue
        out.append({header[i]: str(raw[i]).strip() for i in range(min(len(header), len(raw)))})
    return out


def _rows_from_xlsx(p: pathlib.Path, sheet: str | int | None) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        if sheet is None or sheet == "":
            ws = wb.worksheets[0]
        elif isinstance(sheet, int) or str(sheet).isdigit():
            ws = wb.worksheets[int(sheet)]
        else:
            ws = wb[str(sheet)]
        it = ws.iter_rows(values_only=True)
        try:
            header = [normalise_header(c) for c in next(it)]
        except StopIteration:
            return []
        out = []
        for raw in it:
            if raw is None or all(c is None or str(c).strip() == "" for c in raw):
                continue
            out.append(
                {
                    header[i]: ("" if raw[i] is None else str(raw[i]).strip())
                    for i in range(min(len(header), len(raw)))
                }
            )
        return out
    finally:
        wb.close()


def source_fingerprint(path: str | pathlib.Path) -> str:
    """First 12 hex of the file's sha256 -- provenance, so a resolution names its evidence.

    Short on purpose: it rides in ``SyncConflict.resolution_note``, a CharField(255) that
    also has to hold the field names. Twelve hex characters distinguish the rosters a
    school will ever hold without crowding out the part a human reads.
    """
    digest = hashlib.sha256(pathlib.Path(path).read_bytes()).hexdigest()
    return digest[:12]


# -- indexing ------------------------------------------------------------------


#: A clause matches by equality, or by the live row's tokens being CONTAINED in the
#: source's. Containment is not a convenience -- it is forced by the data. StudentProfile
#: has ``first_name`` and ``last_name`` and no ``middle_name``, so a three-token roster
#: name cannot survive the trip: ``first_last`` keeps tokens 0 and -1, ``last_first`` keeps
#: tokens 1 and 0, and each DISCARDS a token the other kept. The two nodes therefore do not
#: hold the same tokens as each other, and neither holds the same tokens as the roster.
#: Equality would join nothing. What is still true is that each node's tokens are a subset
#: of the roster's, which is what SUBSET tests -- and it holds under either variant, so the
#: join survives the disagreement it exists to adjudicate.
EQUAL = "equal"
SUBSET = "subset"
MODES = (EQUAL, SUBSET)


class KeySpec:
    """One clause of the match key: source column, model field(s), invariant, and mode.

    ``field`` may name several model attributes joined by ``+``. That is not sugar: the
    roster's one ``Name`` column corresponds to two columns on the row, and a clause that
    could only read one attribute could not express the correspondence at all.
    """

    __slots__ = ("column", "fields", "matcher_name", "mode", "_fn")

    def __init__(
        self, column: str, field: str, matcher_name: str = "exact", mode: str = EQUAL
    ):
        self.column = normalise_header(column)
        self.fields = tuple(f.strip() for f in str(field).split("+") if f.strip())
        if not self.fields:
            raise ValueError("match clause for column %r names no model field" % column)
        self.matcher_name = str(matcher_name).strip().lower()
        self.mode = str(mode).strip().lower() or EQUAL
        if self.mode not in MODES:
            raise ValueError(
                "unknown mode %r; known: %s" % (self.mode, ", ".join(MODES))
            )
        self._fn = matcher(self.matcher_name)
        if self.mode == SUBSET and self.matcher_name != "name_tokens":
            raise ValueError(
                "mode 'subset' needs a matcher that yields tokens; got %r"
                % self.matcher_name
            )

    @property
    def field(self) -> str:
        return "+".join(self.fields)

    def of_source(self, row: dict) -> Any:
        return self._fn(row.get(self.column))

    def of_instance(self, instance: Any) -> Any:
        parts = [
            str(getattr(instance, f, "") or "").strip()
            for f in self.fields
        ]
        return self._fn(" ".join(p for p in parts if p))

    def __repr__(self) -> str:  # pragma: no cover - diagnostics only
        return "KeySpec(%s->%s:%s:%s)" % (
            self.column, self.field, self.matcher_name, self.mode
        )


def source_key(row: dict, specs: Sequence[KeySpec]) -> tuple | None:
    """The composite EQUAL key of a source row, or None when any clause is unusable.

    All-or-nothing on purpose. A partial key would match on the clauses that happened to
    be present and ignore the one clause that would have told two students apart.
    """
    parts = []
    for spec in specs:
        value = spec.of_source(row)
        if value is None:
            return None
        parts.append(value)
    return tuple(parts)


def instance_key(instance: Any, specs: Sequence[KeySpec]) -> tuple | None:
    parts = []
    for spec in specs:
        value = spec.of_instance(instance)
        if value is None:
            return None
        parts.append(value)
    return tuple(parts)


class SourceIndex:
    """Looks a live row up in the roster and returns EVERY line it could be.

    A list, never a winner. Collapsing duplicates would pick by file order and call it a
    fact; a key that reaches two lines has not identified a student, and the caller has to
    be able to see that instead of being handed the first one.
    """

    def __init__(self, rows: Iterable[dict], specs: Sequence[KeySpec]):
        self.rows = list(rows)
        self.specs = tuple(specs)
        if not self.specs:
            raise ValueError("a match needs at least one clause")
        self._equal = tuple(s for s in self.specs if s.mode == EQUAL)
        self._subset = tuple(s for s in self.specs if s.mode == SUBSET)

        # Equal clauses index to a composite key; subset clauses index each token to the
        # rows carrying it, so a lookup intersects postings instead of scanning the file.
        self._by_key: dict[tuple, list[int]] = {}
        for i, row in enumerate(self.rows):
            key = source_key(row, self._equal) if self._equal else ()
            if key is not None:
                self._by_key.setdefault(key, []).append(i)

        self._postings: list[dict[str, set[int]]] = []
        for spec in self._subset:
            postings: dict[str, set[int]] = {}
            for i, row in enumerate(self.rows):
                for tok in frozenset(spec.of_source(row) or ()):
                    postings.setdefault(tok, set()).add(i)
            self._postings.append(postings)

    def lookup(self, instance: Any) -> list[dict] | None:
        """Matching rows, or None when the row cannot produce a key at all.

        None is not "no match". A row that cannot be keyed was never asked about, and
        reporting that as "the roster does not mention this student" would invite an
        operator to conclude something the run never established.
        """
        if self._equal:
            key = instance_key(instance, self._equal)
            if key is None:
                return None
            candidates = set(self._by_key.get(key, ()))
        else:
            candidates = set(range(len(self.rows)))

        for pos, spec in enumerate(self._subset):
            want = frozenset(spec.of_instance(instance) or ())
            if not want:
                return None
            # INTERSECTING the postings of every wanted token IS the containment test: a
            # row survives only by carrying all of them, which is exactly want <= row. A
            # union here would return every line sharing ANY token -- one shared given
            # name would be enough -- so the operator would be handed another student's
            # admission number as though the roster had said so.
            postings = self._postings[pos]
            narrowed: set[int] | None = None
            for tok in want:
                hits = postings.get(tok, set())
                narrowed = set(hits) if narrowed is None else (narrowed & hits)
                if not narrowed:
                    break
            candidates &= narrowed or set()
            if not candidates:
                break

        return [self.rows[i] for i in sorted(candidates)]


def build_index(rows: Iterable[dict], specs: Sequence[KeySpec]) -> dict[tuple, list[dict]]:
    """``{key: [rows]}`` for equality-only matching. Kept for callers that never narrow."""
    index: dict[tuple, list[dict]] = {}
    for row in rows:
        key = source_key(row, specs)
        if key is None:
            continue
        index.setdefault(key, []).append(row)
    return index
