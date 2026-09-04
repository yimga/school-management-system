"""End-to-end: Mama Novi 3-file bundle (student + subjects + specialties) lands clean."""

from __future__ import annotations

import io
import types
from pathlib import Path

from django.test import TestCase

from apps.migration_cloud import artifact_blob_store as store
from apps.migration_cloud.models import (
    ArtifactFormat,
    BundleStatus,
    IntakeMethod,
    MigrationArtifact,
    MigrationBundle,
)
from apps.migration_cloud.orchestrator import apply_bundle
from apps.migration_cloud.pipeline import advance_bundle
from apps.schools.models import School

MAMA_NOVI = Path(
    r"c:/Users/yimga/Documents/HY_DOC_MAINPC/Docs for Others_Friends_family/Gilead Tech High/Mama Novi"
)


def _payload(data: bytes):
    return types.SimpleNamespace(content_opener=lambda: io.BytesIO(data))


def _add_xlsx(bundle, filename: str) -> None:
    from openpyxl import Workbook, load_workbook

    src = MAMA_NOVI / filename
    if not src.is_file():
        return
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = Workbook()
    out_ws = out.active
    for row in rows:
        out_ws.append(list(row))
    buf = io.BytesIO()
    out.save(buf)
    raw = buf.getvalue()
    art = MigrationArtifact.objects.create(
        bundle=bundle,
        path_within_bundle=filename,
        filename=filename,
        detected_format=ArtifactFormat.XLSX,
        byte_size=len(raw),
        sha256="b" * 64,
    )
    store.capture_artifact_blob(art, _payload(raw))


class MamaNoviThreeFileImportTests(TestCase):
    def test_all_three_files_classify_and_land_without_mass_quarantine(self):
        if not (MAMA_NOVI / "student_2026.xlsx").is_file():
            self.skipTest("Mama Novi fixture files not present on this machine")

        from apps.academics.models import Subject, Specialty
        from apps.people.models import StudentProfile

        school = School.objects.create(
            name="Gilead Mama Novi",
            subdomain="gilead-mama-novi",
            country_code="CM",
            settings={"school_type": "tvet", "migration_gap_fill_provisioning": True},
        )
        bundle = MigrationBundle.objects.create(
            label="mama-novi-e2e",
            intake_method=IntakeMethod.FILE_UPLOAD,
            idempotency_key="mama-novi-e2e",
            status=BundleStatus.INGESTING,
            school=school,
        )
        for fn in ("specialties_2026.xlsx", "subjects_2026.xlsx", "student_2026.xlsx"):
            _add_xlsx(bundle, fn)

        advance_bundle(bundle_id=bundle.pk, use_accelerator=True)
        bundle.refresh_from_db()
        per_art = (bundle.discovery_summary or {}).get("per_artifact_domain") or {}
        domains = {
            path.rsplit("/", 1)[-1]: meta.get("domain")
            for path, meta in per_art.items()
        }
        self.assertEqual(domains.get("student_2026.xlsx"), "students", per_art)
        self.assertEqual(domains.get("subjects_2026.xlsx"), "academics", per_art)
        self.assertEqual(domains.get("specialties_2026.xlsx"), "specialties", per_art)

        apply_bundle(bundle_id=bundle.pk, workers=1)
        bundle.refresh_from_db()
        totals = (bundle.mapping_summary or {}).get("apply_totals") or {}
        quarantined = int(totals.get("quarantined") or 0)
        created = int(totals.get("created") or 0)
        updated = int(totals.get("updated") or 0)

        student_count = StudentProfile.objects.filter(school=school).count()
        subject_count = Subject.objects.filter(school=school).count()
        specialty_count = Specialty.objects.filter(school=school).count()

        self.assertGreaterEqual(student_count, 400, msg=f"totals={totals} domains={domains}")
        self.assertGreaterEqual(subject_count, 100, msg=f"totals={totals}")
        self.assertGreaterEqual(specialty_count, 8, msg=f"totals={totals}")
        self.assertLess(
            quarantined,
            20,
            msg=f"expected near-zero quarantine for 3-file TVET bundle, got {quarantined}; "
            f"created={created} updated={updated} totals={totals}",
        )
