"""Live repair simulation proxy for production bundle #84 (Mama Novi 3-file set).

Bundle pk=84 is not present in local dev SQLite; this test uses the same
artifacts and classification path verified against verbatim bundle-84
production candidate lists in ``test_filename_led_fallback_real_bundle_84``.
"""

from __future__ import annotations

import io
import types
from pathlib import Path

from django.test import TransactionTestCase

from apps.migration_cloud import artifact_blob_store as store
from apps.migration_cloud.classifiers.domain import DomainCandidate, _filename_led_fallback, classify_domain
from apps.migration_cloud.models import (
    ArtifactFormat,
    BundleStatus,
    IntakeMethod,
    MigrationArtifact,
    MigrationBundle,
)
from apps.migration_cloud.orchestrator import apply_bundle
from apps.migration_cloud.pipeline import advance_bundle, refresh_bundle_inference
from apps.migration_cloud.quarantine_resolution import (
    pending_quarantine_count,
    quarantine_breakdown,
)
from apps.migration_cloud.repair import repair_readiness, unresolved_issue_count
from apps.schools.models import School

MAMA_NOVI = Path(
    r"c:/Users/yimga/Documents/HY_DOC_MAINPC/Docs for Others_Friends_family/Gilead Tech High/Mama Novi"
)


def _payload(data: bytes):
    return types.SimpleNamespace(content_opener=lambda: io.BytesIO(data))


def _add_xlsx(bundle, filename: str) -> MigrationArtifact:
    from openpyxl import Workbook, load_workbook

    src = MAMA_NOVI / filename
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = Workbook()
    out_ws = out.active
    for row in rows:
        out_ws.append(list(row))
    buf = io.BytesIO()
    out.save(buf)
    raw = buf.getvalue()
    art = MigrationArtifact.objects.create(
        bundle=bundle,
        path_within_bundle=filename,
        filename=filename,
        detected_format=ArtifactFormat.XLSX,
        byte_size=len(raw),
        sha256="d" * 64,
    )
    store.capture_artifact_blob(art, _payload(raw))
    return art


class Bundle84RepairSimulationTests(TransactionTestCase):
    def test_repair_simulation_quarantine_breakdown(self):
        if not (MAMA_NOVI / "student_2026.xlsx").is_file():
            self.skipTest("Mama Novi fixture files not present")

        from apps.academics.models import Subject, Specialty
        from apps.people.models import StudentProfile

        school = School.objects.create(
            name="Gilead Bundle 84 Sim",
            subdomain="gilead-bundle-84-sim",
            country_code="CM",
            settings={"school_type": "tvet", "migration_gap_fill_provisioning": True},
        )
        bundle = MigrationBundle.objects.create(
            label="bundle-84-repair-sim",
            intake_method=IntakeMethod.FILE_UPLOAD,
            idempotency_key="bundle-84-repair-sim",
            status=BundleStatus.INGESTING,
            school=school,
        )
        for fn in ("specialties_2026.xlsx", "subjects_2026.xlsx", "student_2026.xlsx"):
            _add_xlsx(bundle, fn)

        advance_bundle(bundle_id=bundle.pk, use_accelerator=True)
        bundle.refresh_from_db()
        per_before = (bundle.discovery_summary or {}).get("per_artifact_domain") or {}

        subjects_art = bundle.artifacts.get(filename="subjects_2026.xlsx")
        raw_cls = classify_domain(artifact=subjects_art)
        ranked = [
            DomainCandidate(
                c["domain"],
                c["confidence"],
                c.get("matched_canonical_fields") or [],
                "",
            )
            for c in (raw_cls.get("candidates") or [])[:5]
        ]
        fixed_domain = _filename_led_fallback("subjects_2026.xlsx", ranked)

        refresh_bundle_inference(bundle_id=bundle.pk, use_accelerator=True)
        bundle.refresh_from_db()
        per_after = (bundle.discovery_summary or {}).get("per_artifact_domain") or {}

        def _domains(per_art: dict) -> dict[str, str | None]:
            return {
                path.rsplit("/", 1)[-1]: (meta or {}).get("domain")
                for path, meta in per_art.items()
            }

        domains_before = _domains(per_before)
        domains_after = _domains(per_after)

        readiness = repair_readiness(bundle)
        apply_bundle(bundle_id=bundle.pk, workers=1)
        bundle.refresh_from_db()

        totals = (bundle.mapping_summary or {}).get("apply_totals") or {}
        breakdown = quarantine_breakdown(bundle, pending_only=True)
        pending = pending_quarantine_count(bundle)

        # Production bundle #84 had subjects -> behavior; fix must route to academics.
        self.assertEqual(fixed_domain, "academics")
        self.assertEqual(domains_after.get("subjects_2026.xlsx"), "academics")
        self.assertEqual(domains_after.get("student_2026.xlsx"), "students")
        self.assertEqual(domains_after.get("specialties_2026.xlsx"), "specialties")

        student_count = StudentProfile.objects.filter(school=school).count()
        subject_count = Subject.objects.filter(school=school).count()
        specialty_count = Specialty.objects.filter(school=school).count()

        self.assertGreaterEqual(student_count, 400)
        self.assertGreaterEqual(subject_count, 100)
        self.assertGreaterEqual(specialty_count, 8)
        self.assertLess(int(totals.get("quarantined") or 0), 20)
        self.assertLess(unresolved_issue_count(bundle), 20)

        # Emit a human-readable report (visible with --verbosity=2).
        lines = [
            "=== Bundle #84 repair simulation (Mama Novi proxy) ===",
            f"OLD subjects filename fallback would pick: {fixed_domain} (was behavior in prod)",
            f"Domains after advance: {domains_before}",
            f"Domains after refresh_bundle_inference: {domains_after}",
            f"repair_readiness: {readiness.repairable}",
            f"After apply — status={bundle.status} totals={totals}",
            f"pending_quarantine={pending} unresolved={unresolved_issue_count(bundle)}",
            f"landed: students={student_count} subjects={subject_count} specialties={specialty_count}",
            "Quarantine breakdown (pending):",
        ]
        if breakdown:
            for row in breakdown:
                lines.append(f"  {row['issue_class']}: {row['count']} — {row['label']}")
        else:
            lines.append("  (none)")
        print("\n".join(lines))
