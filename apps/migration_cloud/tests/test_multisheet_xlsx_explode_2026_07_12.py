"""Multi-sheet XLSX explode: every tab lands, not just the first (data-loss fix).

A school's SIS export is one workbook with several tabs. Every reader in the
pipeline reads only sheetnames[0], so tabs 2+ were silently dropped. The
FILE_UPLOAD adapter now explodes a multi-tab workbook into one TSV artifact per
non-empty sheet; a single-sheet workbook is untouched.
"""

from __future__ import annotations

import csv
import io
import tempfile
from pathlib import Path

from django.test import SimpleTestCase


def _write_workbook(sheets: dict[str, list[list]]) -> Path:
    """Write a temp .xlsx with the given {sheet_name: [rows...]} and return path."""
    from openpyxl import Workbook

    wb = Workbook()
    # Workbook starts with one default sheet; rename/fill it for the first entry.
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = name
        for r in rows:
            ws.append(r)
        first = False
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    import os

    os.close(fd)
    wb.save(tmp)
    return Path(tmp)


def _parse_tsv(data: bytes) -> tuple[list[str], list[list[str]]]:
    text = data.decode("utf-8")
    reader = list(csv.reader(io.StringIO(text), delimiter="\t"))
    if not reader:
        return [], []
    return reader[0], reader[1:]


class ExplodeWorkbookTests(SimpleTestCase):
    def test_explodes_every_nonempty_sheet(self):
        from apps.migration_cloud.xlsx_explode import explode_workbook

        path = _write_workbook(
            {
                "Students": [["admission_no", "first_name"], ["S1", "Ada"], ["S2", "Grace"]],
                "Staff": [["staff_no", "first_name"], ["T1", "Alan"]],
            }
        )
        try:
            result = explode_workbook(path)
        finally:
            path.unlink(missing_ok=True)

        names = [name for name, _ in result]
        self.assertEqual(names, ["Students", "Staff"])
        by_name = {name: _parse_tsv(tsv) for name, tsv in result}
        self.assertEqual(by_name["Students"][0], ["admission_no", "first_name"])
        self.assertEqual(len(by_name["Students"][1]), 2)  # two student rows
        self.assertEqual(by_name["Staff"][1], [["T1", "Alan"]])

    def test_skips_empty_sheets(self):
        from apps.migration_cloud.xlsx_explode import explode_workbook

        path = _write_workbook(
            {
                "Data": [["a", "b"], ["1", "2"]],
                "Blank": [],
            }
        )
        try:
            result = explode_workbook(path)
        finally:
            path.unlink(missing_ok=True)
        self.assertEqual([name for name, _ in result], ["Data"])


class FileIntakeMultiSheetTests(SimpleTestCase):
    def _ctx(self):
        from apps.migration_cloud.intake.base import IntakeContext

        return IntakeContext(bundle_id=0, idempotency_key="ut-xlsx")

    def _adapter(self):
        from apps.migration_cloud.intake import get_adapter
        from apps.migration_cloud.models import IntakeMethod

        return get_adapter(IntakeMethod.FILE_UPLOAD)

    def test_multisheet_upload_yields_one_artifact_per_sheet(self):
        adapter = self._adapter()
        path = _write_workbook(
            {
                "Students": [["admission_no", "first_name"], ["S1", "Ada"]],
                "Teachers": [["staff_no", "first_name"], ["T1", "Alan"]],
                "Classes": [["class_code", "name"], ["JSS1", "Junior 1"]],
            }
        )
        try:
            payloads = list(adapter.iter_artifacts(path, self._ctx()))
        finally:
            path.unlink(missing_ok=True)

        # Three tabs -> three artifacts (was: one, the first sheet only).
        self.assertEqual(len(payloads), 3)
        names = sorted(p.path_within_bundle for p in payloads)
        self.assertTrue(all(n.endswith(".tsv") for n in names))
        self.assertTrue(any("Students" in n for n in names))
        self.assertTrue(any("Teachers" in n for n in names))
        self.assertTrue(any("Classes" in n for n in names))
        # Distinct names -> the (bundle, sha256, path) uniqueness constraint holds.
        self.assertEqual(len(set(names)), 3)
        # Each payload's bytes are the readable TSV for that sheet.
        for p in payloads:
            header, rows = _parse_tsv(p.content_opener().read())
            self.assertTrue(header)
            self.assertTrue(rows)

    def test_single_sheet_upload_unchanged(self):
        adapter = self._adapter()
        path = _write_workbook(
            {"OnlyTab": [["admission_no", "first_name"], ["S1", "Ada"]]}
        )
        try:
            payloads = list(adapter.iter_artifacts(path, self._ctx()))
        finally:
            path.unlink(missing_ok=True)

        # A single-sheet workbook stays one raw .xlsx artifact (Wave-2 path).
        self.assertEqual(len(payloads), 1)
        self.assertTrue(payloads[0].path_within_bundle.endswith(".xlsx"))
