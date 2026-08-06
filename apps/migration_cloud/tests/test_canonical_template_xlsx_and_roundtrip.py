"""The canonical templates must round-trip and carry per-column guidance.

Two gaps this locks:

1. **Round-trip.** The template ships a leading ``# runmycampus-canonical-template``
   comment line (and now a ``#``-commented example row). The intake profiler read
   row 0 as the header row unconditionally, so a school that downloaded a
   template, filled it in and re-uploaded it had the comment parsed as its only
   "column" and the real headers demoted to data. The profiler must skip
   ``#``-comment and blank lines before picking headers.

2. **Guidance.** Downloads were headers-only CSV. A school filling one in had to
   guess each column. The XLSX now carries a Data sheet (headers only, so it
   uploads clean) plus an Instructions sheet describing every column with an
   example — enriched from the canonical ontology where a field aligns.
"""
from __future__ import annotations

import io

from django.test import SimpleTestCase
from django.urls import reverse

from apps.migration_cloud.accelerators.runmycampus_canonical import (
    DOMAIN_CANONICAL_HEADERS,
)
from apps.migration_cloud.profiler import _read_csv_like
from apps.migration_cloud.views import (
    _canonical_template_csv,
    _canonical_template_workbook_all,
    _canonical_template_xlsx,
)


class ProfilerSkipsCommentLinesTests(SimpleTestCase):
    def test_filled_in_template_round_trips_to_the_real_headers(self):
        """A downloaded-then-filled template must profile to its real headers."""
        headers = DOMAIN_CANONICAL_HEADERS["students"]
        template = _canonical_template_csv("students", headers)
        # Simulate a school filling in one data row under the template headers.
        sorted_headers = sorted(headers)
        filled = template + ",".join(["x"] * len(sorted_headers)) + "\n"
        rows, detected_headers, _enc = _read_csv_like(
            io.BytesIO(filled.encode("utf-8")), "utf-8"
        )
        self.assertEqual(
            detected_headers,
            sorted_headers,
            "the leading '#' comment line was parsed as the header row",
        )
        self.assertNotIn("# runmycampus-canonical-template", detected_headers[0])
        # The real filled data row is present; the '#'-comment example is skipped.
        self.assertEqual(len(rows), 1)

    def test_blank_leading_lines_are_skipped(self):
        text = "\n\n# a comment\nexternal_id,first_name\n1,Ada\n"
        rows, headers, _enc = _read_csv_like(io.BytesIO(text.encode("utf-8")), "utf-8")
        self.assertEqual(headers, ["external_id", "first_name"])
        self.assertEqual(rows, [["1", "Ada"]])


class CanonicalTemplateCsvContractTests(SimpleTestCase):
    def test_csv_is_marker_line_plus_sorted_headers_only(self):
        text = _canonical_template_csv("students", DOMAIN_CANONICAL_HEADERS["students"])
        lines = text.splitlines()
        # Contract: comment marker line 0, headers line 1 (sorted), nothing else.
        self.assertTrue(lines[0].startswith("# runmycampus-canonical-template:"))
        self.assertEqual(lines[1].split(","), sorted(lines[1].split(",")))
        # No data region at all — and crucially no post-header '#' line (which the
        # apply reader would otherwise land as a bogus record).
        self.assertEqual(len(lines), 2)


class OrchestratorRoundTripTests(SimpleTestCase):
    """The apply reader — not just the profiler — must skip the marker line."""

    def test_apply_reader_skips_marker_and_reads_the_real_headers(self):
        import csv

        from apps.migration_cloud.orchestrator import _strip_leading_comment_lines

        headers = DOMAIN_CANONICAL_HEADERS["students"]
        sorted_headers = sorted(headers)
        # A school downloads the template (marker + headers) and fills in a row.
        filled = _canonical_template_csv("students", headers) + ",".join(
            f"v_{h}" for h in sorted_headers
        ) + "\n"
        reader = csv.DictReader(_strip_leading_comment_lines(io.StringIO(filled)))
        self.assertEqual(reader.fieldnames, sorted_headers)
        rows = list(reader)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][sorted_headers[0]], f"v_{sorted_headers[0]}")

    def test_data_value_starting_with_hash_is_not_dropped(self):
        import csv

        from apps.migration_cloud.orchestrator import _strip_leading_comment_lines

        text = "# runmycampus-canonical-template: domain=x version=1.0\nid,name\n#4501,Ada\n7,Zoe\n"
        rows = list(csv.DictReader(_strip_leading_comment_lines(io.StringIO(text))))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["id"], "#4501")  # a '#' data value survives


class CanonicalTemplateXlsxTests(SimpleTestCase):
    def test_single_domain_xlsx_has_data_and_instructions_sheets(self):
        from openpyxl import load_workbook

        data = _canonical_template_xlsx("students", DOMAIN_CANONICAL_HEADERS["students"])
        wb = load_workbook(io.BytesIO(data))
        self.assertEqual(wb.sheetnames, ["Data", "Instructions"])
        # Data sheet is headers-only (row 1 == sorted headers, no data rows).
        data_ws = wb["Data"]
        first_row = [c.value for c in next(data_ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(first_row, sorted(DOMAIN_CANONICAL_HEADERS["students"]))
        self.assertEqual(data_ws.max_row, 1)
        # Instructions sheet has the guidance columns + one row per header.
        info_ws = wb["Instructions"]
        header_row = [c.value for c in next(info_ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(header_row, ["Column", "Description", "Required?", "Example"])
        self.assertEqual(
            info_ws.max_row, len(DOMAIN_CANONICAL_HEADERS["students"]) + 1
        )
        # A required field is flagged.
        required_flags = {
            row[0].value: row[2].value for row in info_ws.iter_rows(min_row=2)
        }
        self.assertEqual(required_flags.get("first_name"), "Yes")

    def test_every_domain_xlsx_generates_without_error(self):
        from openpyxl import load_workbook

        data = _canonical_template_workbook_all()
        wb = load_workbook(io.BytesIO(data))
        self.assertIn("Instructions", wb.sheetnames)
        # One data sheet per domain (plus the Instructions overview).
        self.assertGreaterEqual(len(wb.sheetnames), len(DOMAIN_CANONICAL_HEADERS))

    def test_previously_blank_domain_now_has_examples(self):
        """The 8 header-only domains (e.g. athletics) had blank Example cells."""
        from openpyxl import load_workbook

        data = _canonical_template_xlsx(
            "athletics_teams", DOMAIN_CANONICAL_HEADERS["athletics_teams"]
        )
        info = load_workbook(io.BytesIO(data))["Instructions"]
        examples = {row[0].value: row[3].value for row in info.iter_rows(min_row=2)}
        self.assertTrue(examples.get("sport"), "athletics_teams example still blank")
        self.assertTrue(examples.get("team_name"))


class CanonicalTemplateXlsxRouteTests(SimpleTestCase):
    def test_xlsx_route_resolves_in_both_namespaces(self):
        for ns in ("migration_cloud_super", "migration_cloud_portal"):
            with self.subTest(ns=ns):
                url = reverse(f"{ns}:canonical_template_xlsx", kwargs={"domain": "students"})
                self.assertTrue(url.endswith("students.xlsx"))
