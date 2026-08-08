"""Migration Cloud wizard views (operator + tenant surfaces).

Routes a single bundle through its lifecycle states with one URL per
state transition. All views are class-based, login-required, and respect
both operator (super) and tenant-admin (portal) entry points via the
``shell`` URL kwarg.

This is the Phase U6 scaffold — the surface that Phase U6 finishes with
drag-and-drop, side-by-side preview, and Apple-tier grammar. The
endpoints + routing + state-machine wiring all land here so the UI team
can iterate templates without touching views.

URL grammar (one router, two mount points):

    /super/migration/                      — operator console (list)
    /super/migration/<bundle_id>/          — detail / mapping wizard
    /super/migration/<bundle_id>/apply/    — POST → apply step (dry-run by default)
    /super/migration/<bundle_id>/reconcile/— POST → generate reconciliation report
    /super/migration/<bundle_id>/feedback/ — POST → operator override → record_feedback

    /portal/configure/migration/           — tenant mirror (plan-gated)
    ...same suffixes...
"""

from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from typing import Any

from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.utils.decorators import method_decorator
from django.views import View

from apps.platform_runtime.workflow_tracker import track_workflow
from apps.migration_cloud.schema_binding import resolve_school_schema_name

from .ai_bridge import AIProposal, record_operator_feedback, remember_mapping_decision
from .reliability import (
    idempotent_post,
    safe_500,
    with_progress_fallback,
)
from .models import (
    AssetStatus,
    BundleStatus,
    ConflictResolution,
    IntakeMethod,
    MigrationBundle,
    MigrationConflict,
    MigrationIdMapping,
    SlaTier,
)

logger = logging.getLogger(__name__)


# Intake methods exposed in the wizard. Three buckets:
#   * upload  — operator drops file(s); we save them and hand paths to the adapter.
#   * url     — operator pastes a URL/SFTP/S3 link; we hand the string to the adapter.
#   * pending — adapter exists but requires extra plumbing (OAuth, mailbox, live DB);
#               the wizard accepts it but only stages a bundle for follow-up.
_INTAKE_WIZARD_METHODS = (
    (IntakeMethod.FILE_UPLOAD, "upload", "Direct file upload (CSV / XLSX / JSON / XML / TXT)"),
    (IntakeMethod.ARCHIVE, "upload", "Compressed archive (zip / tar / gz / 7z)"),
    (IntakeMethod.SQL_DUMP, "upload", "SQL dump file (.sql)"),
    (IntakeMethod.PDF, "upload", "PDF transcript stack"),
    (IntakeMethod.ACCESS_DB, "upload", "Microsoft Access database (.mdb / .accdb)"),
    (IntakeMethod.URL, "url", "Remote URL / signed link (http / https)"),
    (IntakeMethod.SFTP, "url", "SFTP path (sftp://user:pass@host/path)"),
    (IntakeMethod.S3, "url", "S3 prefix (s3://bucket/key)"),
    (IntakeMethod.DATABASE, "pending", "Live database connection (operator wires credentials)"),
    (IntakeMethod.OAUTH_FOLDER, "pending", "Google Drive / OneDrive / Dropbox folder (OAuth handoff)"),
    (IntakeMethod.EMAIL, "pending", "Email mailbox with attachments"),
    (IntakeMethod.API_PULL, "pending", "Vendor API accelerator"),
)


def _intake_method_kind(method: str) -> str:
    for value, kind, _label in _INTAKE_WIZARD_METHODS:
        if value == method:
            return kind
    return "upload"


def _mc_base_for_shell(shell: str) -> str:
    """Base template so a migration_cloud page renders in the correct shell.

    Tenant mounts render inside the portal shell via the _mc_base_portal adapter
    (re-homes cp_shell_page → portal content); operator mounts use the
    control-plane base natively. This is the operator-shell-leak fix — a
    migration page authored with a ``{% extends mc_base|default:... %}`` line and
    a ``cp_shell_page`` block renders correctly in either chrome. tenant⟂operator.
    """
    return (
        "migration_cloud/_mc_base_portal.html"
        if shell == "portal"
        else "control_plane_base.html"
    )


def _enforce_portal_entitlement(request, shell: str) -> JsonResponse | None:
    """Block tenant entry without the migration_cloud entitlement.

    Operator shell (``shell == "super"``) is always allowed for staff.
    Portal shell consults ``apps.billing`` if available; missing billing
    app degrades to "allow" so dev / pre-billing environments still work.
    Returns ``None`` when the request is allowed; a 402 JsonResponse to
    short-circuit otherwise.
    """
    if shell != "portal":
        return None
    school = getattr(request, "school", None) or getattr(request, "tenant", None)
    if school is None:
        return None
    try:
        from apps.billing.entitlements import can as has_capability
    except Exception:  # noqa: BLE001
        return None
    try:
        if has_capability(school, "migration_cloud"):
            return None
    except Exception:  # noqa: BLE001 — broken billing should never lock out the wizard
        return None
    return JsonResponse(
        {"error": "migration_cloud entitlement required for tenant access"},
        status=402,
    )


def _tenant_scoped_bundle(request, bundle_id: int, shell: str):
    """Resolve a bundle by id, tenant-scoped in portal shell.

    Operator shell (``super``) sees everything. Portal shell can only resolve
    bundles bound to the caller's active tenant — a mismatch is reported as
    404, never 403, so an ID-enumeration attacker cannot distinguish "exists
    elsewhere" from "doesn't exist".
    """
    from django.http import Http404
    bundle = get_object_or_404(MigrationBundle, pk=bundle_id)
    if shell == "portal":
        school = getattr(request, "school", None) or getattr(request, "tenant", None)
        school_pk = getattr(school, "pk", None)
        if school_pk is None or bundle.school_id != school_pk:
            raise Http404("bundle not found")
    return bundle


class MigrationCloudIntakeView(LoginRequiredMixin, View):
    """GET → render intake wizard. POST → create the bundle and ingest artifacts.

    Accepts every intake method the platform registers (file upload,
    archive, URL/SFTP/S3, SQL dump, PDF, Access DB), plus stage-only
    methods (live DB, OAuth folder, email, vendor API) that capture the
    operator's intent and create an empty bundle for follow-up. The
    ``source_hint`` field is deliberately free-text and may be left blank
    — the classifier handles known platforms, unknown competitors, and
    custom in-house products via its ``unknown_custom`` fallback path,
    with unmappable columns landing in ``custom_fields.*``.
    """

    template_name = "migration_cloud/intake_new.html"

    def get(self, request, shell: str = "super"):
        gate = _enforce_portal_entitlement(request, shell)
        if gate is not None:
            return gate

        # v3.17 (2026-05-17): honour pre-fill query params from the onboarding
        # handoff page. Keeps the bundle-creation contract intact (POST still
        # validates everything) while letting the operator land on a form
        # already populated with their vendor and school choice — one click
        # from there to "Create bundle and start".
        prefill_keys = (
            "vendor", "source_hint", "profile", "school_id", "label", "intake_method", "sla_tier", "domains",
        )
        prefill = {k: (request.GET.get(k) or "").strip() for k in prefill_keys}
        # `vendor` is our own slug; the form's `source_hint` accepts either
        # a marketing-friendly vendor slug OR the MigrationProfile.SourceSystem
        # value. If vendor is set but source_hint isn't, propagate vendor → hint.
        if prefill["vendor"] and not prefill["source_hint"]:
            prefill["source_hint"] = prefill["vendor"]

        form = None
        if any(prefill.values()):
            form = {
                "intake_method": prefill["intake_method"] or "",
                "label": prefill["label"] or "",
                "source_hint": prefill["source_hint"] or "",
                "sla_tier": prefill["sla_tier"] or "",
                "intake_source_uri": "",
                "school_id": prefill["school_id"] or "",
                "auto_advance": True,
                "apply_atomic": True,
                "diff_mode": "full",
                "diff_since": "",
                "parity_drift_rollback_pct": "",
                "expected_students_count": "",
                "expected_guardians_count": "",
                "expected_invoice_count": "",
                "expected_invoice_total_amount": "",
            }

        ctx = self._context(request=request, shell=shell, errors=None, form=form)
        # Surface pre-fill provenance so the template can render an inline hint
        # ("Pre-filled from your onboarding choice") and a discreet reset link.
        ctx["prefill_from_onboarding"] = bool(prefill["vendor"] or prefill["profile"])
        ctx["prefill_vendor_slug"] = prefill["vendor"]
        ctx["prefill_profile_slug"] = prefill["profile"]
        ctx["prefill_domains"] = [
            d for d in (prefill["domains"].split(",") if prefill["domains"] else []) if d
        ]
        return render(request, self.template_name, ctx)

    @idempotent_post
    @safe_500
    def post(self, request, shell: str = "super"):
        gate = _enforce_portal_entitlement(request, shell)
        if gate is not None:
            return gate

        from django.contrib import messages
        from django.shortcuts import redirect
        from django.urls import reverse

        form = {
            "intake_method": (request.POST.get("intake_method") or IntakeMethod.FILE_UPLOAD).strip(),
            "label": (request.POST.get("label") or "").strip(),
            "source_hint": (request.POST.get("source_hint") or "").strip(),
            "sla_tier": (request.POST.get("sla_tier") or SlaTier.SMALL).strip(),
            "intake_source_uri": (request.POST.get("intake_source_uri") or "").strip(),
            "school_id": (request.POST.get("school_id") or "").strip(),
            "auto_advance": request.POST.get("auto_advance") in ("1", "on", "true"),
            "apply_atomic": request.POST.get("apply_atomic") in ("1", "on", "true"),
            "diff_mode": (request.POST.get("diff_mode") or "full").strip().lower(),
            "diff_since": (request.POST.get("diff_since") or "").strip(),
            "parity_drift_rollback_pct": (
                request.POST.get("parity_drift_rollback_pct") or ""
            ).strip(),
            "expected_students_count": (
                request.POST.get("expected_students_count") or ""
            ).strip(),
            "expected_guardians_count": (
                request.POST.get("expected_guardians_count") or ""
            ).strip(),
            "expected_invoice_count": (
                request.POST.get("expected_invoice_count") or ""
            ).strip(),
            "expected_invoice_total_amount": (
                request.POST.get("expected_invoice_total_amount") or ""
            ).strip(),
        }
        errors: list[str] = []

        method = form["intake_method"]
        kind = _intake_method_kind(method)
        valid_methods = {value for value, _kind, _label in _INTAKE_WIZARD_METHODS}
        if method not in valid_methods:
            errors.append(f"Unknown intake method: {method!r}")

        # Double-submit guard: a per-user, per-method cache lock for 8s.
        # Belt-and-suspenders alongside the deterministic idempotency key.
        from django.core.cache import cache

        user_pk = getattr(request.user, "pk", None) or "anon"
        lock_key = f"mc:intake:lock:{user_pk}:{method}"
        if cache.add(lock_key, "1", timeout=8) is False:
            errors.append("A migration is already being created. Wait a few seconds and try again.")

        # Resolve target school.
        school = getattr(request, "school", None) or getattr(request, "tenant", None)
        school_id: int | None = getattr(school, "pk", None)
        if shell != "portal" and form["school_id"]:
            try:
                school_id = int(form["school_id"])
            except ValueError:
                errors.append("school_id must be an integer (or blank).")

        intake_options = self._clean_intake_options(form=form, errors=errors)

        # Build handle per kind.
        handle: Any = None
        intake_source_uri = form["intake_source_uri"]
        saved_paths: list[str] = []
        file_digests: list[str] = []
        total_upload_bytes = 0

        if kind == "upload":
            # upload-validation-allow: schema-agnostic SIS export (CSV/TSV/TXT/JSON/JSONL/XLS/XLSX/ZIP/PDF) has no single magic-byte type — re-sniffed + structure-validated by the profiler at parse time; streaming byte-size cap enforced below; a full-buffer AV read would defeat the GB-scale streaming-to-disk design
            files = [f for f in request.FILES.getlist("artifacts") if f and f.size > 0]
            empty_count = len(request.FILES.getlist("artifacts")) - len(files)
            if not files:
                errors.append("Attach at least one non-empty file to upload.")
            elif empty_count:
                errors.append(f"{empty_count} zero-byte file(s) skipped — re-export and retry.")
            else:
                max_bytes = self._max_upload_bytes()
                oversized = [f.name for f in files if f.size > max_bytes]
                if oversized:
                    errors.append(
                        f"File(s) exceed the {max_bytes:,}-byte cap: {', '.join(oversized)}. "
                        "Use the URL/SFTP/S3 method for larger drops."
                    )
                else:
                    total_upload_bytes = sum(f.size for f in files)
                    saved_paths, file_digests = self._persist_uploads(files)
                    intake_source_uri = (
                        saved_paths[0]
                        if len(saved_paths) == 1
                        else f"{len(saved_paths)} files staged ({total_upload_bytes:,} bytes)"
                    )
                    handle = saved_paths[0] if len(saved_paths) == 1 else saved_paths

        elif kind == "url":
            if not intake_source_uri:
                errors.append("Paste the remote URL / SFTP path / S3 prefix.")
            elif not self._looks_like_supported_url(intake_source_uri):
                errors.append("URL must start with http://, https://, sftp://, or s3://.")
            else:
                handle = intake_source_uri

        elif kind == "pending":
            # Stage-only: create the bundle so the operator can attach
            # credentials / OAuth tokens / queue API jobs out-of-band.
            handle = None

        if errors:
            # Release the lock so the operator can correct + resubmit immediately.
            cache.delete(lock_key)
            return render(
                request,
                self.template_name,
                self._context(request=request, shell=shell, errors=errors, form=form),
                status=400,
            )

        # Stable idempotency key. Same operator submitting the same payload
        # twice (double-click / browser-back / network retry) collapses to the
        # same bundle row instead of duplicating.
        idem_key = self._derive_idempotency_key(
            shell=shell,
            method=method,
            school_id=school_id,
            user_pk=user_pk,
            file_digests=file_digests,
            intake_source_uri=intake_source_uri,
        )

        # Pending bundles bypass adapter dispatch — create the row directly
        # and let the operator wire the intake source from the detail page.
        if kind == "pending":
            bundle, _created = MigrationBundle.objects.get_or_create(
                idempotency_key=idem_key,
                defaults={
                    "school_id": school_id,
                    "schema_name": resolve_school_schema_name(school),
                    "label": form["label"] or f"{method} (staged)",
                    "intake_method": method,
                    "intake_source_uri": intake_source_uri,
                    "source_hint": form["source_hint"],
                    "sla_tier": form["sla_tier"],
                    "triggered_by_id": getattr(request.user, "pk", None),
                    **intake_options,
                },
            )
            self._apply_intake_options(bundle=bundle, values=intake_options)
            bundle_id = bundle.pk
            messages.info(
                request,
                "Bundle staged. Attach the live source (credentials, OAuth token, "
                "or API endpoint) from this page to continue.",
            )
        else:
            from .services import BundleIngestionService, BundleSpec

            spec = BundleSpec(
                intake_method=method,
                handle=handle,
                school_id=school_id,
                schema_name=resolve_school_schema_name(school),
                label=form["label"],
                source_hint=form["source_hint"],
                sla_tier=form["sla_tier"],
                idempotency_key=idem_key,
                intake_source_uri=intake_source_uri,
                triggered_by_id=getattr(request.user, "pk", None),
            )
            try:
                result = BundleIngestionService().ingest(spec)
                bundle_id = result.bundle_id
                bundle = MigrationBundle.objects.get(pk=bundle_id, school_id=school_id)
                self._apply_intake_options(bundle=bundle, values=intake_options)
                # Persist the operator's per-file domain tags BEFORE auto-advance
                # so the classify step routes each file to the domain the
                # operator chose (students / teachers / subjects / …).
                self._store_operator_domains(request=request, bundle=bundle)
            except Exception as exc:  # noqa: BLE001 — surface intake failures inline
                logger.exception("migration_cloud.views: intake failed for method=%s", method)
                cache.delete(lock_key)
                errors.append(f"Intake failed: {type(exc).__name__}: {exc}")
                return render(
                    request,
                    self.template_name,
                    self._context(request=request, shell=shell, errors=errors, form=form),
                    status=500,
                )

            if result.artifacts_registered == 0 and result.artifacts_skipped_duplicate == 0:
                messages.warning(
                    request,
                    "Intake completed but no artifacts were registered. The source may be empty or unreadable.",
                )
            elif result.artifacts_skipped_duplicate and not result.artifacts_registered:
                messages.info(
                    request,
                    f"Duplicate submission detected — opened the existing bundle "
                    f"({result.artifacts_skipped_duplicate} artifact(s) already on file).",
                )
            else:
                messages.success(
                    request,
                    f"Intake complete — {result.artifacts_registered} artifact(s) registered.",
                )

            if form["auto_advance"]:
                self._dispatch_advance(request, bundle_id)

        cache.delete(lock_key)
        detail_url = reverse(
            f"migration_cloud_{'portal' if shell == 'portal' else 'super'}:bundle_detail",
            kwargs={"bundle_id": bundle_id},
        )
        return redirect(detail_url)

    def _store_operator_domains(self, *, request, bundle) -> None:
        """Record the operator's per-file domain tags on the bundle.

        The multi-file uploader posts a JSON ``artifact_domain_map`` of
        ``{filename: canonical_domain}``. We validate each tag against the
        canonical domain registry (an unknown/garbage tag is dropped, never
        routed) and stash the map under
        ``discovery_summary['operator_assigned_domains']``, which the pipeline's
        classify step reads to override inference. Best-effort — a malformed map
        never blocks the intake.
        """
        import json

        raw = (request.POST.get("artifact_domain_map") or "").strip()
        if not raw:
            return
        try:
            parsed = json.loads(raw)
        except (ValueError, TypeError):
            return
        if not isinstance(parsed, dict):
            return
        from .accelerators.runmycampus_canonical import is_valid_canonical_domain

        cleaned: dict[str, str] = {}
        for fname, domain in parsed.items():
            tag = (str(domain).strip() if domain is not None else "")
            if tag and tag != "auto" and is_valid_canonical_domain(tag):
                cleaned[str(fname)[:255]] = tag  # magic-number-allow: matches MigrationArtifact.filename max_length
        if not cleaned:
            return
        summary = dict(bundle.discovery_summary or {})
        existing = dict(summary.get("operator_assigned_domains") or {})
        existing.update(cleaned)
        summary["operator_assigned_domains"] = existing
        bundle.discovery_summary = summary
        bundle.save(update_fields=["discovery_summary", "updated_at"])

    def _context(self, *, request, shell: str, errors, form):
        defaults = {
            "intake_method": "",
            "label": "",
            "source_hint": "",
            "sla_tier": "",
            "intake_source_uri": "",
            "school_id": "",
            "auto_advance": True,
            "apply_atomic": True,
            "diff_mode": "full",
            "diff_since": "",
            "parity_drift_rollback_pct": "",
            "expected_students_count": "",
            "expected_guardians_count": "",
            "expected_invoice_count": "",
            "expected_invoice_total_amount": "",
        }
        from .accelerators.runmycampus_canonical import canonical_domain_choices

        form_data = {**defaults, **(form or {})}
        max_upload_bytes = self._max_upload_bytes()
        intake_methods = [
            {"value": value, "kind": kind, "label": label}
            for value, kind, label in _INTAKE_WIZARD_METHODS
        ]
        method_groups = [
            {
                "key": "upload",
                "label": "Upload files",
                "summary": "CSV, XLSX, JSON, XML, SQL, PDFs, Access DB, or archives.",
                "icon": "bi-cloud-arrow-up",
                "methods": [m for m in intake_methods if m["kind"] == "upload"],
            },
            {
                "key": "url",
                "label": "Pull from a location",
                "summary": "HTTP(S), SFTP, and S3 sources without local re-upload.",
                "icon": "bi-link-45deg",
                "methods": [m for m in intake_methods if m["kind"] == "url"],
            },
            {
                "key": "pending",
                "label": "Stage a live source",
                "summary": "Database, OAuth folder, mailbox, or vendor API handoff.",
                "icon": "bi-diagram-3",
                "methods": [m for m in intake_methods if m["kind"] == "pending"],
            },
        ]
        return {
            "shell": shell,
            "mc_base": _mc_base_for_shell(shell),
            "page_title": "Start a new migration",
            "intake_methods": intake_methods,
            "intake_method_groups": method_groups,
            "sla_tiers": list(SlaTier.choices),
            "default_intake_method": IntakeMethod.FILE_UPLOAD,
            "default_sla_tier": SlaTier.SMALL,
            "errors": errors or [],
            "form": form_data,
            "is_super_shell": shell != "portal",
            "max_upload_bytes": max_upload_bytes,
            "max_upload_mb": max(1, round(max_upload_bytes / (1024 * 1024))),  # magic-number-allow: byte-size-cap
            "upload_guardrails": [
                "Per-file cap is enforced before storage.",
                "SHA-256 idempotency prevents double-submit duplicates.",
                "Archives preserve folder lineage and child artifacts.",
                "Unmapped fields remain queryable under custom_fields.*.",
            ],
            "pipeline_steps": [
                {"label": "Land", "detail": "Save and fingerprint every artifact."},
                {"label": "Profile", "detail": "Read headers, rows, encodings, and formats."},
                {"label": "Classify", "detail": "Detect source system and domains."},
                {"label": "Map", "detail": "Bind columns to canonical fields with AI recall."},
                {"label": "Validate", "detail": "Run totals, conflict, and rollback guardrails."},
                {"label": "Apply", "detail": "Dry-run first; live apply requires confirmation."},
            ],
            "source_playbooks": [
                "Student roster, guardians, enrollments, sections, attendance, grades, invoices.",
                "Export one ZIP when possible; folder structure is preserved.",
                "Include a row-count or financial control total when the source provides one.",
                "Use remote URL/SFTP/S3 for exports that exceed browser upload limits.",
            ],
            # Per-file domain tagger: the operator uploads many canonical CSVs at
            # once and tells us which record type each file is (students /
            # teachers / subjects / …). Auto-detected from the filename, always
            # overridable, and authoritative over inference at apply time.
            "canonical_domains": canonical_domain_choices(),
        }

    def _clean_intake_options(self, *, form: dict[str, Any], errors: list[str]) -> dict[str, Any]:
        expected_totals: dict[str, str] = {}

        def clean_count(field: str, key: str, label: str) -> None:
            raw = str(form.get(field) or "").replace(",", "").strip()
            if not raw:
                return
            try:
                value = int(raw)
            except ValueError:
                errors.append(f"{label} must be a whole number.")
                return
            if value < 0:
                errors.append(f"{label} cannot be negative.")
                return
            expected_totals[key] = str(value)

        def clean_money(field: str, key: str, label: str) -> None:
            raw = str(form.get(field) or "").replace(",", "").replace("$", "").strip()
            if not raw:
                return
            try:
                value = Decimal(raw)
            except (InvalidOperation, ValueError):
                errors.append(f"{label} must be a valid amount.")
                return
            if value < Decimal("0"):
                errors.append(f"{label} cannot be negative.")
                return
            expected_totals[key] = format(value, "f")

        clean_count("expected_students_count", "students.count", "Expected students")
        clean_count("expected_guardians_count", "guardians.count", "Expected guardians")
        clean_count("expected_invoice_count", "finance.invoice_count", "Expected invoices")
        clean_money(
            "expected_invoice_total_amount",
            "finance.invoice_total_amount",
            "Expected invoice total",
        )

        diff_mode = str(form.get("diff_mode") or "full").lower()
        if diff_mode not in ("full", "since"):
            errors.append("Diff mode must be full or since.")
            diff_mode = "full"

        diff_since = None
        diff_since_raw = str(form.get("diff_since") or "").strip()
        if diff_since_raw:
            parsed = parse_datetime(diff_since_raw)
            if parsed is None:
                parsed_date = parse_date(diff_since_raw)
                if parsed_date is not None:
                    from datetime import datetime, time

                    parsed = datetime.combine(parsed_date, time.min)
            if parsed is None:
                errors.append("Diff since must be a valid date or ISO timestamp.")
            else:
                if timezone.is_naive(parsed):
                    parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
                diff_since = parsed
                diff_mode = "since"

        rollback_pct = 0.0
        rollback_raw = str(form.get("parity_drift_rollback_pct") or "").strip()
        if rollback_raw:
            try:
                rollback_pct = float(rollback_raw)
            except ValueError:
                errors.append("Rollback parity threshold must be a number.")
                rollback_pct = 0.0
            else:
                if rollback_pct < 0 or rollback_pct > 100:
                    errors.append("Rollback parity threshold must be between 0 and 100.")
                    rollback_pct = 0.0

        return {
            "expected_totals": expected_totals,
            "diff_mode": diff_mode,
            "diff_since": diff_since,
            "apply_atomic": bool(form.get("apply_atomic")),
            "parity_drift_rollback_pct": rollback_pct,
        }

    def _apply_intake_options(self, *, bundle: MigrationBundle, values: dict[str, Any]) -> None:
        update_fields: list[str] = []
        for field in (
            "expected_totals",
            "diff_mode",
            "diff_since",
            "apply_atomic",
            "parity_drift_rollback_pct",
        ):
            if getattr(bundle, field) != values.get(field):
                setattr(bundle, field, values.get(field))
                update_fields.append(field)
        if update_fields:
            update_fields.append("updated_at")
            bundle.save(update_fields=update_fields)

    def _persist_uploads(self, files) -> tuple[list[str], list[str]]:
        """Stream uploaded files to durable MEDIA storage under a per-day prefix.

        Returns ``(absolute_paths, sha256_digests)``. Files land under
        ``MEDIA_ROOT/migration_cloud/intake/YYYY-MM-DD/<random>/<filename>``
        so subsequent profiler runs can re-open them after a process restart
        (a per-request tempdir would not survive). Hashing happens during
        the write loop — single pass, bounded memory.
        """
        import hashlib
        import secrets
        from datetime import date
        from pathlib import Path

        from django.conf import settings

        media_root = Path(getattr(settings, "MEDIA_ROOT", "media"))
        day_prefix = date.today().isoformat()
        slot = secrets.token_urlsafe(12).replace("/", "_").replace("=", "")
        dest_dir = media_root / "migration_cloud" / "intake" / day_prefix / slot
        dest_dir.mkdir(parents=True, exist_ok=True)

        saved_paths: list[str] = []
        digests: list[str] = []
        for f in files:
            safe_name = Path(f.name).name or "upload.bin"
            target = dest_dir / safe_name
            digest = hashlib.sha256()
            with target.open("wb") as out:
                for chunk in f.chunks():
                    out.write(chunk)
                    digest.update(chunk)
            saved_paths.append(str(target))
            digests.append(digest.hexdigest())
        return saved_paths, digests

    def _derive_idempotency_key(
        self,
        *,
        shell: str,
        method: str,
        school_id: int | None,
        user_pk: Any,
        file_digests: list[str],
        intake_source_uri: str,
    ) -> str:
        """Stable key so double-submit collapses to one bundle.

        Composition: shell + method + tenant + user + content signature.
        Content signature is the joined SHA256s for uploads (deterministic
        across retries), or the URI itself for url/pending kinds. Falls
        back to a random token only if neither signal exists.
        """
        import hashlib
        import secrets

        if file_digests:
            signature = "|".join(sorted(file_digests))
        elif intake_source_uri:
            signature = intake_source_uri
        else:
            signature = secrets.token_urlsafe(16)

        composite = f"{shell}|{method}|{school_id or '-'}|{user_pk or '-'}|{signature}"
        return "mc-" + hashlib.sha256(composite.encode("utf-8")).hexdigest()[:48]

    def _max_upload_bytes(self) -> int:
        """Per-file upload cap. Reuses the intake adapter cap so all paths agree."""
        try:
            from . import defaults as mc_defaults

            return int(mc_defaults.get("migration_cloud.intake.max_artifact_bytes"))
        except Exception:  # noqa: BLE001 — defensive default
            return 1024 * 1024 * 1024  # 1 GiB  # magic-number-allow: byte-size-cap

    def _looks_like_supported_url(self, value: str) -> bool:
        v = value.strip().lower()
        return v.startswith(("http://", "https://", "sftp://", "s3://"))

    def _dispatch_advance(self, request, bundle_id: int) -> None:
        """Queue post-intake pipeline on durable outbox (never sync on HTTP)."""
        from django.contrib import messages

        from .celery_tasks import enqueue_advance

        result = enqueue_advance(bundle_id, use_accelerator=True)
        job_id = getattr(result, "outbox_id", None) or getattr(result, "id", None)
        messages.info(
            request,
            (
                f"Auto-advance queued (job {job_id}). Refresh the bundle page "
                "shortly to see profile / classify / map progress."
                if job_id
                else "Auto-advance queued. Refresh the bundle page shortly."
            ),
        )


class MigrationCloudConsoleView(LoginRequiredMixin, View):
    """List recent bundles for the active shell + intake CTA."""

    template_name = "migration_cloud/console.html"

    def get(self, request, shell: str = "super"):
        gate = _enforce_portal_entitlement(request, shell)
        if gate is not None:
            return gate
        bundles_qs = (
            MigrationBundle.objects
            .order_by("-created_at")
            .select_related("school", "triggered_by")
        )
        if shell == "portal":
            # Tenant scope: limit to the active school only.
            school = getattr(request, "school", None) or getattr(request, "tenant", None)
            if school is not None:
                bundles_qs = bundles_qs.filter(school=school)
        bundles = bundles_qs[:50]
        # Operator-shell-leak fix: the tenant mount renders inside the tenant
        # portal shell, the operator mount inside the control-plane shell.
        template = (
            "migration_cloud/console_portal.html"
            if shell == "portal"
            else self.template_name
        )
        return render(
            request,
            template,
            {
                "shell": shell,
                "bundles": bundles,
                "page_title": "Migration Cloud",
                "is_super_shell": shell != "portal",
            },
        )


def _rollback_runs_for_bundle(bundle):
    """Per-(domain,artifact) ``MigrationRun`` rows for this bundle that can still
    be rolled back (a rollback snapshot is present and the run was not already
    reverted). Empty when the automation app is absent / unmigrated.

    Runs carry no bundle FK — they are school-scoped — but
    ``orchestrator._create_audit_run`` stamps ``execution_summary["bundle_id"]``
    at create time, so we associate by that plus the school. ``can_rollback`` is
    a Python property (not a DB field), so the final filter runs in memory.
    """
    if bundle.school_id is None or bundle.status not in (
        BundleStatus.APPLIED,
        BundleStatus.RECONCILED,
    ):
        return []
    try:
        from apps.automation.models import MigrationRun
    except Exception:  # noqa: BLE001 — automation app absent / unmigrated → no rollback surface
        return []
    run_qs = MigrationRun.objects.filter(
        school_id=bundle.school_id,
        dry_run=False,
        execution_summary__bundle_id=bundle.pk,
    ).order_by("-started_at")
    return [r for r in run_qs if r.can_rollback]


class MigrationCloudBundleDetailView(LoginRequiredMixin, View):
    """Show one bundle's profile / classification / mapping / reconciliation surfaces."""

    template_name = "migration_cloud/bundle_detail.html"
    _ARTIFACTS_PAGE_SIZE = 20

    def get(self, request, bundle_id: int, shell: str = "super"):
        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        artifacts_qs = bundle.artifacts.all().order_by("path_within_bundle")
        artifacts_page_obj = Paginator(
            artifacts_qs, self._ARTIFACTS_PAGE_SIZE
        ).get_page(request.GET.get("artifact_page") or 1)
        per_artifact_domain = (
            (bundle.discovery_summary or {}).get("per_artifact_domain") or {}
        )
        per_artifact_mappings = (bundle.mapping_summary or {}).get("per_artifact") or {}
        pending_methods = {value for value, kind, _ in _INTAKE_WIZARD_METHODS if kind == "pending"}
        return render(
            request,
            self.template_name,
            {
                "mc_base": _mc_base_for_shell(shell),
                "shell": shell,
                "bundle": bundle,
                "artifacts": list(artifacts_page_obj.object_list),
                "artifacts_page_obj": artifacts_page_obj,
                "per_artifact_domain": per_artifact_domain,
                "per_artifact_mappings": per_artifact_mappings,
                "reconciliation": bundle.reconciliation_summary or {},
                "page_title": bundle.label or bundle.idempotency_key,
                "can_apply": bundle.status == BundleStatus.MAPPED,
                "can_reconcile": bundle.status == BundleStatus.APPLIED,
                "can_advance": bundle.status in (
                    BundleStatus.PENDING,
                    BundleStatus.INGESTING,
                    BundleStatus.PROFILED,
                    BundleStatus.CLASSIFIED,
                ),
                # Pending-method bundle without a source attached yet.
                "needs_source": (
                    bundle.intake_method in pending_methods
                    and not (bundle.intake_source_uri or "").strip()
                ),
                # Pre-tenant staging — bundle not bound to a school yet.
                "needs_school_binding": bundle.school_id is None and shell == "super",
                # Rollback surface: this bundle's still-revertible apply runs.
                "rollback_runs": _rollback_runs_for_bundle(bundle),
            },
        )


@method_decorator(
    track_workflow(
        "migration_bundle_advance",
        steps=("profile", "classify", "map"),
        expected_duration_seconds=900,  # magic-number-allow: workflow-expected-duration-seconds
        email_on_failure=True,
    ),
    name="post",
)
class MigrationCloudAdvanceView(LoginRequiredMixin, View):
    """POST endpoint: advance bundle through profile → classify → map.

    Reliability contract (v3.17):
      - HTTP-level idempotency: a duplicate POST with the same Idempotency-Key
        header within 24h returns the cached response (X-Idempotency-Replay: true).
      - Uncaught exceptions → structured JSON 500 with a request_id for support.
    """

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        from django.http import Http404

        # Tenant scope: the portal shell must only drive its OWN bundles. This
        # view (and the DRF action that delegates here) previously ran the
        # pipeline on the raw pk with no tenant check — a cross-tenant IDOR.
        try:
            _tenant_scoped_bundle(request, bundle_id, shell)
        except Http404:
            return JsonResponse({"error": "bundle not found"}, status=404)
        try:
            from .celery_tasks import enqueue_advance

            result = enqueue_advance(bundle_id, use_accelerator=True)
            return JsonResponse(
                {
                    "ok": True,
                    "queued": True,
                    "durable_outbox": True,
                    "outbox_id": getattr(result, "outbox_id", None)
                    or getattr(result, "id", None),
                    "bundle_id": bundle_id,
                }
            )
        except MigrationBundle.DoesNotExist:
            return JsonResponse({"error": "bundle not found"}, status=404)
        return JsonResponse({"error": "enqueue_failed"}, status=500)


@method_decorator(
    track_workflow(
        "migration_bundle_apply",
        steps=("prepare", "apply_waves", "finalize"),
        expected_duration_seconds=1800,  # magic-number-allow: workflow-expected-duration-seconds
        email_on_failure=True,
    ),
    name="post",
)
class MigrationCloudApplyView(LoginRequiredMixin, View):
    """POST endpoint: apply the MAPPED bundle to the tenant.

    Safety contract — **default is dry-run**. Live apply requires
    ``?confirm=1`` explicitly; ``?dry_run=1`` (or any truthy dry_run
    value) also forces dry-run regardless of confirm. This means API
    misuse and accidental form replays never overwrite a live tenant.

    Reliability contract (v3.17):
      - HTTP-level idempotency: a duplicate POST with the same Idempotency-Key
        header within 24h returns the cached response (X-Idempotency-Replay: true).
        Critical here: live applies are LARGE writes; a doubled POST without
        this guard could double-create records (the orchestrator catches it
        via bundle.idempotency_key, but request-level is the first line of
        defense).
      - Uncaught exceptions → structured JSON 500 with a request_id.
    """

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        from django.http import Http404

        from apps.accounts.decorators import user_is_tenant_admin

        # Role gate: LoginRequired alone admitted any membership (S2 residual).
        # Portal → tenant-admin tier for the request school; operator shell →
        # staff/superuser (control-plane) before any live or dry-run apply.
        # Return JSON 403 (do not raise) — ``@safe_500`` would otherwise turn
        # PermissionDenied into an opaque internal_error envelope.
        if shell == "portal":
            school = getattr(request, "school", None) or getattr(request, "tenant", None)
            if not user_is_tenant_admin(request.user, school):
                return JsonResponse({"error": "forbidden"}, status=403)
        elif not (
            getattr(request.user, "is_staff", False)
            or getattr(request.user, "is_superuser", False)
        ):
            return JsonResponse({"error": "forbidden"}, status=403)

        # Tenant scope before a LIVE apply — a portal caller must never apply
        # another tenant's bundle (cross-tenant IDOR closed here + on the DRF
        # action that delegates to this view).
        try:
            _tenant_scoped_bundle(request, bundle_id, shell)
        except Http404:
            return JsonResponse({"error": "bundle not found"}, status=404)
        dry_run_explicit = str(request.GET.get("dry_run", "")).lower() in ("1", "true", "yes")
        confirmed = str(request.GET.get("confirm", "")).lower() in ("1", "true", "yes")
        # Live apply iff the caller explicitly confirmed AND did not set dry_run=1.
        dry_run = dry_run_explicit or not confirmed
        try:
            from .celery_tasks import enqueue_apply

            result = enqueue_apply(bundle_id, dry_run=dry_run)
            return JsonResponse(
                {
                    "ok": True,
                    "queued": True,
                    "durable_outbox": True,
                    "dry_run": dry_run,
                    "outbox_id": getattr(result, "outbox_id", None)
                    or getattr(result, "id", None),
                    "bundle_id": bundle_id,
                }
            )
        except MigrationBundle.DoesNotExist:
            return JsonResponse({"error": "bundle not found"}, status=404)
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=409)


class MigrationCloudReconcileView(LoginRequiredMixin, View):
    """POST endpoint: compute reconciliation report for an APPLIED bundle.

    Reliability contract (v3.17): idempotency-key replay + structured 500.
    """

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json

        from django.http import Http404

        from .reconciliation import reconcile_bundle

        # Tenant scope — reconcile returns per-domain counts/parity/field names;
        # a portal caller must only reconcile their own bundle.
        try:
            _tenant_scoped_bundle(request, bundle_id, shell)
        except Http404:
            return JsonResponse({"error": "bundle not found"}, status=404)
        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            payload = {}
        cohort = payload.get("cohort") if isinstance(payload.get("cohort"), dict) else None

        try:
            report = reconcile_bundle(bundle_id=bundle_id, cohort=cohort)
        except MigrationBundle.DoesNotExist:
            return JsonResponse({"error": "bundle not found"}, status=404)
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=409)
        return JsonResponse({
            "bundle_id": report.bundle_id,
            "overall_parity_pct": report.overall_parity_pct,
            "per_domain": [
                {
                    "domain": d.domain,
                    "source_count": d.source_count,
                    "target_created": d.target_created,
                    "target_updated": d.target_updated,
                    "quarantined": d.quarantined,
                    "parity_pct": d.parity_pct,
                    "fill_rate_by_field": d.fill_rate_by_field,
                    "sample_count": len(d.sample_rows),
                }
                for d in report.per_domain
            ],
            "notes": report.notes,
        })


class MigrationCloudFeedbackView(LoginRequiredMixin, View):
    """POST endpoint: record an operator's accept/override decision on an AI proposal.

    Flows back into ``services.ai_gateway.record_feedback`` so the
    platform's daily ``AIGatewayMetric`` rollup picks up acceptance rates
    + manual-correction counts per task type. This is how the AI gets
    *functional* over time — we measure operator trust and iterate
    prompts on the slowest-converging domains.
    """

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json

        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "invalid JSON"}, status=400)

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        accepted = bool(payload.get("accepted", False))
        manual_correction = bool(payload.get("manual_correction", False))
        prompt_type = str(payload.get("prompt_type") or "migration_cloud.field_mapper")
        tier = str(payload.get("tier") or "unknown")
        confidence = float(payload.get("confidence") or 0.0)
        reasoning = str(payload.get("reasoning") or "")

        proposal = AIProposal(
            answer=payload.get("answer"),
            confidence=confidence,
            reasoning=reasoning,
            raw="",
            provider_meta={"tier": tier},
        )
        record_operator_feedback(
            school=bundle.school,
            proposal=proposal,
            prompt_type=prompt_type,
            accepted=accepted,
            manual_correction=manual_correction,
            request_id=payload.get("request_id"),
        )

        # When the operator accepts a mapping (or supplies a manual correction
        # the operator wants reused), persist into AIEmbeddingStore so the
        # next bundle for this tenant can skip the AI tiebreaker entirely.
        remembered = False
        if (accepted or manual_correction) and prompt_type.endswith("field_mapper"):
            mapping_meta = payload.get("mapping") or {}
            source_column = str(mapping_meta.get("source_column") or "")
            canonical_field = str(payload.get("answer") or mapping_meta.get("canonical_field") or "")
            if source_column and canonical_field:
                source_system = ((bundle.discovery_summary or {}).get("source") or {}).get("chosen")
                remembered = remember_mapping_decision(
                    school=bundle.school,
                    source_column=source_column,
                    sample_values=list(mapping_meta.get("sample_values") or []),
                    canonical_field=canonical_field,
                    domain=str(mapping_meta.get("domain") or ""),
                    confidence=max(confidence, 0.90 if accepted else 0.85),
                    method="operator_accept" if accepted else "operator_correction",
                    transformer=mapping_meta.get("transformer"),
                    source_system=source_system,
                )

        # T15 correctness fix: an operator OVERRIDE (drag-and-drop or the
        # "Override" button — manual_correction=True with a canonical_field
        # that differs from the AI proposal) must also rewrite THIS bundle's
        # mapping_summary so ``apply`` uses the operator's choice. Previously
        # this view only recorded feedback + next-bundle recall, so the UI
        # showed a "saved" toast while apply silently used the original AI
        # mapping. Mirrors MigrationCloudAIRebindView's in-place edit.
        applied_to_bundle = False
        artifacts_updated: list[str] = []
        reapply_required = False
        if manual_correction and prompt_type.endswith("field_mapper"):
            mapping_meta = payload.get("mapping") or {}
            src_col = str(mapping_meta.get("source_column") or "")
            new_canon = str(payload.get("answer") or mapping_meta.get("canonical_field") or "")
            if src_col and new_canon:
                per_artifact = (bundle.mapping_summary or {}).get("per_artifact") or {}
                for path, mappings in per_artifact.items():
                    for m in (mappings or []):
                        if (
                            str(m.get("source_column") or "") == src_col
                            and str(m.get("canonical_field") or "") != new_canon
                        ):
                            m["canonical_field"] = new_canon
                            m["confidence"] = max(float(m.get("confidence") or 0.0), 0.95)
                            m["method"] = "operator_override"
                            m["reasoning"] = "Operator override via the mapping editor."
                            if mapping_meta.get("transformer"):
                                m["transformer"] = mapping_meta.get("transformer")
                            artifacts_updated.append(path)
                if artifacts_updated:
                    bundle.mapping_summary = {
                        **(bundle.mapping_summary or {}),
                        "per_artifact": per_artifact,
                    }
                    bundle.save(update_fields=["mapping_summary", "updated_at"])
                    applied_to_bundle = True
                    # If the bundle was already applied, the operator must
                    # re-run Apply for the override to reach landed data.
                    reapply_required = bundle.status in (
                        BundleStatus.APPLIED,
                        BundleStatus.RECONCILED,
                    )

        return JsonResponse({
            "bundle_id": bundle.pk,
            "recorded": True,
            "remembered_for_recall": remembered,
            "prompt_type": prompt_type,
            "accepted": accepted,
            "manual_correction": manual_correction,
            "applied_to_bundle": applied_to_bundle,
            "artifacts_updated": artifacts_updated,
            "reapply_required": reapply_required,
        })


class MigrationCloudShadowView(LoginRequiredMixin, View):
    """POST endpoint cluster for the shadow-mode lifecycle.

    Action picked from ``?action=`` query string:

        ?action=start     — open a shadow window
        ?action=refresh   — take a tick (compute drift, maybe auto-cutover)
        ?action=close     — seal the window (body ``{"accepted": true/false}``)
        ?action=status    — GET the current state (read-only via POST)

    Source-side counts are operator-supplied in the body
    (``{"source_counts": {"students": 1240, ...}}``) because the platform
    can't assume how the old SIS exposes itself.
    """

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json

        from . import shadow

        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            payload = {}
        action = (request.GET.get("action") or payload.get("action") or "status").strip().lower()
        source_counts = payload.get("source_counts") if isinstance(payload.get("source_counts"), dict) else None
        source_pull = (lambda c=source_counts: dict(c)) if source_counts is not None else None

        # Tenant-scope EVERY shadow action. `status` re-resolves below via
        # _tenant_scoped_bundle, but start/refresh/close MUTATE cutover state
        # (open the window, arm auto-cutover, seal + advance to RECONCILED) and
        # previously passed the raw bundle_id straight into shadow.* — so a
        # portal caller could drive another tenant's migration. Fail closed here
        # before any shadow.* call touches a bundle it doesn't own.
        if shell == "portal":
            school = getattr(request, "school", None) or getattr(request, "tenant", None)
            school_pk = getattr(school, "pk", None)
            if school_pk is None or not MigrationBundle.objects.filter(
                pk=bundle_id, school_id=school_pk
            ).exists():
                return JsonResponse({"error": "bundle not found"}, status=404)

        try:
            if action == "start":
                state = shadow.start_shadow_window(
                    bundle_id=bundle_id,
                    target_parity_pct=float(payload.get("target_parity_pct", 99.0)),
                    max_window_hours=int(payload.get("max_window_hours", 168)),  # magic-number-allow: window-duration-hours
                    auto_cutover_armed=bool(payload.get("auto_cutover_armed", False)),
                    source_pull=source_pull,
                )
            elif action == "refresh":
                state = shadow.refresh_shadow(
                    bundle_id=bundle_id, source_pull=source_pull,
                )
            elif action == "close":
                state = shadow.close_shadow(
                    bundle_id=bundle_id,
                    accepted=bool(payload.get("accepted", False)),
                )
            elif action == "status":
                bundle = _tenant_scoped_bundle(request, bundle_id, shell)
                state = (bundle.reconciliation_summary or {}).get("shadow") or {}
                return JsonResponse({"bundle_id": bundle.pk, "shadow": state})
            else:
                return JsonResponse({"error": f"unknown action {action!r}"}, status=400)
        except MigrationBundle.DoesNotExist:
            return JsonResponse({"error": "bundle not found"}, status=404)
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=409)
        except Exception as exc:  # noqa: BLE001
            logger.exception("migration_cloud.views: shadow %s failed for bundle %s", action, bundle_id)
            return JsonResponse({"error": str(exc)}, status=500)

        from dataclasses import asdict
        return JsonResponse({
            "bundle_id": bundle_id,
            "action": action,
            "shadow": asdict(state),
        })


class MigrationCloudRollbackView(LoginRequiredMixin, View):
    """POST endpoint: roll back one child ``MigrationRun`` of a bundle."""

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, run_id: int, shell: str = "super"):
        from django.http import Http404

        try:
            from apps.automation.models import MigrationRun
        except ImportError:
            return JsonResponse({"error": "automation app not available"}, status=500)

        # Tenant isolation. Rollback is *destructive* (it deletes migrated tenant
        # rows keyed on ``run.school``). Without scoping, a logged-in portal user
        # of school A could delete school B's migrated students/grades simply by
        # enumerating ``run_id`` — a cross-tenant destructive IDOR. Resolve the
        # bundle tenant-checked (portal shell 404s a cross-tenant bundle_id), then
        # require the run to belong to the same school as that bundle. Operator
        # (``super``) shell is intentionally cross-tenant but still pins run↔bundle
        # school consistency. See docs/MIGRATION_CLOUD_AUDIT_2026_07_24.md (BLOCKER 4).
        gate = _enforce_portal_entitlement(request, shell)
        if gate is not None:
            return gate
        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        run = get_object_or_404(MigrationRun, pk=run_id)
        if run.school_id != bundle.school_id:
            raise Http404("run not found")
        rollback_run, result = run.trigger_rollback(user=request.user)
        return JsonResponse({
            "bundle_id": bundle_id,
            "run_id": run.pk,
            "rollback_run_id": getattr(rollback_run, "pk", None),
            "result": result,
        })


class MigrationCloudSaveProfileView(LoginRequiredMixin, View):
    """POST endpoint: distill a MAPPED bundle's accepted mappings into a reusable
    ``apps.automation.MigrationProfile`` row.

    Curated profiles beat AI tiebreaker forever: subsequent bundles from the
    same source pre-load the saved column→canonical map and skip discovery
    for known shapes. Tenant-scoped via the bundle's school.
    """

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json
        import re

        try:
            from apps.automation.models import MigrationProfile
        except Exception:  # noqa: BLE001
            return JsonResponse({"error": "automation app not available"}, status=500)

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "invalid JSON"}, status=400)

        name = str(payload.get("name") or "").strip() or (
            bundle.label or f"Profile from bundle {bundle.pk}"
        )
        description = str(payload.get("description") or "").strip()
        source_chosen = ((bundle.discovery_summary or {}).get("source") or {}).get("chosen") or "other"
        domain_hint = str(payload.get("domain") or "generic_sis").lower()

        source_map = {
            "powerschool": MigrationProfile.SourceSystem.POWERSCHOOL,
            "blackbaud": MigrationProfile.SourceSystem.BLACKBAUD,
            "veracross": MigrationProfile.SourceSystem.VERACROSS,
            "infinite_campus": MigrationProfile.SourceSystem.INFINITE_CAMPUS,
            "facts": MigrationProfile.SourceSystem.FACTS,
            "skyward": MigrationProfile.SourceSystem.SKYWARD,
            "alma": MigrationProfile.SourceSystem.ALMA,
        }
        domain_map = {
            "students": MigrationProfile.Domain.STUDENTS,
            "finance": MigrationProfile.Domain.FINANCE,
            "attendance": MigrationProfile.Domain.ATTENDANCE,
            "grades": MigrationProfile.Domain.GRADES,
        }
        source_system_value = source_map.get(source_chosen, MigrationProfile.SourceSystem.OTHER)
        domain_value = domain_map.get(domain_hint, MigrationProfile.Domain.GENERIC_SIS)

        per_artifact = (bundle.mapping_summary or {}).get("per_artifact") or {}
        curated_columns: dict[str, dict[str, Any]] = {}
        for _path, mappings in per_artifact.items():
            for m in (mappings or []):
                src = m.get("source_column")
                cf = m.get("canonical_field")
                if not src or not cf or cf.startswith("custom_fields."):
                    continue
                if float(m.get("confidence") or 0.0) < 0.65:
                    continue
                curated_columns.setdefault(src, {
                    "canonical_field": cf,
                    "transformer": m.get("transformer"),
                    "domain": m.get("domain") or domain_hint,
                    "confidence": m.get("confidence"),
                })

        slug_base = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")[:48] or f"bundle-{bundle.pk}"
        slug = slug_base
        counter = 1
        while MigrationProfile.objects.filter(slug=slug).exists():
            counter += 1
            slug = f"{slug_base}-{counter}"[:64]

        config: dict[str, Any] = {
            "saved_from_bundle_id": bundle.pk,
            "source_chosen": source_chosen,
            "tenant_school_id": str(getattr(bundle.school, "pk", "")) if bundle.school else "",
            "columns": curated_columns,
            "ontology_version": "v1",
        }

        profile = MigrationProfile.objects.create(
            slug=slug,
            source_system=source_system_value,
            profile_category=MigrationProfile.ProfileCategory.STRATEGY,
            name=name,
            description=description or f"Curated from bundle {bundle.label or bundle.pk}",
            format=MigrationProfile.Format.GENERIC_SIS,
            domain=domain_value,
            config=config,
            is_active=True,
        )

        return JsonResponse({
            "profile_slug": profile.slug,
            "profile_id": profile.pk,
            "columns_saved": len(curated_columns),
            "source_system": source_system_value,
            "domain": domain_value,
        })


class MigrationCloudAnomalyNudgeView(LoginRequiredMixin, View):
    """GET endpoint: operator review queue for a bundle.

    Surfaces (a) low-confidence column mappings flagged for human review,
    (b) quarantine records from apply, and (c) reconciliation drift hotspots
    — the three things an operator actually needs to act on. This is the
    "anomaly nudge" surface: actionable items only, no chatter.
    """

    template_name = "migration_cloud/anomaly_nudge.html"

    def get(self, request, bundle_id: int, shell: str = "super"):
        from apps.migration_cloud import defaults as mc_defaults

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        threshold = float(mc_defaults.get("migration_cloud.mapper.field_min_confidence"))

        low_conf_mappings: list[dict[str, Any]] = []
        per_artifact = (bundle.mapping_summary or {}).get("per_artifact") or {}
        for path, mappings in per_artifact.items():
            for m in (mappings or []):
                conf = float(m.get("confidence") or 0.0)
                if conf < threshold or str(m.get("canonical_field") or "").startswith("custom_fields."):
                    low_conf_mappings.append({
                        "artifact": path,
                        "source_column": m.get("source_column"),
                        "canonical_field": m.get("canonical_field"),
                        "confidence": conf,
                        "method": m.get("method"),
                        "transformer": m.get("transformer"),
                        "reasoning": m.get("reasoning"),
                        "domain": m.get("domain"),
                    })

        quarantine_rows: list[dict[str, Any]] = []
        try:
            from apps.automation.models import MigrationQuarantineRecord, MigrationRun

            run_ids = list(
                # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
                # Runs are linked to the bundle via execution_summary["bundle_id"]
                # (orchestrator._create_audit_run) — there is no parent_bundle FK,
                # so the old parent_bundle_id filter raised FieldError (swallowed)
                # and this surface always rendered empty.
                MigrationRun.objects.filter(
                    execution_summary__bundle_id=bundle.pk
                ).values_list("pk", flat=True)
            )
            # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
            for q in MigrationQuarantineRecord.objects.filter(
                migration_run_id__in=run_ids
            ).order_by("-id")[:200]:
                payload = q.payload if isinstance(q.payload, dict) else {}
                quarantine_rows.append({
                    "id": q.pk,
                    "run_id": q.migration_run_id,
                    "domain": q.domain,
                    "row_index": q.row_index,
                    "issue_class": q.issue_class,
                    "reason": payload.get("error", "") or q.issue_class,
                    "raw_row": q.payload,
                    "ack_status": q.status,
                })
        except Exception:  # noqa: BLE001
            logger.debug("anomaly_nudge: quarantine fetch failed", exc_info=True)

        reconciliation = bundle.reconciliation_summary or {}
        drift_domains = [
            d for d in (reconciliation.get("per_domain") or [])
            if float(d.get("parity_pct") or 100.0) < 99.0
        ]

        return render(
            request,
            self.template_name,
            {
                "mc_base": _mc_base_for_shell(shell),
                "shell": shell,
                "bundle": bundle,
                "low_conf_mappings": low_conf_mappings,
                "quarantine_rows": quarantine_rows,
                "drift_domains": drift_domains,
                "threshold": threshold,
                "page_title": f"Review queue — {bundle.label or bundle.idempotency_key}",
            },
        )


class MigrationCloudAttachSourceView(LoginRequiredMixin, View):
    """Attach a live source to a PENDING bundle (OAuth folder / DB / mailbox / API).

    GET renders a per-method form. POST writes the operator-supplied
    handle onto ``MigrationBundle.intake_source_uri`` and ``config``
    JSON (credentials are stored encrypted at-rest by the platform's
    secrets layer; this view never logs them). If the method is one we
    can immediately ingest from (URL-shaped: SFTP/S3/HTTP), we also
    auto-dispatch the advance pipeline.
    """

    template_name = "migration_cloud/attach_source.html"

    def get(self, request, bundle_id: int, shell: str = "super"):
        gate = _enforce_portal_entitlement(request, shell)
        if gate is not None:
            return gate
        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        return render(request, self.template_name, self._context(shell, bundle, errors=None))

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        from django.contrib import messages
        from django.shortcuts import redirect
        from django.urls import reverse

        gate = _enforce_portal_entitlement(request, shell)
        if gate is not None:
            return gate
        bundle = _tenant_scoped_bundle(request, bundle_id, shell)

        method = bundle.intake_method
        errors: list[str] = []
        source_uri = (request.POST.get("intake_source_uri") or "").strip()
        notes = (request.POST.get("notes") or "").strip()

        if not source_uri:
            errors.append("Provide the source URI / endpoint / mailbox address.")

        if method == IntakeMethod.DATABASE:
            # Validate DSN shape lightly — drivers do real validation at connect time.
            if not any(source_uri.startswith(p) for p in (
                "postgres://", "postgresql://", "mysql://", "mssql://", "sqlite://", "oracle://"
            )):
                errors.append(
                    "Database DSN must use a known scheme "
                    "(postgres / postgresql / mysql / mssql / sqlite / oracle)."
                )
        elif method == IntakeMethod.EMAIL:
            if "@" not in source_uri:
                errors.append("Email address required (operator wires the mailbox via IMAP credentials separately).")
        elif method == IntakeMethod.OAUTH_FOLDER:
            if not source_uri.startswith(("gdrive://", "onedrive://", "dropbox://", "https://")):
                errors.append("OAuth folder requires gdrive:// / onedrive:// / dropbox:// / https:// handle.")
        elif method == IntakeMethod.API_PULL:
            if not source_uri.startswith(("http://", "https://")):
                errors.append("API endpoint must be an http(s) URL.")

        if errors:
            return render(
                request, self.template_name,
                self._context(shell, bundle, errors=errors, form={"intake_source_uri": source_uri, "notes": notes}),
                status=400,
            )

        # Capture method-specific live-source CREDENTIALS into the encrypted-at-
        # rest connector_secret (Fernet). These are the only secret fields; the
        # non-secret url/provider/folder_id ride along so pipeline.build_connector_
        # handle can reconstruct the adapter handle at ingest. NEVER logged, never
        # returned. Attach can be partial (URL now, token later) — ingest simply
        # no-ops until the required pieces are present.
        secret_payload: dict[str, Any] = {}
        if method == IntakeMethod.API_PULL:
            api_token = (request.POST.get("api_token") or "").strip()
            if api_token:
                secret_payload = {
                    "url": source_uri,
                    "api_token": api_token,
                    "artifact_name": (request.POST.get("artifact_name") or "").strip()
                    or "api_export.json",
                }
        elif method == IntakeMethod.OAUTH_FOLDER:
            access_token = (request.POST.get("access_token") or "").strip()
            provider = (request.POST.get("oauth_provider") or "").strip().lower()
            folder_id = (request.POST.get("folder_id") or "").strip()
            if access_token and provider and folder_id:
                secret_payload = {
                    "provider": provider,
                    "folder_id": folder_id,
                    "access_token": access_token,
                }

        bundle.intake_source_uri = source_uri
        existing_summary = bundle.size_summary or {}
        existing_summary["attached_source_notes"] = notes[:500]
        bundle.size_summary = existing_summary
        update_fields = ["intake_source_uri", "size_summary", "updated_at"]
        if secret_payload:
            bundle.connector_secret = secret_payload
            update_fields.append("connector_secret")
        bundle.save(update_fields=update_fields)

        messages.success(
            request,
            "Source attached. Click Advance to pull, profile, and map it.",
        )

        detail_url = reverse(
            f"migration_cloud_{'portal' if shell == 'portal' else 'super'}:bundle_detail",
            kwargs={"bundle_id": bundle.pk},
        )
        return redirect(detail_url)

    def _context(self, shell, bundle, errors=None, form=None):
        return {
            "mc_base": _mc_base_for_shell(shell),
            "shell": shell,
            "bundle": bundle,
            "errors": errors or [],
            "form": form or {},
            "page_title": f"Attach source — {bundle.label or bundle.idempotency_key}",
        }


class MigrationCloudBindSchoolView(LoginRequiredMixin, View):
    """Bind a pre-tenant bundle to a School after the tenant is provisioned.

    Super-shell only — a tenant cannot bind a bundle staged before its
    own provisioning (that would be a privilege boundary violation).
    """

    template_name = "migration_cloud/bind_school.html"

    def get(self, request, bundle_id: int, shell: str = "super"):
        if shell != "super":
            return JsonResponse({"error": "Operator-only action."}, status=403)
        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        return render(request, self.template_name, self._context(shell, bundle, errors=None))

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        from django.contrib import messages
        from django.shortcuts import redirect
        from django.urls import reverse

        if shell != "super":
            return JsonResponse({"error": "Operator-only action."}, status=403)
        bundle = _tenant_scoped_bundle(request, bundle_id, shell)

        raw = (request.POST.get("school_id") or "").strip()
        errors: list[str] = []
        if not raw:
            errors.append("Provide the target school ID.")
        try:
            school_id = int(raw)
        except ValueError:
            errors.append("school_id must be an integer.")
            school_id = 0

        if not errors:
            try:
                from apps.schools.models import School

                school = School.objects.get(pk=school_id)
            except Exception:  # noqa: BLE001
                errors.append(f"No School row with id={school_id}.")
                school = None
            else:
                bundle.school = school
                bundle.schema_name = resolve_school_schema_name(school) or bundle.schema_name
                bundle.save(update_fields=["school", "schema_name", "updated_at"])
                messages.success(
                    request,
                    f"Bundle bound to school #{school.pk} ({getattr(school, 'name', '?')}).",
                )

        if errors:
            return render(
                request, self.template_name,
                self._context(shell, bundle, errors=errors),
                status=400,
            )

        detail_url = reverse(
            "migration_cloud_super:bundle_detail",
            kwargs={"bundle_id": bundle.pk},
        )
        return redirect(detail_url)

    def _context(self, shell, bundle, errors=None):
        return {
            "mc_base": _mc_base_for_shell(shell),
            "shell": shell,
            "bundle": bundle,
            "errors": errors or [],
            "page_title": f"Bind to school — {bundle.label or bundle.idempotency_key}",
        }


class MigrationCloudAIPlanView(LoginRequiredMixin, View):
    """GET endpoint: AI-generated migration plan for the bundle.

    Returns a structured plan based on the bundle's discovery + mapping
    summary so the operator can preview "what will happen" before
    applying. Renders a partial via HTMX if requested, full JSON otherwise.
    """

    def get(self, request, bundle_id: int, shell: str = "super"):
        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        from .ai_bridge import generate_migration_plan

        plan = generate_migration_plan(bundle=bundle)
        return JsonResponse({"bundle_id": bundle.pk, "plan": plan})


class MigrationCloudAIExplainView(LoginRequiredMixin, View):
    """POST endpoint: AI explains a quarantine row in plain language."""

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json

        from .ai_bridge import explain_quarantine_row

        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "invalid JSON"}, status=400)

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        row = payload.get("row") or {}
        reason = (payload.get("reason") or "").strip()
        if not isinstance(row, dict) and not isinstance(row, list):
            return JsonResponse({"error": "row must be an object or array"}, status=400)
        if not reason:
            return JsonResponse({"error": "reason required"}, status=400)

        explanation = explain_quarantine_row(school=bundle.school, row=row, reason=reason)
        return JsonResponse({
            "bundle_id": bundle.pk,
            "explanation": explanation.answer if explanation else None,
            "confidence": explanation.confidence if explanation else 0.0,
            "ai_available": explanation is not None,
        })


class MigrationCloudAIRebindView(LoginRequiredMixin, View):
    """POST endpoint: parse a natural-language mapping command and apply it.

    Operator types ``"set Student_Number as student.external_id"`` in
    the UI textbox. We parse it (regex first, AI fallback), write the
    edit into ``bundle.mapping_summary["per_artifact"]``, and record an
    operator-correction signal through the existing feedback path so the
    decision feeds AIEmbeddingStore for next-bundle recall.
    """

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json

        from .ai_bridge import parse_mapping_command, remember_mapping_decision

        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "invalid JSON"}, status=400)

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        command = (payload.get("command") or "").strip()
        if not command:
            return JsonResponse({"error": "command required"}, status=400)

        per_artifact = (bundle.mapping_summary or {}).get("per_artifact") or {}
        available_source_columns: list[str] = []
        available_canonical_fields: set[str] = set()
        for mappings in per_artifact.values():
            for m in mappings or []:
                if m.get("source_column"):
                    available_source_columns.append(str(m["source_column"]))
                if m.get("canonical_field"):
                    available_canonical_fields.add(str(m["canonical_field"]))

        proposal = parse_mapping_command(
            school=bundle.school,
            command=command,
            available_source_columns=available_source_columns,
            available_canonical_fields=sorted(available_canonical_fields),
        )
        if proposal is None or not isinstance(proposal.answer, dict):
            return JsonResponse({
                "bundle_id": bundle.pk,
                "applied": False,
                "ai_available": False,
                "reason": "Could not parse the command (or AI unavailable). Try the "
                          "drag-and-drop UI, or rephrase as 'set <column> as <canonical_field>'.",
            }, status=422)

        src = proposal.answer.get("source_column") or ""
        canon = proposal.answer.get("canonical_field") or ""

        # Apply the edit in-place.
        updated_artifacts: list[str] = []
        for path, mappings in per_artifact.items():
            for m in mappings or []:
                if str(m.get("source_column") or "") == src:
                    m["canonical_field"] = canon
                    m["confidence"] = max(float(m.get("confidence") or 0.0), 0.95)
                    m["method"] = "operator_command"
                    m["reasoning"] = proposal.reasoning
                    updated_artifacts.append(path)
        if not updated_artifacts:
            return JsonResponse({
                "bundle_id": bundle.pk,
                "applied": False,
                "ai_available": True,
                "parsed": {"source_column": src, "canonical_field": canon},
                "reason": f"Source column {src!r} not found in any artifact's mapping.",
            }, status=404)

        bundle.mapping_summary = {
            **(bundle.mapping_summary or {}),
            "per_artifact": per_artifact,
        }
        bundle.save(update_fields=["mapping_summary", "updated_at"])

        # Feed the embedding-store so the next bundle for this tenant
        # recalls this decision without re-asking the AI.
        try:
            source_system = ((bundle.discovery_summary or {}).get("source") or {}).get("chosen") or ""
            remember_mapping_decision(
                school=bundle.school,
                source_column=src,
                sample_values=[],
                canonical_field=canon,
                domain="",
                confidence=0.95,
                method="operator_command",
                transformer=None,
                source_system=source_system,
            )
        except Exception:  # noqa: BLE001 — recall persistence is best-effort
            logger.debug("ai_rebind: remember_mapping_decision failed", exc_info=True)

        return JsonResponse({
            "bundle_id": bundle.pk,
            "applied": True,
            "ai_available": True,
            "parsed": {"source_column": src, "canonical_field": canon},
            "artifacts_updated": updated_artifacts,
            "confidence": proposal.confidence,
            "method": proposal.provider_meta.get("method", "ai_bridge"),
            "reasoning": proposal.reasoning,
        })


class MigrationCloudIntakeAIAskView(LoginRequiredMixin, View):
    """POST: conversational help on the intake wizard (pre-bundle)."""

    @safe_500
    def post(self, request, shell: str = "super"):
        import json

        from .ai_bridge import answer_intake_question

        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "invalid JSON"}, status=400)
        question = (payload.get("question") or "").strip()
        if not question:
            return JsonResponse({"error": "question required"}, status=400)
        if len(question) > 500:
            return JsonResponse({"error": "question too long (max 500 chars)"}, status=400)
        school = getattr(request, "school", None)
        ctx = {
            "intake_method": (payload.get("intake_method") or "")[:64],
            "vendor": (payload.get("vendor") or "")[:64],
            "screen": "intake_new",
        }
        proposal = answer_intake_question(
            school=school,
            question=question,
            intake_context=ctx,
        )
        return JsonResponse({
            "question": question,
            "answer": proposal.answer if proposal else None,
            "confidence": proposal.confidence if proposal else 0.0,
            "ai_available": proposal is not None,
        })


class MigrationCloudAIAskView(LoginRequiredMixin, View):
    """POST endpoint: AI Q&A grounded in the bundle's profile + classification.

    Structural questions only — counts / breakdowns / classifier guesses /
    column shapes. Does NOT load raw artifact bytes; the apply path is
    the right surface for full-row computation.
    """

    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json

        from .ai_bridge import answer_bundle_question

        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "invalid JSON"}, status=400)

        question = (payload.get("question") or "").strip()
        if not question:
            return JsonResponse({"error": "question required"}, status=400)
        if len(question) > 500:
            return JsonResponse({"error": "question too long (max 500 chars)"}, status=400)

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        proposal = answer_bundle_question(
            school=bundle.school, bundle=bundle, question=question,
        )
        return JsonResponse({
            "bundle_id": bundle.pk,
            "question": question,
            "answer": proposal.answer if proposal else None,
            "confidence": proposal.confidence if proposal else 0.0,
            "ai_available": proposal is not None,
        })


class MigrationCloudAINarrateReconciliationView(LoginRequiredMixin, View):
    """GET endpoint: AI-generated school-facing summary of the reconciliation report."""

    def get(self, request, bundle_id: int, shell: str = "super"):
        from .ai_bridge import narrate_reconciliation

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        report = bundle.reconciliation_summary or {}
        if not report:
            return JsonResponse({
                "bundle_id": bundle.pk,
                "narrative": None,
                "ai_available": False,
                "reason": "Reconcile the bundle first; no report to summarise.",
            }, status=409)

        proposal = narrate_reconciliation(school=bundle.school, reconciliation_summary=report)
        return JsonResponse({
            "bundle_id": bundle.pk,
            "narrative": proposal.answer if proposal else None,
            "confidence": proposal.confidence if proposal else 0.0,
            "overall_parity_pct": report.get("overall_parity_pct"),
            "ai_available": proposal is not None,
        })


class MigrationCloudAIVendorFromImageView(LoginRequiredMixin, View):
    """POST endpoint: identify the source vendor from an uploaded screenshot / PDF page.

    Multipart upload of one image (PNG / JPG / PDF). We OCR locally then
    pass the extracted text to the AI gateway with the same allow-list
    the classifier uses. Returns a vendor guess + confidence; the operator
    can then pre-fill the source_hint on the intake wizard.
    """

    MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MiB  # magic-number-allow: byte-size-cap
    ALLOWED_SUFFIXES = (".png", ".jpg", ".jpeg", ".pdf", ".gif", ".bmp", ".tiff", ".webp")

    @idempotent_post
    @safe_500
    def post(self, request, shell: str = "super"):
        from .ai_bridge import identify_vendor_from_image
        from .classifiers.signatures import SOURCE_HEADER_SIGNATURES

        upload = request.FILES.get("image")
        if upload is None:
            return JsonResponse({"error": "image field required (multipart upload)"}, status=400)
        if upload.size > self.MAX_UPLOAD_BYTES:
            return JsonResponse({
                "error": f"image too large ({upload.size:,} bytes; max {self.MAX_UPLOAD_BYTES:,})",
            }, status=413)
        name_lower = (upload.name or "").lower()
        if not name_lower.endswith(self.ALLOWED_SUFFIXES):
            return JsonResponse({
                "error": f"unsupported file type; allowed: {self.ALLOWED_SUFFIXES}",
            }, status=415)

        # A-4: route through the shared upload validator (magic-byte sniff +
        # size cap + malware-scan hook). The declared name / content-type are
        # ignored, so an executable renamed .png is caught by content here —
        # a stricter gate than the suffix check above, which is kept as a
        # cheap first pass. Cursor is restored to 0, so the read below works.
        from apps.security.upload_validation import (
            RASTER_IMAGE_MIMES,
            UploadValidationError,
            validate_uploaded_file,
        )
        try:
            validate_uploaded_file(
                upload,
                allowed_mimes=RASTER_IMAGE_MIMES | {"application/pdf"},
                max_bytes=self.MAX_UPLOAD_BYTES,
            )
        except UploadValidationError as exc:
            return JsonResponse({"error": str(exc)}, status=415)

        # School context for AI policy — portal shell binds request.school,
        # super shell can hit this without one (vendor ID is school-agnostic).
        school = getattr(request, "school", None) or getattr(request, "tenant", None)

        # Read bytes with bounded memory. UploadedFile.read() respects the
        # configured FILE_UPLOAD_MAX_MEMORY_SIZE; for larger files it
        # comes off disk via chunks. Either way the read is one-shot here.
        try:
            image_bytes = upload.read()
        except Exception as exc:  # noqa: BLE001
            return JsonResponse({"error": f"upload read failed: {exc}"}, status=400)

        proposal = identify_vendor_from_image(
            school=school,
            image_bytes=image_bytes,
            image_filename=upload.name or "upload.png",
            known_sources=sorted(SOURCE_HEADER_SIGNATURES.keys()),
        )
        from .tier3 import ocr_confidence_warning

        ocr_chars = proposal.provider_meta.get("ocr_chars") if proposal else 0
        confidence = proposal.confidence if proposal else 0.0
        warning = ocr_confidence_warning(ocr_chars=int(ocr_chars or 0), vendor_confidence=confidence)
        return JsonResponse({
            "vendor": proposal.answer if proposal else None,
            "confidence": confidence,
            "reasoning": proposal.reasoning if proposal else None,
            "ocr_chars": ocr_chars,
            "ai_available": proposal is not None,
            "warning": warning,
        })


# --- Tier 1 / Tier 2 / Tier 3 views (sms-v3.7) -----------------------------

class MigrationCloudExpectedTotalsView(LoginRequiredMixin, View):
    """GET/POST endpoint: operator sets financial control totals on a bundle.

    POST body::

        {"finance.invoice_total_amount": "125000.00", "students.count": 1240}

    Stored on ``bundle.expected_totals``; enforced by the orchestrator
    after the apply step lands rows.
    """

    def get(self, request, bundle_id: int, shell: str = "super"):
        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        return JsonResponse({
            "bundle_id": bundle.pk,
            "expected_totals": bundle.expected_totals or {},
            "financial_guardrail": (bundle.mapping_summary or {}).get("financial_guardrail"),
        })

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "invalid JSON"}, status=400)
        if not isinstance(payload, dict):
            return JsonResponse({"error": "payload must be an object"}, status=400)
        cleaned: dict[str, str] = {}
        for key, value in payload.items():
            if not isinstance(key, str) or value in (None, ""):
                continue
            cleaned[key[:64]] = str(value)[:64]
        bundle.expected_totals = cleaned
        bundle.save(update_fields=["expected_totals", "updated_at"])
        return JsonResponse({"bundle_id": bundle.pk, "expected_totals": cleaned})


class MigrationCloudGuardrailCheckView(LoginRequiredMixin, View):
    """POST endpoint: run the financial guardrail manually (read-only check).

    Returns the comparison report without changing bundle status. Use
    before flipping APPLIED to preview whether the guardrail will pass.
    """

    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        from .guardrails import compute_observed_totals, evaluate_expected_totals

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        observed = compute_observed_totals(bundle=bundle)
        report = evaluate_expected_totals(bundle=bundle, observed=observed)
        return JsonResponse(report.to_dict())


class MigrationCloudIdMappingLookupView(LoginRequiredMixin, View):
    """GET endpoint: answer "what's the new ID for old ID X?" for an operator.

    Query string: ``?legacy_id=PS-1029`` (optional ``?namespace=powerschool``).
    Tenant-scoped — only mappings for the operator's school are returned in
    portal shell; super shell sees all.
    """

    def get(self, request, shell: str = "super"):
        legacy_id = (request.GET.get("legacy_id") or "").strip()
        namespace = (request.GET.get("namespace") or "").strip()
        if not legacy_id:
            return JsonResponse({"error": "legacy_id required"}, status=400)
        # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
        qs = MigrationIdMapping.objects.filter(legacy_id=legacy_id)
        if namespace:
            qs = qs.filter(legacy_namespace=namespace)
        if shell == "portal":
            school = getattr(request, "school", None) or getattr(request, "tenant", None)
            school_pk = getattr(school, "pk", None)
            if school_pk is None:
                # Fail closed: an unresolved tenant must NOT see cross-tenant
                # id-mappings. Mirrors `_tenant_scoped_bundle` (which 404s on a
                # None school) — return an empty match set rather than leaking
                # every school's legacy→canonical id map.
                return JsonResponse({"legacy_id": legacy_id, "matches": []})
            qs = qs.filter(school_id=school_pk)
        return JsonResponse({
            "legacy_id": legacy_id,
            "matches": [
                {
                    "namespace": m.legacy_namespace,
                    "canonical_model": m.canonical_model,
                    "canonical_pk": m.canonical_pk,
                    "domain": m.domain,
                    "bundle_id": m.bundle_id,
                    "school_id": m.school_id,
                    "created_at": m.created_at.isoformat(),
                }
                for m in qs[:25]
            ],
        })


def _can_reveal_pii(user) -> bool:
    """Only a platform superuser may see staged PII verbatim in the review UI.

    Everyone else — tenant admins and operators included — sees masked values, so
    raw SSN / DOB / medical / financial data never reaches an unprivileged
    browser (screen-share, shoulder-surf, browser cache). A granular per-tenant
    ``reveal_pii`` permission is the documented next refinement.
    """
    return bool(getattr(user, "is_superuser", False))


class MigrationCloudConflictsView(LoginRequiredMixin, View):
    """GET endpoint: list pending conflicts. POST endpoint: resolve a conflict.

    POST body::

        {"conflict_id": 42, "resolution": "OVERWRITE" | "PRESERVE" | "MERGE"}
        {"action": "fix_all", "resolution": "OVERWRITE" | "PRESERVE" | "MERGE"}
    """

    template_name = "migration_cloud/conflicts.html"
    _CONFLICTS_PAGE_SIZE = 20

    def get(self, request, bundle_id: int, shell: str = "super"):
        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        pending_qs = bundle.conflicts.filter(resolution=ConflictResolution.PENDING).order_by(
            "-created_at"
        )
        resolved_qs = bundle.conflicts.exclude(
            resolution=ConflictResolution.PENDING
        ).order_by("-resolved_at")
        pending_page_obj = Paginator(pending_qs, self._CONFLICTS_PAGE_SIZE).get_page(
            request.GET.get("page") or 1
        )
        resolved_page_obj = Paginator(resolved_qs, self._CONFLICTS_PAGE_SIZE).get_page(
            request.GET.get("resolved_page") or 1
        )
        pending = list(pending_page_obj.object_list)
        # PII masking: staged conflict rows can hold SSN / DOB / medical / financial
        # values. Mask them IN MEMORY (never saved) for any viewer without an
        # explicit reveal right, so raw PII never leaves the server for the review
        # surface. Applied here so it covers BOTH the JSON and HTML renders below.
        reveal_pii = _can_reveal_pii(request.user)
        if not reveal_pii:
            from .pii_display import mask_dict
            for _c in pending:
                _c.existing_values = mask_dict(_c.existing_values)
                _c.incoming_values = mask_dict(_c.incoming_values)
        pending_count = pending_qs.count()
        resolved = list(resolved_page_obj.object_list)
        pending_extra = request.GET.copy()
        pending_extra.pop("page", None)
        resolved_extra = request.GET.copy()
        resolved_extra.pop("resolved_page", None)
        if request.GET.get("format") == "json":
            return JsonResponse({
                "bundle_id": bundle.pk,
                "pending_count": pending_count,
                "page": pending_page_obj.number,
                "num_pages": pending_page_obj.paginator.num_pages,
                "pending": [
                    {
                        "id": c.pk,
                        "domain": c.domain,
                        "canonical_model": c.canonical_model,
                        "canonical_pk": c.canonical_pk,
                        "legacy_id": c.legacy_id,
                        "changed_fields": c.changed_fields,
                        "existing_values": c.existing_values,
                        "incoming_values": c.incoming_values,
                    } for c in pending
                ],
            })
        return render(
            request, self.template_name,
            {
                "mc_base": _mc_base_for_shell(shell),
                "shell": shell,
                "bundle": bundle,
                "pending": pending,
                "pending_count": pending_count,
                "page_obj": pending_page_obj,
                "pagination_extra_query": pending_extra.urlencode(),
                "resolved": resolved,
                "resolved_page_obj": resolved_page_obj,
                "resolved_pagination_extra_query": resolved_extra.urlencode(),
                "can_reveal_pii": reveal_pii,
                "page_title": f"Conflict review — {bundle.label or bundle.idempotency_key}",
            },
        )

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json
        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            payload = request.POST.dict()
        from django.http import Http404

        resolution = (payload.get("resolution") or "").upper()
        if resolution not in {c[0] for c in ConflictResolution.choices}:
            return JsonResponse({"error": "invalid resolution"}, status=400)
        # Scope by tenant FIRST (both the batch and single paths). The GET
        # resolves via _tenant_scoped_bundle; the POST must too, then act only
        # WITHIN that bundle — else a portal caller could flip another tenant's
        # conflict resolution (which drives whether apply overwrites/preserves
        # rows). MigrationConflict has no school_id, so the view-layer bundle
        # scope is the only guard here.
        try:
            bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        except Http404:
            return JsonResponse({"error": "bundle not found"}, status=404)

        # One-click "Fix All": resolve EVERY pending conflict for this bundle
        # with the chosen resolution in a single action. Bundle-scoped, so it can
        # never touch another tenant's queue.
        if (payload.get("action") or "").lower() == "fix_all":
            resolver = request.user if request.user.is_authenticated else None
            resolved_count = bundle.conflicts.filter(
                resolution=ConflictResolution.PENDING
            ).update(
                resolution=resolution,
                resolved_by=resolver,
                resolved_at=timezone.now(),
            )
            return JsonResponse(
                {"action": "fix_all", "resolution": resolution, "resolved_count": resolved_count}
            )

        conflict_id = payload.get("conflict_id")
        try:
            conflict_pk = int(conflict_id)
        except (TypeError, ValueError):
            return JsonResponse({"error": "conflict_id required"}, status=400)
        conflict = get_object_or_404(MigrationConflict, pk=conflict_pk, bundle=bundle)
        conflict.resolution = resolution
        conflict.resolved_by = request.user if request.user.is_authenticated else None
        conflict.resolved_at = timezone.now()
        conflict.save(update_fields=["resolution", "resolved_by", "resolved_at"])
        return JsonResponse({"conflict_id": conflict.pk, "resolution": resolution})


class MigrationCloudProgressView(LoginRequiredMixin, View):
    """GET endpoint: live DAG-style progress snapshot for a bundle.

    Returns the per-stage breakdown the UI renders as a timeline.

    Reliability contract (v3.17): wrapped with :func:`with_progress_fallback`
    so any snapshot-computation failure renders a degraded surface instead of
    a 500. The operator always sees their bundle's identity + a request_id
    + remediation copy, never a stack trace.
    """

    template_name = "migration_cloud/progress.html"

    @with_progress_fallback
    def get(self, request, bundle_id: int, shell: str = "super"):
        from .progress import refresh_snapshot

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        snapshot = refresh_snapshot(bundle=bundle)
        recent_events = list(
            bundle.progress_events.order_by("-created_at").values(
                "id", "kind", "stage", "message", "detail", "created_at",
            )[:50]
        )
        if request.GET.get("format") == "json":
            return JsonResponse({
                "bundle_id": bundle.pk,
                "snapshot": snapshot,
                "recent_events": [
                    {**ev, "created_at": ev["created_at"].isoformat()}
                    for ev in recent_events
                ],
            })
        return render(
            request, self.template_name,
            {
                "mc_base": _mc_base_for_shell(shell),
                "shell": shell,
                "bundle": bundle,
                "snapshot": snapshot,
                "recent_events": recent_events,
                "page_title": f"Progress — {bundle.label or bundle.idempotency_key}",
            },
        )


class MigrationCloudProgressStreamView(LoginRequiredMixin, View):
    """GET SSE endpoint: streams MigrationProgressEvent rows as they arrive.

    Real-time progress so the operator doesn't need to refresh. Client
    sends ``?after_id=<n>`` to resume; we send up to 500 events per response
    and let the client reconnect (browsers retry SSE automatically).
    """

    def get(self, request, bundle_id: int, shell: str = "super"):
        from .progress import stream_events_since

        try:
            after_id = int(request.GET.get("after_id") or 0)
        except ValueError:
            after_id = 0
        _tenant_scoped_bundle(request, bundle_id, shell)

        def _events():
            yield ": connected\n\n"
            for event_id, payload in stream_events_since(bundle_id=bundle_id, after_id=after_id):
                import json as _json
                yield f"id: {event_id}\ndata: {_json.dumps(payload, default=str)}\n\n"

        response = StreamingHttpResponse(_events(), content_type="text/event-stream")
        response["Cache-Control"] = "no-cache"
        response["X-Accel-Buffering"] = "no"
        return response


class MigrationCloudPreflightView(LoginRequiredMixin, View):
    """POST endpoint: run the pre-flight gate before APPLYING.

    Mutates ``bundle.size_summary['preflight']`` so the report sticks
    between requests; hence ``@idempotent_post`` even though the response
    is shaped like a read-only check.
    """

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        from .preflight import run_all

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        report = run_all(bundle=bundle)
        summary = dict(bundle.size_summary or {})
        summary["preflight"] = report.to_dict()
        bundle.size_summary = summary
        bundle.save(update_fields=["size_summary", "updated_at"])
        return JsonResponse(report.to_dict())


class MigrationCloudAssetsView(LoginRequiredMixin, View):
    """GET endpoint: list assets for a bundle. POST endpoint: trigger fetch worker."""

    template_name = "migration_cloud/assets.html"

    def get(self, request, bundle_id: int, shell: str = "super"):
        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        assets = bundle.assets.order_by("-created_at")[:500]
        counts = {
            status: bundle.assets.filter(status=status).count()
            for status, _ in AssetStatus.choices
        }
        if request.GET.get("format") == "json":
            return JsonResponse({
                "bundle_id": bundle.pk,
                "counts": counts,
                "assets": [
                    {
                        "id": a.pk,
                        "entity_kind": a.entity_kind,
                        "legacy_id": a.legacy_id,
                        "asset_kind": a.asset_kind,
                        "status": a.status,
                        "stored_path": a.stored_path,
                        "byte_size": a.byte_size,
                        "sha256": a.sha256,
                    } for a in assets
                ],
            })
        return render(
            request, self.template_name,
            {
                "mc_base": _mc_base_for_shell(shell),
                "shell": shell,
                "bundle": bundle,
                "assets": assets,
                "counts": counts,
                "page_title": f"Assets — {bundle.label or bundle.idempotency_key}",
            },
        )

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        from .celery_tasks import enqueue_fetch_assets
        from .asset_pipeline import fetch_pending_assets

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        async_result = enqueue_fetch_assets(bundle.pk, max_batch=int(request.POST.get("max_batch") or 100))
        if async_result is None:
            counts = fetch_pending_assets(bundle_id=bundle.pk)
            return JsonResponse({"bundle_id": bundle.pk, "mode": "inline", **counts})
        return JsonResponse({"bundle_id": bundle.pk, "mode": "queued"})


class MigrationCloudSandboxView(LoginRequiredMixin, View):
    """POST endpoint: clone / promote / discard sandbox bundles.

    ``?action=clone``    — clone the bundle into a sandbox
    ``?action=promote``  — promote a sandbox back to its origin
    ``?action=discard``  — discard a sandbox
    """

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        from .sandbox import clone_bundle_to_sandbox, discard_sandbox, promote_sandbox_to_origin

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        action = (request.GET.get("action") or "clone").lower()
        try:
            if action == "clone":
                clone = clone_bundle_to_sandbox(bundle=bundle)
                return JsonResponse({"sandbox_bundle_id": clone.pk, "schema_name": clone.schema_name})
            if action == "promote":
                return JsonResponse(promote_sandbox_to_origin(sandbox=bundle))
            if action == "discard":
                return JsonResponse(discard_sandbox(sandbox=bundle))
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=409)
        return JsonResponse({"error": f"unknown action {action!r}"}, status=400)


class MigrationCloudDiffModeView(LoginRequiredMixin, View):
    """POST endpoint: configure diff-mode re-ingest for a bundle.

    Body::
        {"diff_mode": "since", "diff_since": "2026-04-01T00:00:00"}
    """

    def get(self, request, bundle_id: int, shell: str = "super"):
        from .diff_mode import recommended_diff_since

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        source = (bundle.discovery_summary or {}).get("source", {}).get("chosen") or ""
        suggested = recommended_diff_since(school_id=bundle.school_id, source_system=source)
        return JsonResponse({
            "bundle_id": bundle.pk,
            "diff_mode": bundle.diff_mode,
            "diff_since": bundle.diff_since.isoformat() if bundle.diff_since else None,
            "suggested_diff_since": suggested.isoformat() if suggested else None,
        })

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json
        from datetime import datetime

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "invalid JSON"}, status=400)
        mode = (payload.get("diff_mode") or "full").lower()
        if mode not in ("full", "since"):
            return JsonResponse({"error": "diff_mode must be 'full' or 'since'"}, status=400)
        bundle.diff_mode = mode
        since_raw = payload.get("diff_since")
        if mode == "since" and since_raw:
            try:
                bundle.diff_since = datetime.fromisoformat(str(since_raw).replace("Z", "+00:00"))
            except ValueError:
                return JsonResponse({"error": "diff_since must be ISO-8601"}, status=400)
        elif mode == "full":
            bundle.diff_since = None
        bundle.save(update_fields=["diff_mode", "diff_since", "updated_at"])
        return JsonResponse({
            "bundle_id": bundle.pk,
            "diff_mode": bundle.diff_mode,
            "diff_since": bundle.diff_since.isoformat() if bundle.diff_since else None,
        })


class MigrationCloudBundleSettingsView(LoginRequiredMixin, View):
    """POST endpoint: flip the apply_atomic + parity_drift_rollback_pct flags."""

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "invalid JSON"}, status=400)
        if "apply_atomic" in payload:
            bundle.apply_atomic = bool(payload["apply_atomic"])
        if "parity_drift_rollback_pct" in payload:
            try:
                bundle.parity_drift_rollback_pct = float(payload["parity_drift_rollback_pct"])
            except (TypeError, ValueError):
                return JsonResponse({"error": "parity_drift_rollback_pct must be a number"}, status=400)
        bundle.save(update_fields=[
            "apply_atomic", "parity_drift_rollback_pct", "updated_at",
        ])
        return JsonResponse({
            "bundle_id": bundle.pk,
            "apply_atomic": bundle.apply_atomic,
            "parity_drift_rollback_pct": bundle.parity_drift_rollback_pct,
        })


class MigrationCloudCostEstimateView(LoginRequiredMixin, View):
    """GET endpoint: pre-flight AI token spend estimate."""

    def get(self, request, bundle_id: int, shell: str = "super"):
        from .tier3 import estimate_token_spend

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        est = estimate_token_spend(bundle=bundle)
        return JsonResponse({
            "bundle_id": bundle.pk,
            "artifact_count": est.artifact_count,
            "column_count_total": est.column_count_total,
            "estimated_ai_calls": est.estimated_ai_calls,
            "estimated_tokens": est.estimated_tokens,
            "estimated_usd": est.estimated_usd,
        })


class MigrationCloudProfileSuggestView(LoginRequiredMixin, View):
    """GET endpoint: cross-tenant MigrationProfile suggestions for a source.

    Query string: ``?source_system=powerschool&domain=students``.
    """

    def get(self, request, shell: str = "super"):
        from .tier3 import suggest_profiles_for

        source_system = (request.GET.get("source_system") or "").strip()
        domain = (request.GET.get("domain") or "").strip()
        return JsonResponse({
            "source_system": source_system,
            "domain": domain,
            "profiles": suggest_profiles_for(source_system=source_system, domain=domain),
        })


class MigrationCloudHandoffDocView(LoginRequiredMixin, View):
    """GET endpoint: auto-generated migration receipt for the school's IT lead."""

    template_name = "migration_cloud/handoff_doc.html"

    def get(self, request, bundle_id: int, shell: str = "super"):
        from .tier3 import generate_handoff_doc

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        doc = generate_handoff_doc(bundle=bundle)
        if request.GET.get("format") == "json":
            return JsonResponse(doc)
        return render(
            request, self.template_name,
            {
                "mc_base": _mc_base_for_shell(shell),
                "shell": shell,
                "bundle": bundle,
                "doc": doc,
                "page_title": doc.get("title"),
            },
        )


class MigrationCloudLegacyLockoutView(LoginRequiredMixin, View):
    """POST endpoint: confirm the operator has flipped the legacy SIS to read-only."""

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json
        from .tier3 import lockout_legacy_source

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            payload = {}
        notes = str(payload.get("instructions") or "")
        return JsonResponse({"bundle_id": bundle.pk, "lockout": lockout_legacy_source(bundle=bundle, instructions=notes)})


class MigrationCloudExportCanonicalView(LoginRequiredMixin, View):
    """GET endpoint: tenant data export for the no-lock-in promise.

    Returns a zip of canonical CSVs (one per domain). Tenant-scoped via the
    operator's request.school in portal shell.
    """

    def get(self, request, shell: str = "super"):
        import io
        import zipfile

        from .tier3 import export_tenant_to_canonical

        school = getattr(request, "school", None) or getattr(request, "tenant", None)
        if school is None and shell == "portal":
            return JsonResponse({"error": "tenant required"}, status=400)
        if school is None:
            try:
                from apps.schools.models import School
                school_id = int(request.GET.get("school_id") or 0)
                school = School.objects.get(pk=school_id) if school_id else None
            except Exception:  # noqa: BLE001
                return JsonResponse({"error": "school_id required for operator shell"}, status=400)
        if school is None:
            return JsonResponse({"error": "no school resolved"}, status=400)
        dumps = export_tenant_to_canonical(school=school)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for domain, csv_text in dumps.items():
                zf.writestr(f"{domain}.csv", csv_text)
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type="application/zip")
        resp["Content-Disposition"] = f'attachment; filename="canonical-export-{school.pk}.zip"'
        return resp


class MigrationCloudRolloutPlanView(LoginRequiredMixin, View):
    """POST endpoint: set or advance a multi-stage rollout plan."""

    @idempotent_post
    @safe_500
    def post(self, request, bundle_id: int, shell: str = "super"):
        import json
        from .tier3 import advance_rollout_stage, stage_rollout_plan

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        action = (request.GET.get("action") or "set").lower()
        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            payload = {}
        try:
            if action == "set":
                stages = payload.get("stages") or []
                if not isinstance(stages, list) or not stages:
                    return JsonResponse({"error": "stages list required"}, status=400)
                return JsonResponse(stage_rollout_plan(bundle=bundle, stages=stages))
            if action == "advance":
                return JsonResponse(advance_rollout_stage(bundle=bundle))
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        return JsonResponse({"error": f"unknown action {action!r}"}, status=400)


class MigrationCloudSlaTargetsView(LoginRequiredMixin, View):
    """GET endpoint: SLA tier targets + elapsed time + escalate flag."""

    def get(self, request, bundle_id: int, shell: str = "super"):
        from .tier3 import sla_tier_targets

        bundle = _tenant_scoped_bundle(request, bundle_id, shell)
        return JsonResponse({"bundle_id": bundle.pk, **sla_tier_targets(bundle=bundle)})


class MigrationCloudMergeBundlesView(LoginRequiredMixin, View):
    """POST endpoint: merge N bundles into a single parent bundle for joint apply."""

    @idempotent_post
    @safe_500
    def post(self, request, shell: str = "super"):
        import json
        from .tier3 import merge_bundles

        try:
            payload = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "invalid JSON"}, status=400)
        bundle_ids = payload.get("bundle_ids") or []
        if not isinstance(bundle_ids, list) or len(bundle_ids) < 2:
            return JsonResponse({"error": "bundle_ids must be a list of at least 2 IDs"}, status=400)
        # Tenant-scope the merge set. In the portal shell a caller may only merge
        # bundles owned by their OWN school; without this a tenant could POST
        # another school's bundle ids and fold them into a joint parent for a
        # cross-tenant apply (IDOR). Operator (super) shell is control-plane-gated
        # by the URL mount and may span tenants, matching `_tenant_scoped_bundle`.
        # tenant-isolation-allow: operator-shell-spans-tenants-portal-adds-school-filter-below
        merge_qs = MigrationBundle.objects.filter(pk__in=bundle_ids)
        if shell == "portal":
            school = getattr(request, "school", None) or getattr(request, "tenant", None)
            school_pk = getattr(school, "pk", None)
            if school_pk is None:
                return JsonResponse({"error": "one or more bundles not found"}, status=404)
            merge_qs = merge_qs.filter(school_id=school_pk)
        bundles = list(merge_qs)
        if len(bundles) != len(bundle_ids):
            return JsonResponse({"error": "one or more bundles not found"}, status=404)
        parent = merge_bundles(bundles=bundles, label=str(payload.get("label") or ""))
        return JsonResponse({"merged_bundle_id": parent.pk, "merged_from": bundle_ids})


class MigrationCloudCanonicalTemplateView(LoginRequiredMixin, View):
    """GET endpoint: download the RunMyCampus canonical-template CSV(s).

    The "Shopify CSV import" front door for the long tail. Operators with
    data in Excel / Google Sheets / MS Access / in-house apps / any
    vendor not signature-matched download these empty templates, fill
    them in with what they have, and upload through the standard intake
    wizard. The accelerator at
    ``apps.migration_cloud.accelerators.runmycampus_canonical`` then
    short-circuits classification + mapping for the resulting bundle.

    Routes:
        GET /…/template/                — zip of every canonical-domain CSV (+ README)
        GET /…/template/?format=xlsx    — one Excel workbook covering every domain
        GET /…/template/<domain>.csv    — single-domain CSV (headers + version marker)
        GET /…/template/<domain>.xlsx   — single-domain Excel (Data + Instructions sheets)

    No tenant data is touched — this is a static template generator. Auth
    is required to gate against scraping the canonical schema.
    """

    _XLSX_CONTENT_TYPE = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    def get(self, request, domain: str | None = None, shell: str = "super", fmt: str = "csv"):
        from .accelerators.runmycampus_canonical import DOMAIN_CANONICAL_HEADERS

        # Format from the route (.xlsx suffix) or an explicit ?format= override.
        want_xlsx = fmt == "xlsx" or request.GET.get("format", "").lower() == "xlsx"

        if domain is not None:
            domain_key = domain.strip().lower()
            headers = DOMAIN_CANONICAL_HEADERS.get(domain_key)
            if headers is None:
                return JsonResponse(
                    {"error": f"unknown canonical domain {domain_key!r}",
                     "known_domains": sorted(DOMAIN_CANONICAL_HEADERS.keys())},
                    status=404,
                )
            if want_xlsx:
                resp = HttpResponse(
                    _canonical_template_xlsx(domain_key, headers),
                    content_type=self._XLSX_CONTENT_TYPE,
                )
                resp["Content-Disposition"] = f'attachment; filename="{domain_key}.xlsx"'
                return resp
            csv_text = _canonical_template_csv(domain_key, headers)
            resp = HttpResponse(csv_text, content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{domain_key}.csv"'
            return resp

        # No domain → the whole catalogue. XLSX = one workbook (Instructions +
        # a sheet per domain); CSV = a ZIP of one file per domain + a README.
        if want_xlsx:
            resp = HttpResponse(
                _canonical_template_workbook_all(),
                content_type=self._XLSX_CONTENT_TYPE,
            )
            resp["Content-Disposition"] = (
                'attachment; filename="runmycampus-canonical-template.xlsx"'
            )
            return resp

        import io
        import zipfile

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for domain_key, headers in sorted(DOMAIN_CANONICAL_HEADERS.items()):
                zf.writestr(f"{domain_key}.csv", _canonical_template_csv(domain_key, headers))
            zf.writestr("README.txt", _canonical_template_readme())
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type="application/zip")
        resp["Content-Disposition"] = 'attachment; filename="runmycampus-canonical-template.zip"'
        return resp


class MigrationCloudCanonicalTemplatePickerView(LoginRequiredMixin, View):
    """GET endpoint: full-page canonical-template picker.

    Surfaces every canonical domain with header count, required-field
    summary, and a 1-row sample so operators can preview what the
    template expects before downloading. Pairs with
    ``MigrationCloudCanonicalTemplateView`` (the actual CSV/zip
    download endpoint) — this view is pure UI discovery, no tenant data
    touched. Auth-gated to match the download endpoint.

    Routed at:
        GET /super/migration/template/picker/        (operator shell)
        GET /portal/configure/migration/template/picker/   (tenant shell)
    """

    template_name = "migration_cloud/canonical_template_picker.html"

    def get(self, request, shell: str = "super"):
        from .accelerators.runmycampus_canonical import DOMAIN_CANONICAL_HEADERS

        domains: list[dict[str, Any]] = []
        for slug, headers in sorted(DOMAIN_CANONICAL_HEADERS.items()):
            sorted_headers = sorted(headers)
            required = _canonical_required_fields(slug)
            sample_row = _canonical_sample_row(slug, sorted_headers)
            domains.append({
                "slug": slug,
                "headers": sorted_headers,
                "header_count": len(sorted_headers),
                "required": required,
                "sample_row": sample_row,
            })
        ctx = {
            "mc_base": _mc_base_for_shell(shell),
            "shell": shell,
            "domains": domains,
            "page_title": "Canonical template picker",
        }
        return render(request, self.template_name, ctx)


# Per-domain required-field summary — mirrors the README contract in
# ``_canonical_template_readme()`` so the picker UI and the zip's README
# stay in lockstep. Domains absent from this map have no strictly
# required fields (operator fills in what they have).
_CANONICAL_REQUIRED_FIELDS: dict[str, tuple[str, ...]] = {
    "students": ("external_id", "first_name", "last_name"),
    "staff": ("staff_external_id", "first_name", "last_name"),
    "guardians": ("guardian_external_id", "first_name", "last_name", "student_external_id"),
    "enrollment": ("student_external_id",),
    "attendance": ("student_external_id", "date"),
    "grades": ("student_external_id", "subject_code", "term"),
    "finance": ("reference",),
}


def _canonical_required_fields(domain: str) -> list[str]:
    return list(_CANONICAL_REQUIRED_FIELDS.get(domain, ()))


# Plausible sample values for canonical headers. Kept generic and
# culturally-neutral: a different given name per domain, ISO-format
# dates, three-letter currency, neutral subject/section codes. The point
# is to communicate the SHAPE of each row, not to model a specific
# school. Operators replace these with their own data.
_CANONICAL_SAMPLE_VALUES: dict[str, dict[str, str]] = {
    "students": {
        "external_id": "STD-001", "first_name": "Maria", "last_name": "Garcia",
        "middle_name": "Elena", "date_of_birth": "2010-04-12", "gender": "F",
        "email": "maria.garcia@example.edu", "phone": "+1-555-0101",
        "grade_level": "9", "enrollment_status": "active",
        "admission_number": "ADM-2026-001", "address": "12 Oak Lane",
    },
    "staff": {
        "staff_external_id": "STF-001", "first_name": "Aiden", "last_name": "Okonkwo",
        "email": "aiden.okonkwo@example.edu", "role": "teacher",
        "department": "Mathematics", "phone": "+1-555-0102",
    },
    "guardians": {
        "guardian_external_id": "GRD-001", "first_name": "Priya", "last_name": "Sharma",
        "email": "priya.sharma@example.com", "phone": "+1-555-0103",
        "relationship": "mother", "is_primary": "true",
        "student_external_id": "STD-001",
    },
    "enrollment": {
        "student_external_id": "STD-001", "grade_level": "9",
        "enrollment_status": "active", "enrollment_date": "2026-08-19",
        "exit_date": "", "section": "9A",
    },
    "sections": {
        "section_external_id": "SEC-9A-MATH", "subject_code": "MATH101",
        "subject_name": "Algebra I", "term": "fall", "academic_year": "2026-2027",
        "teacher_external_id": "STF-001",
    },
    "attendance": {
        "student_external_id": "STD-001", "date": "2026-08-19",
        "status": "present", "code": "P", "notes": "",
    },
    "grades": {
        "student_external_id": "STD-001", "subject_code": "MATH101",
        "term": "fall", "score": "87", "letter_grade": "B+", "comments": "Strong on linear systems",
    },
    "behavior": {
        "student_external_id": "STD-001", "date": "2026-09-04",
        "category": "tardy", "description": "Late to homeroom",
        "action_taken": "verbal reminder",
    },
    "finance": {
        "reference": "INV-2026-0001", "student_external_id": "STD-001",
        "amount": "1250.00", "currency": "USD", "issued_date": "2026-08-01",
        "due_date": "2026-08-31", "status": "open",
        "description": "Tuition — Fall 2026",
    },
    "transcripts": {
        "student_external_id": "STD-001", "academic_year": "2025-2026",
        "term": "spring", "subject_code": "ENG081", "final_grade": "A-",
        "credits_earned": "1.0",
    },
    "health": {
        "student_external_id": "STD-001", "record_date": "2026-08-15",
        "category": "immunization", "description": "Tdap booster",
        "provider": "City Health Clinic", "follow_up": "",
    },
    "payroll": {
        "staff_external_id": "STF-001", "pay_period": "2026-09",
        "gross_amount": "4500.00", "net_amount": "3420.00",
        "currency": "USD", "issued_date": "2026-09-30",
    },
    "communications": {
        "recipient_external_id": "GRD-001", "channel": "email",
        "subject": "Welcome to the new school year",
        "body": "We're delighted to have you back…",
        "sent_at": "2026-08-12T09:00:00Z", "status": "delivered",
    },
    "events": {
        "title": "Open House", "category": "community",
        "starts_at": "2026-09-12T17:00:00Z",
        "ends_at": "2026-09-12T19:00:00Z",
        "location": "Main Auditorium",
        "description": "Tour the campus and meet teachers",
    },
    "library": {
        "item_external_id": "LIB-00451", "title": "A Brief History of Time",
        "author": "Stephen Hawking", "isbn": "978-0553380163",
        "category": "non-fiction", "status": "available",
    },
    "transport": {
        "student_external_id": "STD-001", "route": "Route 4",
        "stop": "Oak Lane & Main", "pickup_time": "07:25",
        "dropoff_time": "15:40", "vehicle": "Bus 12",
    },
    "hostel": {
        "student_external_id": "STD-001", "room": "B-204", "bed": "1",
        "checkin_date": "2026-08-18", "checkout_date": "",
    },
    "cafeteria": {
        "student_external_id": "STD-001", "meal_plan": "standard",
        "balance": "85.00", "currency": "USD", "dietary_notes": "vegetarian",
    },
    "alumni": {
        "external_id": "ALM-2018-042", "first_name": "Yusuf", "last_name": "Adeyemi",
        "graduation_year": "2018", "email": "yusuf.adeyemi@example.com",
        "phone": "+1-555-0104", "current_employer": "Atlas Labs",
        "current_role": "Software Engineer",
    },
    "compliance": {
        "subject_external_id": "STF-001", "category": "background_check",
        "status": "complete", "due_date": "2026-07-15",
        "completed_date": "2026-07-08", "notes": "Renewed",
    },
    "structure": {
        "academic_year": "2026/2027", "year_start": "2026-09-01", "year_end": "2027-07-15",
        "year_is_active": "true", "term": "First", "term_label": "First Term",
        "term_position": "1", "term_start": "2026-09-01", "term_end": "2026-12-18",
        "department": "Sciences", "specialty": "General", "classroom": "Form 4A",
        "subject": "Mathematics", "coefficient": "4", "teacher_ref": "STF-001",
        "teacher_first_name": "Amina", "teacher_last_name": "Okoro",
        "teacher_email": "amina.okoro@example.edu",
    },
    "academics": {
        "subject_code": "MATH-101", "subject_name": "Mathematics", "code": "SCI",
        "name": "Sciences", "department": "Sciences", "credits": "4",
    },
    "transport_assignments": {
        "student_external_id": "STD-001", "route": "Route 4", "stop": "Oak Lane & Main",
        "pickup_time": "07:25", "dropoff_time": "15:40",
    },
    "hostel_assignments": {
        "student_external_id": "STD-001", "hostel": "Unity House", "room": "B-204",
        "checkin_date": "2026-08-18", "checkout_date": "",
    },
    "cafeteria_assignments": {
        "student_external_id": "STD-001", "meal_plan": "standard", "balance": "85.00",
        "currency": "USD", "dietary_notes": "vegetarian",
    },
    "athletics_teams": {
        "team_name": "Senior Football", "sport": "Football", "season": "2026/2027",
        "gender": "mixed", "level": "senior", "home_venue": "Main Field",
        "roster_cap": "22", "status": "active",
    },
    "athletics_memberships": {
        "student_external_id": "STD-001", "team_name": "Senior Football",
        "position": "Midfielder", "jersey_number": "8", "joined_date": "2026-09-05",
        "status": "active",
    },
    "athletics_fixtures": {
        "team_name": "Senior Football", "opponent_name": "Riverside Academy",
        "fixture_type": "league", "venue": "Main Field",
        "scheduled_start": "2026-10-03T15:00:00Z", "scheduled_end": "2026-10-03T16:45:00Z",
        "home_score": "2", "away_score": "1", "status": "completed",
    },
}


def _canonical_sample_row(domain: str, sorted_headers: list[str]) -> str:
    """Render a single CSV sample row for the given domain.

    Header order matches the CSV download (alphabetical), so the sample
    lines up column-by-column with the empty template the operator
    downloads. Missing values render as empty cells.
    """
    values = _CANONICAL_SAMPLE_VALUES.get(domain, {})
    row = [values.get(h, "") for h in sorted_headers]
    return ",".join(sorted_headers) + "\n" + ",".join(row)


def _canonical_template_csv(domain: str, headers: set[str]) -> str:
    """Render a single canonical-domain CSV (headers only, no data rows).

    Headers are sorted alphabetically for stability across releases so
    diff-tools work cleanly when operators version their filled-in
    templates. A leading ``#``-commented row carries the canonical-template
    contract version so future schema bumps are detectable; the intake profiler
    AND the apply-path CSV reader both skip leading ``#``-comment/blank lines, so
    a filled-in template round-trips (the marker never becomes the header row).
    A worked example per column lives on the XLSX Instructions sheet — never in
    the CSV data region, where it would be read back as a bogus record.
    """
    sorted_headers = sorted(headers)
    return (
        f"# runmycampus-canonical-template: domain={domain} version=1.0\n"
        + ",".join(sorted_headers) + "\n"
    )


def _canonical_field_guidance(domain: str, sorted_headers: list[str]) -> list[dict[str, str]]:
    """Per-column guidance for a domain: description, required?, example.

    Enriches the bare canonical headers with the rich metadata from
    ``CANONICAL_ONTOLOGY`` (description + value examples) where a field aligns,
    falling back to the sample-value table. This is what turns a headers-only
    template into one a school can actually fill in without guessing.
    """
    try:
        from .ontology.catalog import CANONICAL_ONTOLOGY

        onto_fields = CANONICAL_ONTOLOGY.get(domain, {}) or {}
    except Exception:  # noqa: BLE001 — ontology is advisory enrichment, never fatal
        onto_fields = {}
    required = set(_canonical_required_fields(domain))
    samples = _CANONICAL_SAMPLE_VALUES.get(domain, {})
    guidance: list[dict[str, str]] = []
    for header in sorted_headers:
        meta = onto_fields.get(header) or {}
        description = str(meta.get("description") or "")
        example = samples.get(header) or ""
        if not example:
            examples = meta.get("value_examples") or []
            example = str(examples[0]) if examples else ""
        guidance.append(
            {
                "column": header,
                "description": description,
                "required": "Yes" if header in required else "",
                "example": example,
            }
        )
    return guidance


def _canonical_template_xlsx(domain: str, headers: set[str]) -> bytes:
    """A single-domain XLSX: a Data sheet (headers only) + an Instructions sheet.

    The Data sheet stays headers-only so the operator fills it in and uploads it
    verbatim; the Instructions sheet carries the per-column description / required
    flag / example, so guidance never contaminates the data the importer reads.
    """
    import io

    from openpyxl import Workbook

    sorted_headers = sorted(headers)
    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "Data"
    data_ws.append(sorted_headers)

    info_ws = wb.create_sheet(title="Instructions")
    info_ws.append(["Column", "Description", "Required?", "Example"])
    for row in _canonical_field_guidance(domain, sorted_headers):
        info_ws.append([row["column"], row["description"], row["required"], row["example"]])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


def _canonical_template_workbook_all() -> bytes:
    """One workbook covering every domain: an overview Instructions sheet + one
    headers-only Data sheet per domain (sheet titles capped at Excel's 31 chars)."""
    import io

    from openpyxl import Workbook

    from .accelerators.runmycampus_canonical import DOMAIN_CANONICAL_HEADERS

    wb = Workbook()
    overview = wb.active
    overview.title = "Instructions"
    overview.append(["Domain", "Column", "Description", "Required?", "Example"])
    for domain_key, headers in sorted(DOMAIN_CANONICAL_HEADERS.items()):
        sorted_headers = sorted(headers)
        for row in _canonical_field_guidance(domain_key, sorted_headers):
            overview.append(
                [domain_key, row["column"], row["description"], row["required"], row["example"]]
            )
        ws = wb.create_sheet(title=domain_key[:31] or "sheet")
        ws.append(sorted_headers)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


def _canonical_template_readme() -> str:
    return (
        "RunMyCampus canonical migration template\n"
        "========================================\n"
        "\n"
        "One CSV per canonical domain. Fill in the rows you have; leave\n"
        "columns blank where you don't. Save and upload through the\n"
        "Migration Cloud wizard. Files keep their canonical names so the\n"
        "accelerator pre-classifies them automatically — do not rename.\n"
        "\n"
        "Required fields per domain (rows missing these will quarantine):\n"
        "  students:   external_id, first_name, last_name\n"
        "  staff:      staff_external_id, first_name, last_name\n"
        "  guardians:  guardian_external_id, first_name, last_name,\n"
        "              student_external_id\n"
        "  enrollment: student_external_id\n"
        "  attendance: student_external_id, date\n"
        "  grades:     student_external_id, subject_code, term\n"
        "  finance:    reference\n"
        "\n"
        "Extra columns past the canonical header set are preserved and\n"
        "land as DynamicFieldValues on the matching record. Re-uploading\n"
        "the same external_id updates the existing record (idempotent).\n"
    )
