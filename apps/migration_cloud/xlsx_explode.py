"""Explode a multi-worksheet XLSX into one TSV per sheet.

A school's SIS export almost always arrives as ONE workbook with several tabs
— Students, Staff, Classes, Grades. Every reader in the pipeline (profiler,
orchestrator) only ever reads ``sheetnames[0]``, so tabs 2+ were silently
dropped: a school uploaded four tabs and only the first landed. That is real,
invisible data loss.

This turns each non-empty worksheet into its own tabular artifact (a TSV) so
the workbook lands as its constituent record types, each independently
classified + landed — "we found Students, Staff and Grades in your workbook"
instead of "we imported Students and quietly lost the rest".

Everything degrades to an empty result when neither reader (openpyxl for
``.xlsx``/``.xlsm``, xlrd for legacy ``.xls``) is available or the file is
unreadable, so the caller can fall back to handling the workbook as a single
artifact rather than failing.
"""

from __future__ import annotations

import csv
import io
import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def _stringify(cell: Any) -> str:
    """Coerce a cell to the string shape the CSV/TSV readers expect.

    Integers stored as floats (``36.0``) render as ``"36"`` to match what the
    same value looks like in a CSV export.
    """
    if cell is None:
        return ""
    if isinstance(cell, str):
        return cell
    if isinstance(cell, float) and cell.is_integer():
        return str(int(cell))
    return str(cell)


def _sheet_to_tsv(ws) -> str:
    """Serialise a worksheet to TSV, skipping fully-blank rows. ``""`` if empty.

    Uses ``csv.writer`` (tab delimiter) so cell values containing tabs, quotes
    or newlines are properly quoted and survive the downstream CSV reader.
    """
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter="\t", lineterminator="\n")
    wrote_any = False
    for row in ws.iter_rows(values_only=True):
        cells = [_stringify(c) for c in row]
        if not any(c.strip() for c in cells):
            continue  # skip fully-blank row (common trailing padding)
        writer.writerow(cells)
        wrote_any = True
    return buf.getvalue() if wrote_any else ""


def count_nonempty_sheets(source) -> int:
    """Number of worksheets with at least one non-blank row. 0 if unreadable."""
    return len(explode_workbook(source))


def _as_workbook_arg(source):
    """Coerce a path / bytes / stream to something ``load_workbook`` accepts.

    openpyxl takes a path string or a file-like object, so raw bytes (an
    archive member read into memory) are wrapped in a BytesIO.
    """
    if isinstance(source, (bytes, bytearray)):
        return io.BytesIO(bytes(source))
    if isinstance(source, (str, Path)):
        return str(source)
    return source  # already a file-like object


def explode_workbook(source) -> list[tuple[str, bytes]]:
    """Return ``[(sheet_name, tsv_bytes), ...]`` for every non-empty sheet.

    ``source`` may be a path, raw bytes, or a binary file-like object — so both
    a FILE_UPLOAD workbook (path) and an archive member (bytes) can be exploded
    by the same code. Modern ``.xlsx``/``.xlsm`` workbooks are read via openpyxl;
    a legacy ``.xls`` (OLE2, which openpyxl cannot open) falls back to xlrd so
    its tabs 2+ are not silently dropped either. Empty list when neither reader
    is available or the file is unreadable — the caller then treats the workbook
    as a single artifact.
    """
    out = _explode_via_openpyxl(source)
    if out:
        return out
    # openpyxl returned nothing: it may be unavailable, the file unreadable, or
    # (the case this fallback exists for) a legacy .xls it structurally cannot
    # open. Try xlrd so a multi-sheet .xls explodes exactly like a .xlsx.
    return _explode_via_xlrd(source)


def _explode_via_openpyxl(source) -> list[tuple[str, bytes]]:
    """Explode a modern ``.xlsx``/``.xlsm`` workbook via openpyxl. ``[]`` on failure."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except Exception:  # noqa: BLE001
        return []
    try:
        wb = load_workbook(filename=_as_workbook_arg(source), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return []
    out: list[tuple[str, bytes]] = []
    try:
        for name in wb.sheetnames:
            try:
                ws = wb[name]
                tsv = _sheet_to_tsv(ws)
            except Exception:  # noqa: BLE001 — one bad sheet must not sink the rest
                logger.warning("xlsx_explode: could not read sheet %r", name)
                continue
            if tsv.strip():
                out.append((name, tsv.encode("utf-8")))
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass
    return out


def _explode_via_xlrd(source) -> list[tuple[str, bytes]]:
    """Explode a legacy ``.xls`` (OLE2) workbook via xlrd. ``[]`` on failure.

    openpyxl cannot open a legacy ``.xls``, so without this a multi-sheet ``.xls``
    would drop tabs 2+ exactly like the bug this module fixes for ``.xlsx``.
    Mirrors the openpyxl path: empty list when xlrd is unavailable or the file
    is unreadable, so the caller falls back to a single-artifact workbook.
    """
    try:
        import xlrd  # type: ignore[import-not-found]
    except Exception:  # noqa: BLE001
        return []
    data = _as_bytes_arg(source)
    if data is None:
        return []
    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception:  # noqa: BLE001
        return []
    out: list[tuple[str, bytes]] = []
    try:
        for sheet in wb.sheets():
            try:
                tsv = _xls_sheet_to_tsv(sheet)
            except Exception:  # noqa: BLE001 — one bad sheet must not sink the rest
                logger.warning(
                    "xlsx_explode: could not read .xls sheet %r",
                    getattr(sheet, "name", "?"),
                )
                continue
            if tsv.strip():
                out.append((sheet.name, tsv.encode("utf-8")))
    finally:
        try:
            wb.release_resources()
        except Exception:  # noqa: BLE001
            pass
    return out


def _xls_sheet_to_tsv(sheet) -> str:
    """Serialise an xlrd worksheet to TSV, skipping fully-blank rows. ``""`` if empty.

    Mirrors :func:`_sheet_to_tsv` (openpyxl) so an ``.xls`` sheet lands in the
    identical TSV shape the rest of the pipeline expects.
    """
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter="\t", lineterminator="\n")
    wrote_any = False
    for r in range(sheet.nrows):
        cells = [_stringify(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
        if not any(c.strip() for c in cells):
            continue  # skip fully-blank row (common trailing padding)
        writer.writerow(cells)
        wrote_any = True
    return buf.getvalue() if wrote_any else ""


def _as_bytes_arg(source) -> bytes | None:
    """Coerce a path / bytes / stream to raw bytes for xlrd. ``None`` on failure.

    xlrd reads a whole-file ``file_contents`` buffer, so a path is read from
    disk and a (possibly already-consumed) stream is rewound and read.
    """
    if isinstance(source, (bytes, bytearray)):
        return bytes(source)
    if isinstance(source, (str, Path)):
        try:
            return Path(source).read_bytes()
        except Exception:  # noqa: BLE001
            return None
    read = getattr(source, "read", None)
    if not callable(read):
        return None
    try:
        seek = getattr(source, "seek", None)
        if callable(seek):
            source.seek(0)
        data = read()
    except Exception:  # noqa: BLE001
        return None
    return bytes(data) if isinstance(data, (bytes, bytearray)) else None


def first_sheet_name(source) -> str | None:
    """Name of the workbook's FIRST worksheet (sheet index 0), or ``None``.

    The profiler + orchestrator only ever read sheet 0, so when a workbook
    explodes to exactly ONE non-empty sheet the caller compares that sheet's
    name against this to tell whether it IS sheet 0 (the raw reader already
    lands it) or a later sheet (sheet 0 is blank → the raw reader lands zero
    rows, so the exploded TSV must be substituted). Handles ``.xlsx``/``.xlsm``
    (openpyxl) and legacy ``.xls`` (xlrd); ``None`` when unreadable.
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except Exception:  # noqa: BLE001
        load_workbook = None  # type: ignore[assignment]
    if load_workbook is not None:
        try:
            wb = load_workbook(filename=_as_workbook_arg(source), read_only=True, data_only=True)
        except Exception:  # noqa: BLE001
            wb = None
        if wb is not None:
            try:
                return wb.sheetnames[0] if wb.sheetnames else None
            finally:
                try:
                    wb.close()
                except Exception:  # noqa: BLE001
                    pass
    # Legacy .xls fallback.
    try:
        import xlrd  # type: ignore[import-not-found]
    except Exception:  # noqa: BLE001
        return None
    data = _as_bytes_arg(source)
    if data is None:
        return None
    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception:  # noqa: BLE001
        return None
    try:
        return wb.sheet_by_index(0).name if wb.nsheets else None
    finally:
        try:
            wb.release_resources()
        except Exception:  # noqa: BLE001
            pass
