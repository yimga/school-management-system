"""Phase U5 — apply step orchestrator.

Walks every artifact in a MAPPED bundle, streams rows through their
column mappings + transformers, and lands them via the registered
per-domain lander under the tenant schema. Records one
``apps.automation.MigrationRun`` per (domain, artifact) for audit +
rollback.

Lifecycle: MAPPED → APPLYING → APPLIED (or FAILED on hard error).

Invariants:
    * **Tenant scoping.** All persistence runs inside
      ``django_tenants.utils.schema_context(bundle.schema_name)`` so writes
      land in the right tenant schema. Bundle metadata + MigrationRun audit
      stays in the public schema.
    * **Idempotency.** Re-running an APPLIED bundle is a no-op
      (lifecycle guard); operators rollback explicitly via
      ``MigrationRun.trigger_rollback()`` if they need to redo.
    * **No data loss.** Rows that fail transformation or upsert land in
      ``apps.automation.MigrationQuarantineRecord`` with the source row +
      error attached, viewable in the wizard.
    * **AI gateway is never called from here.** All AI happens earlier
      (mapper, classifier). Apply is deterministic.

Public surface:
    apply_bundle(bundle_id, *, dry_run=False, workers=None) → ApplyResult
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

from django.utils import timezone

from django.db import transaction

from apps.migration_cloud import defaults as mc_defaults

from .guardrails import enforce_financial_guardrail
from .landers import LanderError, LanderResult, get_lander
from .landers.reason_codes import classify_message, normalize_reason_code
from .models import (
    BundleStatus,
    FinancialMismatchError,
    MigrationArtifact,
    MigrationBundle,
)
from .progress import emit as _emit_progress, refresh_snapshot
from .services.lifecycle_events import (
    EVENT_BUNDLE_APPLIED,
    EVENT_BUNDLE_FAILED,
    emit_bundle_lifecycle_event,
)
from .transformers import TransformerContext, TransformerError, get_transformer

logger = logging.getLogger(__name__)


@dataclass
class ArtifactApplyOutcome:
    artifact_id: int
    path_within_bundle: str
    domain: str
    migration_run_id: int | None
    result: LanderResult = field(default_factory=LanderResult)
    status: str = "PENDING"
    error: str = ""


@dataclass
class ApplyResult:
    bundle_id: int
    dry_run: bool
    per_artifact: list[ArtifactApplyOutcome]
    total_created: int = 0
    total_updated: int = 0
    total_quarantined: int = 0
    status: str = ""


def _heartbeat_apply(bundle_id: int) -> None:
    """Prove a live apply is still making progress by bumping ``updated_at``.

    A running apply otherwise leaves ``updated_at`` frozen at the MAPPED->APPLYING
    flip for its whole duration (progress lands on the MigrationProgressEvent stream
    and the once-at-the-end snapshot, never the bundle row). Without this pulse a
    large-but-healthy apply would look identical to a crashed one, and the
    durable-retry self-heal (which reclaims a stale APPLYING) could reclaim a still-
    running import into a concurrent second apply. A direct UPDATE (not save) keeps
    it cheap and side-effect-free — no auto_now surprise, no signals, no lock.
    """
    try:
        MigrationBundle.objects.filter(  # tenant-isolation-allow: heartbeat bumps updated_at on the bundle already being applied by this worker
            pk=bundle_id, status=BundleStatus.APPLYING
        ).update(updated_at=timezone.now())
    except Exception:  # noqa: BLE001 — a heartbeat write must never break the apply
        logger.debug("orchestrator: apply heartbeat failed for bundle %s", bundle_id, exc_info=True)


def _pulse_apply_progress(
    bundle_id: int,
    outcomes: list,
    *,
    wave_index: int,
    jobs_total: int,
    tracker: Any | None = None,
    bundle: MigrationBundle | None = None,
) -> None:
    """Emit running created / updated / held so the kickoff page can paint live."""
    try:
        if tracker is not None:
            tracker.on_artifact_complete(outcomes, wave_index=wave_index)
            return
        totals = _summarize_outcomes(outcomes)
        done = len(outcomes)
        total = max(int(jobs_total or 0), 1)
        pct = int(round(100 * done / total))
        rows = int(totals["created"]) + int(totals["updated"]) + int(totals["quarantined"])
        _emit_progress(
            bundle_id=bundle_id,
            kind="artifact_progress",
            stage="APPLYING",
            message=(
                f"Imported {totals['created']} new, {totals['updated']} updated, "
                f"{totals['quarantined']} held ({done}/{total} files)"
            ),
            detail={
                "pct": pct,
                "rows": rows,
                "created": totals["created"],
                "updated": totals["updated"],
                "quarantined": totals["quarantined"],
                "artifacts_done": done,
                "artifacts_total": total,
                "wave": wave_index,
            },
        )
        if bundle is not None:
            from .unified_progress import pulse_apply_progress, expected_row_total

            pulse_apply_progress(
                bundle_id=bundle_id,
                bundle=bundle,
                message=(
                    f"Imported {totals['created']} new, {totals['updated']} updated, "
                    f"{totals['quarantined']} held ({done}/{total} files)"
                ),
                rows_processed=rows,
                rows_expected=expected_row_total(bundle),
                artifacts_done=done,
                artifacts_total=total,
                created=int(totals["created"]),
                updated=int(totals["updated"]),
                quarantined=int(totals["quarantined"]),
                wave=wave_index,
            )
    except Exception:  # noqa: BLE001 — live pulse must never break the apply
        logger.debug(
            "orchestrator: apply progress pulse failed for bundle %s",
            bundle_id,
            exc_info=True,
        )


def apply_bundle(
    *,
    bundle_id: int,
    dry_run: bool = False,
    workers: int | None = None,
) -> ApplyResult:
    """Apply a MAPPED bundle to its tenant.

    Sentry custom transaction `migration.bundle_apply` backs the
    homonymous SLO in `apps/observability/slo.py`.
    """
    from apps.observability.tracing import (
        finish_transaction, set_transaction_status, start_named_transaction,
    )
    from apps.platform_runtime.workflow_tracker import ensure_workflow_run

    _txn = start_named_transaction("migration.bundle_apply", bundle_id=bundle_id)
    try:
        # Track the apply as a WorkflowRun so a wedged apply is VISIBLE to the stuck
        # / abandoned watchdogs. The outbox drain (production path) and repair call
        # this function directly — bypassing the @track_workflow-decorated Celery
        # task — so without this wrap no run exists and every watchdog is blind. The
        # orchestrator already pulses "prepare"/"apply_waves"/"finalize" against
        # active_workflow_run(); this gives those pulses a run to land on.
        with ensure_workflow_run(
            "migration_bundle_apply",
            steps=("prepare", "apply_waves", "finalize"),
            expected_duration_seconds=1800,  # magic-number-allow: workflow-expected-duration-seconds (matches celery_tasks.apply_bundle_task)
            payload={"bundle_id": bundle_id, "dry_run": bool(dry_run)},
        ):
            return _apply_bundle_inner(bundle_id=bundle_id, dry_run=dry_run, workers=workers)
    except Exception:
        set_transaction_status(_txn, "internal_error")
        raise
    finally:
        finish_transaction(_txn)


def _apply_bundle_inner(
    *,
    bundle_id: int,
    dry_run: bool = False,
    workers: int | None = None,
) -> ApplyResult:
    # Claim the bundle atomically under a row lock: read status, re-check, and
    # flip MAPPED->APPLYING inside one transaction so two concurrent applies
    # (HeavyWorkOutbox double-dispatch, or a manual apply racing the outbox)
    # cannot both observe MAPPED and both run every lander -> parallel user
    # provisioning double-creates guardians/staff. select_for_update is a no-op
    # on SQLite (tests) and locks the row on Postgres (prod); the second apply
    # blocks until this commits, then sees APPLYING and refuses below.
    # A dry run is a PREVIEW: it must never mutate the durable lifecycle status.
    # Flipping the real bundle to APPLYING for a dry run made repair.py / the
    # tenant progress card see a live import in flight, and a dry-run worker crash
    # left the bundle wedged at APPLYING — so keep it at MAPPED; the dry-run
    # results land in size_summary["last_dry_run"] at the end.
    from .schema_binding import ensure_bundle_schema_name

    with transaction.atomic():
        bundle = MigrationBundle.objects.select_for_update().get(pk=bundle_id)  # tenant-isolation-allow: PK lookup by internal id from caller

        if bundle.status == BundleStatus.APPLIED and not dry_run:
            logger.info("migration_cloud.apply: bundle %s already APPLIED — no-op", bundle_id)
            return _empty_result(bundle, dry_run, BundleStatus.APPLIED)

        # Self-heal a WEDGED apply. A prior worker died mid-apply (SIGKILL / OOM /
        # deploy restart) before the except-handler below could mark it FAILED, so
        # the durable retry (HeavyWorkOutbox re-dispatch) re-enters here with the
        # bundle still APPLYING — which the MAPPED guard would reject with a
        # ValueError, dead-lettering the retry and stranding the import forever
        # (only a manual repair click could rescue it). Reclaim it: a LIVE apply
        # heartbeats updated_at every wave/artifact, so an APPLYING bundle whose
        # updated_at is stale past the threshold means its worker stopped writing —
        # reset to MAPPED and fall through to apply. Under the select_for_update row
        # lock a genuinely concurrent apply is still refused below, because its
        # updated_at is fresh (< threshold). Apply is upsert-by-external-id, so
        # re-applying never duplicates already-landed rows. Time-only staleness (not
        # repair.py's in-flight variant): the retry's OWN outbox row reads as
        # in-flight here and would mask the dead prior apply.
        if not dry_run and bundle.status == BundleStatus.APPLYING:
            from .repair import applying_stale_by_time

            if applying_stale_by_time(bundle):
                # Count the attempts. Retrying a dead worker is right; retrying
                # forever hides a bundle that finishes and never settles, which
                # is indistinguishable from a hang to everyone watching it.
                _reclaims = wedged_reclaims_so_far(bundle.size_summary)
                if wedged_reclaim_budget_exhausted(bundle.size_summary):
                    logger.error(
                        "orchestrator: bundle %s has been reclaimed from a wedged APPLYING "
                        "state %s times without ever settling — refusing to retry again and "
                        "marking FAILED so it stops looping and becomes repairable",
                        bundle_id,
                        _reclaims,
                    )
                    bundle.mark_status(
                        BundleStatus.FAILED,
                        summary_patch={
                            "error": (
                                "This import kept restarting without ever finishing. It has "
                                "been stopped so it cannot loop. Your data is unchanged — "
                                "records already imported were updated in place, never "
                                "duplicated. Use Repair to try again."
                            ),
                            "wedged_apply_reclaim_ceiling_hit_at": timezone.now().isoformat(),
                        },
                    )
                    return _empty_result(bundle, dry_run, BundleStatus.FAILED)
                logger.warning(
                    "orchestrator: bundle %s wedged at APPLYING with no heartbeat past "
                    "the stale threshold (prior worker died mid-apply) — reclaiming to "
                    "MAPPED for retry (attempt %s of %s)",
                    bundle_id,
                    _reclaims + 1,
                    _MAX_WEDGED_APPLY_RECLAIMS,
                )
                bundle.mark_status(
                    BundleStatus.MAPPED,
                    summary_patch={
                        "reclaimed_wedged_apply_at": timezone.now().isoformat(),
                        "wedged_apply_reclaims": _reclaims + 1,
                    },
                )

        if bundle.status != BundleStatus.MAPPED:
            raise ValueError(
                f"Bundle {bundle_id} is in status {bundle.status}; must be MAPPED to apply."
            )

        effective_schema = ensure_bundle_schema_name(bundle)
        if bundle.school_id and not effective_schema:
            raise ValueError(
                f"Bundle {bundle_id} is bound to school_id={bundle.school_id} but has no "
                "tenant schema_name — refuse apply so rows are not written off-tenant. "
                "Re-bind the school or repair Client.schema_name, then retry."
            )

        if not dry_run and effective_schema:
            from .tenant_schema_readiness import (
                assess_tenant_schema_readiness,
                schema_drift_summary_patch,
            )

            schema_ready = assess_tenant_schema_readiness(
                effective_schema, attempt_repair=True
            )
            if not schema_ready.ready:
                logger.error(
                    "migration_cloud.apply: tenant schema drift blocks bundle=%s "
                    "schema=%s missing=%s",
                    bundle_id,
                    effective_schema,
                    schema_ready.missing_labels,
                )
                bundle.mark_status(
                    BundleStatus.FAILED,
                    summary_patch=schema_drift_summary_patch(schema_ready),
                )
                return _empty_result(bundle, dry_run, BundleStatus.FAILED)

        if not dry_run:
            # Drop per-domain / notes from a prior apply so a successful retry
            # cannot keep showing stale "held for review" copy if reconcile is
            # slow or fails. Preserve shadow-mode state nested under the same JSON.
            prior_recon = dict(bundle.reconciliation_summary or {})
            shadow = prior_recon.get("shadow")
            bundle.reconciliation_summary = {"shadow": shadow} if shadow else {}
            bundle.save(update_fields=["reconciliation_summary", "updated_at"])
            # Open a new progress run at the same moment the status flips. The
            # event stream is append-only for the life of the bundle and APPLYING
            # is the stage a retry RE-RUNS, so without this boundary the snapshot
            # keeps serving the previous apply's ratcheted pct and live totals.
            from .progress import mark_apply_run_start

            mark_apply_run_start(bundle)
            bundle.mark_status(BundleStatus.APPLYING)
    bundle.refresh_from_db()
    if not dry_run:
        # This apply is about to regenerate every held row from scratch, so the
        # previous apply's PENDING rows are superseded, not history. Without this
        # each re-apply appended a fresh copy: one bundle re-applied 128 times and
        # accumulated 40,448 records that were 128 copies of the same 316.
        _clear_superseded_quarantine(bundle)
    _emit_progress(bundle_id=bundle_id, kind="stage_started", stage="APPLYING",
                   message=f"Apply started (dry_run={dry_run}, atomic={bundle.apply_atomic})")

    worker_count = workers or int(
        mc_defaults.get("migration_cloud.orchestrator.worker_count")
    )

    per_artifact_jobs = _build_jobs(bundle)

    # Honesty gate (defense in depth with profiler.profile_bundle): a real apply
    # of a bundle that HAS artifacts but zero WORKABLE jobs (every file quarantined
    # or an archive shell) would otherwise stamp a *green* APPLIED with all-zero
    # totals — the operator believes rows landed when none did. Mark FAILED
    # (repairable) so "APPLIED" always means real rows were considered. A genuinely
    # empty bundle (no artifacts at all) keeps its prior no-op behaviour.
    if not dry_run and not per_artifact_jobs and bundle.artifacts.exists():
        logger.warning(
            "orchestrator: bundle %s reached apply with no workable artifacts "
            "(all quarantined) — marking FAILED instead of a 0-row APPLIED",
            bundle_id,
        )
        bundle.mark_status(
            BundleStatus.FAILED,
            summary_patch={
                "no_workable_artifacts": True,
                "error": "No workable artifacts to apply — every file was quarantined.",
            },
        )
        _emit_progress(bundle_id=bundle_id, kind="warning", stage="APPLYING",
                       message="Apply aborted — no workable artifacts (all quarantined).")
        try:
            emit_bundle_lifecycle_event(
                bundle, EVENT_BUNDLE_FAILED,
                {"reason": "no_workable_artifacts"},
            )
        except Exception:  # noqa: BLE001 — event emission never blocks
            pass
        return _empty_result(bundle, dry_run, BundleStatus.FAILED)

    try:
        from apps.platform_runtime.workflow_tracker import active_workflow_run, pulse_workflow_step

        pulse_workflow_step(
            active_workflow_run(),
            "prepare",
            payload={
                "bundle_id": bundle_id,
                "dry_run": dry_run,
                "artifacts": len(per_artifact_jobs),
            },
        )
    except Exception:
        pass
    outcomes: list[ArtifactApplyOutcome] = []
    failed = False

    # FK dependency DAG: students + staff + sections must land before
    # enrollment / attendance / grades / guardians / behavior / finance
    # so child rows can resolve their parent FKs. custom_fields runs last
    # since it references the entity that owns the dynamic value.
    waves = _partition_jobs_by_dependency(per_artifact_jobs)
    jobs_total = max(sum(len(w) for w in waves), 1)
    from .unified_progress import ApplyProgressTracker

    apply_tracker = ApplyProgressTracker(bundle=bundle, jobs_total=jobs_total)

    def _stall_watchdog_heartbeat(watchdog: Any) -> None:
        totals = _summarize_outcomes(outcomes)
        watchdog.heartbeat(
            current_pointer=len(outcomes),
            mutations_count=int(totals["created"])
            + int(totals["updated"])
            + int(totals["quarantined"]),
            rows_processed=apply_tracker.rows_global,
        )
        if not dry_run:
            _heartbeat_apply(bundle_id)

    def _run_waves() -> None:
        from .apply_stall import (
            reset_stall_pulse_hook,
            resolve_stall_timeout_seconds,
            set_stall_pulse_hook,
        )
        from .loop_watchdog import LoopWatchdog

        stall_timeout = resolve_stall_timeout_seconds(bundle)
        with LoopWatchdog(
            max_stall_iterations=3,
            timeout_seconds=stall_timeout,
            workflow_identifier=f"migration.apply bundle={bundle_id}",
        ) as apply_watchdog:
            apply_tracker.on_stall_heartbeat = lambda: _stall_watchdog_heartbeat(
                apply_watchdog
            )
            _stall_pulse_token = set_stall_pulse_hook(
                lambda: _stall_watchdog_heartbeat(apply_watchdog)
            )
            try:
                if not dry_run:
                    _stall_watchdog_heartbeat(apply_watchdog)
                for wave_index, wave_jobs in enumerate(waves):
                    if not wave_jobs:
                        continue
                    if not dry_run:
                        _stall_watchdog_heartbeat(apply_watchdog)
                    if (
                        not dry_run
                        and any(job.domain in _DEPENDENT_STRUCTURE_DOMAINS for job in wave_jobs)
                    ):
                        with _bundle_schema_context(bundle):
                            try:
                                from apps.migration_cloud.post_apply_provision import (
                                    provision_structure_before_dependent_domains,
                                )

                                provision_structure_before_dependent_domains(bundle=bundle)
                            except Exception as exc:  # noqa: BLE001 — never abort apply
                                logger.warning(
                                    "orchestrator: mid-apply structure provision errored "
                                    "for bundle %s: %s",
                                    bundle_id,
                                    exc,
                                    exc_info=True,
                                )
                    if not dry_run:
                        _stall_watchdog_heartbeat(apply_watchdog)
                    if not dry_run:
                        _heartbeat_apply(bundle_id)
                    _emit_progress(
                        bundle_id=bundle_id,
                        kind="artifact_progress",
                        stage="APPLYING",
                        message=f"Wave {wave_index} starting ({len(wave_jobs)} artifact(s))",
                        detail={"wave": wave_index, "artifacts": len(wave_jobs)},
                    )
                    try:
                        from apps.platform_runtime.workflow_tracker import (
                            active_workflow_run,
                            pulse_workflow_step,
                        )

                        pulse_workflow_step(
                            active_workflow_run(),
                            "apply_waves",
                            payload={"wave": wave_index, "artifacts": len(wave_jobs)},
                        )
                    except Exception:
                        pass
                    jobs_total = max(sum(len(w) for w in waves), 1)
                    if worker_count <= 1 or len(wave_jobs) <= 1:
                        for job in wave_jobs:
                            outcomes.append(
                                _apply_artifact(
                                    bundle,
                                    job,
                                    dry_run=dry_run,
                                    progress_tracker=apply_tracker,
                                )
                            )
                            if not dry_run:
                                _heartbeat_apply(bundle_id)
                                totals = _summarize_outcomes(outcomes)
                                apply_watchdog.heartbeat(
                                    current_pointer=len(outcomes),
                                    mutations_count=int(totals["created"])
                                    + int(totals["updated"])
                                    + int(totals["quarantined"]),
                                    rows_processed=apply_tracker.rows_global,
                                )
                                _pulse_apply_progress(
                                    bundle_id,
                                    outcomes,
                                    wave_index=wave_index,
                                    jobs_total=jobs_total,
                                    tracker=apply_tracker,
                                )
                    else:
                        with ThreadPoolExecutor(max_workers=worker_count) as pool:
                            futures = {
                                pool.submit(
                                    _apply_artifact,
                                    bundle,
                                    job,
                                    dry_run=dry_run,
                                    progress_tracker=apply_tracker,
                                ): job
                                for job in wave_jobs
                            }
                            for future in as_completed(futures):
                                outcomes.append(future.result())
                                if not dry_run:
                                    _heartbeat_apply(bundle_id)
                                    totals = _summarize_outcomes(outcomes)
                                    apply_watchdog.heartbeat(
                                        current_pointer=len(outcomes),
                                        mutations_count=int(totals["created"])
                                        + int(totals["updated"])
                                        + int(totals["quarantined"]),
                                        rows_processed=apply_tracker.rows_global,
                                    )
                                    _pulse_apply_progress(
                                        bundle_id,
                                        outcomes,
                                        wave_index=wave_index,
                                        jobs_total=jobs_total,
                                        tracker=apply_tracker,
                                    )

            finally:
                reset_stall_pulse_hook(_stall_pulse_token)
    # Finance MUST be all-or-nothing. In non-atomic mode finance rows commit
    # (autocommit) BEFORE the financial guardrail runs; a control-total mismatch
    # then marks the bundle FAILED and calls _rollback_all_runs — but finance has
    # no rollback handler, so the mismatched ledger stays committed while the
    # operator believes nothing landed. Forcing atomic makes the guardrail's abort
    # real: transaction.atomic() DB-rolls-back every finance write on
    # FinancialMismatchError. Mirrors repair.repair_readiness's finance-requires-
    # atomic gate. See docs/MIGRATION_CLOUD_AUDIT_2026_07_24.md (BLOCKER 3).
    finance_present = any(job.domain == "finance" for job in per_artifact_jobs)
    atomic_mode = (bool(getattr(bundle, "apply_atomic", False)) or finance_present) and not dry_run
    if atomic_mode:
        # Worker threads open their own DB connections and would NOT join the
        # outer transaction.atomic(), silently breaking all-or-nothing. Force
        # single-threaded so the atomic block actually wraps every write.
        worker_count = 1
    try:
        if atomic_mode:
            with transaction.atomic():
                _run_waves()
                _maybe_check_financial_guardrail(bundle, outcomes)
        else:
            _run_waves()
            if not dry_run:
                _maybe_check_financial_guardrail(bundle, outcomes)
    except FinancialMismatchError as exc:
        logger.warning("orchestrator: financial guardrail aborted apply: %s", exc)
        _emit_progress(bundle_id=bundle_id, kind="warning", stage="APPLYING",
                       message=f"Financial guardrail failure — apply aborted: {exc}")
        bundle.mark_status(
            BundleStatus.FAILED,
            summary_patch={"financial_guardrail_failed": True, "error": str(exc)},
        )
        if not atomic_mode:
            _rollback_all_runs(outcomes)
        # Partner lifecycle event (G-5): fires on BOTH the API and UI paths.
        emit_bundle_lifecycle_event(
            bundle, EVENT_BUNDLE_FAILED,
            {"reason": "financial_guardrail", "error": str(exc)},
        )
        raise
    except Exception as exc:  # noqa: BLE001 — ANY apply failure must mark FAILED, never wedge at APPLYING
        # The bundle was set to APPLYING at the top of this function. Letting an
        # unexpected error (a lander crash, a DB error, an OOM) propagate from here
        # would leave it APPLYING forever: re-apply requires MAPPED and repair refused
        # APPLYING as "still running", so the import became unrecoverable without DB
        # surgery. Mark it FAILED — which repair_readiness treats as repairable — as a
        # best-effort, then re-raise the original error so the caller / outbox still
        # sees the failure. (A worker SIGKILL never reaches this handler; repair.py
        # reclaims that stale-APPLYING case by timeout.)
        logger.exception("orchestrator: apply failed for bundle %s — marking FAILED", bundle_id)
        # A dry-run crash must NOT corrupt the durable status. The bundle was left
        # at MAPPED (the APPLYING flip above is gated on ``not dry_run``), so a
        # preview that blows up simply re-raises to the caller and the operator can
        # retry the preview — the real bundle is untouched.
        if not dry_run:
            try:
                summary_patch: dict[str, Any] = {"error": f"{type(exc).__name__}: {exc}"}
                from .loop_watchdog import SystemicStallError

                if isinstance(exc, SystemicStallError):
                    summary_patch["systemic_stall"] = True
                bundle.mark_status(
                    BundleStatus.FAILED,
                    summary_patch=summary_patch,
                )
            except Exception:  # noqa: BLE001 — marking FAILED must not mask the original error
                logger.exception("orchestrator: could not mark bundle %s FAILED", bundle_id)
        try:
            _emit_progress(bundle_id=bundle_id, kind="error", stage="APPLYING",
                           message=f"Apply failed — {type(exc).__name__}")
            if not atomic_mode:
                _rollback_all_runs(outcomes)
            if not dry_run:
                emit_bundle_lifecycle_event(
                    bundle, EVENT_BUNDLE_FAILED,
                    {"reason": "apply_exception", "error": str(exc)},
                )
        except Exception:  # noqa: BLE001 — best-effort cleanup, never mask the original error
            logger.exception("orchestrator: post-failure cleanup errored for bundle %s", bundle_id)
        raise

    # Post-apply structural gap-fill (S): scaffold the academic year / default
    # department+specialty / cycle nodes / teaching grid a running school needs
    # but the upload didn't carry. Idempotent + deduped against what landed;
    # gated to only fire when roster/catalog data actually landed. Best-effort —
    # it must never turn a successful apply into a failure.
    gap_fill_summary: dict | None = None
    if not dry_run:
        # Post-apply hooks query tenant models; without schema_context they hit
        # public (stale shadow tables) and raise UndefinedColumn even when the
        # tenant schema is healthy.
        with _bundle_schema_context(bundle):
            try:
                from apps.migration_cloud.post_apply_provision import gap_fill_after_apply

                gap_fill_summary = gap_fill_after_apply(
                    bundle=bundle, outcomes=outcomes, dry_run=dry_run
                )
            except Exception as exc:  # noqa: BLE001 — gap-fill is additive; never break apply
                from .tenant_schema_readiness import post_apply_step_error

                logger.warning(
                    "orchestrator: post-apply gap-fill errored for bundle %s",
                    bundle_id,
                    exc_info=True,
                )
                if gap_fill_summary is None:
                    gap_fill_summary = {}
                gap_fill_summary["gap_fill_error"] = post_apply_step_error(exc)
            try:
                from apps.migration_cloud.guardian_directory import (
                    promote_unlinked_guardian_hints,
                )

                directory_summary = promote_unlinked_guardian_hints(
                    school=getattr(bundle, "school", None),
                )
                if gap_fill_summary is None:
                    gap_fill_summary = {}
                if isinstance(gap_fill_summary, dict):
                    gap_fill_summary["guardian_directory"] = directory_summary
            except Exception as exc:  # noqa: BLE001 — directory promote is additive; never break apply
                from .tenant_schema_readiness import post_apply_step_error

                logger.warning(
                    "orchestrator: guardian-directory promote errored for bundle %s",
                    bundle_id,
                    exc_info=True,
                )
                if gap_fill_summary is None:
                    gap_fill_summary = {}
                gap_fill_summary["guardian_directory"] = post_apply_step_error(exc)
            try:
                from apps.migration_cloud.staff_role_map import promote_imported_staff_roles

                role_summary = promote_imported_staff_roles(
                    school=getattr(bundle, "school", None),
                )
                if gap_fill_summary is None:
                    gap_fill_summary = {}
                if isinstance(gap_fill_summary, dict):
                    gap_fill_summary["staff_role_backfill"] = role_summary
            except Exception as exc:  # noqa: BLE001 — role backfill is additive; never break apply
                from .tenant_schema_readiness import post_apply_step_error

                logger.warning(
                    "orchestrator: staff-role backfill errored for bundle %s",
                    bundle_id,
                    exc_info=True,
                )
                if gap_fill_summary is None:
                    gap_fill_summary = {}
                gap_fill_summary["staff_role_backfill"] = post_apply_step_error(exc)

    totals = _summarize_outcomes(outcomes)
    # Landers wrote operator-review data (dedup_candidates / dedup_links) STRAIGHT to
    # the bundle row's mapping_summary during the waves, using their OWN re-fetched
    # bundle instance (LanderContext carries bundle_id, not this object — see
    # student_lander._surface_dedup_candidates / staff_lander). This in-memory instance
    # still holds the PRE-wave mapping_summary, so rebuilding from it and saving
    # update_fields=["mapping_summary"] would clobber those DB writes and lose the
    # operator's duplicate-review queue (#7). Refresh first so the merge preserves them;
    # apply_totals + gap_fill are layered on top. Only this one field is refreshed, so
    # the orchestrator's other in-memory state (status, size_summary) is untouched.
    if not dry_run:
        bundle.refresh_from_db(fields=["mapping_summary"])
    bundle.mapping_summary = {
        **(bundle.mapping_summary or {}),
        "apply_totals": {
            "created": totals["created"],
            "updated": totals["updated"],
            "quarantined": totals["quarantined"],
            "errors": totals["errors"],
            "dry_run": dry_run,
            "applied_at": timezone.now().isoformat(),
        },
        **({"gap_fill_provisioning": gap_fill_summary} if gap_fill_summary is not None else {}),
    }
    bundle.save(update_fields=["mapping_summary", "updated_at"])

    failed = bundle_apply_failed(outcomes=outcomes, totals=totals)
    new_status = BundleStatus.FAILED if failed else BundleStatus.APPLIED
    if dry_run:
        # Dry-run never advances past MAPPED — operator still needs to apply.
        bundle.mark_status(BundleStatus.MAPPED, summary_patch={"last_dry_run": totals})
    else:
        # wedged_apply_reclaims resets here: a bundle that settles has, by
        # definition, stopped being wedged, so a later genuine worker death gets
        # the full retry budget again rather than inheriting an old tally.
        bundle.mark_status(
            new_status,
            summary_patch={"apply_totals": totals, "wedged_apply_reclaims": 0},
        )
        # A non-atomic bundle where one artifact COMMITTED rows (autocommit) and
        # another FAILED would otherwise read FAILED while the committed rows
        # stayed LIVE — breaking the "FAILED = nothing landed" contract the
        # atomic path guarantees (and the except handlers above already honour).
        # Roll the succeeded runs back CHILD-FIRST so a FAILED bundle really
        # leaves nothing behind. Gated on ``not atomic_mode`` to mirror those
        # handlers: atomic mode is the finance / all-or-nothing lane whose
        # boundary the audit deemed solid and which this fix must not touch. This
        # normal-return branch is mutually exclusive with the except handlers
        # (no exception propagated to reach here), so their _rollback_all_runs
        # never double-fires with this one.
        if failed and not atomic_mode:
            _rollback_all_runs(outcomes)
        # Partner lifecycle event (G-5): emitted here at the SERVICE layer so a
        # migration run from the connector/customer UI fires the same webhook an
        # API-driven apply does (the REST viewset no longer emits — avoids double).
        emit_bundle_lifecycle_event(
            bundle,
            EVENT_BUNDLE_APPLIED if new_status == BundleStatus.APPLIED else EVENT_BUNDLE_FAILED,
            {
                "created": totals.get("created", 0),
                "updated": totals.get("updated", 0),
                "quarantined": totals.get("quarantined", 0),
                "status": str(new_status),
            },
        )

    _emit_progress(
        bundle_id=bundle.pk, kind="stage_finished", stage="APPLYING",
        message=f"Apply finished: {totals.get('created')} created, "
                f"{totals.get('updated')} updated, {totals.get('quarantined')} quarantined",
        detail={"totals": totals},
    )
    refresh_snapshot(bundle=bundle)

    if not dry_run:
        # Re-read the status instead of trusting the write above. An apply that
        # announces "finished" and leaves the bundle at APPLYING is completely
        # silent: no exception, no failed row, nothing in the log. Thirty minutes
        # later the stale detector reclaims it and the entire import runs again.
        # One live bundle re-ran a 44-second import roughly 48 times in 24 hours
        # this way. Verifying here turns that into a single loud line naming the
        # status that actually survived, and forces the terminal state so the
        # loop cannot start.
        _settled = (
            MigrationBundle.objects.filter(pk=bundle.pk)  # tenant-isolation-allow: PK re-read of the bundle this apply already holds
            .values_list("status", flat=True)
            .first()
        )
        if _settled not in _TERMINAL_BUNDLE_STATUSES:
            logger.error(
                "orchestrator: bundle %s finished its apply (%s created, %s updated, "
                "%s quarantined) but the persisted status is %r instead of the %r just "
                "written — something overwrote it. Forcing the terminal status so the "
                "bundle cannot be reclaimed into an endless re-apply loop.",
                bundle.pk,
                totals.get("created"),
                totals.get("updated"),
                totals.get("quarantined"),
                _settled,
                str(new_status),
            )
            # .update() deliberately: it bypasses the in-memory instance (whose
            # state we have just proved untrustworthy) and does not re-stamp
            # updated_at, so it cannot re-arm the staleness heartbeat.
            MigrationBundle.objects.filter(pk=bundle.pk).update(status=new_status)  # tenant-isolation-allow: PK forced settle of the bundle this apply already holds

    try:
        from apps.platform_runtime.workflow_tracker import active_workflow_run, pulse_workflow_step

        pulse_workflow_step(
            active_workflow_run(),
            "finalize",
            payload={"status": new_status, "created": totals.get("created", 0)},
        )
    except Exception:
        pass

    try:
        from apps.migration_cloud.models_audit import MigrationCloudAuditEventType
        from apps.migration_cloud.services.bundle_lifecycle_audit import safe_bundle_audit

        safe_bundle_audit(
            bundle,
            MigrationCloudAuditEventType.BUNDLE_APPLIED.value,
            payload_summary={
                "bundle_id": str(bundle.pk),
                "dry_run": bool(dry_run),
                "status": str(new_status if not dry_run else BundleStatus.MAPPED),
                "created": int(totals.get("created", 0)),
                "updated": int(totals.get("updated", 0)),
                "quarantined": int(totals.get("quarantined", 0)),
                # Field-level rollup: per domain, the counts + the NAMES of fields
                # overwritten on existing records (names as list values; raw
                # values never leave the tenant schema).
                "domains": _field_level_apply_summary(outcomes),
            },
        )
    except Exception:  # noqa: BLE001
        logger.warning(
            "migration_cloud.orchestrator: apply audit failed bundle_id=%s",
            bundle_id,
            exc_info=True,
        )

    if not dry_run:
        # Did this apply actually move anything? A live apply whose totals are
        # identical to the previous one, having created no rows, is the livelock
        # signature that pinned a worker for four days on bundle 84. Recorded here
        # -- the single point every apply path returns through -- so the enqueue
        # guard sees it no matter which of the seven callers queued the work.
        from .apply_progress_guard import record_apply_outcome

        record_apply_outcome(
            bundle,
            created=totals["created"],
            updated=totals["updated"],
            quarantined=totals["quarantined"],
            status=new_status,
        )
        if new_status == BundleStatus.APPLIED:
            from .reconciliation import run_post_apply_verification

            run_post_apply_verification(bundle_id=bundle.pk)
            refresh_snapshot(bundle=bundle)

    return ApplyResult(
        bundle_id=bundle.pk,
        dry_run=dry_run,
        per_artifact=outcomes,
        total_created=totals["created"],
        total_updated=totals["updated"],
        total_quarantined=totals["quarantined"],
        status=new_status if not dry_run else BundleStatus.MAPPED,
    )


def _maybe_check_financial_guardrail(
    bundle: MigrationBundle, outcomes: list["ArtifactApplyOutcome"],
) -> None:
    """Run the financial guardrail if any finance domain landed and expected_totals is set."""
    finance_landed = any(
        o.domain == "finance" and o.status in ("SUCCESS", "PARTIAL") for o in outcomes
    )
    students_landed = any(
        o.domain == "students" and o.status in ("SUCCESS", "PARTIAL") for o in outcomes
    )
    if not bundle.expected_totals:
        # #4b: money landed with NO operator control totals = UNVERIFIED. Don't pass
        # it off silently. Warn loudly by default (recorded on the bundle + logged);
        # a deployment can hard-require totals on sensitive tenants via the settings
        # flag, in which case an unverified finance import is refused (FAILED + rolled
        # back through the orchestrator's FinancialMismatchError handler).
        if finance_landed:
            _handle_unverified_finance(bundle)
        return
    # Only enforce when something happened in a domain the guardrail observes.
    if not (finance_landed or students_landed):
        return
    bundle.refresh_from_db()
    report = enforce_financial_guardrail(bundle=bundle)
    bundle.mapping_summary = {
        **(bundle.mapping_summary or {}),
        "financial_guardrail": report.to_dict(),
    }
    bundle.save(update_fields=["mapping_summary", "updated_at"])


def _handle_unverified_finance(bundle: MigrationBundle) -> None:
    """Finance landed with no operator control totals (#4b).

    Default: WARN — record it on the bundle (visible to the operator) and log it, so
    "money landed unverified" is never silent, but the import still applies (matching
    prior behaviour for the many finance imports that don't supply totals). When
    ``RMC_MIGRATION_REQUIRE_FINANCE_TOTALS`` is on, REFUSE instead: raise
    FinancialMismatchError so the orchestrator marks the bundle FAILED and rolls back.
    """
    from django.conf import settings

    require = bool(getattr(settings, "RMC_MIGRATION_REQUIRE_FINANCE_TOTALS", False))
    message = (
        "Finance rows landed but no expected_totals control totals were provided — "
        "the money was NOT verified against operator-supplied totals."
    )
    if require:
        raise FinancialMismatchError(
            f"Bundle {bundle.pk}: {message} (RMC_MIGRATION_REQUIRE_FINANCE_TOTALS is enabled)."
        )
    logger.warning("orchestrator: bundle %s — %s", bundle.pk, message)
    try:
        bundle.mapping_summary = {
            **(bundle.mapping_summary or {}),
            "finance_landed_unverified": True,
        }
        bundle.save(update_fields=["mapping_summary", "updated_at"])
    except Exception:  # noqa: BLE001 — surfacing the warning must never break the apply
        logger.debug("orchestrator: could not record finance_landed_unverified", exc_info=True)


def _rollback_all_runs(outcomes: list["ArtifactApplyOutcome"]) -> None:
    """Roll back every MigrationRun produced by this apply, CHILD-FIRST.

    Runs are rolled back most-recent-first (``order_by("-started_at")``) — the
    REVERSE of the dependency-wave order they applied in. This is load-bearing:
    later-wave rows PROTECT earlier-wave rows (``Evaluation.student`` /
    ``Evaluation.teacher`` are ``on_delete=PROTECT``), so deleting a wave-1
    student BEFORE its wave-3 grades raises ``IntegrityError`` — that student's
    rollback is marked FAILED and swallowed while the grades roll back
    afterwards, leaving the student/teacher ORPHANED and live though the bundle
    reads FAILED. Iterating ``outcomes`` in append (wave) order did exactly this.
    Rolling back child rows first (grades before students) is the only safe
    order; ``connector_rollback`` / ``reconciliation`` auto-rollback get it
    implicitly via ``MigrationRun.Meta.ordering`` — here it is explicit.
    """
    try:
        from apps.automation.models import MigrationRun
    except ImportError:
        return
    run_ids = [o.migration_run_id for o in outcomes if o.migration_run_id]
    if not run_ids:
        return
    # tenant-isolation-allow: PK-set lookup by internal run ids from this apply
    runs = list(MigrationRun.objects.filter(pk__in=run_ids).order_by("-started_at"))
    for run in runs:
        try:
            run.trigger_rollback(user=None)
        except Exception:  # noqa: BLE001
            logger.debug(
                "orchestrator: rollback failed for run %s", run.pk, exc_info=True,
            )


# Operator-chosen date reading -> the strptime format the date transformer uses.
# "" means "leave the existing inference alone" (profiler vote, then country
# profile), so an unset preference behaves exactly as before.
_DATE_ORDER_FORMATS = {
    "day_first": "%d/%m/%Y",
    "month_first": "%m/%d/%Y",
    "year_first": "%Y-%m-%d",
}


def operator_date_format(bundle) -> str:
    """The date format the operator explicitly chose on the review page, or "".

    An unrecognised value is ignored rather than trusted -- a bad preference
    must never become a strptime format that silently misreads every date.
    """
    prefs = (getattr(bundle, "mapping_summary", None) or {}).get("transform_prefs") or {}
    order = str(prefs.get("date_order") or "").strip().lower()
    return _DATE_ORDER_FORMATS.get(order, "")


def apply_operator_date_override(locale_hints: dict, bundle) -> None:
    """Let an explicit operator choice outrank every inferred date format.

    The profiler's per-column vote and the country profile are both inferences;
    this is the school telling us what its file actually is. A wrong date reading
    is the worst class of import defect because it is SILENT -- every row lands,
    every date is wrong, and nothing is quarantined to hint at it -- so the one
    party who can actually know must be able to say.
    """
    chosen = operator_date_format(bundle)
    if chosen:
        locale_hints["date_format"] = chosen


def artifact_outcome_status(result) -> str:
    """The verdict for ONE artifact: SUCCESS, PARTIAL or REJECTED.

    REJECTED means rows were rejected and nothing landed at all. It is
    deliberately NOT "FAILED": the run it writes still reads Failed to the
    operator (see ``_RUN_STATUS_MAP``), which is what stops a 0-created import
    reporting "succeeded" -- but the bundle-level rule treats FAILED as
    "everything must be rolled back", and one unimportable file must never
    discard the files that imported cleanly beside it.

    ``quarantined`` is required for REJECTED: an artifact that legitimately had
    nothing to do -- a header-only file, or a re-run where every record was
    already current -- reports 0/0/0 and stays SUCCESS.
    """
    if result.quarantined and not (result.created or result.updated):
        return "REJECTED"
    return "PARTIAL" if result.quarantined else "SUCCESS"


def bundle_apply_failed(*, outcomes, totals) -> bool:
    """Whether the BUNDLE failed — which also decides whether rows get rolled back.

    Two ways to fail, and rejecting some rows is not one of them:

      * a hard failure in any artifact (a lander crash, a DB error), which the
        rollback contract exists for; or
      * nothing landed ANYWHERE while rows were rejected -- the 431-of-431 case,
        where reporting success is what cost a live tenant hours.

    A bundle that imported four files and could not import a fifth keeps the
    four. It stays repairable regardless: ``repair._has_unresolved_issues``
    already returns True for any bundle carrying quarantined rows, so
    repairability never depended on the FAILED status.
    """
    if any(getattr(o, "status", "") == "FAILED" for o in outcomes):
        return True
    landed = int(totals.get("created") or 0) + int(totals.get("updated") or 0)
    rejected = int(totals.get("quarantined") or 0)
    return bool(outcomes) and landed == 0 and rejected > 0


def _run_status_map():
    from apps.automation.models import MigrationRun

    return {
        "SUCCESS": MigrationRun.Status.SUCCESS,
        "PARTIAL": MigrationRun.Status.PARTIAL,
        "FAILED": MigrationRun.Status.FAILED,
        # No REJECTED member on the model, and none is needed: a file that
        # imported nothing should read "Failed" to whoever is looking.
        "REJECTED": MigrationRun.Status.FAILED,
    }


class _LazyRunStatusMap(dict):
    """Populated on first use — the model import must stay lazy in this module."""

    def _ensure(self):
        if not self:
            self.update(_run_status_map())

    def get(self, key, default=None):
        self._ensure()
        return dict.get(self, key, default)

    def __getitem__(self, key):
        self._ensure()
        return dict.__getitem__(self, key)


_RUN_STATUS_MAP = _LazyRunStatusMap()


# --- Dependency-DAG wave partitioning ------------------------------------

# Ordered waves; jobs within a wave run in parallel, waves run serially so
# the next wave sees its parent rows already in the tenant schema.
_DEPENDENCY_WAVES: tuple[frozenset[str], ...] = (
    frozenset({"structure", "academic_sessions", "specialties"}),   # wave 0: academic scaffold (SPLIT provisioning + OneRoster years/terms + trade/stream catalog) — MUST precede students/enrollment/grades
    frozenset({"students", "staff", "sections", "academics", "alumni"}),  # wave 1: independent roots (academics = Subject catalog, precedes grades; alumni upserts StudentProfile so guardians/finance/grades can resolve alumni students)
    frozenset({"enrollment", "guardians", "schedule"}),             # wave 2: depend on wave 1
    frozenset({"attendance", "grades", "behavior", "finance", "transcripts",  # wave 3: depend on wave 2
               "health", "library", "transport", "hostel", "cafeteria",
               "athletics_teams"}),                                 # athletics_teams precedes its roster/fixtures
    frozenset({"custom_fields"}),                                   # wave 4: catch-all (athletics_memberships/fixtures, *_assignments) — after their parents
)

# Wave-3+ domains need SubjectAssignments / terms / classrooms that gap-fill used
# to create only AFTER this wave finished — so Repair from the UI re-quarantined
# the same rows every pass. Mid-apply provisioning runs once roster/catalog landed.
_DEPENDENT_STRUCTURE_DOMAINS = frozenset({
    "attendance", "grades", "behavior", "finance", "transcripts",
})


def _partition_jobs_by_dependency(jobs: list["_ArtifactJob"]) -> list[list["_ArtifactJob"]]:
    """Bucket jobs into FK-safe waves so child rows can resolve their parents.

    Any domain not in ``_DEPENDENCY_WAVES`` lands in the final catch-all
    wave alongside custom_fields, so adding a new domain is non-breaking.
    """
    waves: list[list["_ArtifactJob"]] = [[] for _ in _DEPENDENCY_WAVES]
    catch_all_idx = len(_DEPENDENCY_WAVES) - 1
    for job in jobs:
        placed = False
        for i, wave_domains in enumerate(_DEPENDENCY_WAVES):
            if job.domain in wave_domains:
                waves[i].append(job)
                placed = True
                break
        if not placed:
            waves[catch_all_idx].append(job)
    return waves


# --- Job model + builder ------------------------------------------------

@dataclass
class _ArtifactJob:
    artifact: MigrationArtifact
    domain: str
    mappings: list[dict[str, Any]]


def _build_jobs(bundle: MigrationBundle) -> list[_ArtifactJob]:
    mapping_summary = bundle.mapping_summary or {}
    per_artifact_mappings = mapping_summary.get("per_artifact") or {}
    per_artifact_domain = (
        (bundle.discovery_summary or {}).get("per_artifact_domain") or {}
    )

    jobs: list[_ArtifactJob] = []
    for artifact in bundle.artifacts.filter(quarantined=False):
        # Skip parent archive shells — they hold no rows.
        if artifact.detected_format == "archive":
            continue
        mappings = per_artifact_mappings.get(artifact.path_within_bundle) or []
        domain_entry = per_artifact_domain.get(artifact.path_within_bundle) or {}
        # Tenant/operator per-file override on the artifact row wins over discovery.
        assigned = (getattr(artifact, "assigned_domain", None) or "").strip()
        domain = assigned or domain_entry.get("domain", "custom_fields")
        if not mappings:
            # If U4 never ran for this artifact, default everything to custom_fields
            # unless the tenant explicitly assigned a domain (still land via that
            # lander with empty mappings → lander quarantines bad rows).
            if not assigned:
                domain = "custom_fields"
        jobs.append(_ArtifactJob(artifact=artifact, domain=domain, mappings=mappings))
    return jobs


# --- Per-artifact apply -------------------------------------------------

def _apply_artifact(
    bundle: MigrationBundle,
    job: _ArtifactJob,
    *,
    dry_run: bool,
    progress_tracker: Any | None = None,
) -> ArtifactApplyOutcome:
    outcome = ArtifactApplyOutcome(
        artifact_id=job.artifact.pk,
        path_within_bundle=job.artifact.path_within_bundle,
        domain=job.domain,
        migration_run_id=None,
    )

    run = _create_audit_run(bundle, job, dry_run=dry_run)
    outcome.migration_run_id = getattr(run, "pk", None)

    lander = get_lander(job.domain) or get_lander("custom_fields")
    if lander is None:
        outcome.status = "FAILED"
        outcome.error = f"No lander registered for domain {job.domain!r} (or custom_fields fallback)"
        _finalize_audit_run(run, outcome, status="FAILED")
        return outcome

    try:
        from .apply_stall import maybe_stall_pulse

        maybe_stall_pulse()
        rows_iter = _iter_canonical_rows(job)
        if progress_tracker is not None and not dry_run:
            label = job.artifact.filename or job.artifact.path_within_bundle or job.domain
            rows_iter = progress_tracker.wrap_rows(rows_iter, artifact_label=label)
        result = _run_lander_under_schema(
            lander=lander,
            rows_iter=rows_iter,
            bundle=bundle,
            artifact=job.artifact,
            dry_run=dry_run,
        )
    except LanderError as exc:
        outcome.status = "FAILED"
        outcome.error = str(exc)
        _finalize_audit_run(run, outcome, status="FAILED")
        return outcome
    except Exception as exc:  # noqa: BLE001
        logger.exception(
            "migration_cloud.apply: unexpected error applying artifact %s",
            job.artifact.pk,
        )
        outcome.status = "FAILED"
        outcome.error = f"{type(exc).__name__}: {exc}"
        _finalize_audit_run(run, outcome, status="FAILED")
        return outcome

    outcome.result = result
    # Total rejection is a FAILURE, not a partial success. An artifact that
    # quarantined rows and landed NOTHING wrote nothing at all, yet the old
    # `PARTIAL if quarantined else SUCCESS` reported it green: a live tenant saw
    # five consecutive repairs report `succeeded` while every one of them wrote
    # 0 created / 0 updated / 431 quarantined, so nobody looked for hours. Zero
    # rows landed against a non-empty input is exactly the signal an operator
    # needs raised, and marking it FAILED also makes repair_readiness treat the
    # bundle as repairable once the source is corrected.
    #
    # `quarantined` is required for this branch: an artifact that legitimately had
    # nothing to do (header-only file, every row already current) reports 0/0/0 and
    # must stay SUCCESS rather than being called a failure.
    outcome.status = artifact_outcome_status(result)
    if outcome.status == "REJECTED":
        outcome.error = (
            f"Every row was rejected ({result.quarantined} of {result.quarantined}); "
            "nothing was imported from this file."
        )
    _finalize_audit_run(run, outcome, status=outcome.status)
    _quarantine_errors(
        bundle=bundle, run=run, artifact=job.artifact, domain=job.domain, result=result
    )
    return outcome


# --- Row iteration + transformer application ----------------------------

# Formats we CAN stream into rows. A row-bearing artifact that reaches apply with
# NO readable bytes (no captured blob, no top-level source file) is a genuine
# source-availability failure, not an empty success — see _iter_canonical_rows.
# Container / non-tabular formats (archive is dropped in _build_jobs; image / sql /
# sqlite / parquet do not stream rows through this path) are deliberately excluded
# so they keep the lenient empty-yield instead of false-failing.
_ROW_BEARING_FORMATS = frozenset({"csv", "tsv", "unknown", "json", "jsonl", "xlsx", "xls", "pdf"})


def _iter_canonical_rows(job: _ArtifactJob) -> Iterator[dict[str, Any]]:
    """Stream the artifact's bytes, apply mappings + transformers, yield canonical rows.

    When the parent bundle has ``diff_mode='since'`` set with ``diff_since``,
    rows older than the threshold are filtered out before reaching the lander.
    """
    from .apply_stall import maybe_stall_pulse, read_with_stall_pulse

    maybe_stall_pulse()
    artifact = job.artifact
    mapping_index = {m["source_column"]: m for m in job.mappings}
    locale_hints = dict(artifact.locale_hints or {})
    diff_threshold = None
    if getattr(artifact.bundle, "diff_mode", "full") == "since":
        diff_threshold = getattr(artifact.bundle, "diff_since", None)

    # Surface the tenant's country to every transformer so country-aware
    # transformers (grading_scale_to_canonical, name_split_locale,
    # attendance_code_rewrite) can resolve scale / dialect / name-order
    # without needing per-mapping options.
    # An explicit operator choice outranks BOTH the profiler's vote and the
    # country default below -- it is the only one of the three that is knowledge
    # rather than inference.
    apply_operator_date_override(locale_hints, artifact.bundle)

    school = getattr(artifact.bundle, "school", None)
    country = str(getattr(school, "country_code", "") or "").upper()
    if country:
        locale_hints.setdefault("country", country)
        # Seed locale + date format from the tenant's country profile so the date
        # transformer disambiguates DD/MM vs MM/DD. The profiler's own per-column
        # date_format vote (evidence from the data) is already in locale_hints and
        # WINS — we only FILL when it produced none, so an all-days-<=12 US export
        # is not misread as EU. See docs/MIGRATION_CLOUD_AUDIT_2026_07_24.md (B-5).
        locale_hints.setdefault("locale", country)
        if "date_format" not in locale_hints:
            try:
                from .country_profiles import resolved_country_profile
                prof = resolved_country_profile(country)
                if prof is not None and getattr(prof, "date_format", ""):
                    locale_hints["date_format"] = prof.date_format
            except Exception:  # noqa: BLE001 — hint seeding is best-effort
                pass

    # Phase U5 content store (gap #2): if the source bytes were captured at
    # ingest, stream them from the encrypted blob — this is what lets archive
    # members + remote / OAuth-folder pulls apply real rows instead of nothing.
    from .artifact_blob_store import open_artifact_blob_stream

    blob_stream, blob_encoding = open_artifact_blob_stream(artifact)
    if blob_stream is not None:
        try:
            raw_bytes = read_with_stall_pulse(blob_stream)
        finally:
            try:
                blob_stream.close()
            except Exception:  # noqa: BLE001
                pass
        maybe_stall_pulse()
        fmt = artifact.detected_format
        if fmt in ("csv", "tsv", "unknown", "json", "jsonl"):
            text = raw_bytes.decode(blob_encoding or "utf-8", errors="replace")
            maybe_stall_pulse()
            if fmt == "json":
                raw_iter = _iter_json_rows_text(text, mapping_index, locale_hints)
            elif fmt == "jsonl":
                raw_iter = _iter_jsonl_rows_stream(io.StringIO(text), mapping_index, locale_hints)
            else:
                raw_iter = _iter_csv_rows_stream(io.StringIO(text), mapping_index, locale_hints)
        elif fmt in ("xlsx", "xls"):
            # Binary spreadsheet: read from bytes, NOT the decoded text (which
            # would be garbage). openpyxl/xlrd degrade to no rows if absent.
            raw_iter = _iter_spreadsheet_rows_bytes(raw_bytes, fmt, mapping_index, locale_hints)
        elif fmt == "pdf":
            # PDF: extract + tabularise via the shared extractor, then read the
            # resulting TSV. Digital PDFs land rows; scanned-without-OCR → none.
            raw_iter = _iter_pdf_rows_bytes(raw_bytes, mapping_index, locale_hints)
        else:
            return iter(())
        return _emit_canonical_rows(raw_iter, domain=job.domain, diff_threshold=diff_threshold)

    bundle_uri = artifact.bundle.intake_source_uri or ""
    path = Path(bundle_uri) if bundle_uri else None
    if path is None or not path.exists() or artifact.path_within_bundle != path.name:
        # No captured blob AND no top-level local file to fall back to: we have NO
        # bytes for this artifact. Silently yielding nothing here stamped the
        # artifact a green SUCCESS with 0 rows, so the operator believed a file
        # imported when nothing was even readable (#5 — "no source-blob reports
        # SUCCESS"). For a row-bearing format that is a real, repairable failure —
        # the file's contents never reached apply (ingest blob-capture gap / lost
        # local file) — so raise; _apply_artifact marks the outcome FAILED with an
        # honest message and the bundle stays repairable (re-ingest, then retry).
        # Non-row / container formats keep the lenient empty-yield: they were never
        # going to produce rows through this path, so failing them would be noise.
        if artifact.detected_format in _ROW_BEARING_FORMATS:
            raise LanderError(
                f"No source data available for {artifact.path_within_bundle!r}: its "
                "bytes were not captured at ingest and no source file is present, so "
                "nothing could be imported. Re-ingest the bundle to re-capture the "
                "file's contents, then retry the import."
            )
        return iter(())

    encoding = artifact.encoding or "utf-8"

    if artifact.detected_format in ("csv", "tsv", "unknown"):
        raw_iter = _iter_csv_rows(path, encoding, mapping_index, locale_hints)
    elif artifact.detected_format == "json":
        raw_iter = _iter_json_rows(path, encoding, mapping_index, locale_hints)
    elif artifact.detected_format == "jsonl":
        raw_iter = _iter_jsonl_rows(path, encoding, mapping_index, locale_hints)
    elif artifact.detected_format in ("xlsx", "xls"):
        maybe_stall_pulse()
        raw_iter = _iter_spreadsheet_rows_bytes(
            path.read_bytes(), artifact.detected_format, mapping_index, locale_hints
        )
    elif artifact.detected_format == "pdf":
        maybe_stall_pulse()
        raw_iter = _iter_pdf_rows_bytes(path.read_bytes(), mapping_index, locale_hints)
    else:
        return iter(())
    return _emit_canonical_rows(raw_iter, domain=job.domain, diff_threshold=diff_threshold)


def _emit_canonical_rows(raw_iter, *, domain: str, diff_threshold):
    """Apply diff filter then promote leftover headers onto empty canonical fields."""
    if diff_threshold is not None:
        from .diff_mode import row_passes_diff_filter

        raw_iter = (
            row for row in raw_iter if row_passes_diff_filter(row=row, threshold=diff_threshold)
        )
    domain = (domain or "").strip()
    if not domain or domain == "custom_fields":
        return raw_iter
    from .mapper import bind_residual_headers

    return (bind_residual_headers(row, domain) for row in raw_iter)


def _iter_csv_rows(
    path: Path,
    encoding: str,
    mapping_index: dict[str, dict[str, Any]],
    locale_hints: dict[str, Any],
) -> Iterator[dict[str, Any]]:
    with path.open("r", encoding=encoding, errors="replace", newline="") as fh:
        # ``yield from`` keeps the generator frame — and thus the open file —
        # alive until the caller has consumed every row.
        yield from _iter_csv_rows_stream(fh, mapping_index, locale_hints)


def _strip_leading_comment_lines(fh: Any) -> Iterator[str]:
    """Yield lines from ``fh``, dropping ONLY leading blank + ``#``-comment lines.

    The RunMyCampus canonical template ships a leading
    ``# runmycampus-canonical-template: ...`` marker line. Without dropping it,
    ``csv.DictReader`` takes that comment as the sole fieldname and every real
    row misroutes to quarantine on apply — so a school that downloaded a
    template, filled it in and re-uploaded it lost its data at import (the
    profiler was fixed but this apply-path reader was not). Only LEADING lines
    are dropped; once real content starts, every row is yielded unchanged, so a
    data value that begins with ``#`` is never lost and quoted newlines still
    reassemble (csv pulls more lines from this iterator inside a quote).
    """
    header_seen = False
    for line in fh:
        if not header_seen:
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            header_seen = True
        yield line


def _iter_csv_rows_stream(
    fh: Any,
    mapping_index: dict[str, dict[str, Any]],
    locale_hints: dict[str, Any],
) -> Iterator[dict[str, Any]]:
    """Core CSV row iterator over any seekable text stream (file OR blob-backed StringIO)."""
    sample = fh.read(4096)  # magic-number-allow: file-read-chunk-bytes
    fh.seek(0)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(_strip_leading_comment_lines(fh), dialect=dialect)
    for raw_row in reader:
        yield _transform_row(raw_row, mapping_index, locale_hints)


def _iter_json_rows(
    path: Path,
    encoding: str,
    mapping_index: dict[str, dict[str, Any]],
    locale_hints: dict[str, Any],
) -> Iterator[dict[str, Any]]:
    text = path.read_text(encoding=encoding, errors="replace")
    yield from _iter_json_rows_text(text, mapping_index, locale_hints)


def _iter_json_rows_text(
    text: str,
    mapping_index: dict[str, dict[str, Any]],
    locale_hints: dict[str, Any],
) -> Iterator[dict[str, Any]]:
    """Core JSON row iterator over already-decoded text (file OR blob-backed)."""
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        return
    for raw_row in data:
        if isinstance(raw_row, dict):
            yield _transform_row(raw_row, mapping_index, locale_hints)


def _iter_jsonl_rows(
    path: Path,
    encoding: str,
    mapping_index: dict[str, dict[str, Any]],
    locale_hints: dict[str, Any],
) -> Iterator[dict[str, Any]]:
    with path.open("r", encoding=encoding, errors="replace") as fh:
        yield from _iter_jsonl_rows_stream(fh, mapping_index, locale_hints)


def _iter_jsonl_rows_stream(
    fh: Any,
    mapping_index: dict[str, dict[str, Any]],
    locale_hints: dict[str, Any],
) -> Iterator[dict[str, Any]]:
    """Core JSONL row iterator over any line-iterable text stream (file OR blob-backed StringIO)."""
    for line in fh:
        line = line.strip()
        if not line:
            continue
        try:
            raw_row = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(raw_row, dict):
            yield _transform_row(raw_row, mapping_index, locale_hints)


def _stringify_cell(cell: Any) -> str:
    """Coerce a spreadsheet cell to the string shape downstream expects.

    CSV rows arrive as strings; transformers assume strings. Integers stored
    as floats by the reader (``36.0``) are rendered as ``"36"`` to match what
    the same value would look like in a CSV export.
    """
    if cell is None:
        return ""
    if isinstance(cell, str):
        return cell
    if isinstance(cell, float) and cell.is_integer():
        return str(int(cell))
    return str(cell)


def _xlsx_rows(raw_bytes: bytes) -> tuple[list[Any], Iterator[Any]]:
    """Return ``(header_row, data_row_iterator)`` for an in-memory XLSX.

    ``([], iter(()))`` when openpyxl is unavailable or the file is unreadable,
    so apply degrades to zero rows instead of raising. The workbook is held
    open until the data iterator is exhausted (read-only, streaming).
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except Exception:  # noqa: BLE001
        return [], iter(())
    try:
        wb = load_workbook(filename=io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return [], iter(())
    ws = wb[wb.sheetnames[0]] if wb.sheetnames else None
    if ws is None:
        _close_quietly(wb)
        return [], iter(())
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        _close_quietly(wb)
        return [], iter(())

    def _gen() -> Iterator[Any]:
        try:
            for row in rows_iter:
                yield row
        finally:
            _close_quietly(wb)

    return list(header_row), _gen()


def _xls_rows(raw_bytes: bytes) -> tuple[list[Any], Iterator[Any]]:
    """Return ``(header_row, data_row_iterator)`` for a legacy in-memory XLS.

    ``([], iter(()))`` when xlrd is unavailable or the file is unreadable.
    """
    try:
        import xlrd  # type: ignore[import-not-found]
    except Exception:  # noqa: BLE001
        return [], iter(())
    try:
        wb = xlrd.open_workbook(file_contents=raw_bytes)
    except Exception:  # noqa: BLE001
        return [], iter(())
    if wb.nsheets == 0:
        return [], iter(())
    sh = wb.sheet_by_index(0)
    if sh.nrows == 0:
        return [], iter(())
    header_row = [sh.cell_value(0, c) for c in range(sh.ncols)]

    def _gen() -> Iterator[Any]:
        for r in range(1, sh.nrows):
            yield [sh.cell_value(r, c) for c in range(sh.ncols)]

    return header_row, _gen()


def _close_quietly(wb: Any) -> None:
    try:
        wb.close()
    except Exception:  # noqa: BLE001
        pass


def _iter_spreadsheet_rows_bytes(
    raw_bytes: bytes,
    fmt: str,
    mapping_index: dict[str, dict[str, Any]],
    locale_hints: dict[str, Any],
) -> Iterator[dict[str, Any]]:
    """Yield canonical rows from an in-memory XLSX/XLS workbook.

    First worksheet, first row = headers. All cells are stringified so the
    downstream transformers see the same shape they get from CSV. This is the
    apply-time counterpart to the profiler's ``_read_xlsx`` / ``_read_xls`` —
    without it, an Excel upload classifies but lands zero rows.
    """
    header_row, data_rows = _xls_rows(raw_bytes) if fmt == "xls" else _xlsx_rows(raw_bytes)
    if not header_row:
        return
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    for row in data_rows:
        raw_row: dict[str, Any] = {}
        for h, cell in zip(headers, row):
            if not h:
                continue  # unnamed column — nothing to map onto
            raw_row[h] = _stringify_cell(cell)
        if not any(str(v).strip() for v in raw_row.values()):
            continue  # skip fully-blank trailing rows
        yield _transform_row(raw_row, mapping_index, locale_hints)


def _iter_pdf_rows_bytes(
    raw_bytes: bytes,
    mapping_index: dict[str, dict[str, Any]],
    locale_hints: dict[str, Any],
) -> Iterator[dict[str, Any]]:
    """Yield canonical rows from a PDF's extracted + tabularised text.

    Digitally-generated PDFs (pdfplumber) yield rows; scanned PDFs with no OCR
    binaries yield nothing (the review surface explains why). Reuses the CSV
    row iterator over the TSV the shared extractor produces, so a PDF lands
    exactly like the equivalent CSV would.
    """
    from .pdf_extract import extract_pdf_tsv

    tsv = extract_pdf_tsv(raw_bytes)
    if not tsv.strip():
        return
    yield from _iter_csv_rows_stream(io.StringIO(tsv), mapping_index, locale_hints)


# Value-hygiene for messy exports (pandas / spreadsheet dumps). A pandas
# ``to_csv`` writes NaN as the literal ``nan`` and None as ``None``; a numeric
# id read as a float is written ``241904748.0``. Stored verbatim these become a
# student's parent named "None", an admission number "nan", or an id that no
# longer round-trips. Normalised centrally so EVERY reader (CSV/XLSX/JSON) and
# EVERY domain gets the same clean value.
_NULL_LITERALS: frozenset[str] = frozenset(
    {"nan", "none", "null", "n/a", "#n/a", "nil", "(null)", "\\n", "\\N"}
)
_INT_FLOAT_RE = re.compile(r"^-?\d+\.0+$")


def _normalize_source_value(value: Any) -> Any:
    """Fold export sentinels to a real null and repair spreadsheet float-ids.

    ``nan``/``None``/``null``/``N/A`` → ``""`` (a genuine empty, not a literal
    string); an integer-valued float string (``"241904748.0"``) → its integer
    form so a text id round-trips exactly. Genuine decimals (``"36.5"``) and
    non-string values are returned untouched.
    """
    if not isinstance(value, str):
        return value
    s = value.strip()
    if not s:
        return value
    if s.lower() in _NULL_LITERALS:
        return ""
    if _INT_FLOAT_RE.match(s):
        return s.split(".", 1)[0]
    return value


def _transform_row(
    raw_row: dict[str, Any],
    mapping_index: dict[str, dict[str, Any]],
    locale_hints: dict[str, Any],
) -> dict[str, Any]:
    """Apply column mappings + transformers; return canonical-keyed dict."""
    canonical: dict[str, Any] = {}
    for source_col, raw_value in raw_row.items():
        raw_value = _normalize_source_value(raw_value)
        mapping = mapping_index.get(source_col)
        if mapping is None and isinstance(source_col, str):
            # The PROFILER strips surrounding whitespace from headers (so the
            # mapping's source_column is "TEACHER UNIQUE ID"), but the apply-path
            # csv.DictReader keys the row by the RAW header (" TEACHER UNIQUE ID"
            # with the leading space real African/TVET exports carry). Without
            # this whitespace-tolerant retry the padded column never joins its
            # mapping, lands in _unmapped, and a required id (staff_external_id)
            # is silently lost — quarantining every row. BOM is deliberately NOT
            # stripped here: the profiler keeps it ("﻿NAME"), so the exact
            # match above already covers it and both sides stay consistent.
            stripped = source_col.strip()
            if stripped != source_col:
                mapping = mapping_index.get(stripped)
        if mapping is None:
            # Unmapped column — drop into custom_fields key for the lander to pick up.
            canonical[f"_unmapped.{source_col}"] = raw_value
            continue

        canonical_field = mapping.get("canonical_field", "")
        if canonical_field.startswith("custom_fields."):
            canonical[canonical_field] = raw_value
            continue

        transformer_name = mapping.get("transformer")
        if not transformer_name:
            canonical[canonical_field] = raw_value
            continue

        transformer = get_transformer(transformer_name)
        if transformer is None:
            canonical[canonical_field] = raw_value
            continue

        ctx = TransformerContext(
            canonical_field=canonical_field,
            hints=locale_hints,
            options=mapping.get("transformer_options") or {},
        )
        try:
            canonical[canonical_field] = transformer.transform(raw_value, ctx)
        except TransformerError as exc:
            # Per-row transformer failure: store raw + flag for quarantine via
            # a sentinel error key the lander recognises.
            canonical[canonical_field] = raw_value
            canonical.setdefault("_transformer_errors", []).append(
                f"{canonical_field}: {exc!s}"
            )
    return canonical


# --- Tenant-schema wrapper ---------------------------------------------

# Upper bound on residual (custom_fields.*/_unmapped.*) values the no-loss net
# buffers per artifact before it stops accumulating. A guard against a pathological
# file; overflow is reported on the LanderResult (never silently dropped-and-hidden).
_RESIDUAL_MAX_ENTRIES = 200_000


class _ResidualCapture:
    """No-data-loss net behind a non-sweeping lander.

    The orchestrator emits every unmapped/below-threshold source column as a
    ``_unmapped.<col>`` / ``custom_fields.<slug>`` key on the canonical row
    (``_transform_row``). A lander that does not read those keys would drop them
    — and ``_apply_artifact`` runs exactly ONE lander per artifact with no
    fallback (``get_lander(domain) or get_lander("custom_fields")`` only reaches
    the fallback for domains with no registered lander). This tee captures those
    residual values as the rows stream to the lander, then persists them to
    ``DynamicFieldValue`` under ``migration_residual:<domain>`` so "ingest
    everything" holds for EVERY domain — including ones added in the future,
    which default to ``sweeps_custom_columns = False``.

    It is additive, not duplicative: the landers that call ``persist_dfv_extras``
    directly persist only *canonical* extras (e.g. ``grade_level``), never the
    ``custom_fields.*``/``_unmapped.*`` residual keys captured here.
    """

    def __init__(self, domain: str) -> None:
        self.domain = domain
        self._buffer: list[tuple[int, str, dict[str, Any]]] = []
        self._row_index = 0
        self.truncated = 0

    def wrap(self, rows_iter: Iterator[dict[str, Any]]) -> Iterator[dict[str, Any]]:
        for row in rows_iter:
            self._capture(row)
            yield row

    def _capture(self, row: dict[str, Any]) -> None:
        self._row_index += 1
        residual: dict[str, Any] = {}
        external_id = ""
        for key, value in row.items():
            if value in (None, ""):
                continue
            if key.startswith("custom_fields.") or key.startswith("_unmapped."):
                clean_key = key.split(".", 1)[1] if "." in key else key
                if clean_key:
                    residual[clean_key] = value
            elif not external_id and (key == "external_id" or key.endswith("_external_id")):
                external_id = str(value)
        if not residual:
            return
        if len(self._buffer) >= _RESIDUAL_MAX_ENTRIES:
            self.truncated += 1
            return
        self._buffer.append((self._row_index, external_id, residual))

    def flush(self, *, ctx, result: LanderResult) -> None:
        if not self._buffer and not self.truncated:
            return
        from .landers._helpers import persist_dfv_extras

        entity_type = f"migration_residual:{self.domain}"
        for row_index, external_id, residual in self._buffer:
            entity_id = external_id or f"a{ctx.artifact_id}r{row_index}"
            persist_dfv_extras(
                ctx=ctx,
                entity_type=entity_type,
                entity_id=entity_id,
                extras=residual,
                result=result,
            )
        if self.truncated and result is not None:
            result.errors.append(
                f"residual-capture: {self.truncated} rows exceeded the "
                f"{_RESIDUAL_MAX_ENTRIES}-entry cap and were not swept"
            )


def _bundle_schema_context(bundle: MigrationBundle):
    """Enter ``bundle.schema_name`` for tenant ORM writes (landers + post-apply hooks)."""
    from contextlib import nullcontext

    schema_name = (getattr(bundle, "schema_name", None) or "").strip()
    if not schema_name:
        return nullcontext()
    try:
        from django_tenants.utils import schema_context
    except ImportError:
        return nullcontext()
    from django.db import connection

    if not hasattr(connection, "set_schema"):
        return nullcontext()
    return schema_context(schema_name)


def _run_lander_under_schema(
    *,
    lander,
    rows_iter,
    bundle: MigrationBundle,
    artifact: MigrationArtifact,
    dry_run: bool,
) -> LanderResult:
    """Wrap the lander call in ``schema_context(bundle.schema_name)`` for tenant scoping.

    A residual-capture net runs behind any lander that does not itself sweep the
    ``custom_fields.*``/``_unmapped.*`` pass-through keys, so no source column is
    ever dropped regardless of which domain (or a future lander) processes it.
    """
    from .landers.base import LanderContext  # local to keep import surface clean

    ctx = LanderContext(
        school=bundle.school,
        schema_name=bundle.schema_name,
        bundle_id=bundle.pk,
        artifact_id=artifact.pk,
        dry_run=dry_run,
        # Operator transform preferences chosen on the review page (currently the
        # combined-name order). LanderContext has always carried this field; it
        # was never populated, so a school could not correct a mis-split name --
        # e.g. a roster written surname-first, where auto-detection reads the
        # family name as the given name for every student in the file.
        # getattr: the residual-net tests hand in a lightweight bundle stub, and a
        # missing mapping_summary must degrade to "no preference", never raise.
        transformer_options=dict(
            (getattr(bundle, "mapping_summary", None) or {}).get("transform_prefs") or {}
        ),
    )

    capture: _ResidualCapture | None = None
    if not dry_run and not getattr(lander, "sweeps_custom_columns", False):
        capture = _ResidualCapture(lander.domain or "unknown")
        rows_iter = capture.wrap(rows_iter)

    def _land() -> LanderResult:
        result = lander.land(canonical_rows=rows_iter, ctx=ctx)
        if capture is not None:
            capture.flush(ctx=ctx, result=result)
        return result

    with _bundle_schema_context(bundle):
        return _land()


# --- Audit + quarantine -------------------------------------------------

def _create_audit_run(bundle: MigrationBundle, job: _ArtifactJob, *, dry_run: bool):
    """Create one ``apps.automation.MigrationRun`` per (domain, artifact)."""
    try:
        from apps.automation.models import MigrationRun
    except ImportError:
        return None

    return MigrationRun.objects.create(
        school=bundle.school,
        migration_type=f"{job.domain}:{job.artifact.path_within_bundle}"[:64],
        dry_run=dry_run,
        status=MigrationRun.Status.PENDING,
        triggered_by=bundle.triggered_by,
        execution_summary={
            "bundle_id": bundle.pk,
            "artifact_id": job.artifact.pk,
            "domain": job.domain,
            "mapping_count": len(job.mappings),
        },
    )


def _json_safe(value: Any) -> Any:
    """Make lander audit payloads JSONField-safe (no model instances)."""
    from .landers._helpers import json_field_safe

    return json_field_safe(value)


def _finalize_audit_run(run, outcome: ArtifactApplyOutcome, *, status: str) -> None:
    if run is None:
        return
    try:
        from apps.automation.models import MigrationRun
    except ImportError:
        return

    run.mark_completed(
        status=_RUN_STATUS_MAP.get(status, MigrationRun.Status.FAILED),
        created_count=outcome.result.created,
        updated_count=outcome.result.updated,
        error_count=outcome.result.quarantined,
        error_message=outcome.error[:2000] if outcome.error else "",
        summary={
            **(run.execution_summary or {}),
            "created_ids": outcome.result.created_ids[:200],  # cap for size
            "updated_ids_with_old_values": _json_safe(
                outcome.result.updated_ids_with_old_values[:200]
            ),
            "errors_sample": outcome.result.errors[:20],
        },
    )
    # Rollback snapshot: minimal enough to revert. Must be JSON-safe — landers
    # often snapshot FK fields via getattr(obj, "school") which is a model inst.
    run.rollback_snapshot = _json_safe(
        {
            "created_ids": outcome.result.created_ids,
            "updated_ids_with_old_values": outcome.result.updated_ids_with_old_values,
            "domain": outcome.domain,
            "artifact_id": outcome.artifact_id,
        }
    )
    run.save(update_fields=["rollback_snapshot"])


# Per-artifact ceiling on durable quarantine records. A runaway lander must not
# be able to write a million rows, but exceeding this is a REPORTED event, never
# a silent one — see _quarantine_errors.
# Bundle 84 held 326 student row errors; the old cap of 200 dropped 126 with no record.
QUARANTINE_RECORD_CAP = 2000  # magic-number-allow: quarantine-records-per-artifact

# How many partial-write notes to name in one log line. The count is always
# exact; this only bounds the sample, so one pathological artifact cannot write a
# megabyte into the log.
_LANDER_NOTE_LOG_CAP = 20  # magic-number-allow: log-line-sample-cap


def _classify_quarantine_issue(err: str) -> str:
    """Bucket a lander error STRING into a ``MigrationQuarantineRecord.issue_class``.

    The legacy path, kept for rows that arrive without a declared reason code.
    It now delegates to ``landers.reason_codes.classify_message`` so there is one
    implementation of the rule rather than two that can drift — landers were
    already classifying at the point of failure while this classified again at
    write time, from the same string, with a copy of the same regex-free matcher.

    Prefer the declared code. Substring-matching English is why 60 of 106 per-row
    failure sites read as ``lander_error`` — "a person must look at this" — when
    11 of them were plainly a missing field or an unresolvable reference.
    """
    return classify_message(err)


def _clear_superseded_quarantine(bundle: MigrationBundle) -> int:
    """Drop PENDING quarantine rows from earlier applies of this same bundle.

    An apply regenerates its held rows from scratch, so every re-apply used to
    append a whole fresh copy. Bundle 84 re-applied 128 times and accumulated
    40,448 quarantine records that were 128 identical copies of the same 316 --
    growing by 316 every thirty minutes, with no upper bound.

    Only PENDING rows from PRIOR runs are removed: anything already REPAIRED or
    FAILED is a resolution someone or something reached, and deleting that would
    destroy the audit trail the whole review surface depends on.
    """
    try:
        from apps.automation.models import MigrationQuarantineRecord, MigrationRun

        # Runs link to a bundle only through execution_summary["bundle_id"] --
        # there is no FK (see views.py: the old parent_bundle_id filter raised a
        # swallowed FieldError and left this surface permanently empty).
        prior_run_ids = list(
            MigrationRun.objects.filter(  # tenant-isolation-allow: bundle pk is globally unique and the bundle is already tenant-scoped
                execution_summary__bundle_id=bundle.pk
            ).values_list("pk", flat=True)
        )
        if not prior_run_ids:
            return 0
        deleted, _detail = MigrationQuarantineRecord.objects.filter(  # tenant-isolation-allow: scoped transitively by the bundle's own runs
            migration_run_id__in=prior_run_ids,
            status=MigrationQuarantineRecord.Status.PENDING,
        ).delete()
    except Exception:  # noqa: BLE001 - housekeeping must never block an apply
        logger.warning(
            "migration_cloud.apply: superseded-quarantine sweep failed for bundle %s",
            bundle.pk,
            exc_info=True,
        )
        return 0
    if deleted:
        logger.info(
            "migration_cloud.apply: cleared %s superseded PENDING quarantine row(s) "
            "for bundle %s before re-applying",
            deleted,
            bundle.pk,
        )
    return int(deleted or 0)


def _quarantine_errors(
    *,
    bundle: MigrationBundle,
    run,
    artifact: MigrationArtifact,
    domain: str,
    result: LanderResult,
) -> None:
    """Persist per-row failures to ``apps.automation.MigrationQuarantineRecord``.

    Every held row carries the SOURCE ROW that failed, the reason as a code, and
    the offending field where the lander knew it — because a row you did not keep
    is a row you cannot replay, and a reason you have to re-derive from English at
    read time is a reason you will re-derive wrongly.

    **Rows are paired to errors BY INDEX, not by message.** The previous
    implementation built a ``{error_string: row}`` dict, so two rows failing with
    the same message collapsed onto one entry and every row but the last silently
    lost its snapshot. Most messages do not interpolate the row, so most
    multi-row failures hit exactly that. ``record_row_error`` appends to
    ``errors`` and ``error_rows`` in lockstep; index alignment cannot collide.

    A lander that appends a bare string still works — it just lands with no row,
    no field, and a reason guessed from its text. Those are counted and logged so
    the remaining backlog is a number rather than an impression.

    NOTE: an older implementation wrote ``row_snapshot=``/``reason=`` — fields
    that do not exist on the model — so every ``create()`` raised ``TypeError``
    swallowed by the guard below, and quarantine rows were NEVER persisted (silent
    data loss). This writes the real model fields.
    """
    _record_lander_notes(bundle=bundle, artifact=artifact, domain=domain, result=result)
    if not result.errors:
        return
    try:
        from apps.automation.models import MigrationQuarantineRecord
    except ImportError:
        return

    # Positional pairing. Only trust it when the two lists actually correspond:
    # a lander that mixes record_row_error with a bare append would otherwise
    # shift every row onto the wrong error.
    structured: list[dict[str, Any]] = [
        er for er in (getattr(result, "error_rows", None) or []) if isinstance(er, dict)
    ]
    aligned = len(structured) == len(result.errors) and all(
        str(er.get("error")) == str(err)
        for er, err in zip(structured, result.errors)
    )
    if structured and not aligned:
        logger.warning(
            "migration_cloud.apply: %s structured row(s) do not align with %s error(s) "
            "for bundle=%s domain=%s — falling back to message pairing, so rows that "
            "share an error message will lose their snapshot",
            len(structured), len(result.errors), bundle.pk, domain or "",
        )
    row_by_error: dict[str, Any] = {}
    if not aligned:
        for er in structured:
            if "error" in er:
                row_by_error[str(er.get("error"))] = er.get("row")

    domain_label = ((domain or "") or (artifact.path_within_bundle or ""))[:32]

    _undeclared = sum(
        1 for er in structured if (er.get("reason_source") or "fallback") != "declared"
    ) + max(0, len(result.errors) - len(structured))
    if _undeclared:
        logger.info(
            "migration_cloud.apply: %s of %s held row(s) for bundle=%s domain=%s were "
            "classified by matching the error text rather than a declared reason_code "
            "— that is the lander-contract backlog, not a per-run problem",
            _undeclared, len(result.errors), bundle.pk, domain_label,
        )

    # The cap is a runaway guard, not a review policy — but it was SILENT. The
    # board counts every held row (totals["quarantined"]) while only the first
    # QUARANTINE_RECORD_CAP per artifact were ever written, so a tenant is told
    # "442 held for review" and can only ever see some of them. The rest are
    # counted and gone. One live artifact held 326 rows: 126 of them left no
    # record at all. Say so, loudly, rather than letting the two numbers drift.
    _dropped = max(0, len(result.errors) - QUARANTINE_RECORD_CAP)
    if _dropped:
        logger.error(
            "migration_cloud.apply: quarantine truncated for bundle=%s domain=%s — "
            "%s row error(s) held but only %s recorded; %s have NO durable record "
            "and cannot be reviewed or replayed",
            bundle.pk, domain_label, len(result.errors), QUARANTINE_RECORD_CAP, _dropped,
        )
    for idx, err in enumerate(result.errors[:QUARANTINE_RECORD_CAP], start=1):
        payload = {"error": err, "artifact": artifact.path_within_bundle}
        entry = structured[idx - 1] if aligned else {}
        source_row = entry.get("row") if aligned else row_by_error.get(str(err))
        if source_row is not None:
            payload["source_row"] = source_row
        if entry.get("field"):
            payload["field"] = entry["field"]
        # Record HOW the class was decided. A remediation pass must be able to
        # tell a class the lander asserted from one a matcher guessed, and to
        # refuse to act automatically on a guess.
        payload["reason_source"] = entry.get("reason_source") or "fallback"
        issue_class = (
            normalize_reason_code(entry.get("reason_code"))
            or _classify_quarantine_issue(err)
        )
        try:
            # Savepoint: this runs INSIDE the forced-atomic finance apply, and the
            # swallow below would otherwise leave a failed create's needs_rollback set
            # on the whole transaction — poisoning every subsequent write. The
            # savepoint keeps the swallow clean (same rule as landers._helpers).
            with transaction.atomic():
                MigrationQuarantineRecord.objects.create(
                    school=bundle.school,
                    migration_run=run,
                    domain=domain_label,
                    row_index=idx,
                    payload=payload,
                    issue_class=issue_class,
                    status=MigrationQuarantineRecord.Status.PENDING,
                )
        except Exception:  # noqa: BLE001 — quarantine writes never block apply
            logger.warning(
                "migration_cloud.apply: quarantine write skipped for domain=%s row=%s",
                domain_label,
                idx,
                exc_info=True,
            )


def _record_lander_notes(
    *,
    bundle: MigrationBundle,
    artifact: MigrationArtifact,
    domain: str,
    result: LanderResult,
) -> None:
    """Surface partial-write diagnostics WITHOUT counting them as held rows.

    A note means the row landed but something attached to it did not — an extras
    write, a custom-attributes sweep. Twelve such sites appended to
    ``result.errors`` while never incrementing ``result.quarantined``, so each one
    minted a quarantine record the board's "held for review" count did not
    include. The banner and the table disagreed, and a school was shown a
    partial-write warning as though a row had been rejected.

    They are not hidden — that would be reducing a held count by concealing rows,
    which is the one thing the zero-touch standard forbids. They are logged at
    WARNING and stashed on the run's summary, filed as what they actually are.
    """
    notes = [n for n in (getattr(result, "notes", None) or []) if isinstance(n, dict)]
    if not notes:
        return
    domain_label = ((domain or "") or (artifact.path_within_bundle or ""))[:32]
    logger.warning(
        "migration_cloud.apply: %s partial-write note(s) for bundle=%s domain=%s — "
        "the row landed but an attached write did not: %s",
        len(notes),
        bundle.pk,
        domain_label,
        "; ".join(str(n.get("note", ""))[:200] for n in notes[:_LANDER_NOTE_LOG_CAP]),
    )


# --- Helpers ----------------------------------------------------------

def _summarize_outcomes(outcomes: list[ArtifactApplyOutcome]) -> dict[str, Any]:
    totals = {
        "created": sum(o.result.created for o in outcomes),
        "updated": sum(o.result.updated for o in outcomes),
        "quarantined": sum(o.result.quarantined for o in outcomes),
        "errors": sum(len(o.result.errors) for o in outcomes),
        "artifacts_failed": sum(1 for o in outcomes if o.status == "FAILED"),
    }
    return totals


_APPLY_AUDIT_FIELD_CAP = 60  # magic-number-allow: apply-audit-field-name-list-cap


def _field_level_apply_summary(
    outcomes: list[ArtifactApplyOutcome],
) -> list[dict[str, Any]]:
    """Per-domain, field-level rollup for the ``bundle.applied`` audit event.

    Extends the coarse bundle totals with, per domain: the created / updated /
    quarantined counts AND the NAMES of the fields overwritten on existing
    records (the ``old`` keys of each update). This is the security-relevant
    surface — which existing fields a re-apply mutated — recorded tamper-evidently
    without any raw values.

    Field NAMES ride as list VALUES (never dict keys) so ``_sanitize_payload``,
    which rejects sensitive-keyword dict KEYS, never trips on a legitimate field
    name like ``email`` / ``date_of_birth``; the raw values stay in the tenant
    schema and never reach the append-only chain.
    """
    by_domain: dict[str, dict[str, Any]] = {}
    for o in outcomes:
        info = by_domain.setdefault(
            o.domain,
            {"created": 0, "updated": 0, "quarantined": 0, "_fields": set()},
        )
        r = o.result
        info["created"] += int(getattr(r, "created", 0) or 0)
        info["updated"] += int(getattr(r, "updated", 0) or 0)
        info["quarantined"] += int(getattr(r, "quarantined", 0) or 0)
        for entry in getattr(r, "updated_ids_with_old_values", None) or []:
            old = (entry or {}).get("old") or {}
            if isinstance(old, dict):
                info["_fields"].update(str(k) for k in old.keys())

    summary: list[dict[str, Any]] = []
    for domain in sorted(by_domain):
        info = by_domain[domain]
        fields = sorted(info["_fields"])
        summary.append(
            {
                "domain": domain,
                "created": info["created"],
                "updated": info["updated"],
                "quarantined": info["quarantined"],
                "updated_fields": fields[:_APPLY_AUDIT_FIELD_CAP],
                "updated_fields_truncated": len(fields) > _APPLY_AUDIT_FIELD_CAP,
            }
        )
    return summary


# A wedged apply is reclaimed to MAPPED and retried. That self-heal is correct
# for a worker that genuinely died, but it had no ceiling: a bundle whose apply
# COMPLETES and still does not settle gets reclaimed every _APPLYING_STALE_SECONDS
# forever. One live bundle re-ran a 44-second import ~48 times in 24 hours. A
# self-heal that cannot give up is not a self-heal, it is a loop.
_MAX_WEDGED_APPLY_RECLAIMS = 3  # magic-number-allow: wedged-apply-reclaim-ceiling

def wedged_reclaims_so_far(size_summary) -> int:
    """How many times this bundle has already been reclaimed from a wedged apply.

    Tolerant by design: the counter lives in a free-form JSON summary, so a
    missing / null / non-numeric value must read as zero rather than raise and
    take down an apply.
    """
    if not isinstance(size_summary, dict):
        # size_summary is a free-form JSONField: a string or list is a possible
        # (if wrong) shape, and .get would AttributeError inside a live apply.
        return 0
    try:
        return max(0, int(size_summary.get("wedged_apply_reclaims") or 0))
    except (TypeError, ValueError):
        return 0


def wedged_reclaim_budget_exhausted(size_summary) -> bool:
    """True when this bundle has used up its wedged-apply retries."""
    return wedged_reclaims_so_far(size_summary) >= _MAX_WEDGED_APPLY_RECLAIMS


_TERMINAL_BUNDLE_STATUSES = frozenset({
    BundleStatus.APPLIED,
    BundleStatus.RECONCILED,
    BundleStatus.FAILED,
    BundleStatus.ABORTED,
})


def _empty_result(bundle: MigrationBundle, dry_run: bool, status: str) -> ApplyResult:
    return ApplyResult(
        bundle_id=bundle.pk,
        dry_run=dry_run,
        per_artifact=[],
        total_created=0,
        total_updated=0,
        total_quarantined=0,
        status=status,
    )
