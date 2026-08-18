"""Phase U2 — Artifact profiler.

For every ``MigrationArtifact``, produce a deterministic ``ArtifactProfile``
the downstream layers consume: classifier reads it to pick a source +
domain; mapper reads it to propose canonical fields; transformers read it
to know date formats, decimal separators, name orders.

Determinism is the property that makes the profile usable. Re-profiling
the same bytes must produce the same dict, byte-for-byte. We sample
deterministically (first N rows + a stratified middle sample) and skip
any randomness.

Profile shape::

    {
        "schema_version": 1,
        "format": "csv" | "xlsx" | "json" | ...,
        "encoding": "utf-8" | "cp1252" | ...,
        "row_count": 1284,
        "column_count": 17,
        "columns": [
            {
                "name": "First Name",
                "normalized": "first_name",
                "null_rate": 0.03,
                "distinct_count_sample": 1142,
                "inferred_type": "string" | "date" | "int" | "decimal" | "email" | "phone" | "bool" | "enum" | "currency",
                "type_confidence": 0.95,
                "samples": ["Ada", "Maria", "Jin", ...],   # up to 10
                "regex_shapes": ["[A-Za-z'\\-]+"],          # detected shapes
                "is_pii": true,
            },
            ...
        ],
        "locale_hints": {
            "date_format": "%d/%m/%Y",
            "decimal_separator": ",",
            "default_country_code": "+254",
        },
    }

The profiler degrades gracefully when bytes are unreadable, encoding
sniffing fails, or a row blows up — it records what it can and moves on.
Bundles never block on a bad artifact; the artifact is profiled with what
it has and the classifier sees the partial profile.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import re
import tarfile
import unicodedata
import zipfile
from collections import Counter
from pathlib import PurePosixPath
from typing import IO, Any, Iterable

from django.db import transaction

from .models import ArtifactFormat, BundleStatus, MigrationArtifact, MigrationBundle

logger = logging.getLogger(__name__)

_SAMPLE_ROWS = 200            # rows read per artifact for profiling
_SAMPLE_VALUES_PER_COL = 10   # values stored in the profile per column
_PROFILE_SCHEMA_VERSION = 1

_EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")
_INT_RE = re.compile(r"^-?\d+$")
_DECIMAL_RE = re.compile(r"^-?\d+[.,]\d+$")
_PHONE_RE = re.compile(r"^[\+\d][\d\s\-\(\)]{6,}$")
_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_US_DATE_RE = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")
_EU_DATE_RE = re.compile(r"^\d{1,2}[./-]\d{1,2}[./-]\d{4}$")
_CURRENCY_RE = re.compile(r"^[\$€£¥₹]?\s*-?[\d,\.]+\s*[A-Z]{0,3}$")
_BOOL_VALUES = {"true", "false", "yes", "no", "y", "n", "1", "0", "t", "f"}

_PII_NAME_HINTS = (
    "name", "first", "last", "given", "family", "email", "phone", "mobile",
    "address", "dob", "birth", "ssn", "passport", "national", "tin",
    "contact",
)


_PROFILER_REASON_MAX = 400  # magic-number-allow: quarantine-reason-detail-cap-chars

# Parser exceptions are structural ("File is not a zip file", "Bad magic number"),
# so the operator-facing hint below turns them into something a school can act on
# without needing the traceback.
_PROFILER_HINTS = (
    ("not a zip file", "This file is not a valid archive — it may have been renamed to .zip, or the download was truncated. Re-export it, or upload the CSV/Excel files directly."),
    ("bad magic number", "The file is corrupt or incomplete — re-download it from the source system and upload again."),
    ("encrypted", "This file is password-protected. Remove the password and upload it again."),
    ("password", "This file is password-protected. Remove the password and upload it again."),
    ("no such file", "The uploaded file could not be read back from storage — please upload it again."),
)


def _profiler_failure_reason(exc: Exception) -> str:
    """Operator-readable quarantine reason for a file the profiler could not read.

    Keeps the machine-readable ``profiler_error`` prefix (nothing else in the tree
    matches on it, but callers may) and appends the real cause plus, where the
    message is recognisable, a plain-language next step.
    """
    detail = " ".join(str(exc).split())[:_PROFILER_REASON_MAX]
    label = f"profiler_error: {type(exc).__name__}"
    if detail:
        label = f"{label}: {detail}"
    lowered = detail.casefold()
    for needle, hint in _PROFILER_HINTS:
        if needle in lowered:
            return f"{label} — {hint}"
    return label


# --------------------------------------------------------------------------- #
# Archive expansion (every intake method, not just the ARCHIVE adapter)
# --------------------------------------------------------------------------- #
# A school's most natural way to send its data is one .zip of its exports. That
# only ever worked through the dedicated ARCHIVE intake adapter: on the ordinary
# file-upload path the zip was registered as a single opaque artifact,
# _sniff_format returned ARCHIVE ("handled by archive_intake" -- which never runs
# on that path), no reader matched it, and it quarantined. ``has_workable`` then
# EXCLUDES archive artifacts, so the bundle failed "No workable artifacts after
# profiling" and the upload could never succeed. Expanding here closes that hole
# for every intake method at once.
#
# This reads operator-supplied bytes, so the caps below are load-bearing rather
# than decorative: an unbounded expander turns a small upload into an unbounded
# write (zip bomb), and member names are attacker-controlled input.
_ARCHIVE_MAX_MEMBERS = 2000  # magic-number-allow: archive-expansion-member-cap
_ARCHIVE_MAX_MEMBER_BYTES = 256 * 1024 * 1024  # magic-number-allow: archive-expansion-per-member-cap
_ARCHIVE_MAX_TOTAL_BYTES = 1024 * 1024 * 1024  # magic-number-allow: archive-expansion-total-cap
_TAR_SUFFIXES = (".tar", ".tar.gz", ".tgz", ".tar.bz2", ".tbz", ".tar.xz")


class _ArchiveMemberPayload:
    """Payload adapter for :func:`artifact_blob_store.capture_artifact_blob`.

    That function takes an ingest *payload* and calls ``payload.content_opener()``
    for a fresh stream -- it does NOT accept raw bytes, and handed bytes it quietly
    records the artifact as "no content stream" and captures nothing. An archive
    member is already in memory, so this wraps it in the same shape the intake
    adapters present.
    """

    __slots__ = ("_data",)

    def __init__(self, data: bytes) -> None:
        self._data = data

    def content_opener(self):
        return io.BytesIO(self._data)


def _is_safe_archive_member(name: str) -> bool:
    """Reject directory entries and any name that escapes the archive root.

    Member names come from an uploaded file, so ``../../secrets`` and absolute
    paths are untrusted input even when the uploader is. The name is only ever
    stored (never used to open a path), but a traversal name would still let one
    archive's member masquerade as another bundle's on ``path_within_bundle``.
    """
    if not name or name.endswith("/"):
        return False
    pure = PurePosixPath(name.replace("\\", "/"))
    if pure.is_absolute():
        return False
    return not any(part == ".." for part in pure.parts)


def _iter_archive_members(stream, filename):
    """Yield ``(member_name, bytes)`` for a zip/tar, honouring the safety caps.

    Encrypted members and unsupported archive types raise; the caller turns that
    into a real operator-facing reason via ``_profiler_failure_reason``.
    """
    data = stream.read()
    total = 0

    if zipfile.is_zipfile(io.BytesIO(data)):
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for info in zf.infolist()[:_ARCHIVE_MAX_MEMBERS]:
                if info.is_dir() or not _is_safe_archive_member(info.filename):
                    continue
                if info.file_size > _ARCHIVE_MAX_MEMBER_BYTES:
                    continue
                total += info.file_size
                if total > _ARCHIVE_MAX_TOTAL_BYTES:
                    break
                yield info.filename, zf.read(info)
        return

    if (filename or "").lower().endswith(_TAR_SUFFIXES):
        with tarfile.open(fileobj=io.BytesIO(data)) as tf:
            for count, member in enumerate(tf):
                if count >= _ARCHIVE_MAX_MEMBERS:
                    break
                if not member.isfile() or not _is_safe_archive_member(member.name):
                    continue
                if member.size > _ARCHIVE_MAX_MEMBER_BYTES:
                    continue
                total += member.size
                if total > _ARCHIVE_MAX_TOTAL_BYTES:
                    break
                handle = tf.extractfile(member)
                if handle is not None:
                    yield member.name, handle.read()
        return

    raise ValueError("Unsupported archive type - upload the CSV/Excel files directly.")


def expand_archive_artifacts(bundle: MigrationBundle) -> int:
    """Register each member of an uploaded archive as its own child artifact.

    Idempotent: an archive that already has children is skipped, and a duplicate
    member on a re-run is swallowed rather than duplicated. Returns the number of
    child artifacts created.
    """
    from .artifact_blob_store import capture_artifact_blob

    created = 0
    for artifact in bundle.artifacts.filter(quarantined=False).order_by("id"):
        if artifact.profile or artifact.children.exists():
            continue
        stream, _encoding = _resolve_stream(artifact)
        if stream is None:
            continue
        members = []
        try:
            if _sniff_format(artifact, [], []) != ArtifactFormat.ARCHIVE:
                continue
            members = list(_iter_archive_members(stream, artifact.filename))
        except Exception as exc:  # noqa: BLE001 - a bad archive must not block the bundle
            logger.exception(
                "migration_cloud.profiler: failed to expand archive artifact %s",
                artifact.pk,
            )
            artifact.quarantined = True
            artifact.quarantine_reason = _profiler_failure_reason(exc)
            artifact.detected_format = ArtifactFormat.ARCHIVE
            artifact.save(
                update_fields=[
                    "quarantined",
                    "quarantine_reason",
                    "detected_format",
                    "updated_at",
                ]
            )
            continue
        finally:
            try:
                stream.close()
            except Exception:  # noqa: BLE001
                pass

        for name, payload in members:
            try:
                with transaction.atomic():
                    child = MigrationArtifact.objects.create(
                        bundle=bundle,
                        parent_archive=artifact,
                        path_within_bundle=(
                            artifact.path_within_bundle + "/" + name
                        ),
                        filename=PurePosixPath(name).name,
                        byte_size=len(payload),
                        sha256=hashlib.sha256(payload).hexdigest(),
                        locale_hints={},
                        profile={},
                    )
            except Exception:  # noqa: BLE001 - duplicate member on a re-run
                continue
            capture_artifact_blob(child, _ArchiveMemberPayload(payload))
            created += 1

        # The archive is a container, not a data file: stamp the format so
        # ``has_workable`` keeps excluding it, and leave it UNquarantined so the
        # bundle is judged on its children.
        artifact.detected_format = ArtifactFormat.ARCHIVE
        artifact.save(update_fields=["detected_format", "updated_at"])
    return created


def profile_bundle(bundle: MigrationBundle) -> int:
    """Profile every un-profiled artifact in a bundle. Returns count profiled.

    Idempotent: artifacts that already carry a profile (``profile`` JSONField
    non-empty) are skipped. To re-profile, clear the profile and call again.
    Transitions the bundle status to PROFILED when all artifacts have a profile.
    """
    # Expand any uploaded archive into real child artifacts BEFORE profiling,
    # so members are profiled by the ordinary loop below on every intake method.
    expand_archive_artifacts(bundle)
    # Then the same for an uploaded database file, whose tables become CSV
    # children. Runs AFTER archive expansion so a .db inside a .zip is expanded
    # too -- the shape a school actually sends when IT hands over "the database".
    expand_tabular_source_artifacts(bundle)

    candidates = bundle.artifacts.filter(quarantined=False).order_by("id")
    profiled = 0
    for artifact in candidates:
        if artifact.profile:
            continue
        try:
            with transaction.atomic():
                _profile_one(artifact)
        except Exception as exc:  # noqa: BLE001 — never let one artifact block the bundle
            logger.exception(
                "migration_cloud.profiler: failed to profile artifact %s", artifact.pk
            )
            artifact.quarantined = True
            # Record WHY, not just THAT. This reason is the operator's only
            # explanation on the review page: a bare "profiler_error" told a
            # school its upload had failed while withholding the one fact that
            # would let them fix it (a renamed .zip, a password-protected
            # workbook, a truncated download all look identical). The stable
            # ``profiler_error`` prefix is kept so anything matching on it still
            # does; the cause is appended, bounded so a pathological parser
            # message cannot bloat the row.
            artifact.quarantine_reason = _profiler_failure_reason(exc)
            artifact.save(update_fields=["quarantined", "quarantine_reason", "updated_at"])
            continue
        profiled += 1

    # Transition bundle if every non-quarantined artifact now has a profile.
    unprofiled = bundle.artifacts.filter(quarantined=False, profile={}).exists()
    if unprofiled:
        return profiled

    # Honesty gate: a bundle that HAS artifacts but retains ZERO workable ones
    # after profiling (every file quarantined or unreadable — archive shells hold
    # no rows) can never land a single row. Advancing it to PROFILED lets it sail
    # through classify → map → apply and stamp a *green* APPLIED with all-zero
    # totals, so the operator believes the migration succeeded when nothing was
    # imported. Fail honestly instead — the review surface then shows FAILED with
    # the per-file quarantine reasons. apply_bundle refuses any non-MAPPED bundle
    # so this cannot be applied; repair_readiness treats FAILED as repairable once
    # the operator corrects the source files.
    total_artifacts = bundle.artifacts.count()
    has_workable = (
        bundle.artifacts.filter(quarantined=False)
        .exclude(detected_format=ArtifactFormat.ARCHIVE)
        .exists()
    )
    if total_artifacts and not has_workable:
        quarantined_count = bundle.artifacts.filter(quarantined=True).count()
        bundle.mark_status(
            BundleStatus.FAILED,
            summary_patch={
                "artifacts_profiled": profiled,
                "profiler_all_quarantined": True,
                "quarantined_artifacts": quarantined_count,
                "error": (
                    "No workable artifacts after profiling — every file was "
                    "quarantined or unreadable; nothing to apply."
                ),
            },
        )
        return profiled

    bundle.mark_status(
        BundleStatus.PROFILED,
        summary_patch={"artifacts_profiled": profiled},
    )
    return profiled


def _profile_one(artifact: MigrationArtifact) -> None:
    """Profile a single artifact: detect format, sample rows, infer types."""
    fmt = artifact.detected_format
    if fmt == ArtifactFormat.UNKNOWN:
        # Resolve *binary* formats (XLSX / XLS / PDF) from magic bytes BEFORE
        # sampling. A connectionless FILE_UPLOAD artifact arrives UNKNOWN, and
        # those formats need a dedicated reader — reading them as CSV yields
        # garbage columns (and, for PDF, zero real rows). Text formats stay
        # UNKNOWN here and are still resolved by the header heuristic after the
        # CSV-like read below, so the existing CSV/JSON paths are unchanged.
        signature = _sniff_format(artifact, [], [])
        # XML joins the binary formats here: it has a dedicated reader, and
        # letting it fall through to the CSV-like read profiles an XML export as
        # a single garbage column whose header is the first line of markup.
        if signature in (
            ArtifactFormat.XLSX,
            ArtifactFormat.XLS,
            ArtifactFormat.PDF,
            ArtifactFormat.XML,
        ):
            artifact.detected_format = signature
            fmt = signature
    rows, headers, encoding = _read_sample(artifact)
    if fmt == ArtifactFormat.UNKNOWN:
        fmt = _sniff_format(artifact, headers, rows)

    columns_profile = _profile_columns(headers, rows)
    locale_hints = _infer_locale_hints(columns_profile)

    profile = {
        "schema_version": _PROFILE_SCHEMA_VERSION,
        "format": fmt,
        "encoding": encoding,
        "row_count": len(rows),
        "column_count": len(headers),
        "columns": columns_profile,
        "locale_hints": locale_hints,
    }

    # A file that yields nothing AND has no working reader must say so. Left
    # silent, it profiles as a valid empty file and the school is never told why
    # its upload contributed nothing.
    if not headers and not rows:
        reason = unreadable_format_reason(fmt, artifact.filename)
        if reason:
            profile["unreadable_reason"] = reason
            artifact.quarantined = True
            artifact.quarantine_reason = reason[:500]

    artifact.profile = profile
    artifact.detected_format = fmt
    artifact.encoding = encoding or artifact.encoding
    artifact.row_count = len(rows)
    artifact.column_count = len(headers)
    artifact.locale_hints = locale_hints
    artifact.save(
        update_fields=[
            "profile",
            "detected_format",
            "encoding",
            "row_count",
            "column_count",
            "locale_hints",
            "quarantined",
            "quarantine_reason",
            "updated_at",
        ]
    )


# --- Reading --------------------------------------------------------------

def _read_sample(artifact: MigrationArtifact) -> tuple[list[list[Any]], list[str], str]:
    """Return (rows, headers, encoding) sampled from the artifact bytes.

    Phase U1 stored each artifact's bytes via the intake's ``content_opener``;
    the artifact itself doesn't carry a stream reference. The profiler uses
    the orchestrator's lazy reader (Phase U5 will provide ``open_artifact``).
    Until then, the profiler tries to read from the path stored at intake
    time (the orchestrator's ``_register_archive_parent`` stores the path on
    the artifact via the intake adapter). For artifacts whose bytes aren't
    reachable from the profiler's process (remote URL, OAuth folder), the
    profile remains schema-only and the classifier will see a thinner signal.
    """
    # Phase U2 minimum: in-process read for FILE_UPLOAD / ARCHIVE-extracted
    # artifacts. Long-tail intake (URL, OAuth, email) attach a deferred
    # reader in later phases.
    stream, encoding = _resolve_stream(artifact)
    if stream is None:
        return [], [], encoding

    try:
        if artifact.detected_format in (ArtifactFormat.CSV, ArtifactFormat.TSV, ArtifactFormat.UNKNOWN):
            return _read_csv_like(stream, encoding)
        if artifact.detected_format == ArtifactFormat.JSON:
            return _read_json(stream, encoding)
        if artifact.detected_format == ArtifactFormat.JSONL:
            return _read_jsonl(stream, encoding)
        if artifact.detected_format == ArtifactFormat.XLSX:
            return _read_xlsx(stream, encoding)
        if artifact.detected_format == ArtifactFormat.XLS:
            return _read_xls(stream, encoding)
        if artifact.detected_format == ArtifactFormat.PDF:
            return _read_pdf(stream, encoding)
        if artifact.detected_format == ArtifactFormat.XML:
            return _read_xml(stream, encoding)
    finally:
        try:
            stream.close()
        except Exception:  # noqa: BLE001
            pass
    return [], [], encoding


def _resolve_stream(artifact: MigrationArtifact) -> tuple[IO[bytes] | None, str]:
    """Best-effort byte stream for in-process artifacts.

    Phase U5 content store (gap #2): if the artifact's source bytes were
    captured at ingest, read them back here first — this is what lets archive
    members + remote / OAuth-folder pulls profile with real bytes instead of a
    schema-only signal. Falls through to the legacy single-top-level-local-file
    path when no blob is present.
    """
    from .artifact_blob_store import open_artifact_blob_stream

    blob_stream, blob_encoding = open_artifact_blob_stream(artifact)
    if blob_stream is not None:
        return blob_stream, blob_encoding

    bundle = artifact.bundle
    uri = bundle.intake_source_uri or ""
    if not uri:
        return None, ""
    # Only handle local file paths in this iteration.
    from pathlib import Path
    p = Path(uri)
    if not p.exists() or not p.is_file():
        return None, ""
    if artifact.parent_archive_id is not None:
        # Archive member — defer to Phase U5 content store.
        return None, ""
    if artifact.path_within_bundle != p.name:
        # Different file in the same bundle.
        return None, ""

    encoding = _sniff_encoding(p)
    return p.open("rb"), encoding


def _sniff_format(
    artifact: MigrationArtifact,
    headers: list[str],
    rows: list[list[Any]],
) -> str:
    """Resolve a concrete format for artifacts whose intake left them UNKNOWN.

    Resolution order:
        1. Magic-byte sniff against the artifact's bytes (PDF, ZIP/XLSX,
           SQLite, gzip, OLE2 / .mdb, BOM-marked XML).
        2. File-extension heuristic (last-resort; intake adapters are
           expected to set the format directly when they know).
        3. Header-row heuristic — present headers + at least one row = CSV.
        4. Fall through to ``UNKNOWN`` so the orchestrator routes to the
           custom_fields fallback rather than dropping the artifact.

    Pure-Python; no third-party libs. Reads only the first 16 bytes from
    the artifact when bytes are reachable, so this is cheap to call on
    every UNKNOWN artifact in the bundle.
    """
    magic = _read_magic_bytes(artifact)
    if magic.startswith(b"%PDF-"):
        return ArtifactFormat.PDF
    if magic.startswith(b"PK\x03\x04"):
        # ZIP-shaped — XLSX is a zip with [Content_Types].xml; outer-look
        # check defers to the openpyxl path. Treat as XLSX iff extension
        # hints at it; otherwise leave as ARCHIVE (handled by archive_intake).
        ext = (artifact.filename or "").lower()
        if ext.endswith(".xlsx") or ext.endswith(".xlsm"):
            return ArtifactFormat.XLSX
        return ArtifactFormat.ARCHIVE
    if magic.startswith(b"SQLite format 3\x00"):
        return ArtifactFormat.SQLITE
    if magic.startswith(b"\x1f\x8b"):
        return ArtifactFormat.ARCHIVE  # gzip — usually .csv.gz / .json.gz
    if magic.startswith(b"\xd0\xcf\x11\xe0"):
        # OLE2 compound document — legacy .xls and .mdb both use this.
        ext = (artifact.filename or "").lower()
        if ext.endswith((".mdb", ".accdb")):
            return ArtifactFormat.UNKNOWN  # Access handled by access_intake; profile lands later
        if ext.endswith(".xls"):
            return ArtifactFormat.XLS
        return ArtifactFormat.UNKNOWN
    if magic[:5] == b"<?xml":
        return ArtifactFormat.XML

    # Extension last (intake adapters should usually have set this).
    ext = (artifact.filename or "").lower()
    if ext.endswith(".csv"):
        return ArtifactFormat.CSV
    if ext.endswith(".tsv"):
        return ArtifactFormat.TSV
    if ext.endswith((".json",)):
        return ArtifactFormat.JSON
    if ext.endswith((".jsonl", ".ndjson")):
        return ArtifactFormat.JSONL
    if ext.endswith(".xml"):
        return ArtifactFormat.XML
    if ext.endswith(".sql"):
        return ArtifactFormat.SQL
    if ext.endswith(".parquet"):
        return ArtifactFormat.PARQUET
    if ext.endswith(".pdf"):
        return ArtifactFormat.PDF
    if ext.endswith((".jpg", ".jpeg", ".png", ".tif", ".tiff")):
        return ArtifactFormat.IMAGE

    # Header-row heuristic — if the reader returned at least one row with
    # headers, the artifact is some flavour of CSV.
    if headers and rows:
        return ArtifactFormat.CSV
    return ArtifactFormat.UNKNOWN


def _read_magic_bytes(artifact: MigrationArtifact, n: int = 16) -> bytes:
    """Best-effort read of the first ``n`` bytes; empty bytes on failure."""
    stream, _ = _resolve_stream(artifact)
    if stream is None:
        return b""
    try:
        return stream.read(n)
    except Exception:  # noqa: BLE001
        return b""
    finally:
        try:
            stream.close()
        except Exception:  # noqa: BLE001
            pass


def _sniff_encoding(path) -> str:
    """Best-effort encoding sniff with cascading detectors.

    Order:
        1. BOM detection (UTF-8 / UTF-16 LE / UTF-16 BE).
        2. Plain UTF-8 validity check.
        3. ``charset-normalizer`` if installed (covers cp1252 / mac-roman /
           gb2312 / shift_jis / latin1 / windows-125x intelligently).
        4. cp1252 fallback (most likely vendor export from Windows tooling).
    """
    try:
        with path.open("rb") as fh:
            head = fh.read(65536)
    except OSError:
        return "utf-8"
    if head.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if head.startswith(b"\xff\xfe"):
        return "utf-16-le"
    if head.startswith(b"\xfe\xff"):
        return "utf-16-be"
    try:
        head.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        pass
    try:
        from charset_normalizer import from_bytes  # type: ignore[import-not-found]

        result = from_bytes(head).best()
        if result is not None and result.encoding:
            return str(result.encoding)
    except Exception:  # noqa: BLE001
        pass
    return "cp1252"


def _read_csv_like(stream: IO[bytes], encoding: str) -> tuple[list[list[Any]], list[str], str]:
    text = stream.read().decode(encoding or "utf-8", errors="replace")
    # Sniff delimiter from a 4 KiB window.
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    rows = []
    headers: list[str] = []
    data_seen = 0
    for row in reader:
        if not headers:
            # Skip blank + '#'-comment lines ONLY before the header row. The
            # canonical template ships a leading '# runmycampus-canonical-template:'
            # marker; reading row 0 unconditionally made a filled-in template
            # round-trip with that comment as its ONLY column and the real headers
            # demoted to data. Restricting the skip to the pre-header region means a
            # legitimate DATA value that starts with '#' (e.g. an issue ref) is never
            # dropped from the sample.
            if not row or not any(str(c).strip() for c in row):
                continue
            if str(row[0]).lstrip().startswith("#"):
                continue
            headers = [str(c).strip() for c in row]
            continue
        if data_seen >= _SAMPLE_ROWS:
            break
        rows.append(row)
        data_seen += 1
    return rows, headers, encoding


def _read_json(stream: IO[bytes], encoding: str) -> tuple[list[list[Any]], list[str], str]:
    try:
        data = json.loads(stream.read().decode(encoding or "utf-8", errors="replace"))
    except json.JSONDecodeError:
        return [], [], encoding
    if isinstance(data, dict):
        # Single-record JSON — wrap so the profiler still extracts columns.
        records = [data]
    elif isinstance(data, list):
        records = [r for r in data if isinstance(r, dict)][:_SAMPLE_ROWS]
    else:
        return [], [], encoding
    headers = sorted({k for rec in records for k in rec.keys()})
    rows = [[rec.get(h, "") for h in headers] for rec in records]
    return rows, headers, encoding


def _read_jsonl(stream: IO[bytes], encoding: str) -> tuple[list[list[Any]], list[str], str]:
    text = stream.read().decode(encoding or "utf-8", errors="replace")
    records: list[dict[str, Any]] = []
    for i, line in enumerate(text.splitlines()):
        if i >= _SAMPLE_ROWS:
            break
        line = line.strip()
        if not line:
            continue
        try:
            obj = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(obj, dict):
            records.append(obj)
    headers = sorted({k for rec in records for k in rec.keys()})
    rows = [[rec.get(h, "") for h in headers] for rec in records]
    return rows, headers, encoding


# --- Column type / shape inference --------------------------------------

def _profile_columns(headers: list[str], rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not headers:
        return []

    profiles: list[dict[str, Any]] = []
    for col_idx, name in enumerate(headers):
        values = [row[col_idx] for row in rows if col_idx < len(row)]
        clean_values = [str(v).strip() for v in values if str(v).strip() != ""]
        null_rate = 1.0 - (len(clean_values) / max(len(values), 1))
        type_counts = Counter(_infer_value_type(v) for v in clean_values)
        if type_counts:
            inferred_type, hits = type_counts.most_common(1)[0]
            type_confidence = hits / max(len(clean_values), 1)
        else:
            inferred_type, type_confidence = "string", 0.0
        # Promote to "enum" when distinct cardinality is very low.
        if inferred_type == "string" and clean_values:
            distinct = len(set(clean_values))
            if 1 < distinct <= max(8, len(clean_values) // 20):
                inferred_type = "enum"

        samples = list(dict.fromkeys(clean_values))[:_SAMPLE_VALUES_PER_COL]
        shapes = _regex_shapes(samples)
        is_pii = _looks_pii(name, samples)

        profiles.append({
            "name": name,
            "normalized": _normalize_header(name),
            "null_rate": round(null_rate, 4),
            "distinct_count_sample": len(set(clean_values)),
            "inferred_type": inferred_type,
            "type_confidence": round(type_confidence, 3),
            "samples": samples,
            "regex_shapes": shapes,
            "is_pii": is_pii,
        })
    return profiles


def _infer_value_type(raw: str) -> str:
    if not raw:
        return "string"
    low = raw.lower()
    if low in _BOOL_VALUES:
        return "bool"
    if _EMAIL_RE.match(raw):
        return "email"
    if _ISO_DATE_RE.match(raw):
        return "date"
    if _US_DATE_RE.match(raw) or _EU_DATE_RE.match(raw):
        return "date"
    if _INT_RE.match(raw):
        return "int"
    if _DECIMAL_RE.match(raw):
        return "decimal"
    if _CURRENCY_RE.match(raw) and any(c in raw for c in "$€£¥₹"):
        return "currency"
    if _PHONE_RE.match(raw) and sum(c.isdigit() for c in raw) >= 7:
        return "phone"
    return "string"


def _regex_shapes(samples: list[str]) -> list[str]:
    """Reduce a handful of samples to a small set of distinctive shape patterns."""
    shapes: list[str] = []
    seen: set[str] = set()
    for v in samples:
        shape = "".join(
            "A" if c.isalpha() else "9" if c.isdigit() else c
            for c in v[:32]
        )
        if shape and shape not in seen:
            seen.add(shape)
            shapes.append(shape)
        if len(shapes) >= 3:
            break
    return shapes


def _looks_pii(name: str, samples: Iterable[str]) -> bool:
    n = name.lower()
    if any(h in n for h in _PII_NAME_HINTS):
        return True
    for s in list(samples)[:5]:
        if _EMAIL_RE.match(s):
            return True
        if re.search(r"\d{3}-\d{2}-\d{4}", s):
            return True
    return False


# BOM (U+FEFF) + zero-width space/non-joiner/joiner/word-joiner.
_ZERO_WIDTH_CHARS = ("﻿", "​", "‌", "‍", "⁠")


def _normalize_header(name: str) -> str:
    # Strip BOM + zero-width marks FIRST. A UTF-8 BOM (U+FEFF) glued to the
    # first CSV/XLSX header ("﻿NAME", "﻿title") is NOT whitespace, so
    # .strip() leaves it; it then survives NFKD folding (no ASCII decomposition)
    # and trips the "non-Latin script -> return raw" guard below, so the header
    # normalizes to a BOM-prefixed token and matches NO synonym. Every real
    # spreadsheet exported with a BOM would silently mis-map its first column.
    s = name or ""
    for zw in _ZERO_WIDTH_CHARS:
        s = s.replace(zw, "")
    s = s.strip().lower()
    # Fold Latin diacritics (é→e, ç→c, ñ→n, ô→o) so accented French / Spanish /
    # Portuguese headers ("Prénom", "Numéro d'élève", "Décision") normalize
    # identically to their ASCII synonyms. Genuinely non-Latin scripts (CJK,
    # Arabic, Cyrillic, …) have no ASCII decomposition, so after folding they
    # are still non-ASCII and are preserved raw below for locale-aware mapping —
    # the deliberate behaviour the CJK regression test locks in.
    folded = "".join(
        ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch)
    )
    if re.search(r"[^\x00-\x7f]", folded):
        # Still non-ASCII after folding → a non-Latin script; keep it raw.
        return s
    s = re.sub(r"[^a-z0-9]+", "_", folded)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


# --- Locale hint inference ----------------------------------------------

def _infer_locale_hints(columns: list[dict[str, Any]]) -> dict[str, Any]:
    """Look across all date / currency columns to pick locale defaults."""
    date_format = None
    decimal_separator = None

    for col in columns:
        if col["inferred_type"] == "date":
            df = _vote_date_format(col["samples"])
            if df and date_format is None:
                date_format = df
        if col["inferred_type"] in ("decimal", "currency"):
            ds = _vote_decimal_sep(col["samples"])
            if ds and decimal_separator is None:
                decimal_separator = ds

    hints: dict[str, Any] = {}
    if date_format:
        hints["date_format"] = date_format
    if decimal_separator:
        hints["decimal_separator"] = decimal_separator
    return hints


def _vote_date_format(samples: list[str]) -> str | None:
    """Pick day-first vs month-first by looking at all samples in the column."""
    dayfirst_evidence = 0
    monthfirst_evidence = 0
    for s in samples:
        m = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-]\d{4}$", s)
        if not m:
            continue
        a, b = int(m.group(1)), int(m.group(2))
        if a > 12 and b <= 12:
            dayfirst_evidence += 1
        elif b > 12 and a <= 12:
            monthfirst_evidence += 1
    if dayfirst_evidence > monthfirst_evidence:
        return "%d/%m/%Y"
    if monthfirst_evidence > 0:
        return "%m/%d/%Y"
    if any(_ISO_DATE_RE.match(s) for s in samples):
        return "%Y-%m-%d"
    return None


def _vote_decimal_sep(samples: list[str]) -> str | None:
    comma_decimal = 0
    dot_decimal = 0
    for s in samples:
        if re.search(r"\d,\d{2}\b", s):
            comma_decimal += 1
        if re.search(r"\d\.\d{2}\b", s):
            dot_decimal += 1
    if comma_decimal > dot_decimal:
        return ","
    if dot_decimal > 0:
        return "."
    return None


def _read_xlsx(stream, encoding: str) -> tuple[list[list[Any]], list[str], str]:
    """Read the first sheet of an XLSX via ``openpyxl`` if installed.

    Returns ``([], [], encoding)`` if openpyxl is missing — the classifier
    then falls back to filename-only signal. Schools running on stock
    deployments that include openpyxl get full XLSX support automatically.
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except Exception:  # noqa: BLE001
        return [], [], encoding
    try:
        wb = load_workbook(filename=stream, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return [], [], encoding
    try:
        ws = wb[wb.sheetnames[0]] if wb.sheetnames else None
        if ws is None:
            return [], [], encoding
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers_row = next(rows_iter)
        except StopIteration:
            return [], [], encoding
        headers = [str(h) if h is not None else "" for h in headers_row]
        rows: list[list[Any]] = []
        for i, row in enumerate(rows_iter):
            if i >= 200:
                break
            rows.append([("" if c is None else c) for c in row])
        return rows, headers, encoding
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass


def _read_xls(stream, encoding: str) -> tuple[list[list[Any]], list[str], str]:
    """Read the first sheet of a legacy .xls via ``xlrd`` if installed."""
    try:
        import xlrd  # type: ignore[import-not-found]
    except Exception:  # noqa: BLE001
        return [], [], encoding
    try:
        data = stream.read()
        wb = xlrd.open_workbook(file_contents=data)
    except Exception:  # noqa: BLE001
        return [], [], encoding
    if wb.nsheets == 0:
        return [], [], encoding
    sh = wb.sheet_by_index(0)
    headers = [str(sh.cell_value(0, c)) for c in range(sh.ncols)] if sh.nrows else []
    rows: list[list[Any]] = []
    for r in range(1, min(sh.nrows, 201)):
        rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
    return rows, headers, encoding


def _read_pdf(stream, encoding: str) -> tuple[list[list[Any]], list[str], str]:
    """Extract a PDF's text, tabularise it, and sample it like a TSV.

    Digitally-generated PDFs (pdfplumber / pypdf) yield real rows — the common
    case of an exported transcript / fee statement / roster saved as PDF.
    Scanned PDFs with no OCR binaries installed yield ``([], [], encoding)`` so
    the classifier sees a thin signal and the review surface can honestly tell
    the operator the file needs OCR. Shares the extraction heuristics with the
    ``IntakeMethod.PDF`` adapter via :mod:`apps.migration_cloud.pdf_extract`.
    """
    from .pdf_extract import extract_pdf_tsv

    try:
        data = stream.read()
    except Exception:  # noqa: BLE001
        return [], [], encoding
    tsv = extract_pdf_tsv(data)
    if not tsv.strip():
        return [], [], encoding
    return _read_csv_like(io.BytesIO(tsv.encode("utf-8")), "utf-8")


# Formats whose bytes we can detect but not turn into rows, and the shortest
# honest instruction for each. A file that profiles empty in silence tells the
# school nothing; a file that says "save it as CSV" is a five-minute fix.
_UNREADABLE_FORMAT_HINTS = {
    ArtifactFormat.PARQUET: (
        "Parquet files need a converter this server does not carry. Export the "
        "same data as CSV or Excel (.xlsx) and upload that instead."
    ),
    ArtifactFormat.SQL: (
        "This is a SQL dump, which has to be restored into a database before its "
        "tables can be read. Restore it, then export each table as CSV -- or "
        "upload the database file itself (.db / .sqlite3), which we expand for you."
    ),
    ArtifactFormat.IMAGE: (
        "This is a picture of data rather than data. Upload the spreadsheet it "
        "was taken from, or save it as CSV or Excel (.xlsx)."
    ),
}


def unreadable_format_reason(detected_format: str, filename: str = "") -> str:
    """Why this artifact will profile empty, and what to send instead.

    Returns "" for every format the profiler can actually read, so callers can
    treat a non-empty string as "this file will contribute nothing".
    """
    if detected_format == ArtifactFormat.XLS:
        try:
            import xlrd  # noqa: F401
        except Exception:  # noqa: BLE001
            return (
                "Legacy .xls workbooks need a reader this server does not carry. "
                "Save it as Excel (.xlsx) or CSV and upload that instead."
            )
        return ""
    return _UNREADABLE_FORMAT_HINTS.get(detected_format, "")


def _xml_localname(tag) -> str:
    """``{http://ns}student`` -> ``student``."""
    text = str(tag or "")
    return text.rsplit("}", 1)[-1] if "}" in text else text


def _xml_record_parent(root):
    """Descend to the element whose children are the repeated records.

    Real exports wrap their rows to different depths -- ``<students><student/>``
    but also ``<export><data><students><student/>``. Walking down through
    single-child wrappers finds the record list in both without the caller
    having to describe their file's shape.
    """
    node = root
    for _ in range(8):  # magic-number-allow: xml-wrapper-descent-depth
        children = list(node)
        if len(children) == 1 and len(children[0]):
            node = children[0]
            continue
        return node
    return node


def _read_xml(stream, encoding: str) -> tuple[list[list[Any]], list[str], str]:
    """Sample an XML export as rows + headers.

    XML is what older on-premise SIS packages emit, and it detected correctly
    but had no reader, so it profiled as an empty file with no complaint.

    Parsed through ``defusedxml``: these are operator-supplied bytes, so entity
    expansion (billion laughs) and external-entity fetches must be refused
    rather than served. Without defusedxml installed the reader declines to
    parse at all rather than falling back to the unsafe stdlib parser.
    """
    try:
        from defusedxml.ElementTree import fromstring as _safe_fromstring
    except Exception:  # noqa: BLE001 - no safe parser, so no parsing
        return [], [], encoding
    try:
        root = _safe_fromstring(stream.read())
    except Exception:  # noqa: BLE001 - malformed or hostile XML
        return [], [], encoding

    records = list(_xml_record_parent(root))
    if not records:
        return [], [], encoding

    headers: list[str] = []
    flattened: list[dict[str, Any]] = []
    for record in records[:201]:  # magic-number-allow: profiler-sample-rows
        flat: dict[str, Any] = {
            _xml_localname(k): v for k, v in (record.attrib or {}).items()
        }
        for sub in record:
            name = _xml_localname(sub.tag)
            if name in flat:
                continue
            if len(sub):
                flat[name] = " ".join(t.strip() for t in sub.itertext() if t.strip())
            else:
                flat[name] = (sub.text or "").strip()
        if not flat and (record.text or "").strip():
            flat[_xml_localname(record.tag)] = record.text.strip()
        for key in flat:
            if key not in headers:
                headers.append(key)
        flattened.append(flat)

    rows = [[flat.get(h, "") for h in headers] for flat in flattened]
    return rows, headers, encoding


def expand_tabular_source_artifacts(bundle: MigrationBundle) -> int:
    """Turn an uploaded database file into one CSV artifact per table.

    ``DatabaseIntakeAdapter`` has always been able to do this, but only via the
    DATABASE intake method. A school that asks its IT person for "the database"
    and uploads the ``.db`` through the ordinary web form hit a detected format
    with no reader -- the same shape as the archive.zip defect, and just as
    silent. Expanding here covers every intake method at once.

    Idempotent: a source that already has children is skipped, and a duplicate
    table on a re-run is swallowed. Returns the number of child artifacts made.
    """
    import sqlite3
    import tempfile
    from pathlib import Path

    from .artifact_blob_store import capture_artifact_blob

    created = 0
    for artifact in bundle.artifacts.filter(quarantined=False).order_by("id"):
        if artifact.profile or artifact.children.exists():
            continue
        stream, _encoding = _resolve_stream(artifact)
        if stream is None:
            continue
        try:
            if _sniff_format(artifact, [], []) != ArtifactFormat.SQLITE:
                continue
            data = stream.read()
        except Exception:  # noqa: BLE001
            continue
        finally:
            try:
                stream.close()
            except Exception:  # noqa: BLE001
                pass

        tables: list[tuple[str, bytes]] = []
        try:
            with tempfile.TemporaryDirectory() as tmp:
                # sqlite3 needs a real path; the uploaded bytes live in the blob
                # store, so they are materialised for exactly this call.
                path = Path(tmp) / "source.sqlite3"
                path.write_bytes(data)
                from .intake.database_intake import (
                    _resolve_tables,
                    _table_to_csv_bytes,
                )

                for table in _resolve_tables(None, path):
                    tables.append((table, _table_to_csv_bytes(path, table)))
        except Exception as exc:  # noqa: BLE001 - a bad db must not block the bundle
            logger.exception(
                "migration_cloud.profiler: failed to expand database artifact %s",
                artifact.pk,
            )
            artifact.quarantined = True
            artifact.quarantine_reason = _profiler_failure_reason(exc)
            artifact.detected_format = ArtifactFormat.SQLITE
            artifact.save(
                update_fields=[
                    "quarantined",
                    "quarantine_reason",
                    "detected_format",
                    "updated_at",
                ]
            )
            continue

        for table, payload in tables:
            name = f"{table}.csv"
            try:
                with transaction.atomic():
                    child = MigrationArtifact.objects.create(
                        bundle=bundle,
                        parent_archive=artifact,
                        path_within_bundle=artifact.path_within_bundle + "/" + name,
                        filename=name,
                        byte_size=len(payload),
                        sha256=hashlib.sha256(payload).hexdigest(),
                        locale_hints={},
                        profile={},
                    )
            except Exception:  # noqa: BLE001 - duplicate table on a re-run
                continue
            try:
                capture_artifact_blob(child, _ArchiveMemberPayload(payload))
            except Exception:  # noqa: BLE001
                logger.exception(
                    "migration_cloud.profiler: could not store table %s", name
                )
                continue
            created += 1

        if tables:
            artifact.detected_format = ArtifactFormat.SQLITE
            artifact.save(update_fields=["detected_format", "updated_at"])

    return created
