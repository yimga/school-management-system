# -*- coding: utf-8 -*-

import csv
import io
import logging
import os
import tempfile
import time
import zipfile
from typing import Any
from urllib.parse import urlencode

from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.core.management import CommandError, call_command
from django.db import DatabaseError, OperationalError
from django.db.models import Count, Q
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import NoReverseMatch, reverse
from django.utils import timezone
from django.utils.translation import gettext as _
from django.views.decorators.clickjacking import (
    xframe_options_exempt,
    xframe_options_sameorigin,
)

from apps.academics.models import Classroom
from apps.academics.services import get_active_year_and_term
from apps.brand_experience.models import ThemePack
from apps.people.models import StudentProfile
from apps.platform_runtime.helpers import get_platform_defaults
from apps.siteconfig.config_service import get_effective_site_settings
from apps.platform_runtime.structured_logging import log_view_exception
from apps.policies.policy_registry import get_effective_policy
from apps.reports.models import ReportCard
from apps.reports.services import (
    GLOBAL_REPORT_LABELS,
    annual_report_context,
    resolve_report_labels,
    term_report_context,
)
from apps.reports.weasy import render_pdf
from apps.runtime_blueprints.models import ReportCardStyle, ReportTemplate

from types import SimpleNamespace

from .forms import (
    ReportCardStyleAssignmentForm,
    ReportCardStyleForm,
    ReportCardStyleSelectionForm,
    THEME_EXPERIENCE_FIELD_NAMES,
    THEME_PUBLISH_GUARDED_FIELDS,
    ThemeColorsForm,
    UserPreferenceForm,
    build_theme_contrast_report,
    theme_contrast_targets_for_client,
)
from .models import (
    DashboardView,
    ReportCardStyleAssignment,
    build_platform_default_site_settings,
    UserPreference,
    RegionConfig,
    GradingScaleConfig,
    HolidayCalendar,
)
from .theme_palette_groups import THEME_PALETTE_GROUPS, build_theme_pack_groups
from .preview_state import PREVIEW_MODE_SESSION_KEY, ACT_AS_ROLE_SESSION_KEY
from .tenant_config import apply_tenant_settings_overrides
from apps.accounts.decorators import permission_required
from apps.accounts.models import User
from apps.schools.control_plane import require_super_access_with_host
from apps.siteconfig.control_plane_render import (
    default_operator_breadcrumbs,
    operator_cp_breadcrumb,
    render_siteconfig_stem,
)
from services.post_delete_navigation import safe_next_url as _safe_next_url

logger = logging.getLogger(__name__)

CACHE_KEY = "site_settings_v1"
SESSION_KEY = "site_preview_settings"
PORTAL_PREF_PREVIOUS_PAGE = "portal_pref_previous_page"


_GRADE_SCALE_CODE_ALIASES = {
    "LETTER": "a-f",
    "GPA_4": "gpa",
    "PASS_FAIL": "pass_fail",
}

# Fields that can be applied to session for "preview before save" (theme + branding + semantic colors).
PREVIEW_FROM_FORM_KEYS = [
    "site_name",
    "tagline",
    "primary_color",
    "accent_color",
    "header_bg_color",
    "footer_bg_color",
    "success_color",
    "warning_color",
    "danger_color",
    "theme_brightness",
    "use_dark_mode",
    "admin_use_site_primary",
    "backend_console_theme",
    "theme_pack",
    "admin_theme_pack",
    "teacher_theme_pack",
    "parent_theme_pack",
]

PREVIEW_BOOLEAN_KEYS = frozenset(
    {
        "use_dark_mode",
        "admin_use_site_primary",
    }
)


# Color field names for optional hex validation in preview.
PREVIEW_COLOR_KEYS = frozenset(
    {
        "primary_color",
        "accent_color",
        "header_bg_color",
        "footer_bg_color",
        "success_color",
        "warning_color",
        "danger_color",
    }
)


def _is_valid_hex(s):
    """Return True if s looks like a valid hex color (#rgb or #rrggbb)."""
    if not s or not isinstance(s, str):
        return False
    s = s.strip()
    if s.startswith("#"):
        s = s[1:]
    return len(s) in (3, 6) and all(c in "0123456789AaBbCcDdEeFf" for c in s)


def _normalize_preview_value(key, val):
    """Coerce POST values for preview payload (theme and booleans)."""
    if key in PREVIEW_BOOLEAN_KEYS:
        return val in ("on", "true", "1", 1, True)
    if val is None or val == "":
        return None
    if key in (
        "admin_theme_pack",
        "theme_pack",
        "teacher_theme_pack",
        "parent_theme_pack",
    ):
        if isinstance(val, str) and val.strip().isdigit():
            return int(val.strip())
        if isinstance(val, int):
            return val
        return None
    if isinstance(val, str):
        return val.strip()
    return val


def _registry_grade_scale_choices(
    country_code: str | None = None,
) -> list[tuple[str, str]]:
    try:
        from apps.registries.services import get_grade_scale_families

        choices = []
        seen = set()
        for row in get_grade_scale_families(country_code):
            raw_code = str(row.get("code") or "").strip()
            if not raw_code:
                continue
            code = _GRADE_SCALE_CODE_ALIASES.get(raw_code.upper(), raw_code).lower()
            if code in seen:
                continue
            seen.add(code)
            choices.append((code, str(row.get("name") or code)))
        return choices
    except (AttributeError, ImportError, LookupError, TypeError, ValueError):
        return []


def _language_choices_for_school(school) -> list[tuple[str, str]]:
    country_code = ""
    if school is not None:
        try:
            policy = get_effective_policy(school)
            raw_country = policy.get("country_code") or ""
            from apps.siteconfig.global_catalog import GlobalGeoCatalog

            country_code = (
                GlobalGeoCatalog.alpha2_for_country(raw_country)
                or str(raw_country or "").upper()[:2]
            )
        except (AttributeError, LookupError, TypeError, ValueError):
            country_code = ""
    try:
        from apps.registries.services import get_locales_for_country

        choices = []
        seen = set()
        for row in get_locales_for_country(country_code):
            locale_code = str(row.get("code") or "").strip()
            if not locale_code:
                continue
            language_code = locale_code.split("_", 1)[0].split("-", 1)[0].lower()
            if not language_code or language_code in seen:
                continue
            seen.add(language_code)
            choices.append(
                (language_code, str(row.get("name") or language_code.upper()))
            )
        if choices:
            return choices
    except (AttributeError, ImportError, LookupError, TypeError, ValueError):
        pass
    try:
        from apps.siteconfig.translations import SUPPORTED_LANGUAGES

        return [(code, label) for code, label in SUPPORTED_LANGUAGES.items()]
    except (AttributeError, ImportError, LookupError):
        return [("en", "English"), ("fr", "Français")]


def _is_known_currency_code(currency_code: str | None) -> bool:
    try:
        from apps.registries.services import is_known_currency_code

        return is_known_currency_code(currency_code)
    except (AttributeError, ImportError, TypeError, ValueError):
        return bool((currency_code or "").strip())


def _known_scale_types() -> list[str]:
    scale_types = list(
        GradingScaleConfig.objects.order_by("scale_type")
        .values_list("scale_type", flat=True)
        .distinct()
    )
    if not scale_types:
        scale_types = [code for code, _label in _registry_grade_scale_choices()]
    return list(
        dict.fromkeys(
            [
                str(scale_type).strip()
                for scale_type in scale_types
                if str(scale_type).strip()
            ]
        )
    )


def _snapshot_theme_field_values(instance, field_names):
    snapshot = {}
    for field_name in field_names:
        id_attr = f"{field_name}_id"
        if hasattr(instance, id_attr):
            field_id = getattr(instance, id_attr, None)
            if field_id is not None:
                snapshot[field_name] = field_id
                continue
        snapshot[field_name] = getattr(instance, field_name, None)
    return snapshot


@staff_member_required(login_url=settings.LOGIN_URL)
def preview_from_form(request):
    """
    Accept POST with current Site Settings (or theme) form data; validate and stash in session,
    then return redirect_url so the client can open the site in a new tab (live preview before save).
    """
    if request.method != "POST":
        return HttpResponseBadRequest("POST required")
    payload = {}
    errors = []
    for key in PREVIEW_FROM_FORM_KEYS:
        val = request.POST.get(key)
        normalized = _normalize_preview_value(key, val)
        if normalized is not None:
            if (
                key in PREVIEW_COLOR_KEYS
                and normalized
                and not _is_valid_hex(str(normalized))
            ):
                errors.append(f"{key.replace('_', ' ')}: invalid hex color")
                continue
            payload[key] = normalized
    if errors:
        return JsonResponse({"errors": errors}, status=400)
    request.session[SESSION_KEY] = payload
    request.session["preview_mode_enabled"] = True
    request.session.modified = True
    try:
        redirect_url = reverse("accounts:redirect")
    except NoReverseMatch:
        redirect_url = "/"
    # Optional: scroll to the section(s) being previewed. Use query param so it survives redirect.
    # Supports single section or comma-separated (e.g. "footer,header"). preview_keep=1 keeps highlights until dismiss.
    preview_section = (request.POST.get("preview_section") or "").strip().lower()
    preview_keep = request.POST.get("preview_keep") in ("1", "true", "on", "yes")
    query_parts = []
    if preview_section:
        section_map = {
            "footer-content": "footer",
            "footer": "footer",
            "theme-experience": "theme",
            "theme": "theme",
            "login-header-layout": "header",
            "branding": "header",
            "header": "header",
            "login": "login",
            "login-layout": "login",
            "sidebar": "sidebar",
        }
        if "," in preview_section:
            parts = [p.strip() for p in preview_section.split(",") if p.strip()]
            normalized = ",".join(section_map.get(p, p) for p in parts)
            if normalized:
                query_parts.append("preview_section=" + normalized)
        else:
            normalized = section_map.get(preview_section, preview_section)
            if normalized:
                query_parts.append("preview_section=" + normalized)
    if preview_keep:
        query_parts.append("preview_keep=1")
    if query_parts:
        redirect_url += "&" if "?" in redirect_url else "?"
        redirect_url += "&".join(query_parts)
    # Redirect to login page when previewing login section (so login page highlights run)
    if preview_section and "login" in preview_section and "," not in preview_section:
        try:
            redirect_url = reverse("accounts:login")
            if query_parts:
                redirect_url += "?" + "&".join(query_parts)
        except NoReverseMatch:
            pass
    return JsonResponse({"redirect_url": redirect_url})


@staff_member_required(login_url=settings.LOGIN_URL)
def maintenance_view(request):
    return render_siteconfig_stem(
        request,
        "maintenance",
        {},
        cp_title=_("Maintenance preview"),
        cp_meta_description=_(
            "Staff preview of the maintenance page shown to users when maintenance mode is on."
        ),
        cp_page_archetype="maintenance-preview",
        breadcrumbs=default_operator_breadcrumbs(
            operator_cp_breadcrumb(_("Maintenance preview")),
        ),
    )


# Neutral fallback when no region scales (Phase 2: no country names in tenant-facing form)
GRADING_SCALE_CHOICES_NEUTRAL = [
    ("0-20", "Numeric 0–20"),
    ("0-100", "Numeric 0–100"),
    ("0-10", "Numeric 0–10"),
    ("a-f", "Letter Grade (A–F)"),
    ("gpa", "GPA (0–4.0)"),
]


def get_grading_scale_choices_for_school(school):
    """Phase 2: Policy/registry-driven grading choices; no hardcoded country labels in tenant UX."""
    country_code = ""
    if school is not None:
        try:
            policy = get_effective_policy(school)
            raw_country = policy.get("country_code") or ""
            from apps.siteconfig.global_catalog import GlobalGeoCatalog

            country_code = (
                GlobalGeoCatalog.alpha2_for_country(raw_country)
                or str(raw_country or "").upper()[:2]
            )
        except (AttributeError, LookupError, TypeError, ValueError):
            country_code = ""
    registry_choices = _registry_grade_scale_choices(country_code)
    if registry_choices:
        return registry_choices
    if school and getattr(school, "default_region_id", None):
        try:
            from apps.siteconfig.models import GradingScaleConfig

            configs = GradingScaleConfig.objects.filter(
                region_id=school.default_region_id
            ).order_by("scale_type")
            if configs:
                return [
                    (c.scale_type, c.display_format or c.scale_type) for c in configs
                ]
        except (DatabaseError, OperationalError, ImportError, TypeError, ValueError):
            pass
    return list(GRADING_SCALE_CHOICES_NEUTRAL)


@login_required
def grading_settings(request):
    """School grading and default language (Phase 2). Requires request.school and admin-like role."""
    from django.http import HttpResponseForbidden

    school = getattr(request, "school", None)
    if not school:
        messages.warning(
            request, "Select a school (use your school subdomain) to manage grading."
        )
        return redirect("siteconfig:user_preferences")
    role = (getattr(request.user, "role", "") or "").upper()
    if role not in (
        "ADMIN",
        "LEADERSHIP",
        "IT_ADMIN",
        "PRINCIPAL",
        "VICE_PRINCIPAL",
    ) and not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden(
            "You do not have permission to change school grading settings."
        )
    policy = get_effective_policy(school)
    current_grading = (policy.get("grading") or {}).get(
        "grading_scale"
    ) or get_platform_defaults(use_db=False)["grading_scale"]
    current_language = policy.get("default_language") or "en"
    if request.method == "POST":
        new_grading = (request.POST.get("grading_scale") or "").strip() or None
        new_language = (request.POST.get("default_language") or "").strip() or None
        if new_grading or new_language is not None:
            requested_overrides = {}
            if new_grading:
                requested_overrides["grading_scale"] = new_grading
            if new_language is not None:
                requested_overrides["default_language"] = new_language
            result = apply_tenant_settings_overrides(
                school,
                requested_overrides,
                actor_is_superadmin=bool(request.user.is_superuser),
                force_override=False,
                persist=True,
            )
            if result.get("applied"):
                messages.success(request, "Grading and language settings updated.")
            if result.get("blocked"):
                blocked_keys = ", ".join(sorted(result["blocked"].keys()))
                messages.warning(
                    request,
                    f"Some settings were blocked by policy: {blocked_keys}",
                )
            return redirect("siteconfig:grading_settings")
    region = getattr(school, "default_region", None)
    grading_choices = get_grading_scale_choices_for_school(school)
    action_url = reverse("siteconfig:user_preferences")
    return render(
        request,
        "siteconfig/grading_settings.html",
        {
            "school": school,
            "region": region,
            "current_grading": current_grading,
            "current_language": current_language,
            "grading_choices": grading_choices,
            "language_choices": _language_choices_for_school(school),
            "action_url": action_url,
            "action_text": "Back to preferences",
        },
    )


@login_required
def module_market(request):
    """Module market (App Store): list available modules, activate/deactivate for current school (Phase 3)."""
    from django.http import HttpResponseForbidden
    from apps.schools.feature_registry import get_available_modules
    from apps.siteconfig.feature_toggles import set_toggle_state

    school = getattr(request, "school", None)
    if not school:
        messages.warning(request, "Select a school to manage modules.")
        return redirect("siteconfig:user_preferences")
    role = (getattr(request.user, "role", "") or "").upper()
    if role not in ("ADMIN", "LEADERSHIP", "IT_ADMIN", "PRINCIPAL") and not (
        request.user.is_staff or request.user.is_superuser
    ):
        return HttpResponseForbidden(
            "You do not have permission to manage school modules."
        )
    modules = get_available_modules()
    features = getattr(school, "features", None) or {}
    if request.method == "POST":
        action = request.POST.get("action")
        code = (request.POST.get("code") or "").strip()
        if code and any(m.get("code") == code for m in modules):
            features = dict(features)
            if action == "activate":
                if request.POST.get("module_impact_ack") != "1":
                    messages.error(
                        request,
                        "Confirm you have reviewed the module impact (checkbox) before activating.",
                    )
                    return redirect("siteconfig:module_market")
                features[code] = True
            elif action == "deactivate":
                features[code] = False
            school.features = features
            school.save(update_fields=["features", "updated_at"])
            try:
                from apps.policies.policy_registry import invalidate_policy_cache

                invalidate_policy_cache(school)
            except (AttributeError, ImportError, TypeError, ValueError):
                pass
            set_toggle_state(
                f"module.{code}",
                enabled=bool(features[code]),
                school=school,
                user=request.user,
                label=f"Module: {code}",
                description=f"School-level module toggle for '{code}'.",
                category="modules",
                default_enabled=False,
            )
            messages.success(request, f"Module '{code}' updated.")
            return redirect("siteconfig:module_market")
    active_codes = [k for k, v in features.items() if v]
    return render(
        request,
        "siteconfig/module_market.html",
        {
            "school": school,
            "modules": modules,
            "features": features,
            "active_codes": active_codes,
        },
    )


def build_reportcard_builder_context(
    request, *, studio_output_native: bool = False
) -> dict:
    """
    Context for report card builder: full portal page or Output Studio native pane.
    Forms bind to POST when present; workflow_step follows POST form_type or GET step.
    """
    settings_obj = get_effective_site_settings(request=request)
    if settings_obj is None:
        settings_obj = build_platform_default_site_settings()
    styles = list(ReportCardStyle.objects.order_by("name"))
    style_assignment_counts = {
        row["style_id"]: row["total"]
        for row in (
            ReportCardStyleAssignment.objects.values("style_id")
            .annotate(total=Count("classroom_id"))
            .order_by()
        )
    }
    for style in styles:
        style.assignment_count = style_assignment_counts.get(style.id, 0)
    assignments = list(
        ReportCardStyleAssignment.objects.select_related("classroom", "style").order_by(
            "classroom__name"
        )
    )
    all_classrooms = list(Classroom.objects.order_by("name"))
    sample_students = {}
    classroom_ids = [assignment.classroom_id for assignment in assignments]
    if classroom_ids:
        sample_candidates = (
            # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
            StudentProfile.objects.filter(
                classroom_id__in=classroom_ids, is_active=True
            )
            .defer("passport")
            .order_by("classroom_id", "last_name", "first_name")
        )
        for student in sample_candidates:
            sample_students.setdefault(student.classroom_id, student)
    for assignment in assignments:
        assignment.sample_student = sample_students.get(assignment.classroom_id)
    # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
    preview_students = list(
        StudentProfile.objects.filter(is_active=True)
        .defer("passport")
        .select_related("classroom")
        .order_by("last_name", "first_name")[:40]
    )
    default_style = settings_obj.resolve_default_report_style(
        ReportCard.Type.TERM
    ) or settings_obj.resolve_default_report_style(ReportCard.Type.ANNUAL)
    preview_default_style_slug = getattr(default_style, "slug", None) or (
        styles[0].slug if styles else ""
    )
    preview_default_student_id = preview_students[0].id if preview_students else ""
    assigned_classroom_ids = {assignment.classroom_id for assignment in assignments}
    style_form = ReportCardStyleForm(
        request.POST or None, request.FILES or None, prefix="style"
    )
    assignment_form = ReportCardStyleAssignmentForm(
        request.POST or None, prefix="assign"
    )
    from apps.brand_experience.platform_global_branding import PlatformGlobalBranding

    branding_row, _ = PlatformGlobalBranding.objects.get_or_create(pk=1)
    selection_form = ReportCardStyleSelectionForm(
        request.POST or None,
        prefix="selection",
        instance=branding_row,
    )
    workflow_step = (request.GET.get("step") or "style").strip().lower()
    if request.method == "POST":
        form_type_guess = (request.POST.get("form_type") or "").strip().lower()
        if form_type_guess in {"style", "assignment", "selection"}:
            workflow_step = form_type_guess
    if workflow_step not in {"style", "assignment", "selection"}:
        workflow_step = "style"

    scheduled_reports_delivery_hub_url = None
    report_output_history_evidence_url = None
    try:
        scheduled_reports_delivery_hub_url = reverse(
            "siteconfig:scheduled_reports_delivery_hub"
        )
    except NoReverseMatch:
        pass
    try:
        report_output_history_evidence_url = reverse(
            "siteconfig:report_output_history_evidence"
        )
    except NoReverseMatch:
        pass
    report_templates_catalog_evidence_url = None
    try:
        report_templates_catalog_evidence_url = reverse(
            "siteconfig:report_templates_catalog_evidence"
        )
    except NoReverseMatch:
        pass
    from apps.siteconfig.models_tooling import REPORT_EXPORT_HANDLERS
    from apps.siteconfig.tenant_config import get_report_template_family_for_school

    school = getattr(request, "school", None)
    rt_qs = ReportTemplate.objects.filter(is_active=True)
    if school is not None:
        rtf = (get_report_template_family_for_school(school) or "").strip()
        if rtf:
            rt_qs = rt_qs.filter(Q(template_family="") | Q(template_family=rtf))
    report_templates_catalog_count = rt_qs.count()

    return {
        # Effective tenant settings row for builder UI (avoid template name `settings`: confusable with django.conf).
        "site_settings": settings_obj,
        "settings": settings_obj,
        "styles": styles,
        "assignments": assignments,
        "preview_students": preview_students,
        "preview_default_style_slug": preview_default_style_slug,
        "preview_default_student_id": preview_default_student_id,
        "total_classroom_count": len(all_classrooms),
        "assigned_classroom_count": len(assigned_classroom_ids),
        "unassigned_classroom_count": max(
            0, len(all_classrooms) - len(assigned_classroom_ids)
        ),
        "style_form": style_form,
        "assignment_form": assignment_form,
        "selection_form": selection_form,
        "workflow_step": workflow_step,
        "studio_output_native": studio_output_native,
        "scheduled_reports_delivery_hub_url": scheduled_reports_delivery_hub_url,
        "report_output_history_evidence_url": report_output_history_evidence_url,
        "report_templates_catalog_evidence_url": report_templates_catalog_evidence_url,
        "export_handler_registry_count": len(REPORT_EXPORT_HANDLERS),
        "report_templates_catalog_count": report_templates_catalog_count,
    }


def _render_reportcard_builder_page(request, ctx):
    from apps.siteconfig.control_plane_render import (
        default_operator_breadcrumbs,
        operator_cp_breadcrumb,
        render_siteconfig_stem,
    )

    return render_siteconfig_stem(
        request,
        "reportcard_builder",
        ctx,
        cp_title=_("Report card builder"),
        breadcrumbs=default_operator_breadcrumbs(
            operator_cp_breadcrumb(_("Report card builder"), active=True),
        ),
    )


@permission_required("settings.manage")
def reportcard_builder(request):
    post_studio = (
        request.method == "POST"
        and (request.POST.get("studio_output_native") or "").strip() == "1"
    )

    if request.method == "POST":
        ctx = build_reportcard_builder_context(
            request, studio_output_native=post_studio
        )
        form_type = (request.POST.get("form_type") or "").strip().lower()
        style_form = ctx["style_form"]
        assignment_form = ctx["assignment_form"]
        selection_form = ctx["selection_form"]

        if form_type == "style" and style_form.is_valid():
            style_form.save()
            messages.success(request, _("Report card style saved."))
            if post_studio:
                return redirect(
                    reverse("studio_os:output") + "?pane=builder&step=style"
                )
            return redirect("siteconfig:reportcard_builder")
        if form_type == "assignment" and assignment_form.is_valid():
            assignment_form.save()
            messages.success(request, _("Style assignments updated."))
            if post_studio:
                return redirect(
                    reverse("studio_os:output") + "?pane=builder&step=assignment"
                )
            return redirect("siteconfig:reportcard_builder")
        if form_type == "selection" and selection_form.is_valid():
            selection_form.save()
            messages.success(request, _("Default styles saved."))
            if post_studio:
                return redirect(
                    reverse("studio_os:output") + "?pane=builder&step=selection"
                )
            return redirect("siteconfig:reportcard_builder")

        return _render_reportcard_builder_page(request, ctx)

    ctx = build_reportcard_builder_context(request, studio_output_native=False)
    return _render_reportcard_builder_page(request, ctx)


@permission_required("settings.manage")
def scheduled_reports_delivery_hub(request):
    """
    Tenant-facing hub for scheduled report delivery (entitlement: reports_scheduled_delivery or reports).
    """
    school = getattr(request, "school", None)
    schedules = []
    schedule_summary = {
        "schedule_total": 0,
        "active_count": 0,
        "inactive_count": 0,
        "distinct_report_keys": 0,
        "with_last_run": 0,
        "active_past_due_next": 0,
    }
    if school:
        from apps.reports.models import TenantReportSchedule

        qs = TenantReportSchedule.objects.filter(school=school)
        now = timezone.now()
        agg = qs.aggregate(
            schedule_total=Count("id"),
            active_count=Count("id", filter=Q(is_active=True)),
            inactive_count=Count("id", filter=Q(is_active=False)),
            distinct_report_keys=Count("report_key", distinct=True),
            with_last_run=Count("id", filter=Q(last_run__isnull=False)),
            active_past_due_next=Count(
                "id", filter=Q(is_active=True, next_run__lte=now)
            ),
        )
        schedule_summary = {k: int(v or 0) for k, v in agg.items()}
        schedules = list(qs.order_by("next_run")[:100])
    ctx = {"school": school, "schedules": schedules, "schedule_summary": schedule_summary}
    user = getattr(request, "user", None)
    if user is not None and getattr(user, "is_authenticated", False) and getattr(
        user, "is_staff", False
    ):
        try:
            ctx["tenant_report_schedule_admin_url"] = reverse(
                "admin:reports_tenantreportschedule_changelist"
            )
        except NoReverseMatch:
            ctx["tenant_report_schedule_admin_url"] = None
    else:
        ctx["tenant_report_schedule_admin_url"] = None
    try:
        ctx["reports_scheduled_api_list_url"] = reverse("api_v1:reports-scheduled-list")
    except NoReverseMatch:
        ctx["reports_scheduled_api_list_url"] = None
    try:
        ctx["operator_bulk_letters_url"] = reverse("siteconfig:bulk_letters")
    except NoReverseMatch:
        ctx["operator_bulk_letters_url"] = None
    try:
        ctx["operator_reportcard_builder_url"] = reverse("siteconfig:reportcard_builder")
    except NoReverseMatch:
        ctx["operator_reportcard_builder_url"] = None
    try:
        ctx["operator_studio_output_reports_url"] = reverse("studio_os:output") + "?pane=reports"
    except NoReverseMatch:
        ctx["operator_studio_output_reports_url"] = None
    try:
        ctx["operator_term_publish_evidence_url"] = reverse(
            "siteconfig:term_publish_status_evidence"
        )
    except NoReverseMatch:
        ctx["operator_term_publish_evidence_url"] = None
    try:
        ctx["operator_academic_years_evidence_url"] = reverse(
            "siteconfig:academic_years_setup_evidence"
        )
    except NoReverseMatch:
        ctx["operator_academic_years_evidence_url"] = None
    try:
        ctx["operator_departments_setup_evidence_url"] = reverse(
            "siteconfig:departments_setup_evidence"
        )
    except NoReverseMatch:
        ctx["operator_departments_setup_evidence_url"] = None
    try:
        ctx["operator_config_mutation_audit_evidence_url"] = reverse(
            "siteconfig:config_mutation_audit_evidence"
        )
    except NoReverseMatch:
        ctx["operator_config_mutation_audit_evidence_url"] = None
    try:
        ctx["operator_tenant_report_schedules_evidence_url"] = reverse(
            "siteconfig:tenant_report_schedules_evidence"
        )
    except NoReverseMatch:
        ctx["operator_tenant_report_schedules_evidence_url"] = None
    try:
        ctx["operator_report_templates_catalog_url"] = reverse(
            "siteconfig:report_templates_catalog_evidence"
        )
    except NoReverseMatch:
        ctx["operator_report_templates_catalog_url"] = None
    try:
        ctx["operator_report_output_history_evidence_url"] = reverse(
            "siteconfig:report_output_history_evidence"
        )
    except NoReverseMatch:
        ctx["operator_report_output_history_evidence_url"] = None
    try:
        ctx["operator_compliance_exports_url"] = reverse("siteconfig:compliance_exports")
    except NoReverseMatch:
        ctx["operator_compliance_exports_url"] = None
    try:
        ctx["northstar_ai_draft_url"] = reverse("siteconfig:northstar_ai_draft")
    except NoReverseMatch:
        ctx["northstar_ai_draft_url"] = None
    from apps.siteconfig.control_plane_render import (
        default_operator_breadcrumbs,
        operator_cp_breadcrumb,
        render_siteconfig_stem,
    )

    return render_siteconfig_stem(
        request,
        "scheduled_reports_delivery_hub",
        ctx,
        cp_title=_("Scheduled reports"),
        breadcrumbs=default_operator_breadcrumbs(
            operator_cp_breadcrumb(_("Scheduled reports"), active=True),
        ),
    )


def _build_style_metadata(site: Any) -> dict:
    if callable(getattr(site, "get_brand_metadata", None)):
        return site.get_brand_metadata()
    return {
        "school_name": site.site_name,
        "school_code": site.school_code,
        "country": site.country,
        "region": site.region,
        "ministry": site.ministry,
        "tagline": site.tagline,
    }


def _get_preview_platform_config(site: Any) -> dict[str, object]:
    if callable(getattr(site, "get_preview_platform_config", None)):
        return site.get_preview_platform_config()
    return {
        "preview_mode_enabled": getattr(site, "preview_mode_enabled", False),
        "preview_note": getattr(site, "preview_note", ""),
    }


def _get_theme_experience_settings(site: Any) -> dict[str, object]:
    if callable(getattr(site, "get_theme_experience_settings", None)):
        return site.get_theme_experience_settings()
    return {
        "skip_theme_publish_guard": getattr(site, "skip_theme_publish_guard", False),
        "primary_color": getattr(site, "primary_color", ""),
        "accent_color": getattr(site, "accent_color", ""),
        "header_bg_color": getattr(site, "header_bg_color", ""),
        "footer_bg_color": getattr(site, "footer_bg_color", ""),
        "success_color": getattr(site, "success_color", ""),
        "warning_color": getattr(site, "warning_color", ""),
        "danger_color": getattr(site, "danger_color", ""),
    }


class _PreviewTerm(SimpleNamespace):
    def get_name_display(self):
        return getattr(self, "name", "First term")


def _mock_preview_student():
    return SimpleNamespace(
        id=0,
        last_name="Sample",
        first_name="Learner",
        student_code="00SAMPLE",
        classroom=SimpleNamespace(name="Form One"),
        specialty=SimpleNamespace(name="Carpentry"),
    )


def _preview_student_queryset():
    # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
    return StudentProfile.objects.filter(is_active=True).select_related(
        "classroom", "specialty"
    )


def _resolve_preview_student(request):
    student_id = request.GET.get("student_id")
    queryset = _preview_student_queryset()
    if student_id:
        try:
            student = queryset.filter(id=int(student_id)).first()
            if student:
                return student
        except ValueError:
            pass
    return queryset.first() or _mock_preview_student()


def _build_report_context_for_pdf(style: ReportCardStyle, report_type: str, student):
    site = get_effective_site_settings(school=getattr(student, "school", None))
    metadata = _build_style_metadata(site)
    year, term = get_active_year_and_term()
    labels = resolve_report_labels(student=student)
    context = {
        "report_style": style,
        "metadata": metadata,
        "generated_at": timezone.now(),
        "preview_mode": True,
        "student": student,
        "student_name": f"{student.last_name} {student.first_name}",
        "labels": labels,
        # Report templates expect SITE (portal context processor may not run on all render paths).
        "SITE": site,
    }
    if report_type == ReportCard.Type.TERM:
        if year and term:
            term_ctx = term_report_context(student, year, term)
            context.update(term_ctx)
            context.update({"year": year, "term": term})
        else:
            # Preview fallback — compute current academic year label dynamically
            # so reports don't show stale "2025/2026" forever. Term label is
            # translatable (Django i18n picks up the active locale).
            _today = timezone.localdate()
            _year_label = f"{_today.year}/{_today.year + 1}"
            context.update(
                {
                    "year": SimpleNamespace(name=_year_label),
                    "term": _PreviewTerm(name=_("First Term")),
                    "rows": [],
                    "summary": {
                        "average": None,
                        "class_position": None,
                        "class_size": 0,
                        "class_rank_display": "- / -",
                        "specialty_position": None,
                        "specialty_size": 0,
                        "specialty_rank_display": "- / -",
                        "school_position": None,
                        "school_size": 0,
                        "school_rank_display": "- / -",
                        "promotion_status": "PENDING",
                        "teacher_remark": "Pending results.",
                    },
                    # Default grade weights from settings (env-overridable).
                    # Tenants can override via SiteSettings.report_grade_weights JSON.
                    "weights": SimpleNamespace(
                        seq1_weight=getattr(settings, "GRADE_WEIGHT_SEQ1", 20),
                        seq2_weight=getattr(settings, "GRADE_WEIGHT_SEQ2", 20),
                        exam_weight=getattr(settings, "GRADE_WEIGHT_EXAM", 60),
                        mock_weight=getattr(settings, "GRADE_WEIGHT_MOCK", 0),
                        practical_weight=getattr(settings, "GRADE_WEIGHT_PRACTICAL", 0),
                    ),
                    "sequence_cues": [
                        {
                            "key": "seq1",
                            "label": labels.get(
                                "sequence_1", GLOBAL_REPORT_LABELS["sequence_1"]
                            ),
                            "weight": 20,
                        },
                        {
                            "key": "seq2",
                            "label": labels.get(
                                "sequence_2", GLOBAL_REPORT_LABELS["sequence_2"]
                            ),
                            "weight": 20,
                        },
                        {
                            "key": "exam",
                            "label": labels.get("exam", GLOBAL_REPORT_LABELS["exam"]),
                            "weight": 60,
                        },
                    ],
                }
            )
    else:
        annual_ctx = (
            annual_report_context(student, year)
            if year
            else {
                "term_rows": [],
                "annual_average": None,
                "class_position": None,
                "class_size": 0,
                "class_rank_display": "- / -",
                "specialty_position": None,
                "specialty_size": 0,
                "specialty_rank_display": "- / -",
                "school_position": None,
                "school_size": 0,
                "school_rank_display": "- / -",
                "promotion_status": "PENDING",
                "promotion_average": None,
                "demotion_average": None,
                "teacher_remark": "Pending results.",
                "labels": labels,
            }
        )
        context.update(annual_ctx)
        _t = timezone.localdate()
        context.update({"year": year or SimpleNamespace(name=f"{_t.year}/{_t.year + 1}")})
    return context


@permission_required("settings.manage")
@xframe_options_sameorigin
def reportcard_style_preview(request, slug: str):
    style = get_object_or_404(ReportCardStyle, slug=slug)
    site = get_effective_site_settings(request=request)
    year, term = get_active_year_and_term()
    student = _resolve_preview_student(request)
    metadata = _build_style_metadata(site)

    if year and term:
        base_ctx = term_report_context(student, year, term)
        rows = base_ctx["rows"][:6]
        summary = base_ctx["summary"]
        weights = base_ctx["weights"]
        labels = base_ctx.get("labels", resolve_report_labels(student=student))
        sequence_cues = base_ctx.get("sequence_cues", [])
        student_obj = student
        student_name = f"{student.last_name} {student.first_name}"
        year_obj = year
        term_obj = term
    else:
        student_obj = _mock_preview_student()
        student_name = f"{student_obj.last_name} {student_obj.first_name}"
        _t = timezone.localdate()
        year_obj = SimpleNamespace(name=f"{_t.year}/{_t.year + 1}")
        term_obj = _PreviewTerm(name=_("First Term"))
        rows = [
            {
                "subject": "English",
                "coef": 2,
                "seq1": 12.0,
                "seq2": 13.5,
                "exam": 14.0,
                "mock": 0,
                "practical": 0,
                "total": 13.25,
                "remark": "Very good",
                "complete": True,
            },
            {
                "subject": "Mathematics",
                "coef": 4,
                "seq1": 11.0,
                "seq2": 12.0,
                "exam": 15.5,
                "mock": 0,
                "practical": 0,
                "total": 13.74,
                "remark": "Excellent",
                "complete": True,
            },
            {
                "subject": "Physics",
                "coef": 3,
                "seq1": 10.0,
                "seq2": 11.0,
                "exam": 12.0,
                "mock": 0,
                "practical": 0,
                "total": 11.66,
                "remark": "Solid",
                "complete": True,
            },
            {
                "subject": "Technical Drawing",
                "coef": 2,
                "seq1": 9.0,
                "seq2": 10.5,
                "exam": 11.0,
                "mock": 0,
                "practical": 0,
                "total": 10.19,
                "remark": "Improving",
                "complete": True,
            },
            {
                "subject": "ICT",
                "coef": 1,
                "seq1": 13.0,
                "seq2": 14.0,
                "exam": 15.0,
                "mock": 0,
                "practical": 0,
                "total": 14.29,
                "remark": "Strong",
                "complete": True,
            },
            {
                "subject": "Sports",
                "coef": 1,
                "seq1": 14.0,
                "seq2": 14.5,
                "exam": 0,
                "mock": 0,
                "practical": 0,
                "total": 14.25,
                "remark": "Active",
                "complete": False,
            },
        ]
        summary = {
            "average": 13.21,
            "class_position": 2,
            "class_size": 28,
            "class_rank_display": "2 / 28",
            "school_position": 5,
            "school_size": 120,
            "school_rank_display": "5 / 120",
            "promotion_status": "PROMOTED",
            "teacher_remark": "Consistent dedication.",
        }
        # Default grade weights from settings (env-overridable).
        # Tenants can override via SiteSettings.report_grade_weights JSON.
        weights = SimpleNamespace(
            seq1_weight=getattr(settings, "GRADE_WEIGHT_SEQ1", 20),
            seq2_weight=getattr(settings, "GRADE_WEIGHT_SEQ2", 20),
            exam_weight=getattr(settings, "GRADE_WEIGHT_EXAM", 60),
            mock_weight=getattr(settings, "GRADE_WEIGHT_MOCK", 0),
            practical_weight=getattr(settings, "GRADE_WEIGHT_PRACTICAL", 0),
        )
        labels = resolve_report_labels(student=student_obj)
        sequence_cues = [
            {
                "key": "seq1",
                "label": labels.get("sequence_1", GLOBAL_REPORT_LABELS["sequence_1"]),
                "weight": 20,
            },
            {
                "key": "seq2",
                "label": labels.get("sequence_2", GLOBAL_REPORT_LABELS["sequence_2"]),
                "weight": 20,
            },
            {
                "key": "exam",
                "label": labels.get("exam", GLOBAL_REPORT_LABELS["exam"]),
                "weight": 60,
            },
        ]

    context = {
        "report_style": style,
        "student": student_obj,
        "student_name": student_name,
        "year": year_obj,
        "term": term_obj,
        "rows": rows,
        "summary": summary,
        "weights": weights,
        "labels": labels,
        "sequence_cues": sequence_cues,
        "metadata": metadata,
        "generated_at": timezone.now(),
        "preview_mode": True,
        "SITE": site,
    }
    from apps.siteconfig.control_plane_render import (
        default_operator_breadcrumbs,
        operator_cp_breadcrumb,
        render_siteconfig_stem,
    )

    return render_siteconfig_stem(
        request,
        "reportcard_style_preview",
        context,
        cp_title=_("Report card style preview"),
        breadcrumbs=default_operator_breadcrumbs(
            operator_cp_breadcrumb(_("Report card builder"), url=reverse("siteconfig:reportcard_builder")),
            operator_cp_breadcrumb(_("Style preview"), active=True),
        ),
    )


@permission_required("settings.manage")
@xframe_options_sameorigin
def reportcard_style_pdf(request, slug: str, report_type: str):
    style = get_object_or_404(ReportCardStyle, slug=slug)
    report_type = report_type.upper()
    if report_type not in ReportCard.Type.values:
        return HttpResponseBadRequest("Unknown report type.")

    student = _resolve_preview_student(request) or _mock_preview_student()
    context = _build_report_context_for_pdf(style, report_type, student)
    template_name = style.template_for(report_type)
    filename = f"{report_type.lower()}_preview_{style.slug}.pdf"
    return render_pdf(request, template_name, context, filename=filename)


@permission_required("settings.manage")
@xframe_options_sameorigin
def reportcard_style_live_preview(request, slug: str, report_type: str):
    style = get_object_or_404(ReportCardStyle, slug=slug)
    report_type_value = (report_type or "").upper()
    if report_type_value not in ReportCard.Type.values:
        return HttpResponseBadRequest("Unknown report type.")

    student = _resolve_preview_student(request) or _mock_preview_student()
    context = _build_report_context_for_pdf(style, report_type_value, student)
    template_name = style.template_for(report_type_value)
    return render(request, template_name, context)


@permission_required("settings.manage")
@xframe_options_exempt
def reportcard_style_embed_preview(request, slug: str, report_type: str):
    """
    Iframe-focused preview endpoint used only by the report builder.

    We intentionally exempt X-Frame-Options on this one route and enforce
    a same-origin framing policy via CSP to avoid browser-level embed refusals
    while keeping framing locked to this site.
    """
    style = get_object_or_404(ReportCardStyle, slug=slug)
    report_type_value = (report_type or "").upper()
    if report_type_value not in ReportCard.Type.values:
        return HttpResponseBadRequest("Unknown report type.")

    student = _resolve_preview_student(request) or _mock_preview_student()
    context = _build_report_context_for_pdf(style, report_type_value, student)
    context["embed_preview"] = True
    context["preview_token"] = (request.GET.get("preview_token") or "").strip()
    template_name = style.template_for(report_type_value)
    response = render(request, template_name, context)
    response["Content-Security-Policy"] = "frame-ancestors 'self'"
    return response


@permission_required("settings.manage")
def clear_preview(request):
    request.session.pop(SESSION_KEY, None)
    request.session["preview_mode_enabled"] = False
    request.session.modified = True
    messages.info(request, "Preview cleared.")
    fallback = reverse("siteconfig:user_preferences")
    next_url = _safe_next_url(request, request.GET.get("next"), fallback)
    return redirect(next_url)


@login_required
def set_default_dashboard_view(request):
    """Set the user's default dashboard view (Overview, Workflow Center, etc.) and redirect."""
    view_value = request.GET.get("view") or request.POST.get("view")
    success_fallback = reverse("accounts:redirect")
    invalid_fallback = reverse("siteconfig:user_preferences")
    next_candidate = request.GET.get("next") or request.POST.get("next")
    next_url = _safe_next_url(request, next_candidate, success_fallback)
    allowed = {c[0] for c in DashboardView.choices}
    if view_value not in allowed:
        messages.warning(request, "Invalid dashboard view.")
        return redirect(_safe_next_url(request, next_candidate, invalid_fallback))
    preference, _ = UserPreference.objects.get_or_create(user=request.user)
    preference.dashboard_view = view_value
    preference.save()
    messages.success(request, "Default view updated.")
    return redirect(next_url)


@login_required
def user_preferences(request):
    # RBAC: each user controls only their own preferences (preference is always for request.user)
    from apps.siteconfig.user_identity import ensure_user_identity

    identity = ensure_user_identity(request.user, request=request)
    preference = identity.get("portal_preference") or UserPreference.objects.get(
        user=request.user
    )

    if request.method == "GET":
        previous = request.GET.get("next") or request.META.get("HTTP_REFERER")
        previous = _safe_next_url(request, previous, "")
        if previous:
            normalized = previous.split("?")[0]
            if (
                "/siteconfig/preferences" not in normalized
                and "/siteconfig/user_preferences" not in normalized
            ):
                request.session[PORTAL_PREF_PREVIOUS_PAGE] = previous

    if request.method == "POST":
        form = UserPreferenceForm(
            request.POST,
            instance=preference,
            user=request.user,
            request=request,
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Preferences updated.")
            return redirect("siteconfig:user_preferences")
        messages.error(request, "Please fix the errors below.")
    else:
        form = UserPreferenceForm(
            instance=preference,
            user=request.user,
            request=request,
        )

    next_page = _safe_next_url(
        request,
        request.GET.get("next")
        or request.session.pop(PORTAL_PREF_PREVIOUS_PAGE, None)
        or request.META.get("HTTP_REFERER"),
        reverse("accounts:redirect"),
    )
    if next_page and (
        "/siteconfig/preferences" in next_page
        or "/siteconfig/user_preferences" in next_page
    ):
        next_page = reverse("accounts:redirect")

    from apps.siteconfig.control_plane_render import (
        default_operator_breadcrumbs,
        operator_cp_breadcrumb,
        render_siteconfig_stem,
    )

    return render_siteconfig_stem(
        request,
        "user_preferences",
        {
            "form": form,
            "previous_page": next_page,
            "page_title": "Portal preferences",
            "page_subtitle": "Pick your dashboard style, theme, timezone, and how often to refresh data.",
            "action_url": next_page,
        },
        cp_title=_("User preferences"),
        breadcrumbs=default_operator_breadcrumbs(
            operator_cp_breadcrumb(_("User preferences"), active=True),
        ),
    )


def _get_classrooms_queryset():
    """Classrooms for bulk letter dropdown; prefer active year."""
    year, _ = get_active_year_and_term()
    # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
    if year:
        return Classroom.objects.filter(academic_year=year).order_by("name")
    return Classroom.objects.select_related("academic_year").order_by(
        "-academic_year__start_date", "name"
    )


BULK_LETTER_BODY_MAX_LENGTH = 100_000


def _bulk_letters_page_context(request, classroom_list, form_data):
    """Operator markers, related CP links, and read-only summary for bulk letters (1098/1072)."""
    total_students = sum(item["student_count"] for item in classroom_list)
    ctx = {
        "classroom_list": classroom_list,
        "form_data": form_data,
        "bulk_letters_operator_summary": {
            "classrooms": len(classroom_list),
            "students_total": total_students,
        },
    }
    for key, viewname in (
        ("scheduled_reports_delivery_hub_url", "siteconfig:scheduled_reports_delivery_hub"),
        ("report_templates_catalog_evidence_url", "siteconfig:report_templates_catalog_evidence"),
        ("report_output_history_evidence_url", "siteconfig:report_output_history_evidence"),
    ):
        try:
            ctx[key] = reverse(viewname)
        except NoReverseMatch:
            ctx[key] = None
    ctx["admin_report_template_changelist_url"] = None
    user = getattr(request, "user", None)
    if user is not None and getattr(user, "is_superuser", False):
        try:
            ctx["admin_report_template_changelist_url"] = reverse(
                "admin:siteconfig_reporttemplate_changelist"
            )
        except NoReverseMatch:
            ctx["admin_report_template_changelist_url"] = None
    try:
        ctx["northstar_ai_draft_url"] = reverse("siteconfig:northstar_ai_draft")
    except NoReverseMatch:
        ctx["northstar_ai_draft_url"] = None
    return ctx


def _bulk_letters_form_data(request):
    """Extract form data from POST for re-display on validation errors."""
    letter_body = request.POST.get("letter_body") or ""
    # Escape only </textarea> so re-display doesn't break the HTML textarea
    letter_body_display = letter_body.replace("</textarea>", "&lt;/textarea&gt;")
    return {
        "classroom_id": (request.POST.get("classroom_id") or "").strip(),
        "letter_title": (request.POST.get("letter_title") or "").strip(),
        "letter_body": letter_body,
        "letter_body_display": letter_body_display,
        "include_pdf": request.POST.get("include_pdf") == "on",
    }


@permission_required("settings.manage")
def bulk_letters(request):
    # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
    """Generate one ODT letter per student in a classroom (mail-merge style). Requires Pandoc."""
    classrooms = _get_classrooms_queryset()
    student_counts = dict(
        # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
        StudentProfile.objects.filter(classroom__in=classrooms)
        .values("classroom_id")
        .annotate(count=Count("id"))
        .values_list("classroom_id", "count")
    )
    classroom_list = [
        {"room": r, "student_count": student_counts.get(r.id, 0)} for r in classrooms
    ]
    if request.method != "POST":
        return render(
            request,
            "siteconfig/bulk_letters.html",
            _bulk_letters_page_context(request, classroom_list, None),
        )
    form_data = _bulk_letters_form_data(request)
    classroom_id = form_data["classroom_id"]
    letter_title = form_data.get("letter_title") or ""
    letter_body = form_data["letter_body"].strip()
    include_pdf = form_data["include_pdf"]
    if not letter_body:
        messages.warning(request, "Please enter the letter body.")
        return render(
            request,
            "siteconfig/bulk_letters.html",
            _bulk_letters_page_context(request, classroom_list, form_data),
        )
    if len(letter_body) > BULK_LETTER_BODY_MAX_LENGTH:
        messages.warning(
            request,
            f"Letter body is too long (max {BULK_LETTER_BODY_MAX_LENGTH:,} characters).",
        )
        return render(
            request,
            "siteconfig/bulk_letters.html",
            _bulk_letters_page_context(request, classroom_list, form_data),
        )
    classroom = None
    if classroom_id:
        # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
        try:
            classroom = Classroom.objects.get(pk=classroom_id)
        except (ValueError, Classroom.DoesNotExist):
            pass
    if not classroom:
        messages.warning(request, "Please select a classroom.")
        return render(
            request,
            "siteconfig/bulk_letters.html",
            # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
            _bulk_letters_page_context(request, classroom_list, form_data),
        )
    # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
    students = StudentProfile.objects.filter(classroom=classroom).order_by(
        "last_name", "first_name"
    )
    if not students.exists():
        messages.warning(request, f"No students in {classroom.name}.")
        return render(
            request,
            "siteconfig/bulk_letters.html",
            _bulk_letters_page_context(request, classroom_list, form_data),
        )
    try:
        from apps.portal.document_generation import html_to_odt
    except ImportError:
        messages.error(
            request, "Bulk letters require the portal document_generation module."
        )
        return render(
            request,
            "siteconfig/bulk_letters.html",
            _bulk_letters_page_context(request, classroom_list, form_data),
        )
    if include_pdf:
        try:
            from apps.portal.document_service import convert_document
        except ImportError:
            messages.error(
                request, "PDF option requires the portal document_conversion module."
            )
            return render(
                request,
                "siteconfig/bulk_letters.html",
                _bulk_letters_page_context(request, classroom_list, form_data),
            )
    buf = io.BytesIO()
    pdf_skipped = []  # list of "LastName FirstName (reason)" when PDF conversion is skipped
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for student in students:
            first_name = student.first_name or ""
            last_name = student.last_name or ""
            student_code = student.student_code or ""
            classroom_name = classroom.name or classroom.code or ""
            body = (
                letter_body.replace("{{ first_name }}", first_name)
                .replace("{{ last_name }}", last_name)
                .replace("{{ student_code }}", student_code)
                .replace("{{ classroom }}", classroom_name)
            )
            html = (
                "<!DOCTYPE html><html><head><meta charset='utf-8'/></head><body>"
                "<div style='font-family: sans-serif;'>" + body + "</div></body></html>"
            )
            doc_title = letter_title or f"Letter - {last_name} {first_name}"
            try:
                odt_bytes = html_to_odt(html, title=doc_title)
            except RuntimeError as e:
                messages.error(request, f"Pandoc conversion failed: {e}")
                return render(
                    request,
                    "siteconfig/bulk_letters.html",
                    _bulk_letters_page_context(request, classroom_list, form_data),
                )
            safe_name = f"{last_name}_{first_name}_{student_code}".replace(" ", "_")
            zf.writestr(f"letter_{safe_name}.odt", odt_bytes)
            if include_pdf:
                fd, odt_path = tempfile.mkstemp(suffix=".odt", prefix="bulk_letter_")
                try:
                    os.write(fd, odt_bytes)
                    os.close(fd)
                    fd = None
                    pdf_bytes = convert_document(odt_path, target="pdf", family="writer")
                    zf.writestr(f"letter_{safe_name}.pdf", pdf_bytes)
                except RuntimeError as e:
                    pdf_skipped.append(f"{last_name} {first_name} - {e}")
                finally:
                    if fd is not None:
                        try:
                            os.close(fd)
                        except OSError:
                            pass
                    try:
                        os.unlink(odt_path)
                    except OSError:
                        pass
        if pdf_skipped:
            note = (
                "PDF conversion was skipped for the following (ODT included; LibreOffice may be missing or failed):\n\n"
                + "\n".join(pdf_skipped)
            )
            zf.writestr("PDF_CONVERSION_SKIPPED.txt", note.encode("utf-8"))
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = (
        f'attachment; filename="bulk_letters_{classroom.code}.zip"'
    )
    return response


def _theme_experience_canonical_url() -> str:
    """Studio OS Experience mode is the canonical theme entry; keep legacy URL for embed/standalone."""
    try:
        return reverse("studio_os:experience")
    except NoReverseMatch:
        return reverse("siteconfig:theme_colors")


def _theme_post_success_redirect(request):
    """After save, send staff to Studio; others stay on standalone theme form (same permission as POST)."""
    from apps.schools.control_plane import user_can_access_studio_on_request

    if user_can_access_studio_on_request(request):
        try:
            return redirect(reverse("studio_os:experience"))
        except NoReverseMatch:
            pass
    return redirect(f"{reverse('siteconfig:theme_colors')}?standalone=1")


@permission_required("settings.manage")
def theme_colors_page(request):
    """Standalone Color & harmony page: palette studio, presets, preview, and save flows."""
    from apps.siteconfig.theme_builder_plane import assert_theme_colors_request_plane

    assert_theme_colors_request_plane(request)
    if request.method == "GET":
        if request.GET.get("embed") != "1" and request.GET.get("standalone") != "1":
            from apps.schools.control_plane import user_can_access_studio_on_request

            if user_can_access_studio_on_request(request):
                try:
                    return redirect(reverse("studio_os:experience"))
                except NoReverseMatch:
                    pass
            return redirect(f"{reverse('siteconfig:theme_colors')}?standalone=1")
    site = get_effective_site_settings(request=request)
    if site is None:
        site = build_platform_default_site_settings()
    theme_settings = _get_theme_experience_settings(site)
    all_packs = list(
        ThemePack.objects.filter(is_active=True).order_by(
            "-applies_to_admin", "-is_default", "name"
        )
    )
    canonical_admin_slugs = {
        slug
        for _group_name, slugs in THEME_PALETTE_GROUPS
        for slug in slugs
        if slug.startswith("admin-")
    }
    active_admin_count = sum(
        1 for pack in all_packs if getattr(pack, "applies_to_admin", False)
    )
    has_seeded_admin_catalog = any(
        pack.slug in canonical_admin_slugs for pack in all_packs
    )
    # Safety net for environments where predeploy seed command was skipped.
    # We expect several distinct admin packs so the catalog does not collapse.
    if active_admin_count < 6 or not has_seeded_admin_catalog:
        try:
            call_command("seed_admin_dashboard_palettes")
            all_packs = list(
                ThemePack.objects.filter(is_active=True).order_by(
                    "-applies_to_admin", "-is_default", "name"
                )
            )
            logger.info("Theme catalog auto-seeded from Theme & Experience page.")
        except (CommandError, OSError, RuntimeError, TypeError, ValueError):
            log_view_exception(
                request,
                "siteconfig.views.theme_colors_page: unable to auto-seed admin dashboard palettes",
            )

    # Show all active packs (admin and portal) so the catalog does not collapse to a single card.
    admin_theme_packs = all_packs
    admin_theme_packs_by_group = build_theme_pack_groups(
        admin_theme_packs, THEME_PALETTE_GROUPS
    )
    # Slim tenant settings row: ThemeColorsForm.Meta.fields is empty; guard/compare all experience fields.
    tracked_theme_fields = list(THEME_EXPERIENCE_FIELD_NAMES)

    if request.method == "POST":
        baseline_values = _snapshot_theme_field_values(site, tracked_theme_fields)
        form = ThemeColorsForm(request.POST, instance=site, request=request)
        if form.is_valid():
            changed_fields = []
            for field_name in tracked_theme_fields:
                previous_value = baseline_values.get(field_name)
                next_value = form.cleaned_data.get(field_name)
                if hasattr(previous_value, "pk"):
                    previous_value = previous_value.pk
                if hasattr(next_value, "pk"):
                    next_value = next_value.pk
                if str(previous_value) != str(next_value):
                    changed_fields.append(field_name)
            preview_confirmed = request.POST.get("preview_confirmed") in (
                "1",
                "true",
                "on",
            )
            if theme_settings.get("skip_theme_publish_guard", False):
                preview_confirmed = True

            changed_labels = []
            for field_name in changed_fields:
                field_obj = form.fields.get(field_name)
                label = (
                    field_obj.label
                    if field_obj
                    else field_name.replace("_", " ").title()
                )
                changed_labels.append(str(label))
            governed_changes = [
                name for name in changed_fields if name in THEME_PUBLISH_GUARDED_FIELDS
            ]
            governed_labels = []
            for field_name in governed_changes:
                field_obj = form.fields.get(field_name)
                label = (
                    field_obj.label
                    if field_obj
                    else field_name.replace("_", " ").title()
                )
                governed_labels.append(str(label))

            now_label = timezone.localtime().strftime("%Y-%m-%d %H:%M")
            actor_label = (
                request.user.get_username()
                if request.user.is_authenticated
                else "system"
            )
            if governed_changes and not preview_confirmed:
                request.session["theme_recent_change_meta"] = {
                    "status": "blocked",
                    "actor": actor_label,
                    "timestamp": now_label,
                    "changed_count": len(changed_fields),
                    "changed_fields": changed_labels,
                    "governed_count": len(governed_changes),
                    "governed_fields": governed_labels,
                }
                request.session.modified = True
                preview_hint = ", ".join(governed_labels[:4])
                if len(governed_labels) > 4:
                    preview_hint += ", ..."
                messages.error(
                    request,
                    "Live preview confirmation is required before publishing high-impact theme changes: "
                    f"{preview_hint}. Use Live preview, then tick Confirm and publish, then Save again.",
                )
            else:
                form.save()
                if changed_fields:
                    request.session["theme_previous_state"] = {
                        "values": baseline_values,
                        "actor": actor_label,
                        "timestamp": now_label,
                    }
                request.session["theme_recent_change_meta"] = {
                    "status": "saved",
                    "actor": actor_label,
                    "timestamp": now_label,
                    "changed_count": len(changed_fields),
                    "changed_fields": changed_labels,
                    "governed_count": len(governed_changes),
                    "governed_fields": governed_labels,
                    "preview_confirmed": bool(preview_confirmed),
                }
                request.session.modified = True
                if changed_fields and governed_changes:
                    messages.success(
                        request,
                        "Theme & experience settings published after live preview confirmation.",
                    )
                elif changed_fields:
                    messages.success(request, "Theme & experience settings saved.")
                else:
                    messages.info(request, "No theme changes detected.")
                back_url = _safe_next_url(
                    request,
                    request.GET.get("next") or request.META.get("HTTP_REFERER"),
                    "",
                )
                if back_url:
                    return redirect(back_url)
                return _theme_post_success_redirect(request)
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        form = ThemeColorsForm(instance=site, request=request)

    from apps.siteconfig.staff_navigation import site_settings_change_url

    admin_change_url = site_settings_change_url(request, site.pk)
    back_url = _safe_next_url(request, request.GET.get("next"), admin_change_url)
    preview_config = _get_preview_platform_config(site)
    preview_mode_active = bool(
        request.session.get("preview_mode_enabled")
        or preview_config.get("preview_mode_enabled", False)
    )
    theme_recent_change_meta = request.session.get("theme_recent_change_meta")

    contrast_values = {}
    for field_name in (
        "primary_color",
        "accent_color",
        "header_bg_color",
        "footer_bg_color",
        "success_color",
        "warning_color",
        "danger_color",
    ):
        incoming = form.data.get(field_name) if form.is_bound else None
        if incoming in (None, ""):
            incoming = theme_settings.get(field_name, getattr(site, field_name, ""))
        contrast_values[field_name] = incoming
    contrast_report = build_theme_contrast_report(contrast_values)

    from apps.siteconfig.control_plane_render import (
        default_operator_breadcrumbs,
        operator_cp_breadcrumb,
        render_siteconfig_operator_page,
    )

    theme_ctx = {
        "form": form,
        "site_settings": site,
        "preview_mode_active": preview_mode_active,
        "admin_theme_packs": admin_theme_packs,
        "admin_theme_packs_by_group": admin_theme_packs_by_group,
        "admin_change_url": admin_change_url,
        "back_url": back_url,
        "theme_recent_change_meta": theme_recent_change_meta,
        "theme_contrast_report": contrast_report,
        "theme_contrast_targets": theme_contrast_targets_for_client(),
        "theme_publish_guarded_count": len(THEME_PUBLISH_GUARDED_FIELDS),
        "skip_theme_publish_guard": bool(
            theme_settings.get("skip_theme_publish_guard", False)
        ),
    }
    if request.GET.get("embed") == "1":
        from apps.studio_os.embed_render import render_studio_embed_body

        return render_studio_embed_body(
            request,
            "siteconfig/partials/theme_colors_page_body.html",
            theme_ctx,
            title=_("Theme & Experience"),
        )
    return render_siteconfig_operator_page(
        request,
        portal_template="siteconfig/theme_colors.html",
        body_template="siteconfig/partials/theme_colors_page_body.html",
        context=theme_ctx,
        cp_title=_("Theme & Experience"),
        breadcrumbs=default_operator_breadcrumbs(
            operator_cp_breadcrumb(_("Studio"), url=reverse("studio_os:shell")),
            operator_cp_breadcrumb(_("Theme & Experience"), active=True),
        ),
    )


def perform_theme_experience_publish(request):
    """
    Validate and persist theme/experience from request.POST. Used by Studio OS publish API.
    Returns {"ok": True, "redirect_url": "..."} or {"ok": False, "errors": [...]}.
    Caller must enforce staff/permission; this does not check.
    """
    from .forms import ThemeColorsForm

    site = get_effective_site_settings(request=request)
    if site is None:
        site = build_platform_default_site_settings()
    theme_settings = _get_theme_experience_settings(site)
    tracked_theme_fields = list(THEME_EXPERIENCE_FIELD_NAMES)
    baseline_values = _snapshot_theme_field_values(site, tracked_theme_fields)
    form = ThemeColorsForm(request.POST, instance=site, request=request)
    if not form.is_valid():
        errors = []
        for field, errs in form.errors.items():
            for e in errs:
                errors.append(f"{field}: {e}")
        return {"ok": False, "errors": errors}

    changed_fields = []
    for field_name in tracked_theme_fields:
        previous_value = baseline_values.get(field_name)
        next_value = form.cleaned_data.get(field_name)
        if hasattr(previous_value, "pk"):
            previous_value = previous_value.pk
        if hasattr(next_value, "pk"):
            next_value = next_value.pk
        if str(previous_value) != str(next_value):
            changed_fields.append(field_name)
    preview_confirmed = request.POST.get("preview_confirmed") in ("1", "true", "on")
    if theme_settings.get("skip_theme_publish_guard", False):
        preview_confirmed = True
    governed_changes = [n for n in changed_fields if n in THEME_PUBLISH_GUARDED_FIELDS]
    if governed_changes and not preview_confirmed:
        labels = []
        for fn in governed_changes[:4]:
            f = form.fields.get(fn)
            labels.append(f.label if f else fn.replace("_", " ").title())
        if len(governed_changes) > 4:
            labels.append("...")
        return {
            "ok": False,
            "errors": [
                "Live preview confirmation required for high-impact theme changes: "
                + ", ".join(labels)
                + ". Use Preview, then confirm and publish again.",
            ],
        }

    now_label = timezone.localtime().strftime("%Y-%m-%d %H:%M")
    actor_label = (
        request.user.get_username() if request.user.is_authenticated else "system"
    )
    form.save()
    if changed_fields:
        request.session["theme_previous_state"] = {
            "values": baseline_values,
            "actor": actor_label,
            "timestamp": now_label,
        }
    changed_labels = [
        form.fields.get(fn).label
        if form.fields.get(fn)
        else fn.replace("_", " ").title()
        for fn in changed_fields
    ]
    governed_labels = [
        form.fields.get(fn).label
        if form.fields.get(fn)
        else fn.replace("_", " ").title()
        for fn in governed_changes
    ]
    request.session["theme_recent_change_meta"] = {
        "status": "saved",
        "actor": actor_label,
        "timestamp": now_label,
        "changed_count": len(changed_fields),
        "changed_fields": changed_labels,
        "governed_count": len(governed_changes),
        "governed_fields": governed_labels,
        "preview_confirmed": bool(preview_confirmed),
    }
    request.session.modified = True
    from apps.schools.control_plane import user_can_access_studio_on_request

    try:
        if user_can_access_studio_on_request(request):
            redirect_url = reverse("studio_os:experience")
        else:
            redirect_url = f"{reverse('siteconfig:theme_colors')}?standalone=1"
    except NoReverseMatch:
        log_view_exception(
            request,
            "theme/experience save: studio_os:experience reverse failed, falling back to theme_colors",
            extra={"fallback": "siteconfig:theme_colors"},
        )
        redirect_url = f"{reverse('siteconfig:theme_colors')}?standalone=1"
    return {"ok": True, "redirect_url": redirect_url}


def get_theme_colors_context(request):
    """
    Build theme & experience context for Studio OS or other callers.
    Same keys as theme_colors_page (GET path). No permission check (caller must ensure access).
    """
    site = get_effective_site_settings(request=request)
    if site is None:
        site = build_platform_default_site_settings()
    theme_settings = _get_theme_experience_settings(site)
    form = ThemeColorsForm(instance=site, request=request)
    all_packs = list(
        ThemePack.objects.filter(is_active=True).order_by(
            "-applies_to_admin", "-is_default", "name"
        )
    )
    admin_theme_packs = all_packs
    admin_theme_packs_by_group = build_theme_pack_groups(
        admin_theme_packs, THEME_PALETTE_GROUPS
    )
    from apps.siteconfig.staff_navigation import site_settings_change_url

    admin_change_url = site_settings_change_url(request, site.pk)
    back_url = _safe_next_url(request, request.GET.get("next"), admin_change_url or "")
    preview_config = _get_preview_platform_config(site)
    preview_mode_active = bool(
        request.session.get("preview_mode_enabled")
        or preview_config.get("preview_mode_enabled", False)
    )
    theme_recent_change_meta = request.session.get("theme_recent_change_meta")
    contrast_values = {}
    for field_name in (
        "primary_color",
        "accent_color",
        "header_bg_color",
        "footer_bg_color",
        "success_color",
        "warning_color",
        "danger_color",
    ):
        incoming = form.data.get(field_name) if form.is_bound else None
        if incoming in (None, ""):
            incoming = theme_settings.get(field_name, getattr(site, field_name, ""))
        contrast_values[field_name] = incoming
    contrast_report = build_theme_contrast_report(contrast_values)
    return {
        "form": form,
        "site_settings": site,
        "preview_mode_active": preview_mode_active,
        "admin_theme_packs": admin_theme_packs,
        "admin_theme_packs_by_group": admin_theme_packs_by_group,
        "admin_change_url": admin_change_url,
        "back_url": back_url,
        "theme_recent_change_meta": theme_recent_change_meta,
        "theme_contrast_report": contrast_report,
        "theme_contrast_targets": theme_contrast_targets_for_client(),
        "theme_token_values": contrast_values,
        "theme_publish_guarded_count": len(THEME_PUBLISH_GUARDED_FIELDS),
        "skip_theme_publish_guard": bool(
            theme_settings.get("skip_theme_publish_guard", False)
        ),
    }


@permission_required("settings.manage")
def theme_experience_redirect(request):
    """
    Legacy route: default to the dual-plane Theme & Experience hub.

    Use ``?studio=1`` to jump straight to Studio Experience (older bookmarks).
    """
    next_url = _safe_next_url(request, request.GET.get("next"), "")
    if request.GET.get("studio") in ("1", "true", "on"):
        target = _theme_experience_canonical_url()
        if next_url:
            target = f"{target}?{urlencode({'next': next_url})}"
        return redirect(target)
    target = reverse("siteconfig:theme_experience_hub")
    if next_url:
        target = f"{target}?{urlencode({'next': next_url})}"
    return redirect(target)


@permission_required("settings.manage")
def brand_import_from_url_view(request):
    """
    POST url + consent: fetch URL, parse brand (theme-color, og:image, title), apply primary_color and site_name to the tenant settings singleton.
    Non-negotiable: Website/competitor import — HOW_WE_SCOPE_WEBSITE_IMPORT implemented.
    """
    if request.method != "POST":
        return redirect(_theme_experience_canonical_url())
    consent = request.POST.get("consent") in ("1", "true", "on")
    if not consent:
        messages.error(request, "Consent is required to fetch an external URL.")
        return redirect(_theme_experience_canonical_url())
    url = (request.POST.get("url") or "").strip()
    if not url:
        messages.error(request, "URL is required.")
        return redirect(_theme_experience_canonical_url())
    from apps.siteconfig.brand_import import fetch_and_parse_brand_url

    result = fetch_and_parse_brand_url(url)
    if result.get("error"):
        messages.error(request, result["error"])
        return redirect(_theme_experience_canonical_url())
    site = get_effective_site_settings(request=request)
    if site and site.pk:
        field_updates = {}
        if result.get("primary_color"):
            field_updates["primary_color"] = result["primary_color"]
        if result.get("site_name"):
            field_updates["site_name"] = result["site_name"][:120]
        if field_updates:
            site.apply_theme_experience_state(field_updates=field_updates, save=True)
        # N17: Suggested theme + metadata apply uses template gallery (impact preview + session gate).
        apply_theme = request.POST.get("apply_theme") in ("1", "true", "on")
        if apply_theme and result.get("suggested_theme_pack_slug"):
            pack = ThemePack.objects.filter(
                slug=result["suggested_theme_pack_slug"], is_active=True
            ).first()
            if pack:
                messages.info(
                    request,
                    _(
                        "Brand details saved. Review dependency and package impact for the suggested theme in the template gallery, then confirm apply there."
                    ),
                )
                return redirect(
                    f"{reverse('siteconfig:template_gallery')}?{urlencode({'preview_slug': pack.slug})}"
                )
            messages.success(
                request,
                "Brand details applied. Refine colors and logo below if needed.",
            )
        else:
            messages.success(
                request,
                "Brand details applied. Refine colors and logo below if needed.",
            )
    else:
        messages.info(request, "Brand fetched; create or save site settings to apply.")
    return redirect(_theme_experience_canonical_url())


@permission_required("settings.manage")
def template_gallery_page(request):
    """
    Template gallery: list ThemePacks as templates; Preview links to theme_colors; Use applies pack to tenant.
    Non-negotiable: Strategy Report Phase 2 — template gallery with preview-before-publish.
    N17: dependency_graph + preview gate (session) before metadata apply.
    """
    from apps.packages.engine import metadata_apply_preview_bundle

    site = get_effective_site_settings(request=request)
    if site is None:
        site = build_platform_default_site_settings()
    packs = list(
        ThemePack.objects.filter(is_active=True).order_by("-is_default", "name")
    )
    school = getattr(request, "school", None)
    preview_slug = (request.GET.get("preview_slug") or "").strip()
    impact_bundle = None
    if preview_slug and school:
        pack_prev = next((p for p in packs if p.slug == preview_slug), None)
        if pack_prev:
            ver = getattr(pack_prev, "version", "1") or "1"
            pl = {"theme": {"name": pack_prev.name}}
            impact_bundle = metadata_apply_preview_bundle(
                school.pk, pack_prev.slug, ver, pl
            )
            request.session["template_gallery_impact_gate"] = {
                "slug": preview_slug,
                "ts": time.time(),
            }
    if request.method == "POST":
        slug = (request.POST.get("template_slug") or "").strip()
        if not slug:
            return redirect(reverse("siteconfig:template_gallery"))
        if not school:
            messages.error(
                request,
                _("No school context; cannot apply template."),
            )
            return redirect(reverse("siteconfig:template_gallery"))
        if request.POST.get("confirm_metadata_apply") != "1":
            messages.error(
                request,
                _(
                    "Confirm that you reviewed dependency and package impact before applying."
                ),
            )
            return redirect(
                f"{reverse('siteconfig:template_gallery')}?{urlencode({'preview_slug': slug})}"
            )
        gate = request.session.get("template_gallery_impact_gate") or {}
        gate_ts = float(gate.get("ts") or 0)
        if gate.get("slug") != slug or (time.time() - gate_ts > 900):
            messages.error(
                request,
                _(
                    "Open “Review impact” for this template within 15 minutes before applying."
                ),
            )
            return redirect(
                f"{reverse('siteconfig:template_gallery')}?{urlencode({'preview_slug': slug})}"
            )
        pack = ThemePack.objects.filter(slug=slug, is_active=True).first()
        if pack and site.pk:
            site.apply_theme_pack(pack, save=True)
            try:
                from apps.packages.engine import PackageEngine

                PackageEngine.apply_package(
                    tenant_id=getattr(school, "id", None),
                    package_id=pack.slug,
                    version=getattr(pack, "version", "1") or "1",
                    payload_sections={"theme": {"name": pack.name}},
                    mode="production",
                    actor_id=getattr(request.user, "id", None)
                    if getattr(request, "user", None)
                    else None,
                )
            except (
                AttributeError,
                ImportError,
                LookupError,
                OSError,
                RuntimeError,
                TypeError,
                ValueError,
            ):
                pass
            request.session.pop("template_gallery_impact_gate", None)
            messages.success(
                request,
                f'Template "{pack.name}" applied. You can refine colors in Theme & Experience.',
            )
        return redirect(_theme_experience_canonical_url())
    theme_colors_url = _theme_experience_canonical_url()
    impact_graph_json = None
    if impact_bundle:
        dg = impact_bundle.get("dependency_graph") or {}
        impact_graph_json = {
            "center_id": impact_bundle.get("package_id") or "",
            "upstream_package_ids": list(dg.get("upstream_package_ids") or []),
            "downstream_package_ids": list(dg.get("downstream_package_ids") or []),
        }
    return render(
        request,
        "siteconfig/template_gallery.html",
        {
            "templates": packs,
            "theme_colors_url": theme_colors_url,
            "preview_slug": preview_slug,
            "impact_bundle": impact_bundle,
            "impact_graph_json": impact_graph_json,
        },
    )


@staff_member_required(login_url=settings.LOGIN_URL)
def toggle_preview_mode(request):
    enabled = bool(request.session.get(PREVIEW_MODE_SESSION_KEY))
    request.session[PREVIEW_MODE_SESSION_KEY] = not enabled
    status = "enabled" if not enabled else "disabled"
    messages.info(request, f"Preview/sandbox mode {status}.")
    next_url = _safe_next_url(
        request, request.GET.get("next") or request.META.get("HTTP_REFERER"), "/"
    )
    return redirect(next_url)


@staff_member_required(login_url=settings.LOGIN_URL)
def set_act_as_role(request):
    if request.method != "POST":
        next_url = _safe_next_url(
            request, request.GET.get("next") or request.META.get("HTTP_REFERER"), "/"
        )
        return redirect(next_url)

    role_code = request.POST.get("role")
    valid_roles = {code: label for code, label in User.Role.choices}
    previous = request.session.get(ACT_AS_ROLE_SESSION_KEY)

    if role_code in valid_roles:
        request.session[ACT_AS_ROLE_SESSION_KEY] = role_code
        messages.info(request, f"Now acting as {valid_roles[role_code]}.")
        logger.info(
            "User %s acting as %s (was %s)", request.user.username, role_code, previous
        )
    else:
        request.session.pop(ACT_AS_ROLE_SESSION_KEY, None)
        messages.info(request, "Act-as role cleared.")
        logger.info(
            "User %s cleared act-as role (was %s)", request.user.username, previous
        )

    next_url = _safe_next_url(
        request,
        request.POST.get("next")
        or request.GET.get("next")
        or request.META.get("HTTP_REFERER"),
        "/",
    )
    return redirect(next_url)


def _report_format_from_request(request, template):
    """Resolve format: optional ?format= overrides template.preferred_format."""
    raw = (request.GET.get("format") or "").strip().upper()
    if raw in ReportTemplate.ReportFormat.values:
        return raw
    return (template.preferred_format or ReportTemplate.ReportFormat.CSV).upper()


def _report_filename(slug, fmt):
    """File extension for format (EXCEL -> .xlsx)."""
    if fmt == ReportTemplate.ReportFormat.EXCEL:
        return f"{slug}.xlsx"
    if fmt == ReportTemplate.ReportFormat.ODS:
        return f"{slug}.ods"
    if fmt == ReportTemplate.ReportFormat.PDF:
        return f"{slug}.pdf"
    return f"{slug}.csv"


@permission_required("settings.manage")
def download_report(request, slug):
    template = get_object_or_404(ReportTemplate, slug=slug, is_active=True)
    school = getattr(request, "school", None)
    if school and template.template_family:
        from apps.siteconfig.tenant_config import get_report_template_family_for_school

        family = get_report_template_family_for_school(school)
        if family and template.template_family != family:
            from django.http import HttpResponseForbidden

            return HttpResponseForbidden(
                "This report template is not available for your school's configuration."
            )
    headers, rows = template.get_export_data()

    if not headers:
        messages.warning(request, "No export handler registered for this report.")
        return redirect("studio_os:output")

    fmt = _report_format_from_request(request, template)
    filename = _report_filename(template.slug, fmt)

    if fmt == ReportTemplate.ReportFormat.EXCEL:
        return _render_xlsx_response(headers, rows, filename)
    if fmt == ReportTemplate.ReportFormat.ODS:
        return _render_ods_response(headers, rows, filename)
    if fmt == ReportTemplate.ReportFormat.PDF:
        return _render_report_pdf_response(
            request, template.name, headers, rows, filename
        )

    return render_csv_response(headers, rows, filename)


def render_csv_response(headers, rows, filename) -> HttpResponse:
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def _render_xlsx_response(headers, rows, filename) -> HttpResponse:
    """Serve report as Excel (.xlsx) using openpyxl."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse(
            "Excel export requires openpyxl. Install with: pip install openpyxl",
            status=503,
            content_type="text/plain",
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    for col, value in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=value)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _render_ods_response(headers, rows, filename) -> HttpResponse:
    """Serve report as LibreOffice Calc (.ods) using odfpy."""
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
    except ImportError:
        return HttpResponse(
            "ODS export requires odfpy. Install with: pip install odfpy",
            status=503,
            content_type="text/plain",
        )
    from io import BytesIO

    doc = OpenDocumentSpreadsheet()
    table = Table(name="Report")
    doc.spreadsheet.addElement(table)
    # Header row
    tr = TableRow()
    for h in headers:
        tc = TableCell(valuetype="string")
        tc.addElement(P(text=str(h)))
        tr.addElement(tc)
    table.addElement(tr)
    # Data rows
    for row in rows:
        tr = TableRow()
        for val in row:
            tc = TableCell(valuetype="string")
            tc.addElement(P(text=str(val)))
            tr.addElement(tc)
        table.addElement(tr)
    buf = BytesIO()
    doc.write(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.oasis.opendocument.spreadsheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _render_report_pdf_response(
    request, report_name, headers, rows, filename
) -> HttpResponse:
    """Serve report as PDF using WeasyPrint (table template)."""
    from apps.reports.weasy import render_pdf_bytes

    context = {"report_name": report_name, "headers": headers, "rows": rows}
    pdf_bytes = render_pdf_bytes(request, "siteconfig/report_table_pdf.html", context)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ==========================
# REGIONAL CONFIGURATION VIEWS
# ==========================


@require_super_access_with_host
def region_validation_dashboard(request):
    """
    Dashboard showing regional configuration status and validation warnings.
    Displays completeness checks for each region.
    """
    import pytz
    from apps.academics.models import AcademicYear

    # Reverse relation name for annotate matches active migrations/ORM (see validate_regions, admin).
    regions = RegionConfig.objects.annotate(
        grading_scales_count=Count("gradingscaleconfig"),
        holidays_count=Count("holidays"),
    )

    validation_results = []
    issues_count = 0

    for region in regions:
        issues = []
        severity = "success"  # success, warning, danger

        # Check grading scales
        if region.grading_scales_count < 5:
            issues.append(
                {
                    "icon": "❌",
                    "type": "danger",
                    "message": f"Missing grading scales ({region.grading_scales_count}/5)",
                }
            )
            severity = "danger"
            issues_count += 1

        # Check timezone validity
        try:
            pytz.timezone(region.timezone)
        except pytz.exceptions.UnknownTimeZoneError:
            issues.append(
                {
                    "icon": "❌",
                    "type": "danger",
                    "message": f"Invalid timezone: {region.timezone}",
                }
            )
            severity = "danger"
            issues_count += 1

        # Check currency
        if not _is_known_currency_code(region.default_currency):
            issues.append(
                {
                    "icon": "⚠️",
                    "type": "warning",
                    "message": f"Unknown currency: {region.default_currency}",
                }
            )
            if severity == "success":
                severity = "warning"
            issues_count += 1

        # Check portal features
        portal_count = sum(
            [
                region.enable_online_admissions,
                region.enable_parent_portal,
                region.enable_student_portal,
            ]
        )
        if portal_count == 0:
            issues.append(
                {
                    "icon": "⚠️",
                    "type": "warning",
                    "message": "No portal features enabled",
                }
            # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
            )
            if severity == "success":
                severity = "warning"

        # tenant-isolation-allow: view-layer-scoped-via-request-school-or-role-graph
        # Check holiday coverage for current year
        current_year = AcademicYear.objects.filter(is_active=True).order_by(
            "-start_date"
        ).first()
        if current_year:
            holidays_for_year = HolidayCalendar.objects.filter(
                region=region, academic_year=current_year
            ).count()
            if holidays_for_year == 0:
                issues.append(
                    {
                        "icon": "ℹ️",
                        "type": "info",
                        "message": f"No holidays configured for {current_year}",
                    }
                )

        validation_results.append(
            {
                "region": region,
                "issues": issues,
                "severity": severity,
                "status_badge": "✓"
                if severity == "success"
                else ("⚠️" if severity == "warning" else "❌"),
                "grading_scales": region.grading_scales_count,
                "holidays": region.holidays_count,
            }
        )

    context = {
        "validation_results": validation_results,
        "total_regions": regions.count(),
        "complete_regions": sum(
            1 for r in validation_results if r["severity"] == "success"
        ),
        "regions_with_warnings": sum(
            1 for r in validation_results if r["severity"] in ["warning", "danger"]
        ),
        "total_issues": issues_count,
    }

    return render_siteconfig_stem(
        request,
        "region_validation_dashboard",
        context,
        cp_title=_("Region validation"),
        cp_meta_description=_(
            "Regional configuration completeness and validation checks for operators."
        ),
        cp_page_archetype="operator-report",
        breadcrumbs=default_operator_breadcrumbs(
            operator_cp_breadcrumb(_("Region validation")),
        ),
    )


@require_super_access_with_host
def region_comparison_view(request):
    """
    Comparison view for regional configurations.
    Shows side-by-side comparison of settings across regions.
    """
    regions = RegionConfig.objects.all().order_by("code")

    # Prepare comparison data
    comparison_data = {
        "Timezone": [r.timezone for r in regions],
        "Date Format": [r.date_format for r in regions],
        "Grading Scale": [r.grading_scale for r in regions],
        "Currency": [r.default_currency for r in regions],
        "Year Starts (Month)": [r.academic_year_start_month for r in regions],
        "Terms per Year": [r.term_count_per_year for r in regions],
        "Online Admissions": [
            "✓" if r.enable_online_admissions else "✗" for r in regions
        ],
        "Parent Portal": ["✓" if r.enable_parent_portal else "✗" for r in regions],
        "Student Portal": ["✓" if r.enable_student_portal else "✗" for r in regions],
    }

    comparison_rows = [
        {"label": k, "values": v} for k, v in comparison_data.items()
    ]

    context = {
        "regions": regions,
        "comparison_data": comparison_data,
        "comparison_rows": comparison_rows,
    }

    return render_siteconfig_stem(
        request,
        "region_comparison",
        context,
        cp_title=_("Region comparison"),
        cp_meta_description=_(
            "Side-by-side comparison of core regional settings across fleet regions."
        ),
        cp_page_archetype="operator-report",
        breadcrumbs=default_operator_breadcrumbs(
            operator_cp_breadcrumb(_("Region comparison")),
        ),
    )


@require_super_access_with_host
def region_grading_scales_view(request):
    """
    Detailed view of all grading scales across all regions.
    Shows breakpoints and allows comparison between scales.
    """
    scales_by_region = {}

    for region in RegionConfig.objects.all():
        scales_by_region[region] = GradingScaleConfig.objects.filter(
            region=region
        ).order_by("scale_type")

    # Prepare comparison matrix
    scale_types = _known_scale_types()

    context = {
        "scales_by_region": scales_by_region,
        "scale_types": scale_types,
    }

    return render_siteconfig_stem(
        request,
        "region_grading_scales_matrix",
        context,
        cp_title=_("Region grading scales matrix"),
        cp_meta_description=_(
            "Fleet-wide grading scale matrix for cross-region operator triage."
        ),
        cp_page_archetype="operator-report",
        breadcrumbs=default_operator_breadcrumbs(
            operator_cp_breadcrumb(_("Region grading matrix")),
        ),
    )


@login_required
def branding_api(request):
    """
    Return resolved tenant branding from the canonical BrandProfile resolver.
    """
    school = getattr(request, "school", None)
    from .branding import resolve_brand_profile

    brand = resolve_brand_profile(
        school=school, site=get_effective_site_settings(request=request)
    )
    return JsonResponse(
        {
            "logo_url": brand.get("logo_url") or "",
            "logo_dark_url": brand.get("logo_dark_url") or "",
            "favicon_url": brand.get("favicon_url") or "",
            "primary_color": brand.get("primary_color") or "#0d6efd",
            "secondary_color": brand.get("secondary_color") or "",
            "accent_color": brand.get("accent_color") or "#198754",
            "font_family": brand.get("font_family") or "",
            "login_background_url": brand.get("login_background_url") or "",
            "custom_css": brand.get("custom_css") or "",
            "tokens": brand.get("tokens") or {},
            "source": brand.get("source") or "",
        },
        safe=False,
    )


@login_required
def workflow_clues_api(request):
    """
    World Engine: workflow setup suggestions by country (Ollama). GET params: workflow_key, country_code.
    """
    if not (
        request.user.is_superuser
        or request.user.has_feature_permission("settings.manage")
    ):
        return JsonResponse({"error": "Forbidden"}, status=403)
    workflow_key = (request.GET.get("workflow_key") or "").strip()
    country_code = (request.GET.get("country_code") or "").strip()[:10]
    if not workflow_key or not country_code:
        return JsonResponse(
            {"error": "workflow_key and country_code are required"},
            status=400,
        )
    try:
        from apps.portal.ai_provider import get_workflow_clues

        text, meta = get_workflow_clues(
            workflow_key,
            country_code,
            request=request,
            school=getattr(request, "school", None),
        )
        return JsonResponse(
            {"suggestions": text, "meta": meta},
            safe=False,
        )
    except (
        AttributeError,
        ImportError,
        LookupError,
        OSError,
        RuntimeError,
        TypeError,
        ValueError,
    ) as e:
        return JsonResponse(
            {"error": str(e), "suggestions": None},
            status=500,
        )


@login_required
def admission_number_preview_api(request):
    """
    Section 22.2: Return sample admission number for current tenant policy (setup preview).
    GET params (optional): year_2digit, school_code, seq_4digit, spec_code, class_segment.
    """
    school = getattr(request, "school", None)
    if not school:
        return JsonResponse({"error": "No tenant context"}, status=400)
    from .identifier_policy_service import (
        default_school_code_for,
        get_admissions_policy,
        preview_admission_number,
    )

    policy = get_admissions_policy(school)
    from datetime import datetime

    year_2digit = (request.GET.get("year_2digit") or str(datetime.now().year % 100))[
        :2
    ].zfill(2)
    seq_4digit = (request.GET.get("seq_4digit") or "0001")[:4].zfill(4)
    spec_code = (request.GET.get("spec_code") or "XX")[:10]
    class_segment = (request.GET.get("class_segment") or "00")[:10]
    school_code = (
        request.GET.get("school_code")
        or policy.get("school_code")
        or default_school_code_for(school)
    ).upper()
    preview = preview_admission_number(
        school,
        year_2digit=year_2digit,
        school_code=school_code,
        seq_4digit=seq_4digit,
        spec_code=spec_code,
        class_segment=class_segment,
    )
    return JsonResponse(
        {
            "preview": preview,
            "policy": {
                "strategy": policy.get("admission_number_strategy"),
                "school_code": policy.get("school_code"),
            },
        },
        safe=False,
    )


@login_required
@staff_member_required(login_url=settings.LOGIN_URL)
def feedback_roadmap(request):
    """Legacy ProductFeedback roadmap — redirects to apps.feedback operator surface.

    ProductFeedback (siteconfig.models_marketing) remains for admin/historical rows;
    new product voice lives in apps.feedback FeatureRequest / ReleaseNote.
    """
    from django.shortcuts import redirect

    return redirect("feedback:product_feedback")
