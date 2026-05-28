"""
Phase 10 — 2.1: Migration hub views (wizard, run list, rollback, legacy view, data cleaner).
Extracted from accounts/views.py for giant-file decomposition.
"""

import csv
import io
import json

from django.conf import settings
from django.db import DatabaseError
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
from django.views.decorators.http import require_http_methods

from apps.accounts.decorators import permission_required
from apps.accounts.models import User
from apps.marketplace.ecosystem_links import build_phase9_ecosystem_links


# Pass 8: tunable upload row cap. Default 10000 instead of the previous hardcoded
# 500. Schools doing real PowerSchool/Veracross migrations routinely have 2k-8k
# student rows; the old cap silently truncated them.
def _import_max_rows() -> int:
    try:
        value = int(getattr(settings, "IMPORT_MAX_ROWS", 10000))
    except (TypeError, ValueError):
        value = 10000
    return max(100, min(value, 100000))


# Pass 8: parse CSV or XLSX from the uploaded file. Branches on filename
# extension (and trusts a `.xls` magic-bytes hint for old binary files).
# openpyxl is already in requirements.txt; xlrd is not, so .xls falls back to
# a clear error rather than a silent failure.
def _read_uploaded_table(file_obj) -> tuple[list[str], list[dict], str]:
    """
    Return (headers, rows, fmt) where fmt is one of {"csv", "xlsx"}.
    Raises ValueError with a user-readable message on parse failure.
    """
    name = (getattr(file_obj, "name", "") or "").strip().lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError(
                "XLSX support requires openpyxl. Save the file as CSV (UTF-8) and re-upload."
            ) from exc
        try:
            file_obj.seek(0)
            wb = load_workbook(file_obj, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                return [], [], "xlsx"
            headers = [str(c).strip() if c is not None else "" for c in header_row]
            max_rows = _import_max_rows()
            rows: list[dict] = []
            for row in rows_iter:
                if len(rows) >= max_rows:
                    break
                if row is None:
                    continue
                if not any(cell not in (None, "") for cell in row):
                    continue
                record = {}
                for idx, cell in enumerate(row):
                    if idx >= len(headers):
                        break
                    key = headers[idx]
                    if not key:
                        continue
                    record[key] = "" if cell is None else str(cell)
                rows.append(record)
            return headers, rows, "xlsx"
        except (ValueError, TypeError, KeyError, OSError) as exc:
            raise ValueError(f"Could not read the XLSX file. Details: {exc}") from exc
    if name.endswith(".xls"):
        raise ValueError(
            "Legacy .xls (Excel 97-2003) is not supported. Save the file as .xlsx or .csv and re-upload."
        )
    # Default: CSV
    try:
        content = file_obj.read().decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError(
            "Could not decode the CSV file. Use UTF-8 encoding and re-upload."
        ) from exc
    try:
        reader = csv.DictReader(io.StringIO(content))
        headers = list(reader.fieldnames or [])
        max_rows = _import_max_rows()
        rows = []
        for row in reader:
            if len(rows) >= max_rows:
                break
            rows.append(row)
        return headers, rows, "csv"
    except (csv.Error, TypeError, KeyError, ValueError) as exc:
        raise ValueError(f"Could not read the CSV file. Details: {exc}") from exc


def _is_admin_user(user):
    return user.is_authenticated and (
        user.is_superuser
        or user.is_staff
        or getattr(user, "role", None) == User.Role.ADMIN
    )


# Plan XII: Migration Hub — upload → field mapping → preview → run
MIGRATION_TYPES = {
    "students": {
        "label": "Students",
        "target_fields": [
            "first_name",
            "last_name",
            "admission_number",
            "academic_year",
            "classroom",
            "specialty",
            "status",
        ],
        "required": ["first_name", "last_name"],
    },
    "grades": {
        "label": "Grades",
        "target_fields": [
            "student_code",
            "subject_assignment_id",
            "term_id",
            "teacher_username",
            "seq1",
            "seq2",
            "exam",
            "mock",
            "practical",
            "test1",
            "test2",
            "remarks",
        ],
        "required": ["student_code", "subject_assignment_id", "term_id"],
    },
    # Pass 8: new import types unlocking competitor SIS migration.
    "teachers": {
        "label": "Teachers",
        "target_fields": [
            "username",
            "email",
            "first_name",
            "last_name",
            "phone",
            "subject_specialty",
            "employment_status",
        ],
        "required": ["email", "first_name", "last_name"],
    },
    "guardians": {
        "label": "Parents / Guardians",
        "target_fields": [
            "username",
            "email",
            "first_name",
            "last_name",
            "phone",
            "relationship",
            "student_codes",
        ],
        "required": ["email", "first_name", "last_name", "student_codes"],
    },
    "roster": {
        "label": "Roster (classrooms + subjects + assignments)",
        "target_fields": [
            "classroom_name",
            "subject_code",
            "subject_name",
            "term_name",
            "teacher_username",
            "academic_year",
        ],
        "required": ["classroom_name", "subject_code"],
    },
    "attendance": {
        "label": "Attendance",
        "target_fields": [
            "student_code",
            "date",
            "status",
            "session",
            "remarks",
            "recorded_by_username",
        ],
        "required": ["student_code", "date", "status"],
    },
    "fees": {
        "label": "Fees / invoices",
        "target_fields": [
            "invoice_number",
            "student_code",
            "amount",
            "currency",
            "due_date",
            "status",
            "fee_type",
            "academic_year",
            "term_name",
        ],
        "required": ["student_code", "amount"],
    },
    "payments": {
        "label": "Payments",
        "target_fields": [
            "payment_reference",
            "student_code",
            "amount",
            "currency",
            "paid_at",
            "method",
            "invoice_number",
            "notes",
        ],
        "required": ["student_code", "amount", "paid_at"],
    },
}


@permission_required("settings.manage")
@user_passes_test(_is_admin_user)
@require_http_methods(["GET", "POST"])
def migration_wizard(request):
    """
    One-click data migration: upload CSV → optional field mapping → preview → run.
    Backed by existing bulk-preview/bulk-commit (students) and evals apply_import (grades).

    v3.99.24: routes to the Unified Wizard Engine (JSON SOT
    ``apps/setup_studio/wizards/account_migration.json``) by default.
    Pass ``?legacy=1`` to render the legacy session-driven flow.
    """
    from apps.setup_studio.legacy_view_bridge import engine_redirect_response

    engine_resp = engine_redirect_response(request, "account_migration")
    if engine_resp is not None:
        return engine_resp

    session_key = "migration_wizard"
    wizard_data = request.session.get(session_key) or {}

    if request.method == "GET" and request.GET.get("clear") == "1":
        request.session.pop(session_key, None)
        return redirect("accounts:migration_wizard")

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "select_system":
            source_system = (request.POST.get("source_system") or "").strip()
            allowed = (
                "powerschool",
                "blackbaud",
                "veracross",
                "infinite_campus",
                "other",
            )
            if source_system not in allowed:
                source_system = "other"
            request.session[session_key] = {
                **wizard_data,
                "source_system": source_system,
            }
            return redirect("accounts:migration_wizard")

        if action == "upload":
            from apps.automation.models import MigrationProfile

            profile_slug = request.POST.get("profile_slug")
            if profile_slug:
                profile = MigrationProfile.objects.filter(
                    slug=profile_slug, is_active=True
                ).first()
                if profile:
                    # Pass 8: a profile.config may override profile.domain when the
                    # wizard's migration_type doesn't map 1:1 (e.g. roster, teachers,
                    # guardians don't have first-class Domain enum entries yet).
                    config = profile.config if isinstance(profile.config, dict) else {}
                    migration_type = (
                        (config.get("migration_type") or "").strip() or profile.domain
                    )
                else:
                    migration_type = request.POST.get("migration_type") or "students"
                    profile_slug = None
            else:
                migration_type = request.POST.get("migration_type")
                profile_slug = None
            if migration_type not in MIGRATION_TYPES:
                messages.error(request, "Invalid migration type.")
                return redirect("accounts:migration_wizard")
            file_obj = request.FILES.get("file")
            if not file_obj:
                messages.error(request, "Please upload a CSV or XLSX file.")
                return redirect("accounts:migration_wizard")
            try:
                headers, rows, source_fmt = _read_uploaded_table(file_obj)
            except ValueError as e:
                messages.error(request, str(e))
                return redirect("accounts:migration_wizard")
            from apps.accounts.migration_services import detect_source_system_from_headers

            detection = detect_source_system_from_headers(headers)
            chosen_source = wizard_data.get("source_system", "other")
            if (
                detection.get("suggested_system")
                and detection["suggested_system"] != "other"
                and detection.get("score", 0) >= 0.45
                and chosen_source == "other"
            ):
                messages.info(
                    request,
                    (
                        "Detected source system from column headers: "
                        f"{detection['suggested_system'].replace('_', ' ').title()} "
                        f"(confidence {detection.get('score', 0):.0%}). "
                        "You can change this in step 1 if it is wrong."
                    ),
                )

            # Pass 8: one-click vendor auto-detect — when the schema_fingerprint
            # match is ≥0.8 confidence, snap the wizard to that profile so the
            # user lands on the preview/mapping screen with everything filled in.
            auto_applied_profile = None
            if not profile_slug:
                try:
                    from apps.automation.schema_fingerprint import suggest_profiles_from_headers

                    suggestions = suggest_profiles_from_headers(
                        headers,
                        domain=migration_type if migration_type in (
                            "students", "grades", "attendance", "fees", "finance"
                        ) else None,
                        min_confidence=0.8,
                    )
                    if suggestions:
                        top_profile, top_confidence = suggestions[0]
                        profile_slug = top_profile.slug
                        auto_applied_profile = (top_profile.name, top_confidence)
                        messages.success(
                            request,
                            (
                                f"Auto-applied vendor profile '{top_profile.name}' "
                                f"({top_confidence:.0%} confidence). You can change it in step 1."
                            ),
                        )
                except (TypeError, ValueError, AttributeError, DatabaseError, ImportError):
                    pass

            request.session[session_key] = {
                "source_system": chosen_source,
                "migration_type": migration_type,
                "profile_slug": profile_slug,
                "headers": headers,
                "rows": rows,
                "row_count": len(rows),
                "source_format": source_fmt,
                "header_source_detection": detection,
                "auto_applied_profile": auto_applied_profile,
            }
            return redirect("accounts:migration_wizard")

        if action == "dry_run" and wizard_data:
            migration_type = wizard_data.get("migration_type")
            rows = wizard_data.get("rows", [])
            mapping_json = request.POST.get("mapping")
            if not rows:
                messages.error(request, "No data to validate. Upload again.")
                return redirect("accounts:migration_wizard")
            mapping = json.loads(mapping_json) if mapping_json else {}
            transformed = []
            for row in rows:
                t = {}
                for csv_col, target_field in mapping.items():
                    if target_field and target_field != "__skip__":
                        t[target_field] = row.get(csv_col, "")
                transformed.append(t)
            from apps.accounts.migration_services import run_dry_run

            school = getattr(request, "school", None)
            scorecard = run_dry_run(
                school,
                migration_type,
                transformed,
                user=request.user,
                create_audit=True,
                legacy_snapshot=transformed,
            )
            request.session["migration_wizard_scorecard"] = scorecard
            return redirect("accounts:migration_wizard")

        if action == "run" and wizard_data:
            migration_type = wizard_data.get("migration_type")
            rows = wizard_data.get("rows", [])
            mapping_json = request.POST.get("mapping")
            if not rows:
                messages.error(request, "No data to import. Upload again.")
                request.session.pop(session_key, None)
                return redirect("accounts:migration_wizard")
            mapping = json.loads(mapping_json) if mapping_json else {}
            transformed = []
            for row in rows:
                t = {}
                for csv_col, target_field in mapping.items():
                    if target_field and target_field != "__skip__":
                        t[target_field] = row.get(csv_col, "")
                transformed.append(t)
            from apps.accounts.migration_services import (
                run_migration_start,
                run_migration_finish,
            )

            school = getattr(request, "school", None)
            run = run_migration_start(
                school,
                migration_type,
                len(transformed),
                user=request.user,
                legacy_snapshot=transformed,
            )
            result = {"created": 0, "updated": 0, "error_count": 0, "errors": []}
            if migration_type == "students":
                try:
                    from django.test import Client
                    from django.urls import reverse as rev

                    client = Client()
                    client.force_login(request.user)
                    for k, v in request.session.items():
                        client.session[k] = v
                    client.session.save()
                    url = rev("api:entity-student-bulk-commit")
                    resp = client.post(
                        url,
                        data=json.dumps({"rows": transformed}),
                        content_type="application/json",
                    )
                    try:
                        data = json.loads(resp.content.decode("utf-8"))
                    except (
                        json.JSONDecodeError,
                        UnicodeDecodeError,
                        TypeError,
                        ValueError,
                    ):
                        data = {}
                    if resp.status_code in (200, 201):
                        result["created"] = len(data.get("created", []))
                        result["errors"] = data.get("errors", [])
                        result["error_count"] = len(result["errors"])
                        result["rollback_snapshot"] = {
                            "created_ids": data.get("created", [])
                        }
                        run_migration_finish(run, result)
                        p = result.get("parity", {})
                        messages.success(
                            request,
                            f"Students: created {result['created']}, errors {result['error_count']}. Parity: {p.get('total_processed', 0)} rows processed.",
                        )
                    else:
                        result["error_count"] = len(transformed)
                        result["error_message"] = (
                            data.get("error") or "Student import failed."
                        )
                        result["errors"] = [result["error_message"]]
                        run_migration_finish(run, result)
                        messages.error(request, result["error_message"])
                except (
                    ValueError,
                    TypeError,
                    KeyError,
                    ImportError,
                    AttributeError,
                    RuntimeError,
                ) as e:
                    result["error_count"] = len(transformed)
                    result["error_message"] = str(e)
                    result["errors"] = [str(e)]
                    run_migration_finish(run, result)
                    messages.error(
                        request,
                        f"Import failed: {e}. Check your CSV format and mapping.",
                    )
            elif migration_type == "grades":
                from apps.academics.services import get_active_year_and_term

                active_year, _ = get_active_year_and_term()
                if not active_year:
                    result["error_count"] = len(transformed)
                    result["errors"] = ["No active academic year set."]
                    run_migration_finish(run, result)
                    messages.error(
                        request, "No active academic year. Set one in Academics."
                    )
                else:
                    try:
                        from apps.evals.importers import apply_import, preview_import

                        preview = preview_import(transformed)
                        apply_result = apply_import(preview, active_year)
                        result["created"] = apply_result.get("created", 0)
                        result["updated"] = apply_result.get("updated", 0)
                        result["duration_seconds"] = apply_result.get(
                            "duration_seconds", 0
                        )
                        result["error_count"] = (
                            len(transformed) - result["created"] - result["updated"]
                        )
                        result["rollback_snapshot"] = {
                            "created_ids": apply_result.get("created_ids", []),
                            "updated_ids": apply_result.get("updated_ids", []),
                        }
                        run_migration_finish(run, result)
                        p = result.get("parity", {})
                        messages.success(
                            request,
                            f"Grades: created {result['created']}, updated {result['updated']}, errors {result['error_count']}. Parity: {p.get('total_processed', 0)} rows processed.",
                        )
                    except (
                        ValueError,
                        TypeError,
                        KeyError,
                        ImportError,
                        AttributeError,
                        RuntimeError,
                    ) as e:
                        result["error_count"] = len(transformed)
                        result["error_message"] = str(e)
                        result["errors"] = [str(e)]
                        run_migration_finish(run, result)
                        messages.error(request, f"Grade import failed. Details: {e}")
            elif migration_type in (
                "teachers",
                "guardians",
                "roster",
                "attendance",
                "fees",
                "payments",
            ):
                # Pass 8: dispatch to the new importer module.
                try:
                    from apps.accounts.migration_importers import run_importer

                    importer_result = run_importer(
                        migration_type, school, transformed, actor=request.user
                    )
                    result["created"] = importer_result.get("created", 0)
                    result["updated"] = importer_result.get("updated", 0)
                    result["error_count"] = importer_result.get("error_count", 0)
                    result["errors"] = importer_result.get("errors", [])
                    result["rollback_snapshot"] = importer_result.get(
                        "rollback_snapshot", {}
                    )
                    skipped = importer_result.get("skipped", 0)
                    run_migration_finish(run, result)
                    if result["error_count"] == 0 and skipped == 0:
                        messages.success(
                            request,
                            f"{MIGRATION_TYPES[migration_type]['label']}: "
                            f"created {result['created']}, updated {result['updated']}.",
                        )
                    elif result["created"] or result["updated"]:
                        messages.warning(
                            request,
                            f"{MIGRATION_TYPES[migration_type]['label']}: "
                            f"created {result['created']}, updated {result['updated']}, "
                            f"errors {result['error_count']}, skipped {skipped}.",
                        )
                    else:
                        messages.error(
                            request,
                            f"{MIGRATION_TYPES[migration_type]['label']} import did not "
                            f"persist any rows (errors {result['error_count']}, skipped {skipped}).",
                        )
                except (
                    ValueError,
                    TypeError,
                    KeyError,
                    ImportError,
                    AttributeError,
                    RuntimeError,
                    DatabaseError,
                ) as e:
                    result["error_count"] = len(transformed)
                    result["errors"] = [str(e)]
                    run_migration_finish(run, result)
                    messages.error(
                        request, f"{migration_type} import failed. Details: {e}"
                    )
            else:
                run_migration_finish(run, result)
            request.session.pop(session_key, None)
            return redirect("accounts:migration_wizard")

        if request.POST.get("action") == "clear":
            request.session.pop(session_key, None)
            return redirect("accounts:migration_wizard")

    from apps.automation.models import MigrationProfile

    source_system = wizard_data.get("source_system") or "other"
    # Pass 8: surface all generic + vendor profiles, not just students/grades.
    if source_system == "other":
        profiles = list(
            MigrationProfile.objects.filter(
                is_active=True,
                source_system__isnull=True,
            ).order_by("sort_order", "slug")
        )
    else:
        profiles = list(
            MigrationProfile.objects.filter(
                is_active=True, source_system=source_system
            ).order_by("sort_order", "slug")
        )
    if not profiles:
        profiles = list(
            MigrationProfile.objects.filter(
                is_active=True,
            ).order_by("sort_order", "slug")[:12]
        )
    profile_choices = [(p.slug, p.name) for p in profiles]
    has_data = bool(wizard_data.get("rows"))
    config = MIGRATION_TYPES.get(wizard_data.get("migration_type", "")) or {}
    profile_slug = wizard_data.get("profile_slug")
    schema_hints = {}
    headers = wizard_data.get("headers", [])
    if has_data:
        if profile_slug:
            prof = MigrationProfile.objects.filter(slug=profile_slug).first()
            if prof and isinstance(prof.config, dict):
                schema_hints = prof.config.get("schema_hints") or {}
        if not schema_hints and headers and config.get("target_fields"):
            from apps.accounts.migration_services import infer_schema_mapping

            schema_hints = infer_schema_mapping(
                headers, config.get("target_fields", [])
            )
    rows = wizard_data.get("rows", [])[:15]
    preview_matrix = []
    for row in rows:
        preview_matrix.append([str(row.get(h, "")) for h in headers])
    scorecard = request.session.pop("migration_wizard_scorecard", None)
    if scorecard and scorecard.get("validation_issues"):
        vi = scorecard["validation_issues"]
        labels = {
            "duplicates": "Duplicates",
            "missing_required": "Missing required fields",
            "invalid_refs": "Invalid references",
        }
        scorecard["validation_issues_list"] = [
            {"category": cat, "label": labels.get(cat, cat), "issues": vi.get(cat, [])}
            for cat in ("duplicates", "missing_required", "invalid_refs")
            if vi.get(cat)
        ]
    schema_hints_json = json.dumps(schema_hints)
    profile_suggestions: list[dict[str, object]] = []
    if has_data and headers:
        from apps.automation.schema_fingerprint import suggest_profiles_from_headers

        mt = wizard_data.get("migration_type") or ""
        dom = mt if mt in ("students", "grades") else None
        try:
            scored = suggest_profiles_from_headers(
                headers, domain=dom, min_confidence=0.15
            )
            profile_suggestions = [
                {"slug": p.slug, "name": p.name, "confidence": round(c, 2)}
                for p, c in scored[:8]
            ]
        except (TypeError, ValueError, AttributeError, DatabaseError):
            profile_suggestions = []
    return render(
        request,
        "accounts/migration_wizard.html",
        {
            "migration_types": MIGRATION_TYPES,
            "migration_type_choices": profile_choices,
            "wizard_data": wizard_data,
            "has_data": has_data,
            "source_system": source_system,
            "source_system_choices": [
                ("powerschool", "PowerSchool"),
                ("blackbaud", "Blackbaud"),
                ("veracross", "Veracross"),
                ("infinite_campus", "Infinite Campus"),
                ("other", "Other"),
            ],
            "target_fields": config.get("target_fields", []),
            "required_fields": config.get("required", []),
            "preview_matrix": preview_matrix,
            "scorecard": scorecard,
            "schema_hints": schema_hints,
            "schema_hints_json": schema_hints_json,
            "profile_suggestions": profile_suggestions,
            "page_title": _("Data migration wizard"),
            "page_subtitle": _(
                "Choose your current system, upload a CSV, map columns to target fields, preview, then run."
            ),
            "action_url": reverse("studio_os:import_hub"),
            "action_text": _("Back to Import Hub"),
            "phase9_links": build_phase9_ecosystem_links(),
        },
    )


@permission_required("settings.manage")
@user_passes_test(_is_admin_user)
@require_http_methods(["GET"])
def migration_run_list(request):
    """Section 11.1: List migration runs for this school with links to read-only legacy view and rollback UI."""
    school = getattr(request, "school", None)
    if not school:
        messages.warning(request, "No school context.")
        return redirect("accounts:backend_dashboard")
    from apps.automation.models import MigrationRun

    runs = (
        MigrationRun.objects.filter(school=school)
        .select_related("triggered_by")
        .order_by("-started_at")[:50]
    )
    return render(
        request,
        "accounts/migration_run_list.html",
        {
            "runs": runs,
            "school": school,
            "page_title": _("Migration runs"),
            "page_subtitle": _(
                "Audit of migration runs. Open a run to view read-only legacy data."
            ),
            "action_url": reverse("studio_os:workflow_center"),
            "action_text": _("Back to Workflow Center"),
        },
    )


@permission_required("settings.manage")
@user_passes_test(_is_admin_user)
@require_http_methods(["POST"])
def migration_rollback(request, run_id):
    """Section 11.1: Trigger rollback for a migration run (UI for MigrationRun.trigger_rollback)."""
    if request.method != "POST":
        return redirect("accounts:migration_run_list")
    school = getattr(request, "school", None)
    if not school:
        messages.warning(request, "No school context.")
        return redirect("accounts:backend_dashboard")
    from apps.automation.models import MigrationRun

    run = get_object_or_404(MigrationRun, pk=run_id, school=school)
    if not run.can_rollback:
        messages.error(
            request,
            "This run cannot be rolled back (dry run, already rolled back, or no snapshot).",
        )
        return redirect("accounts:migration_run_list")
    rollback_run, result = run.trigger_rollback(user=request.user)
    if result.get("success"):
        messages.success(
            request,
            f"Rollback created (run #{rollback_run.pk}). {result.get('message', '')} Reverted: {result.get('reverted_count', 0)}.",
        )
    else:
        messages.error(request, result.get("message", "Rollback failed."))
    return redirect("accounts:migration_run_list")


@permission_required("settings.manage")
@user_passes_test(_is_admin_user)
@require_http_methods(["GET"])
def migration_legacy_view(request, run_id):
    """Section 11.1: Read-only legacy view — show uploaded rows snapshot for a migration run."""
    school = getattr(request, "school", None)
    if not school:
        messages.warning(request, "No school context.")
        return redirect("accounts:backend_dashboard")
    from apps.automation.models import MigrationRun

    run = get_object_or_404(MigrationRun, pk=run_id, school=school)
    raw_rows = (run.legacy_snapshot or {}).get("rows") or []
    headers = list(raw_rows[0].keys()) if raw_rows else []
    rows_matrix = [[r.get(h, "") for h in headers] for r in raw_rows]
    return render(
        request,
        "accounts/migration_legacy_view.html",
        {
            "run": run,
            "rows_matrix": rows_matrix,
            "headers": headers,
            "page_title": _("Legacy data (read-only)"),
            "page_subtitle": _(
                "Run #%(run_id)s — %(migration_type)s. This view is read-only."
            )
            % {"run_id": run.id, "migration_type": run.migration_type},
            "action_url": reverse("accounts:migration_run_list"),
            "action_text": _("Back to runs"),
        },
    )


@permission_required("settings.manage")
@user_passes_test(_is_admin_user)
@require_http_methods(["GET", "POST"])
def legacy_data_cleaner_view(request):
    """Section 11.1: Legacy data cleaner — detect and optionally clean legacy/invalid data."""
    school = getattr(request, "school", None)
    if not school:
        messages.warning(request, "No school context.")
        return redirect("accounts:backend_dashboard")
    from apps.accounts.legacy_data_cleaner import (
        detect_legacy_issues,
        clean_legacy_data,
    )

    if request.method == "POST" and request.POST.get("action") == "clean":
        dry_run = request.POST.get("dry_run") == "1"
        result = clean_legacy_data(school, dry_run=dry_run)
        messages.success(
            request,
            f"Cleaner run (dry_run={dry_run}). Actions: {len(result.get('actions', []))}.",
        )
        return render(
            request,
            "accounts/legacy_data_cleaner.html",
            {
                "school": school,
                "issues": detect_legacy_issues(school),
                "clean_result": result,
                "page_title": _("Legacy data cleaner"),
                "page_subtitle": _(
                    "Detect duplicate admission numbers, missing required fields, and optionally run safe cleanups."
                ),
                "action_url": reverse("studio_os:workflow_center"),
                "action_text": _("Back to Workflow Center"),
            },
        )
    issues = detect_legacy_issues(school)
    return render(
        request,
        "accounts/legacy_data_cleaner.html",
        {
            "school": school,
            "issues": issues,
            "clean_result": None,
            "page_title": _("Legacy data cleaner"),
            "page_subtitle": _(
                "Detect duplicate admission numbers, missing required fields, and optionally run safe cleanups."
            ),
            "action_url": reverse("studio_os:workflow_center"),
            "action_text": _("Back to Workflow Center"),
        },
    )
